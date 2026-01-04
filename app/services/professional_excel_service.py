"""
Professional Excel Report Generation Service
Creates branded Excel reports with logo, colors, and professional formatting
"""
import logging
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)

# Brand Colors (matching PDF)
PRIMARY_COLOR = "125435"  # Dark green
ACCENT_COLOR = "E8F5E9"   # Light green
SECONDARY_COLOR = "2E7D32"  # Medium green
HEADER_COLOR = "125435"
ALT_ROW_COLOR = "F9FAFB"
BORDER_COLOR = "CCCCCC"

# Logo path
LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', 'logo.png')


def create_professional_excel_workbook(title="Inspection Report", sheet_name="Report"):
    """Create a new workbook with professional styling
    
    Args:
        title: Report title
        sheet_name: Name of the active sheet
        
    Returns:
        tuple: (workbook, worksheet)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Set default column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    
    return wb, ws


def add_logo_and_title(ws, title, subtitle=None, start_row=1):
    """Add logo and title to worksheet
    
    Args:
        ws: Worksheet object
        title: Report title
        subtitle: Optional subtitle
        start_row: Starting row number
        
    Returns:
        int: Next available row number
    """
    from openpyxl.styles import Alignment
    
    current_row = start_row
    
    # Set row heights for logo area
    ws.row_dimensions[current_row].height = 45
    ws.row_dimensions[current_row + 1].height = 20
    
    # Add logo if available
    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            # Resize logo properly
            img.width = 60
            img.height = 60
            ws.add_image(img, f'A{current_row}')
        except Exception as e:
            logger.warning(f"Could not add logo to Excel: {e}")
    
    # Title (next to logo)
    title_cell = ws[f'B{current_row}']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=16, color=PRIMARY_COLOR, name='Calibri')
    title_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    ws.merge_cells(f'B{current_row}:E{current_row}')
    
    # Add company name
    company_cell = ws[f'B{current_row + 1}']
    company_cell.value = "INJAAZ PLATFORM"
    company_cell.font = Font(bold=True, size=10, color=SECONDARY_COLOR, name='Calibri')
    company_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    current_row += 3
    
    # Subtitle if provided
    if subtitle:
        ws.row_dimensions[current_row].height = 25
        subtitle_cell = ws[f'A{current_row}']
        subtitle_cell.value = subtitle
        subtitle_cell.font = Font(bold=True, size=11, color=SECONDARY_COLOR, name='Calibri')
        subtitle_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
    
    current_row += 1  # Add spacing
    return current_row


def add_info_section(ws, info_data, start_row, title="Information"):
    """Add an information section with key-value pairs
    
    Args:
        ws: Worksheet object
        info_data: List of (label, value) tuples
        start_row: Starting row number
        title: Section title
        
    Returns:
        int: Next available row number
    """
    current_row = start_row
    
    # Section title
    title_cell = ws[f'A{current_row}']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, color=PRIMARY_COLOR, name='Calibri')
    title_cell.fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    ws.merge_cells(f'A{current_row}:D{current_row}')
    
    # Add border
    for col in ['A', 'B', 'C', 'D']:
        cell = ws[f'{col}{current_row}']
        cell.border = Border(
            top=Side(style='medium', color=PRIMARY_COLOR),
            bottom=Side(style='medium', color=PRIMARY_COLOR),
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR)
        )
    
    current_row += 1
    
    # Add info rows
    for label, value in info_data:
        # Set row height
        ws.row_dimensions[current_row].height = 25
        
        label_cell = ws[f'A{current_row}']
        label_cell.value = label
        label_cell.font = Font(bold=True, size=10, name='Calibri')
        label_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        label_cell.fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
        label_cell.border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        
        value_cell = ws[f'B{current_row}']
        value_cell.value = str(value)
        value_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        value_cell.border = Border(
            left=Side(style='thin', color=BORDER_COLOR),
            right=Side(style='thin', color=BORDER_COLOR),
            top=Side(style='thin', color=BORDER_COLOR),
            bottom=Side(style='thin', color=BORDER_COLOR)
        )
        ws.merge_cells(f'B{current_row}:E{current_row}')
        
        # Auto-adjust row height for wrapped content
        if value and len(str(value)) > 50:
            ws.row_dimensions[current_row].height = max(25, min(len(str(value)) // 30 * 15 + 25, 100))
        
        current_row += 1
    
    current_row += 1  # Add spacing
    return current_row


def add_data_table(ws, headers, data_rows, start_row, title=None, col_widths=None):
    """Add a professional data table
    
    Args:
        ws: Worksheet object
        headers: List of column headers
        data_rows: List of data rows (each row is a list)
        start_row: Starting row number
        title: Optional table title
        col_widths: Optional dict of column widths {col_letter: width}
        
    Returns:
        int: Next available row number
    """
    current_row = start_row
    
    # Table title if provided
    if title:
        title_cell = ws[f'A{current_row}']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=13, color=SECONDARY_COLOR, name='Calibri')
        current_row += 1
    
    # Set column widths if provided
    if col_widths:
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # Header row
    header_row = current_row
    ws.row_dimensions[header_row].height = 35
    
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11, name='Calibri')
        cell.fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(
            left=Side(style='thin', color="FFFFFF"),
            right=Side(style='thin', color="FFFFFF"),
            top=Side(style='medium', color=PRIMARY_COLOR),
            bottom=Side(style='medium', color=PRIMARY_COLOR)
        )
    
    current_row += 1
    
    # Data rows with alternating colors
    for row_idx, row_data in enumerate(data_rows):
        is_alternate = row_idx % 2 == 1
        fill_color = ALT_ROW_COLOR if is_alternate else "FFFFFF"
        
        # Set row height for data rows (larger for wrapped text)
        ws.row_dimensions[current_row].height = 40
        
        for col_idx, value in enumerate(row_data, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws[f'{col_letter}{current_row}']
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin', color=BORDER_COLOR),
                right=Side(style='thin', color=BORDER_COLOR),
                top=Side(style='thin', color=BORDER_COLOR),
                bottom=Side(style='thin', color=BORDER_COLOR)
            )
        
        current_row += 1
    
    current_row += 1  # Add spacing
    return current_row


def add_section_header(ws, title, start_row, span_columns=4):
    """Add a section header row
    
    Args:
        ws: Worksheet object
        title: Section title
        start_row: Starting row number
        span_columns: Number of columns to span
        
    Returns:
        int: Next available row number
    """
    # Set row height for section header
    ws.row_dimensions[start_row].height = 30
    
    cell = ws[f'A{start_row}']
    cell.value = title
    cell.font = Font(bold=True, size=13, color=PRIMARY_COLOR, name='Calibri')
    cell.fill = PatternFill(start_color=ACCENT_COLOR, end_color=ACCENT_COLOR, fill_type="solid")
    cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Merge cells
    end_col = get_column_letter(span_columns)
    ws.merge_cells(f'A{start_row}:{end_col}{start_row}')
    
    # Add border
    for col_idx in range(1, span_columns + 1):
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{start_row}']
        cell.border = Border(
            left=Side(style='thin', color=PRIMARY_COLOR),
            right=Side(style='thin', color=PRIMARY_COLOR),
            top=Side(style='medium', color=PRIMARY_COLOR),
            bottom=Side(style='medium', color=PRIMARY_COLOR)
        )
    
    return start_row + 1


def add_signature_section(ws, signatures, start_row):
    """Add signature section to worksheet
    
    Args:
        ws: Worksheet object
        signatures: Dict of {role: signature_data}
        start_row: Starting row number
        
    Returns:
        int: Next available row number
    """
    current_row = add_section_header(ws, "Signatures & Approval", start_row)
    
    # Headers row
    ws.row_dimensions[current_row].height = 25
    
    ws[f'A{current_row}'].value = "Role"
    ws[f'A{current_row}'].font = Font(bold=True, size=10, name='Calibri')
    ws[f'A{current_row}'].fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
    ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'B{current_row}'].value = "Status"
    ws[f'B{current_row}'].font = Font(bold=True, size=10, name='Calibri')
    ws[f'B{current_row}'].fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
    ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws[f'C{current_row}'].value = "Signed On"
    ws[f'C{current_row}'].font = Font(bold=True, size=10, name='Calibri')
    ws[f'C{current_row}'].fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")
    ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
    
    current_row += 1
    
    # Signature rows
    signed_date = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    for role, sig_data in signatures.items():
        # Set row height for signature rows
        ws.row_dimensions[current_row].height = 25
        
        ws[f'A{current_row}'].value = role
        
        if sig_data:
            ws[f'B{current_row}'].value = "✓ Signed"
            ws[f'B{current_row}'].font = Font(color="008000", name='Calibri')
            ws[f'C{current_row}'].value = signed_date
        else:
            ws[f'B{current_row}'].value = "Not signed"
            ws[f'B{current_row}'].font = Font(color="999999", name='Calibri')
            ws[f'C{current_row}'].value = "-"
        
        # Add borders
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{current_row}']
            cell.border = Border(
                left=Side(style='thin', color=BORDER_COLOR),
                right=Side(style='thin', color=BORDER_COLOR),
                top=Side(style='thin', color=BORDER_COLOR),
                bottom=Side(style='thin', color=BORDER_COLOR)
            )
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        current_row += 1
    
    current_row += 1
    return current_row


def finalize_workbook(ws):
    """Apply final formatting to worksheet and auto-size columns
    
    Args:
        ws: Worksheet object
    """
    # Auto-size columns based on content
    from openpyxl.utils import get_column_letter
    
    for col in ws.iter_cols(min_col=1, max_col=ws.max_column):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            try:
                if cell.value:
                    # Calculate cell content length
                    cell_value = str(cell.value)
                    # Count wrapped lines (estimate based on width)
                    lines = cell_value.count('\n') + 1
                    # Estimate width needed (characters per line)
                    if hasattr(cell, 'alignment') and cell.alignment and cell.alignment.wrap_text:
                        # For wrapped text, use a reasonable width (35 chars is good for wrapped content)
                        length = min(len(cell_value) // lines + 5, 50)
                    else:
                        length = len(cell_value)
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # Set column width (add 2 for padding, minimum 10, maximum 60)
        adjusted_width = min(max(max_length + 2, 10), 60)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Freeze top rows (typically after logo and title)
    ws.freeze_panes = 'A6'
    
    # Set print settings
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Set margins (in inches)
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    
    # Add header and footer
    ws.oddHeader.center.text = "INJAAZ PLATFORM"
    ws.oddHeader.center.size = 10
    ws.oddFooter.left.text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.oddFooter.right.text = "Page &P of &N"
    ws.oddFooter.left.size = 8
    ws.oddFooter.right.size = 8
