"""
Admin Routes - User management and access control
"""
from flask import Blueprint, request, jsonify, render_template, current_app, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy.orm import joinedload
from sqlalchemy import or_, func
from sqlalchemy.exc import IntegrityError
from app.models import (
    db, User, AuditLog, AdminEditOtp, Device, BDProject, BDFollowUp, BDContact, BDActivity,
    DocHubAccess, MmrChargeableConfig, NotificationConfig, EmailLog, AdminPersonalProject, AdminPersonalProgressStep,
    Technician, KnowledgeBaseEntry, DatabaseBackup,
    Quotation, QuotationItem, TicketProject,
)
from app.middleware import (
    admin_required,
    bd_access_required,
    user_sees_all_bd_deals,
    user_has_bd_access,
    user_can_prepare_quotations,
)
from common.error_responses import error_response, success_response
from common.form_data_utils import shallow_copy_form_data
from common.document_display import (
    STOREKEEPING_MODULE_TYPES,
    build_document_labels,
    get_module_display_name,
)
from common.datetime_utils import utc_now_naive, parse_employment_start_date
from datetime import datetime, timedelta, date
from io import BytesIO, StringIO
import hashlib
import hmac
import json
import os
import secrets
import threading
import uuid

admin_bp = Blueprint('admin_bp', __name__, url_prefix='/api/admin')

# Internal logins that must not appear on Users & Teams (created only for ticket email).
DIRECTORY_HIDDEN_USERNAMES = frozenset({'email_intake'})


def _ensure_sales_ops_columns():
    """Add BD sales/ops columns on existing DBs (SQLite/Postgres-safe)."""
    try:
        from sqlalchemy import inspect, text
        insp = inspect(db.engine)
        tables = set(insp.get_table_names())
        alters = []
        hiring_was_missing = False
        if 'users' in tables:
            cols = {c['name'] for c in insp.get_columns('users')}
            if 'access_sales_manager' not in cols:
                alters.append('ALTER TABLE users ADD COLUMN access_sales_manager BOOLEAN DEFAULT 0')
            if 'access_quotations' not in cols:
                alters.append('ALTER TABLE users ADD COLUMN access_quotations BOOLEAN DEFAULT 0')
            hiring_was_missing = 'access_hiring' not in cols
            if hiring_was_missing:
                alters.append('ALTER TABLE users ADD COLUMN access_hiring BOOLEAN DEFAULT 0')
        if 'bd_projects' in tables:
            cols = {c['name'] for c in insp.get_columns('bd_projects')}
            if 'owner_user_id' not in cols:
                alters.append('ALTER TABLE bd_projects ADD COLUMN owner_user_id INTEGER')
        with db.engine.begin() as conn:
            for sql in alters:
                try:
                    conn.execute(text(sql))
                except Exception:
                    pass
            if hiring_was_missing:
                try:
                    conn.execute(text('UPDATE users SET access_hiring = access_hr'))
                except Exception:
                    pass
        db.create_all()
        try:
            db.session.execute(text(
                'UPDATE bd_projects SET owner_user_id = created_by '
                'WHERE owner_user_id IS NULL AND created_by IS NOT NULL'
            ))
            db.session.commit()
        except Exception:
            db.session.rollback()
    except Exception as exc:
        current_app.logger.warning('sales ops column ensure: %s', exc)


@admin_bp.record_once
def _admin_bp_on_register(state):
    with state.app.app_context():
        _ensure_sales_ops_columns()


def _bd_current_user():
    uid = get_jwt_identity()
    try:
        uid = int(uid)
    except (TypeError, ValueError):
        pass
    return User.query.get(uid)


def _bd_scoped_projects_query(user):
    q = BDProject.query
    if not user_sees_all_bd_deals(user):
        q = q.filter(db.or_(
            BDProject.owner_user_id == user.id,
            BDProject.created_by == user.id,
        ))
    return q


def _resolve_owner_user_id(data, default_user_id=None):
    raw = data.get('owner_user_id', data.get('ownerUserId'))
    if raw is None or raw == '':
        return default_user_id
    try:
        return int(raw)
    except (TypeError, ValueError):
        return default_user_id


def _sync_owner_display(project):
    if project.owner_user_id:
        u = db.session.get(User, project.owner_user_id)
        if u:
            project.owner = u.full_name or u.username
    return project


def _can_access_bd_project(user, project):
    if user_sees_all_bd_deals(user):
        return True
    return project.owner_user_id in (None, user.id) or project.created_by == user.id


def _admin_user_dict(user):
    """User payload for admin management APIs (includes admin-visible password)."""
    return user.to_dict(include_sensitive=True)


OTP_EDIT_REQUIRED_MSG = (
    'This administrator account is locked to prevent unconfirmed profile, password, '
    'or access changes. Verify the one-time code sent to the account email before editing.'
)
OTP_CODE_TTL = timedelta(minutes=10)
OTP_GRANT_TTL = timedelta(minutes=10)
OTP_RESEND_COOLDOWN = timedelta(minutes=1)
OTP_REQUEST_WINDOW = timedelta(minutes=10)
OTP_MAX_REQUESTS = 3
OTP_MAX_ATTEMPTS = 5


def _int_actor_id():
    try:
        return int(get_jwt_identity())
    except (TypeError, ValueError):
        return None


def _generate_otp_code():
    return f'{secrets.randbelow(1_000_000):06d}'


def _otp_hmac_key():
    return str(current_app.config.get('SECRET_KEY') or 'kynvera-otp').encode('utf-8')


def _hash_otp_code(code):
    return hmac.new(_otp_hmac_key(), str(code).encode('utf-8'), hashlib.sha256).hexdigest()


def _otp_code_matches(code, stored_hash):
    if not stored_hash or not code:
        return False
    return hmac.compare_digest(_hash_otp_code(code), stored_hash)


def _mask_email(email):
    raw = (email or '').strip()
    if '@' not in raw:
        return '***'
    local, _, domain = raw.partition('@')
    keep = local[:1] if local else ''
    return f'{keep}***@{domain}'


def _admin_edit_grant_ok(actor_id, grant_target_id):
    if actor_id is None or grant_target_id is None:
        return False
    row = AdminEditOtp.query.filter_by(
        actor_user_id=actor_id,
        target_user_id=grant_target_id,
    ).first()
    if not row or not row.grant_expires_at:
        return False
    return row.grant_expires_at > utc_now_naive()


def _otp_required_response():
    return jsonify({
        'success': False,
        'error': OTP_EDIT_REQUIRED_MSG,
        'otp_required': True,
    }), 403


def _reject_unless_admin_edit_granted(target, *, promoting_to_admin=False):
    actor_id = _int_actor_id()
    if getattr(target, 'role', None) == 'admin':
        if not _admin_edit_grant_ok(actor_id, target.id):
            return _otp_required_response()
        return None
    if promoting_to_admin:
        if not _admin_edit_grant_ok(actor_id, actor_id):
            return _otp_required_response()
    return None


def _admin_for_reassigned_records(exclude_user_id, preferred_id=None):
    """Another administrator who can keep tickets/notes when an account is deleted."""
    if preferred_id and preferred_id != exclude_user_id:
        pref = db.session.get(User, preferred_id)
        if pref and pref.role == 'admin':
            return pref
    return (
        User.query.filter(User.role == 'admin', User.id != exclude_user_id)
        .order_by(User.id.asc())
        .first()
    )


def _detach_user_for_delete(user_id, fallback_admin_id):
    """Move or clear NOT NULL references so deleting the user cannot fail on FKs."""
    from app.models import (
        AssistantPendingAction, Ticket, TicketImage, TicketNote, TicketSupervisorTeam,
    )

    ticket_count = Ticket.query.filter_by(reporter_id=user_id).count()
    note_count = TicketNote.query.filter_by(user_id=user_id).count()
    image_count = TicketImage.query.filter_by(uploaded_by=user_id).count()
    needs_fallback = ticket_count + note_count + image_count
    if needs_fallback and not fallback_admin_id:
        raise ValueError(
            f'This account reported {ticket_count} ticket(s). Keep another administrator '
            'so those tickets can stay in the system, then try again.'
        )
    if ticket_count:
        Ticket.query.filter_by(reporter_id=user_id).update(
            {'reporter_id': fallback_admin_id}, synchronize_session=False
        )
    Ticket.query.filter_by(assigned_to_id=user_id).update(
        {'assigned_to_id': None}, synchronize_session=False
    )
    Ticket.query.filter_by(supervisor_id=user_id).update(
        {'supervisor_id': None}, synchronize_session=False
    )
    Ticket.query.filter_by(technician_id=user_id).update(
        {'technician_id': None}, synchronize_session=False
    )
    if note_count:
        TicketNote.query.filter_by(user_id=user_id).update(
            {'user_id': fallback_admin_id}, synchronize_session=False
        )
    if image_count:
        TicketImage.query.filter_by(uploaded_by=user_id).update(
            {'uploaded_by': fallback_admin_id}, synchronize_session=False
        )
    TicketSupervisorTeam.query.filter(
        or_(
            TicketSupervisorTeam.supervisor_id == user_id,
            TicketSupervisorTeam.technician_id == user_id,
        )
    ).delete(synchronize_session=False)
    for proj in AdminPersonalProject.query.filter_by(user_id=user_id).all():
        db.session.delete(proj)
    AssistantPendingAction.query.filter_by(user_id=user_id).delete(synchronize_session=False)
    return ticket_count


def _otp_row_for(actor_id, target_id):
    row = AdminEditOtp.query.filter_by(
        actor_user_id=actor_id,
        target_user_id=target_id,
    ).first()
    if row:
        return row
    row = AdminEditOtp(actor_user_id=actor_id, target_user_id=target_id, attempt_count=0, request_count=0)
    db.session.add(row)
    db.session.flush()
    return row


def _otp_iso(dt):
    from common.datetime_utils import naive_utc_isoformat_z
    return naive_utc_isoformat_z(dt) if dt else None


def _otp_has_pending_code(row, now):
    return bool(row and row.code_hash and row.expires_at and row.expires_at > now)


def _otp_sent_at(row):
    if not row or not row.expires_at:
        return None
    return row.expires_at - OTP_CODE_TTL


def _otp_resend_at(row):
    sent = _otp_sent_at(row)
    if not sent:
        return None
    return sent + OTP_RESEND_COOLDOWN


def _otp_timing_fields(row, now):
    pending = _otp_has_pending_code(row, now)
    resend_at = _otp_resend_at(row) if pending else now
    after = 0
    if pending and resend_at and resend_at > now:
        after = int((resend_at - now).total_seconds())
    return {
        'has_pending_code': pending,
        'code_expires_at': _otp_iso(row.expires_at) if pending else None,
        'resend_available_at': _otp_iso(resend_at) if pending else _otp_iso(now),
        'resend_after_seconds': max(0, after),
    }


def _issue_edit_grant(row, now):
    row.code_hash = None
    row.attempt_count = 0
    row.grant_expires_at = now + OTP_GRANT_TTL
    return row.grant_expires_at


def _ensure_technicians_supervisor_column():
    """SQLite / existing DBs: add technicians.supervisor_user_id if missing."""
    try:
        from sqlalchemy import inspect, text

        inspector = inspect(db.engine)
        if 'technicians' not in inspector.get_table_names():
            db.create_all()
            return
        colnames = {c['name'] for c in inspector.get_columns('technicians')}
        if 'supervisor_user_id' in colnames:
            return
        with db.engine.begin() as conn:
            conn.execute(text('ALTER TABLE technicians ADD COLUMN supervisor_user_id INTEGER'))
    except Exception as exc:
        current_app.logger.warning('technicians.supervisor_user_id migration: %s', exc)


def _valid_roster_supervisor_user(user_id):
    """Supervisor-ish accounts that may own technician roster rows."""
    if user_id is None:
        return True
    u = db.session.get(User, user_id)
    if not u or not getattr(u, 'is_active', True):
        return False
    des = (getattr(u, 'designation', None) or '').strip().lower()
    return des in frozenset({'supervisor', 'operations_manager', 'general_manager'})


def _coerce_optional_supervisor_user_id(data):
    """Return '__unset__' (no change), None (clear), int id; raise ValueError if invalid."""
    if 'supervisor_user_id' not in data:
        return '__unset__'
    raw = data.get('supervisor_user_id')
    if raw in (None, '', 'null'):
        return None
    try:
        uid = int(raw)
    except (TypeError, ValueError):
        raise ValueError('supervisor_user_id must be numeric or blank') from None
    if not _valid_roster_supervisor_user(uid):
        raise ValueError('Supervisor account not found or not eligible')
    return uid


def _merge_dochub_into_user_dict(user, data_dict, access_map=None):
    """Add can_access_dochub for admin API payloads (DocHub uses dochub_access, not User columns)."""
    if user.role == 'admin':
        data_dict['can_access_dochub'] = True
    elif access_map is not None:
        data_dict['can_access_dochub'] = access_map.get(user.id, True)
    else:
        row = DocHubAccess.query.filter_by(user_id=user.id).first()
        data_dict['can_access_dochub'] = row.can_access if row else True


def _truncate_job_designation(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    return s[:160] if s else None


def _parse_leave_days_field(raw):
    """Parses HR leave allotment; empty clears to None. Validates 0..366."""
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        return None
    try:
        v = int(raw)
    except (TypeError, ValueError):
        raise ValueError('Annual and other leave days must be whole numbers.') from None
    if v < 0 or v > 366:
        raise ValueError('Leave days must be between 0 and 366.')
    return v


def _resolve_reporting_manager_id(raw_mid, exclude_user_id):
    if raw_mid is None:
        return None
    if isinstance(raw_mid, str) and not str(raw_mid).strip():
        return None
    try:
        mid = int(raw_mid)
    except (TypeError, ValueError):
        raise ValueError('Invalid reporting manager.') from None
    if mid <= 0:
        return None
    if exclude_user_id is not None and mid == exclude_user_id:
        raise ValueError('A user cannot be their own reporting manager.')
    if User.query.filter_by(id=mid).first() is None:
        raise ValueError('Reporting manager not found.')
    return mid


def _resolve_operations_manager_id(raw_mid, exclude_user_id):
    """Admin-assigned operations manager; must be an active user with operations_manager designation."""
    if raw_mid is None:
        return None
    if isinstance(raw_mid, str) and not str(raw_mid).strip():
        return None
    try:
        mid = int(raw_mid)
    except (TypeError, ValueError):
        raise ValueError('Invalid operations manager.') from None
    if mid <= 0:
        return None
    if exclude_user_id is not None and mid == exclude_user_id:
        raise ValueError('A user cannot be their own operations manager.')
    target = User.query.filter_by(id=mid).first()
    if target is None:
        raise ValueError('Operations manager not found.')
    if (target.designation or '').strip().lower() != 'operations_manager':
        raise ValueError('Selected user must have the Operations Manager designation.')
    if not target.is_active:
        raise ValueError('Selected operations manager is inactive.')
    return mid


@admin_bp.route('/mmr/chargeable-config', methods=['GET', 'PUT'])
@jwt_required()
@admin_required
def mmr_chargeable_config():
    """Load or save MMR chargeable rules (BaseUnit defaults + substring overrides)."""
    from module_mmr.mmr_service import (
        DEFAULT_MMR_CHARGEABLE_CONFIG,
        merge_builtin_rules_payload,
        _merge_mmr_chargeable_config,
        invalidate_mmr_chargeable_config_cache,
    )
    if request.method == 'GET':
        try:
            row = MmrChargeableConfig.query.first()
            stored = row.config_json if row else None
            return success_response({'config': _merge_mmr_chargeable_config(stored)})
        except Exception as e:
            current_app.logger.error(f"MMR chargeable config GET: {e}", exc_info=True)
            return error_response('Failed to load MMR chargeable settings', status_code=500, error_code='DATABASE_ERROR')

    try:
        admin_id = get_jwt_identity()
        data = request.get_json()
        if data is None:
            return error_response('JSON body required', status_code=400, error_code='VALIDATION_ERROR')

        flag = data.get('non_apartment_baseunit_non_chargeable')
        overrides = data.get('baseunit_overrides')
        if flag is None or not isinstance(flag, bool):
            return error_response(
                'non_apartment_baseunit_non_chargeable (boolean) is required',
                status_code=400,
                error_code='VALIDATION_ERROR',
            )
        if overrides is None:
            overrides = []
        if not isinstance(overrides, list):
            return error_response('baseunit_overrides must be a list', status_code=400, error_code='VALIDATION_ERROR')

        cleaned = []
        for item in overrides:
            if not isinstance(item, dict):
                continue
            pat = (item.get('pattern') or '').strip()
            if not pat:
                continue
            cleaned.append({'pattern': pat, 'chargeable': bool(item.get('chargeable'))})

        row = MmrChargeableConfig.query.first()
        prev = _merge_mmr_chargeable_config(row.config_json if row else None)
        br_in = data.get('builtin_rules')
        if br_in is not None and not isinstance(br_in, dict):
            return error_response(
                'builtin_rules must be an object',
                status_code=400,
                error_code='VALIDATION_ERROR',
            )
        br_merged = merge_builtin_rules_payload(br_in) if isinstance(br_in, dict) else prev['builtin_rules']

        raw_update = {
            **prev,
            'non_apartment_baseunit_non_chargeable': flag,
            'baseunit_overrides': cleaned,
            'builtin_rules': br_merged,
        }
        if 'location_register_state' in data:
            raw_update['location_register_state'] = data.get('location_register_state')
        if 'location_register_enabled' in data:
            raw_update['location_register_enabled'] = bool(data.get('location_register_enabled'))
        merged = _merge_mmr_chargeable_config(raw_update)

        if row:
            row.config_json = merged
        else:
            db.session.add(MmrChargeableConfig(config_json=merged))

        db.session.commit()
        invalidate_mmr_chargeable_config_cache()

        log_audit(admin_id, 'mmr_chargeable_config', 'settings', 'mmr', {
            'non_apartment_baseunit_non_chargeable': flag,
            'override_count': len(cleaned),
            'builtin_rules': br_merged,
            'location_register_enabled': merged.get('location_register_enabled', True),
        })

        return success_response({'config': merged}, message='MMR chargeable settings saved')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"MMR chargeable config PUT: {e}", exc_info=True)
        return error_response('Failed to save MMR chargeable settings', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/notification-config', methods=['GET', 'PUT'])
@jwt_required()
@admin_required
def notification_config():
    """
    Load/save workflow notification recipients used by inspection/hr email triggers.
    Shape:
    {
      "inspection": {"to": [], "cc": [], "include_submitter": true},
      "hr": {"to": [], "cc": [], "include_submitter": true}
    }
    """
    default_cfg = {
        'inspection': {'to': [], 'cc': [], 'include_submitter': True},
        'hr': {'to': [], 'cc': [], 'include_submitter': True},
    }

    def _clean_emails(raw):
        if raw is None:
            return []
        if isinstance(raw, str):
            raw = [x.strip() for x in raw.split(',')]
        if not isinstance(raw, list):
            return []
        cleaned = []
        seen = set()
        for v in raw:
            email = (v or '').strip()
            if not email:
                continue
            key = email.lower()
            if key in seen:
                continue
            seen.add(key)
            cleaned.append(email)
        return cleaned

    def _normalize(payload):
        payload = payload if isinstance(payload, dict) else {}
        normalized = {}
        for module in ('inspection', 'hr'):
            mod = payload.get(module) if isinstance(payload.get(module), dict) else {}
            normalized[module] = {
                'to': _clean_emails(mod.get('to')),
                'cc': _clean_emails(mod.get('cc')),
                'include_submitter': bool(mod.get('include_submitter', True)),
            }
        return normalized

    if request.method == 'GET':
        try:
            row = NotificationConfig.query.first()
            cfg = _normalize(row.config_json) if row and row.config_json else default_cfg
            return success_response({'config': cfg})
        except Exception as e:
            current_app.logger.error(f"Notification config GET: {e}", exc_info=True)
            return error_response('Failed to load notification settings', status_code=500, error_code='DATABASE_ERROR')

    try:
        data = request.get_json(silent=True) or {}
        incoming = data.get('config') if isinstance(data.get('config'), dict) else data
        normalized = _normalize(incoming)

        row = NotificationConfig.query.first()
        if row:
            row.config_json = normalized
        else:
            db.session.add(NotificationConfig(config_json=normalized))
        db.session.commit()

        return success_response({'config': normalized}, message='Notification settings saved')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Notification config PUT: {e}", exc_info=True)
        return error_response('Failed to save notification settings', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/email-logs', methods=['GET'])
@jwt_required()
@admin_required
def list_email_logs():
    """Newest-first outbound email log for the admin Sent Emails panel."""
    allowed_sources = frozenset({
        'inspection', 'hr', 'bd_email', 'mmr', 'ticketing', 'auth', 'other',
    })
    allowed_statuses = frozenset({'sent', 'failed'})

    source = (request.args.get('source') or '').strip().lower()
    status = (request.args.get('status') or '').strip().lower()
    q = (request.args.get('q') or '').strip()
    try:
        limit = int(request.args.get('limit') or 50)
    except (TypeError, ValueError):
        limit = 50
    try:
        offset = int(request.args.get('offset') or 0)
    except (TypeError, ValueError):
        offset = 0
    limit = max(1, min(limit, 200))
    offset = max(0, offset)

    try:
        query = EmailLog.query
        if source in allowed_sources:
            query = query.filter(EmailLog.source == source)
        if status in allowed_statuses:
            query = query.filter(EmailLog.status == status)
        if q:
            term = f'%{q}%'
            query = query.filter(or_(
                EmailLog.subject.ilike(term),
                EmailLog.to_emails.ilike(term),
                EmailLog.cc_emails.ilike(term),
                EmailLog.related_id.ilike(term),
            ))
        total = query.count()
        rows = (
            query.options(joinedload(EmailLog.sent_by))
            .order_by(EmailLog.created_at.desc(), EmailLog.id.desc())
            .offset(offset)
            .limit(limit)
            .all()
        )
        return success_response({
            'items': [row.to_dict() for row in rows],
            'total': total,
            'limit': limit,
            'offset': offset,
        })
    except Exception as e:
        current_app.logger.error(f"Email logs GET: {e}", exc_info=True)
        return error_response('Failed to load email logs', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/mmr/location-register/parse', methods=['POST'])
@jwt_required()
@admin_required
def mmr_location_register_parse():
    """Parse Location Register Excel (or CAFM HTML export): Base Unit, BU Funct Type, Property, Zone."""
    from module_mmr.mmr_service import parse_location_register_bytes

    _MAX_LOC_BYTES = 5 * 1024 * 1024

    try:
        if 'file' not in request.files:
            return error_response('No file provided', status_code=400, error_code='VALIDATION_ERROR')
        f = request.files['file']
        if not f or not f.filename:
            return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')
        fn = (f.filename or '').lower()
        if not (fn.endswith('.xlsx') or fn.endswith('.xls')):
            return error_response(
                'Upload a .xlsx or .xls file (CAFM “Export to Excel”; HTML exports often use a .xls name).',
                status_code=400,
                error_code='VALIDATION_ERROR',
            )
        data = f.read()
        if len(data) > _MAX_LOC_BYTES:
            return error_response('File too large (max 5 MB)', status_code=400, error_code='VALIDATION_ERROR')
        result = parse_location_register_bytes(data, f.filename or 'register.xlsx')
        result['source_filename'] = (f.filename or '')[:255]
        gen = current_app.config.get('GENERATED_DIR')
        if gen:

            def _write_last_copy():
                try:
                    os.makedirs(gen, exist_ok=True)
                    ext = os.path.splitext(f.filename or '')[1] or '.xlsx'
                    if ext.lower() not in ('.xlsx', '.xls', '.xlsm'):
                        ext = '.xlsx'
                    safe_ext = ext[:8]
                    path = os.path.join(gen, f'mmr_location_register_last{safe_ext}')
                    with open(path, 'wb') as out:
                        out.write(data)
                except Exception as e:
                    current_app.logger.warning('MMR location register copy not saved: %s', e)

            threading.Thread(target=_write_last_copy, daemon=True).start()
        return success_response(result)
    except ValueError as e:
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    except Exception as e:
        current_app.logger.error(f'MMR location register parse: {e}', exc_info=True)
        return error_response('Failed to parse file', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/mmr/chargeable-preview', methods=['POST'])
@jwt_required()
@admin_required
def mmr_chargeable_preview():
    """Batch-resolve Chargeable/Non-Chargeable for BaseUnit strings using current form rules (preview)."""
    from module_mmr.mmr_service import (
        _merge_mmr_chargeable_config,
        merge_builtin_rules_payload,
        preview_chargeable_for_base_units,
    )

    try:
        data = request.get_json(silent=True) or {}
        units = data.get('base_units')
        if not isinstance(units, list):
            return error_response('base_units must be a list', status_code=400, error_code='VALIDATION_ERROR')
        if len(units) > 4000:
            return error_response('Too many base_units (max 4000)', status_code=400, error_code='VALIDATION_ERROR')

        raw: dict = {}
        if 'non_apartment_baseunit_non_chargeable' in data:
            raw['non_apartment_baseunit_non_chargeable'] = bool(
                data.get('non_apartment_baseunit_non_chargeable')
            )
        br_in = data.get('builtin_rules')
        if br_in is not None:
            if not isinstance(br_in, dict):
                return error_response('builtin_rules must be an object', status_code=400, error_code='VALIDATION_ERROR')
            raw['builtin_rules'] = merge_builtin_rules_payload(br_in)
        ov = data.get('baseunit_overrides')
        if ov is not None:
            if not isinstance(ov, list):
                return error_response('baseunit_overrides must be a list', status_code=400, error_code='VALIDATION_ERROR')
            cleaned = []
            for item in ov:
                if not isinstance(item, dict):
                    continue
                pat = (item.get('pattern') or '').strip()
                if not pat:
                    continue
                cleaned.append({'pattern': pat, 'chargeable': bool(item.get('chargeable'))})
            raw['baseunit_overrides'] = cleaned
        if 'location_register_enabled' in data:
            raw['location_register_enabled'] = bool(data.get('location_register_enabled'))

        merged = _merge_mmr_chargeable_config(raw if raw else None)
        strings: list[str] = []
        for u in units:
            if u is None:
                continue
            s = u if isinstance(u, str) else str(u)
            strings.append(s.strip())

        results = preview_chargeable_for_base_units(strings, merged)
        return success_response({'results': results})
    except Exception as e:
        current_app.logger.error(f'MMR chargeable preview: {e}', exc_info=True)
        return error_response('Failed to resolve chargeable preview', status_code=500, error_code='DATABASE_ERROR')


# Admin-triggered password reset fallback. Override via env var in production so
# the literal default below is never used to hand out accounts.
DEFAULT_ADMIN_RESET_PASSWORD = os.environ.get('ADMIN_RESET_PASSWORD_DEFAULT', 'ChangeMeNow!@#')


@admin_bp.route('/users/backfill-passwords', methods=['POST'])
@jwt_required()
@admin_required
def backfill_user_passwords():
    """Try to fill admin_visible_password for accounts using known defaults (hash match only)."""
    try:
        from common.password_admin import backfill_admin_visible_passwords
        stats = backfill_admin_visible_passwords()
        return success_response(
            stats,
            message=(
                f"Backfill complete: {stats['updated']} password(s) recorded, "
                f"{stats['skipped']} still unknown (user must log in once, or use Reset password)."
            ),
        )
    except Exception as e:
        current_app.logger.error(f"Password backfill error: {e}", exc_info=True)
        return error_response('Password backfill failed', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users', methods=['GET'])
@jwt_required()
@admin_required
def list_users():
    """Get all users with their details"""
    try:
        # Note: User.submissions, audit_logs, and sessions are dynamic relationships
        # (lazy='dynamic'), so we can't use joinedload. The to_dict() method doesn't
        # include these relationships anyway, so we just query users directly.
        users = User.query.options(
            joinedload(User.reporting_manager),
            joinedload(User.operations_manager),
        ).filter(
            ~User.username.in_(DIRECTORY_HIDDEN_USERNAMES),
        ).order_by(User.created_at.desc()).all()
        access_map = {row.user_id: row.can_access for row in DocHubAccess.query.all()}
        users_data = []
        for user in users:
            d = _admin_user_dict(user)
            _merge_dochub_into_user_dict(user, d, access_map)
            users_data.append(d)

        return success_response({
            'users': users_data,
            'count': len(users_data)
        })
    except Exception as e:
        current_app.logger.error(f"Error listing users: {str(e)}", exc_info=True)
        return error_response('Failed to fetch users', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users', methods=['POST'])
@jwt_required()
@admin_required
def create_user_admin():
    """Create a new user account (admin). Omit password to use the same default as password reset."""
    try:
        from app.auth.routes import validate_email, validate_password

        admin_id = get_jwt_identity()
        data = request.get_json(force=True, silent=True) or {}
        username = (data.get('username') or '').strip()
        email = (data.get('email') or '').strip().lower()
        full_name = (data.get('full_name') or '').strip() or None

        if not username or not email:
            return error_response('Username and email are required', status_code=400, error_code='VALIDATION_ERROR')
        if not validate_email(email):
            return error_response('Invalid email format', status_code=400, error_code='VALIDATION_ERROR')
        if User.query.filter_by(username=username).first():
            return error_response('Username already exists', status_code=409, error_code='DUPLICATE_USERNAME')
        if User.query.filter_by(email=email).first():
            return error_response('Email already in use', status_code=409, error_code='DUPLICATE_EMAIL')

        raw_pw = (data.get('password') or '').strip()
        temp_password = raw_pw or (os.environ.get('ADMIN_RESET_PASSWORD') or '').strip() or DEFAULT_ADMIN_RESET_PASSWORD
        if raw_pw:
            ok, msg = validate_password(raw_pw)
            if not ok:
                return error_response(msg, status_code=400, error_code='WEAK_PASSWORD')
        else:
            ok, msg = validate_password(temp_password)
            if not ok:
                temp_password = DEFAULT_ADMIN_RESET_PASSWORD

        user = User(username=username, email=email, full_name=full_name, role='user')
        if data.get('role') == 'admin':
            user.role = 'admin'
        user.set_password(temp_password)
        user.password_changed = bool(raw_pw)

        valid_designations = {
            'supervisor', 'operations_manager', 'business_development',
            'procurement', 'general_manager', 'hr_manager', 'employee',
            'technician', 'admin',
        }
        des = data.get('designation')
        if des in (None, ''):
            user.designation = None
        elif des in valid_designations:
            user.designation = des
        else:
            return error_response('Invalid designation', status_code=400, error_code='VALIDATION_ERROR')

        if 'employment_start_date' in data:
            try:
                user.employment_start_date = parse_employment_start_date(data.get('employment_start_date'))
            except ValueError as ve:
                return error_response(str(ve), status_code=400, error_code='VALIDATION_ERROR')

        if 'job_designation' in data:
            user.job_designation = _truncate_job_designation(data.get('job_designation'))

        try:
            if 'annual_leave_days' in data:
                user.annual_leave_days = _parse_leave_days_field(data.get('annual_leave_days'))
            if 'other_leave_days' in data:
                user.other_leave_days = _parse_leave_days_field(data.get('other_leave_days'))
            if 'reporting_manager_id' in data:
                user.reporting_manager_id = _resolve_reporting_manager_id(
                    data.get('reporting_manager_id'), exclude_user_id=None
                )
        except ValueError as ve:
            return error_response(str(ve), status_code=400, error_code='VALIDATION_ERROR')

        if user.role != 'admin':
            user.access_hvac = bool(data.get('access_hvac', False))
            user.access_civil = bool(data.get('access_civil', False))
            user.access_cleaning = bool(data.get('access_cleaning', False))
            user.access_hr = bool(data.get('access_hr', False))
            user.access_hiring = bool(data.get('access_hiring', False))
            user.access_procurement_module = bool(data.get('access_procurement_module', False))
            user.access_business_development = bool(data.get('access_business_development', False))
            user.access_sales_manager = bool(data.get('access_sales_manager', False))
            user.access_quotations = bool(data.get('access_quotations', False))
            user.access_report_generation = bool(data.get('access_report_generation', False))
            user.access_submitted_forms = bool(data.get('access_submitted_forms', False))
            user.access_ticketing = bool(data.get('access_ticketing', False))
            user.access_qhsi = bool(data.get('access_qhsi', False))
            user.access_files = bool(data.get('access_files', False))
            user.is_ticket_reporter = bool(data.get('is_ticket_reporter', False))
            if bool(user.access_sales_manager):
                user.access_business_development = True
            if bool(user.access_hiring):
                user.access_hr = True
            if not bool(user.access_hr):
                user.access_hiring = False
            if (user.designation or '') == 'business_development':
                user.access_business_development = True
                if data.get('access_quotations') is None:
                    user.access_quotations = True

        db.session.add(user)
        db.session.flush()

        if user.role != 'admin' and data.get('can_access_dochub') is False:
            row = DocHubAccess.query.filter_by(user_id=user.id).first()
            if not row:
                db.session.add(DocHubAccess(user_id=user.id, can_access=False))
            else:
                row.can_access = False

        db.session.commit()

        log_audit(admin_id, 'create_user', 'user', str(user.id), {'username': username, 'email': email})

        email_sent = False
        try:
            from common.email_service import send_account_created_email
            email_sent = send_account_created_email(
                user_email=user.email,
                username=user.username,
                full_name=user.full_name,
                temp_password=(temp_password if not raw_pw else None),
            )
        except Exception as email_error:
            current_app.logger.warning('Account created email failed: %s', email_error)

        out = {'user': _admin_user_dict(user), 'email_sent': email_sent}
        if not raw_pw:
            out['temp_password'] = temp_password
        message = 'User created successfully'
        if email_sent:
            message += '. A welcome email with the username was sent.'
        return success_response(out, message=message)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating user: {str(e)}", exc_info=True)
        return error_response('Failed to create user', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/<int:user_id>', methods=['GET'])
@jwt_required()
@admin_required
def get_user(user_id):
    """Get specific user details"""
    try:
        user = User.query.options(
            joinedload(User.reporting_manager),
            joinedload(User.operations_manager),
        ).get_or_404(user_id)
        user_payload = _admin_user_dict(user)
        _merge_dochub_into_user_dict(user, user_payload)
        return jsonify({
            'success': True,
            'user': user_payload
        }), 200
    except Exception as e:
        current_app.logger.error(f"Error fetching user: {str(e)}")
        return jsonify({'error': 'User not found'}), 404


@admin_bp.route('/users/<int:user_id>/edit-otp/status', methods=['GET'])
@jwt_required()
@admin_required
def admin_edit_otp_status(user_id):
    target = User.query.get_or_404(user_id)
    actor_id = _int_actor_id()
    required = (target.role == 'admin')
    now = utc_now_naive()
    row = None
    if required and actor_id is not None:
        row = AdminEditOtp.query.filter_by(
            actor_user_id=actor_id, target_user_id=target.id
        ).first()
    unlocked = required and _admin_edit_grant_ok(actor_id, target.id)
    grant_until = None
    if unlocked and row and row.grant_expires_at:
        grant_until = _otp_iso(row.grant_expires_at)
    timing = _otp_timing_fields(row, now) if row else {
        'has_pending_code': False,
        'code_expires_at': None,
        'resend_available_at': _otp_iso(now),
        'resend_after_seconds': 0,
    }
    return jsonify({
        'success': True,
        'otp_required': required,
        'unlocked': unlocked,
        'grant_expires_at': grant_until,
        'email_masked': _mask_email(target.email) if required else None,
        'has_email': bool((target.email or '').strip()) if required else False,
        **timing,
    }), 200


@admin_bp.route('/users/<int:user_id>/edit-otp/request', methods=['POST'])
@jwt_required()
@admin_required
def admin_edit_otp_request(user_id):
    target = User.query.get_or_404(user_id)
    if target.role != 'admin':
        return jsonify({'error': 'A verification code is only required for administrator accounts.'}), 400
    email = (target.email or '').strip()
    if not email:
        return jsonify({'error': 'This administrator account has no email address.'}), 400

    actor_id = _int_actor_id()
    if actor_id is None:
        return jsonify({'error': 'Invalid session'}), 401

    now = utc_now_naive()
    row = _otp_row_for(actor_id, target.id)
    if _otp_has_pending_code(row, now):
        resend_at = _otp_resend_at(row)
        if resend_at and now < resend_at:
            secs = max(1, int((resend_at - now).total_seconds()))
            return jsonify({
                'error': 'Wait before sending another code.',
                'retry_after': secs,
                'resend_after_seconds': secs,
                'resend_available_at': _otp_iso(resend_at),
                'code_expires_at': _otp_iso(row.expires_at),
                'has_pending_code': True,
            }), 429
    if row.window_started_at and (now - row.window_started_at) > OTP_REQUEST_WINDOW:
        row.request_count = 0
        row.window_started_at = now
    if not row.window_started_at:
        row.window_started_at = now
    if (row.request_count or 0) >= OTP_MAX_REQUESTS:
        return jsonify({
            'error': 'Too many verification codes. Wait a few minutes and try again.',
        }), 429

    code = _generate_otp_code()
    row.code_hash = _hash_otp_code(code)
    row.expires_at = now + OTP_CODE_TTL
    row.attempt_count = 0
    row.grant_expires_at = None
    row.request_count = (row.request_count or 0) + 1
    db.session.commit()

    from common.email_service import send_admin_edit_otp_email
    try:
        sent = send_admin_edit_otp_email(email, code, full_name=target.full_name)
    except Exception as email_error:
        current_app.logger.warning('Admin edit OTP email failed: %s', email_error)
        sent = False
    if not sent:
        row.code_hash = None
        db.session.commit()
        return jsonify({
            'error': 'Could not send the verification email. Check mail settings and try again.',
            'password_unlock_allowed': True,
        }), 502

    log_audit(actor_id, 'admin_edit_otp_requested', 'user', str(user_id), {
        'target_user': target.username,
        'sent_to': _mask_email(email),
    })
    timing = _otp_timing_fields(row, now)
    return jsonify({
        'success': True,
        'sent_to': _mask_email(email),
        'message': f'A verification code was sent to {_mask_email(email)}.',
        **timing,
    }), 200


@admin_bp.route('/users/<int:user_id>/edit-otp/verify', methods=['POST'])
@jwt_required()
@admin_required
def admin_edit_otp_verify(user_id):
    target = User.query.get_or_404(user_id)
    if target.role != 'admin':
        return jsonify({'error': 'A verification code is only required for administrator accounts.'}), 400
    actor_id = _int_actor_id()
    if actor_id is None:
        return jsonify({'error': 'Invalid session'}), 401
    data = request.get_json(silent=True) or {}
    code = ''.join(ch for ch in str(data.get('code') or '') if ch.isdigit())
    if len(code) != 6:
        return jsonify({'error': 'Enter the 6-digit code from the email.'}), 400

    now = utc_now_naive()
    row = AdminEditOtp.query.filter_by(
        actor_user_id=actor_id, target_user_id=target.id
    ).first()
    if not row or not row.code_hash or not row.expires_at or row.expires_at < now:
        return jsonify({'error': 'That code has expired. Send a new one.'}), 400
    if (row.attempt_count or 0) >= OTP_MAX_ATTEMPTS:
        row.code_hash = None
        db.session.commit()
        return jsonify({'error': 'Too many attempts. Send a new verification code.'}), 400
    if not _otp_code_matches(code, row.code_hash):
        row.attempt_count = (row.attempt_count or 0) + 1
        if row.attempt_count >= OTP_MAX_ATTEMPTS:
            row.code_hash = None
            db.session.commit()
            return jsonify({'error': 'Too many attempts. Send a new verification code.'}), 400
        db.session.commit()
        left = OTP_MAX_ATTEMPTS - row.attempt_count
        return jsonify({'error': f'Incorrect code. {left} attempt(s) remaining.'}), 400

    grant_until = _issue_edit_grant(row, now)
    db.session.commit()
    log_audit(actor_id, 'admin_edit_otp_verified', 'user', str(user_id), {
        'target_user': target.username,
    })
    return jsonify({
        'success': True,
        'unlocked': True,
        'grant_expires_at': _otp_iso(grant_until),
        'message': 'Editing unlocked for 10 minutes.',
    }), 200


@admin_bp.route('/users/<int:user_id>/edit-otp/lock', methods=['POST'])
@jwt_required()
@admin_required
def admin_edit_otp_lock(user_id):
    target = User.query.get_or_404(user_id)
    if target.role != 'admin':
        return jsonify({'error': 'A verification code is only required for administrator accounts.'}), 400
    actor_id = _int_actor_id()
    if actor_id is None:
        return jsonify({'error': 'Invalid session'}), 401
    row = AdminEditOtp.query.filter_by(
        actor_user_id=actor_id, target_user_id=target.id
    ).first()
    if row:
        row.grant_expires_at = None
        row.code_hash = None
        row.attempt_count = 0
        db.session.commit()
    log_audit(actor_id, 'admin_edit_otp_locked', 'user', str(user_id), {
        'target_user': target.username,
    })
    now = utc_now_naive()
    return jsonify({
        'success': True,
        'unlocked': False,
        'otp_required': True,
        'email_masked': _mask_email(target.email),
        'has_email': bool((target.email or '').strip()),
        **_otp_timing_fields(row, now),
    }), 200


@admin_bp.route('/users/<int:user_id>/edit-otp/unlock-with-password', methods=['POST'])
@jwt_required()
@admin_required
def admin_edit_otp_unlock_with_password(user_id):
    target = User.query.get_or_404(user_id)
    if target.role != 'admin':
        return jsonify({'error': 'A verification code is only required for administrator accounts.'}), 400
    actor_id = _int_actor_id()
    if actor_id is None:
        return jsonify({'error': 'Invalid session'}), 401
    actor = db.session.get(User, actor_id)
    if not actor:
        return jsonify({'error': 'Invalid session'}), 401
    data = request.get_json(silent=True) or {}
    password = data.get('password')
    if not isinstance(password, str) or not password.strip():
        return jsonify({'error': 'Enter your admin password.'}), 400

    now = utc_now_naive()
    row = _otp_row_for(actor_id, target.id)
    if (row.attempt_count or 0) >= OTP_MAX_ATTEMPTS:
        return jsonify({'error': 'Too many attempts. Wait a few minutes and try again.'}), 400
    if not actor.check_password(password):
        row.attempt_count = (row.attempt_count or 0) + 1
        db.session.commit()
        if row.attempt_count >= OTP_MAX_ATTEMPTS:
            return jsonify({'error': 'Too many attempts. Wait a few minutes and try again.'}), 400
        left = OTP_MAX_ATTEMPTS - row.attempt_count
        return jsonify({'error': f'Incorrect password. {left} attempt(s) remaining.'}), 400

    grant_until = _issue_edit_grant(row, now)
    db.session.commit()
    log_audit(actor_id, 'admin_edit_otp_password_unlock', 'user', str(user_id), {
        'target_user': target.username,
    })
    return jsonify({
        'success': True,
        'unlocked': True,
        'grant_expires_at': _otp_iso(grant_until),
        'message': 'Editing unlocked for 10 minutes.',
    }), 200


@admin_bp.route('/users/<int:user_id>/reset-password', methods=['POST'])
@jwt_required()
@admin_required
def reset_user_password(user_id):
    """Reset user password to the standard admin default (Injaaz@123, or ADMIN_RESET_PASSWORD) and email when configured."""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied
        
        raw_reset = (os.environ.get('ADMIN_RESET_PASSWORD') or '').strip()
        temp_password = raw_reset or DEFAULT_ADMIN_RESET_PASSWORD
        user.set_password(temp_password)
        user.password_changed = False  # Force password change on next login
        
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'reset_password', 'user', str(user_id), {
            'target_user': user.username,
            'reset_by': User.query.get(admin_id).username if admin_id else 'system'
        })
        
        # Send email with temporary password
        email_sent = False
        try:
            from common.email_service import send_password_reset_email
            email_sent = send_password_reset_email(
                user_email=user.email,
                username=user.username,
                temp_password=temp_password
            )
        except Exception as email_error:
            current_app.logger.warning(f"Failed to send password reset email: {str(email_error)}")
            # Continue - password is reset even if email fails
        
        if email_sent:
            message = 'Password reset successfully. Temporary password has been sent to user\'s email.'
        else:
            # If email fails, return password in response as fallback (admin only)
            message = 'Password reset successfully. Email delivery failed - password returned in response.'
            current_app.logger.warning(f"Password reset email failed for user {user_id}, password returned in response")
        
        response_data = {
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email
            },
            'temp_password': temp_password,
        }
        
        if not email_sent:
            response_data['warning'] = 'Email delivery failed. Password returned in response (admin only).'
        
        return success_response(response_data, message=message)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error resetting password: {str(e)}")
        return error_response('Failed to reset password', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/<int:user_id>/reset-mfa', methods=['POST'])
@jwt_required()
@admin_required
def reset_user_mfa(user_id):
    """Clear authenticator enrollment so the user can sign in with password only."""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied
        had_mfa = bool(getattr(user, 'mfa_enabled', False) or getattr(user, 'mfa_secret', None))
        user.mfa_enabled = False
        user.mfa_secret = None
        db.session.commit()
        db.session.refresh(user)
        actor = User.query.get(admin_id)
        log_audit(admin_id, 'mfa_reset', 'user', str(user_id), {
            'target_user': user.username,
            'reset_by': actor.username if actor else 'system',
        })
        from common.email_service import notify_user_mfa_email_later
        sent_to = notify_user_mfa_email_later(user, enabled=False, by_admin=True)
        if sent_to is True:
            sent_to = (user.email or '').strip() or None
        elif not sent_to:
            sent_to = None
        message = (
            'Authenticator reset. The user can sign in with password only, then set up a new QR from Profile.'
        )
        if sent_to:
            message += f' A notice was emailed to {sent_to}.'
        elif (user.email or '').strip():
            message += f' Could not email {(user.email or "").strip()}.'
        else:
            message += ' This account has no email address, so no notice was sent.'
        return success_response({
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'mfa_enabled': False,
                'mfa_configured': False,
            },
            'sent_to': sent_to,
            'had_mfa': had_mfa,
        }, message=message)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error('Error resetting MFA: %s', e)
        return error_response('Failed to reset authenticator', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/<int:user_id>/email-login-details', methods=['POST'])
@jwt_required()
@admin_required
def email_user_login_details(user_id):
    """Email the stored username and password to the user's address."""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get(user_id)
        if not user:
            return error_response('User not found.', status_code=404, error_code='NOT_FOUND')
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied
        email = (user.email or '').strip()
        if not email:
            return error_response('This account has no email address.', status_code=400, error_code='VALIDATION_ERROR')
        password = (getattr(user, 'admin_visible_password', None) or '').strip()
        if not password:
            return error_response(
                'No password on file for admin view. Reset password or save a new password first.',
                status_code=400,
                error_code='VALIDATION_ERROR',
            )

        from common.email_service import send_login_details_email
        try:
            email_sent = send_login_details_email(
                user_email=email,
                username=user.username,
                password=password,
                full_name=user.full_name,
            )
        except Exception as email_error:
            current_app.logger.warning('Login details email failed: %s', email_error)
            email_sent = False
        log_audit(admin_id, 'email_login_details', 'user', str(user_id), {
            'target_user': user.username,
            'email_sent': bool(email_sent),
        })
        if not email_sent:
            return error_response(
                'Could not send the email. Check mail settings and try again.',
                status_code=502,
                error_code='EMAIL_FAILED',
            )
        return success_response(
            {'email': email},
            message=f'Login details were emailed to {email}.',
        )
    except Exception as e:
        current_app.logger.error('Error emailing login details: %s', e, exc_info=True)
        return error_response('Failed to email login details', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/<int:user_id>/toggle-active', methods=['POST'])
@jwt_required()
@admin_required
def toggle_user_active(user_id):
    """Activate or deactivate a user"""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied
        
        # Prevent deactivating yourself
        if user_id == admin_id:
            return jsonify({'error': 'Cannot deactivate your own account'}), 400
        
        user.is_active = not user.is_active
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'toggle_user_status', 'user', str(user_id), {
            'target_user': user.username,
            'new_status': 'active' if user.is_active else 'inactive',
            'changed_by': User.query.get(admin_id).username if admin_id else 'system'
        })

        email_sent = False
        if user.email:
            try:
                from common.email_service import send_account_status_email
                email_sent = send_account_status_email(
                    user.email,
                    user.username,
                    is_active=bool(user.is_active),
                    full_name=user.full_name,
                )
            except Exception as email_error:
                current_app.logger.warning('Account status email failed: %s', email_error)

        status_word = 'activated' if user.is_active else 'deactivated'
        message = f'User {status_word} successfully'
        if email_sent:
            message += f'. An email was sent to {user.email}.'
        elif user.email:
            message += '. The status email could not be sent.'
        
        return jsonify({
            'success': True,
            'message': message,
            'email_sent': email_sent,
            'user': _admin_user_dict(user)
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error toggling user status: {str(e)}")
        return jsonify({'error': 'Failed to update user status'}), 500


@admin_bp.route('/users/<int:user_id>/update-access', methods=['POST'])
@jwt_required()
@admin_required
def update_user_access(user_id):
    """Update user's module access permissions"""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied
        
        # Parse JSON with better error handling
        current_app.logger.info(f"Update access request for user {user_id}, Content-Type: {request.content_type}")
        
        try:
            data = request.get_json(force=True, silent=True)
        except Exception as json_error:
            current_app.logger.error(f"JSON parsing error: {json_error}")
            return jsonify({'error': 'Invalid JSON format'}), 400
        
        if data is None:
            body_text = request.get_data(as_text=True)
            current_app.logger.error(f"Invalid JSON in update-access request for user {user_id}. Content-Type: {request.content_type}, Body: {body_text[:200]}")
            return jsonify({'error': 'Invalid JSON or missing request body'}), 400
        
        current_app.logger.info(f"Received access data: {data}")
        
        # Get access values from request, defaulting to current values if not provided
        # Handle case where columns might not exist yet (use getattr with default)
        try:
            current_access_hvac = getattr(user, 'access_hvac', False)
            current_access_civil = getattr(user, 'access_civil', False)
            current_access_cleaning = getattr(user, 'access_cleaning', False)
        except AttributeError as attr_error:
            current_app.logger.error(f"User model missing access attributes: {attr_error}")
            return jsonify({'error': 'Database schema error - access columns missing. Please run migration.'}), 500
        
        # Get values from request, use current values as defaults
        access_hvac = data.get('access_hvac', current_access_hvac)
        access_civil = data.get('access_civil', current_access_civil)
        access_cleaning = data.get('access_cleaning', current_access_cleaning)
        
        # Convert to boolean (handles string "true"/"false", None, etc.)
        user.access_hvac = bool(access_hvac)
        user.access_civil = bool(access_civil)
        user.access_cleaning = bool(access_cleaning)
        
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'update_user_access', 'user', str(user_id), {
            'target_user': user.username,
            'access_hvac': user.access_hvac,
            'access_civil': user.access_civil,
            'access_cleaning': user.access_cleaning,
            'changed_by': User.query.get(admin_id).username if admin_id else 'system'
        })
        
        return jsonify({
            'success': True,
            'message': 'User access updated successfully',
            'user': _admin_user_dict(user)
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating user access: {str(e)}", exc_info=True)
        return jsonify({'error': f'Failed to update user access: {str(e)}'}), 500


@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_user(user_id):
    """Delete a user account"""
    try:
        admin_id = get_jwt_identity()
        try:
            admin_id_int = int(admin_id)
        except (TypeError, ValueError):
            admin_id_int = None

        # Prevent deleting yourself (JWT identity is stored as a string)
        if admin_id_int is not None and user_id == admin_id_int:
            return jsonify({'error': 'Cannot delete your own account'}), 400
        
        user = User.query.get_or_404(user_id)
        denied = _reject_unless_admin_edit_granted(user)
        if denied:
            return denied

        if user.is_active:
            return jsonify({
                'error': 'Deactivate this account before deleting it.',
            }), 400
        
        # Prevent deleting the last admin
        if user.role == 'admin':
            admin_count = User.query.filter_by(role='admin', is_active=True).count()
            if admin_count <= 1:
                return jsonify({'error': 'Cannot delete the last active admin user'}), 400
        
        username = user.username
        display_name = (user.full_name or username or 'this user').strip()

        fallback_admin = _admin_for_reassigned_records(user_id, preferred_id=admin_id_int)
        fallback_id = fallback_admin.id if fallback_admin else None
        fallback_name = (
            (fallback_admin.full_name or fallback_admin.username).strip()
            if fallback_admin else 'another administrator'
        )
        try:
            ticket_count = _detach_user_for_delete(user_id, fallback_id)
        except ValueError as blocked:
            return jsonify({'error': str(blocked)}), 409

        User.query.filter_by(reporting_manager_id=user_id).update(
            {'reporting_manager_id': None},
            synchronize_session=False,
        )
        User.query.filter_by(operations_manager_id=user_id).update(
            {'operations_manager_id': None},
            synchronize_session=False,
        )
        DocHubAccess.query.filter_by(updated_by=user_id).update(
            {'updated_by': None},
            synchronize_session=False,
        )
        access_row = user.dochub_access_entry
        if access_row is not None:
            db.session.delete(access_row)
        AdminEditOtp.query.filter(
            or_(
                AdminEditOtp.actor_user_id == user_id,
                AdminEditOtp.target_user_id == user_id,
            )
        ).delete(synchronize_session=False)

        db.session.flush()
        db.session.expire(user)

        # Delete user (cascade will handle related records)
        db.session.delete(user)
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'delete_user', 'user', str(user_id), {
            'deleted_user': username,
            'deleted_by': User.query.get(admin_id).username if admin_id else 'system',
            'tickets_reassigned': ticket_count,
            'tickets_reassigned_to': fallback_id,
        })

        message = f'User {username} deleted successfully'
        if ticket_count:
            noun = 'ticket' if ticket_count == 1 else 'tickets'
            message = (
                f'{display_name} was deleted. {ticket_count} {noun} stayed in the system '
                f'and the reporter was set to {fallback_name}.'
            )
        
        return jsonify({
            'success': True,
            'message': message,
            'tickets_reassigned': ticket_count,
        }), 200
    except IntegrityError as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting user: {str(e)}", exc_info=True)
        detail = str(getattr(e, 'orig', None) or e)
        if 'reporter_id' in detail:
            return jsonify({
                'error': 'Cannot delete this account because it is still the reporter on one or more tickets.',
            }), 409
        return jsonify({
            'error': 'Cannot delete this account because other records still reference it.',
        }), 409
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting user: {str(e)}", exc_info=True)
        reason = str(e).split('\n')[0].strip()[:240]
        return jsonify({
            'error': f'Cannot delete this account. {reason}' if reason else 'Cannot delete this account.',
        }), 500


@admin_bp.route('/users/<int:user_id>/activity', methods=['GET'])
@jwt_required()
@admin_required
def get_user_activity(user_id):
    """Get user's submission and review activity"""
    try:
        from app.models import Submission
        
        user = User.query.get_or_404(user_id)
        
        # Get submissions created by this user (as supervisor)
        submitted_forms = []
        if user.designation == 'supervisor':
            submissions = Submission.query.filter_by(supervisor_id=user.id).order_by(Submission.created_at.desc()).all()
            for sub in submissions:
                submitted_forms.append({
                    'id': sub.id,
                    'submission_id': sub.submission_id,
                    'module_type': sub.module_type,
                    'site_name': sub.site_name or 'N/A',
                    'visit_date': sub.visit_date.isoformat() if sub.visit_date else None,
                    'status': sub.status,
                    'workflow_status': getattr(sub, 'workflow_status', 'submitted'),
                    'created_at': sub.created_at.isoformat() + 'Z' if sub.created_at else None
                })
        
        # Get forms reviewed by this user based on their designation
        reviewed_forms = []
        designation = user.designation
        reviews = []
        
        if designation == 'operations_manager':
            reviews = Submission.query.filter(
                Submission.operations_manager_id == user.id
            ).order_by(Submission.updated_at.desc()).all()
        elif designation == 'business_development':
            reviews = Submission.query.filter(
                Submission.business_dev_id == user.id
            ).order_by(Submission.updated_at.desc()).all()
        elif designation == 'procurement':
            reviews = Submission.query.filter(
                Submission.procurement_id == user.id
            ).order_by(Submission.updated_at.desc()).all()
        elif designation == 'general_manager':
            reviews = Submission.query.filter(
                Submission.general_manager_id == user.id
            ).order_by(Submission.updated_at.desc()).all()
        
        for sub in reviews:
            # Get the supervisor info
            supervisor = User.query.get(sub.supervisor_id) if sub.supervisor_id else None
            reviewed_forms.append({
                'id': sub.id,
                'submission_id': sub.submission_id,
                'module_type': sub.module_type,
                'site_name': sub.site_name or 'N/A',
                'visit_date': sub.visit_date.isoformat() if sub.visit_date else None,
                'status': sub.status,
                'workflow_status': getattr(sub, 'workflow_status', 'submitted'),
                'created_at': sub.created_at.isoformat() + 'Z' if sub.created_at else None,
                'supervisor': supervisor.full_name or supervisor.username if supervisor else 'Unknown'
            })
        
        return success_response({
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.full_name,
                'designation': user.designation,
                'role': user.role
            },
            'submitted_forms': submitted_forms,
            'submitted_count': len(submitted_forms),
            'reviewed_forms': reviewed_forms,
            'reviewed_count': len(reviewed_forms)
        })
    except Exception as e:
        current_app.logger.error(f"Error fetching user activity: {str(e)}", exc_info=True)
        return error_response('Failed to fetch user activity', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_user(user_id):
    """Update user details"""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        
        data = request.get_json() or {}
        promoting = (data.get('role') == 'admin' and user.role != 'admin')
        denied = _reject_unless_admin_edit_granted(user, promoting_to_admin=promoting)
        if denied:
            return denied
        
        # Update allowed fields
        if 'full_name' in data:
            user.full_name = data['full_name']
        if 'email' in data:
            email = (data.get('email') or '').strip().lower()
            if not email:
                return jsonify({'error': 'Email is required'}), 400
            existing = User.query.filter(func.lower(User.email) == email).first()
            if existing and existing.id != user_id:
                return jsonify({'error': 'Email already in use'}), 400
            user.email = email
        if 'username' in data:
            # Check if username is already taken
            existing = User.query.filter_by(username=data['username']).first()
            if existing and existing.id != user_id:
                return jsonify({'error': 'Username already in use'}), 400
            user.username = data['username']
        if 'role' in data and data['role'] in ['admin', 'user']:
            # Prevent changing your own role
            if user_id == admin_id and data['role'] != 'admin':
                return jsonify({'error': 'Cannot change your own role'}), 400
            user.role = data['role']
        
        # Update designation if provided (includes hr_manager for HR module)
        if 'designation' in data:
            valid_designations = {
                'supervisor', 'operations_manager', 'business_development',
                'procurement', 'general_manager', 'hr_manager',
                'employee', 'technician', 'admin',
            }
            d = data['designation']
            if d in (None, ''):
                user.designation = None
            elif d in valid_designations:
                user.designation = d
            else:
                return jsonify({'error': 'Invalid designation'}), 400

        if 'default_signature' in data:
            sig = data['default_signature']
            if sig in (None, ''):
                user.default_signature = None
            elif isinstance(sig, str):
                if len(sig) > 2_000_000:
                    return jsonify({'error': 'Signature data too large'}), 400
                user.default_signature = sig
            else:
                return jsonify({'error': 'Invalid signature format'}), 400

        if 'default_comment' in data:
            from common.utils import normalize_approval_comment
            import re
            c = data['default_comment']
            if c in (None, ''):
                user.default_comment = None
            else:
                cleaned = str(c).strip()[:5000]
                if re.match(r'^Signed\s*(?:&|and)\s*Verified\.?$', cleaned, re.I):
                    user.default_comment = None
                else:
                    user.default_comment = normalize_approval_comment(cleaned) or None

        if 'employment_start_date' in data:
            try:
                user.employment_start_date = parse_employment_start_date(data.get('employment_start_date'))
            except ValueError as ve:
                return jsonify({'error': str(ve)}), 400

        if 'job_designation' in data:
            user.job_designation = _truncate_job_designation(data.get('job_designation'))

        try:
            if 'annual_leave_days' in data:
                user.annual_leave_days = _parse_leave_days_field(data.get('annual_leave_days'))
            if 'other_leave_days' in data:
                user.other_leave_days = _parse_leave_days_field(data.get('other_leave_days'))
            if 'reporting_manager_id' in data:
                user.reporting_manager_id = _resolve_reporting_manager_id(
                    data.get('reporting_manager_id'), exclude_user_id=user_id
                )
        except ValueError as ve:
            return jsonify({'error': str(ve)}), 400

        if user.role != 'admin':
            for key, col in (
                ('access_hvac', 'access_hvac'),
                ('access_civil', 'access_civil'),
                ('access_cleaning', 'access_cleaning'),
                ('access_hr', 'access_hr'),
                ('access_hiring', 'access_hiring'),
                ('access_procurement_module', 'access_procurement_module'),
                ('access_business_development', 'access_business_development'),
                ('access_sales_manager', 'access_sales_manager'),
                ('access_quotations', 'access_quotations'),
                ('access_report_generation', 'access_report_generation'),
                ('access_submitted_forms', 'access_submitted_forms'),
                ('access_ticketing', 'access_ticketing'),
                ('access_qhsi', 'access_qhsi'),
                ('access_files', 'access_files'),
                ('is_ticket_reporter', 'is_ticket_reporter'),
            ):
                if key in data:
                    setattr(user, col, bool(data[key]))
            if bool(getattr(user, 'access_sales_manager', False)):
                user.access_business_development = True
            if bool(getattr(user, 'access_hiring', False)):
                user.access_hr = True
            if not bool(getattr(user, 'access_hr', False)):
                user.access_hiring = False
            if (user.designation or '') == 'business_development':
                user.access_business_development = True
                if 'access_quotations' not in data:
                    user.access_quotations = True

        password_updated = False
        if 'password' in data:
            from app.auth.routes import validate_password
            raw_pw = (data.get('password') or '').strip()
            if raw_pw:
                ok, msg = validate_password(raw_pw)
                if not ok:
                    return jsonify({'error': msg}), 400
                user.set_password(raw_pw)
                user.password_changed = True
                password_updated = True
        
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'update_user', 'user', str(user_id), {
            'target_user': user.username,
            'changed_by': User.query.get(admin_id).username if admin_id else 'system'
        })

        email_sent = False
        if password_updated and user.email:
            try:
                from common.email_service import send_password_updated_email
                email_sent = send_password_updated_email(
                    user.email, user.username, by_admin=True, full_name=user.full_name
                )
            except Exception as email_error:
                current_app.logger.warning('Password updated email failed: %s', email_error)

        message = 'User updated successfully'
        if password_updated and email_sent:
            message += '. A password-updated email was sent.'
        
        return jsonify({
            'success': True,
            'message': message,
            'email_sent': email_sent,
            'user': _admin_user_dict(user)
        }), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating user: {str(e)}")
        return jsonify({'error': 'Failed to update user'}), 500


@admin_bp.route('/documents', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_documents():
    """Delete one or multiple documents (submissions)"""
    try:
        admin_id = get_jwt_identity()
        data = request.get_json()
        
        if not data or 'submission_ids' not in data:
            return error_response('submission_ids array is required', status_code=400, error_code='VALIDATION_ERROR')
        
        submission_ids = data.get('submission_ids', [])
        if not isinstance(submission_ids, list) or len(submission_ids) == 0:
            return error_response('submission_ids must be a non-empty array', status_code=400, error_code='VALIDATION_ERROR')
        
        from app.models import Submission
        
        # Get submissions to delete
        submissions = Submission.query.filter(Submission.submission_id.in_(submission_ids)).all()
        
        if len(submissions) != len(submission_ids):
            found_ids = [s.submission_id for s in submissions]
            missing = [sid for sid in submission_ids if sid not in found_ids]
            return error_response(f'Some submissions not found: {missing}', status_code=404, error_code='NOT_FOUND')
        
        deleted_count = 0
        deleted_ids = []
        
        for submission in submissions:
            submission_id = submission.submission_id
            deleted_ids.append(submission_id)
            db.session.delete(submission)
            deleted_count += 1
        
        db.session.commit()
        
        # Log the action
        log_audit(admin_id, 'delete_documents', 'submission', ','.join(deleted_ids), {
            'count': deleted_count,
            'deleted_by': User.query.get(admin_id).username if admin_id else 'system'
        })
        
        return success_response({
            'deleted_count': deleted_count,
            'deleted_ids': deleted_ids
        }, message=f'Successfully deleted {deleted_count} document(s)')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting documents: {str(e)}", exc_info=True)
        return error_response('Failed to delete documents', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/documents', methods=['GET'])
@jwt_required()
@admin_required
def list_documents():
    """Get all submissions/documents with their details"""
    try:
        from app.models import Submission, Job, File
        from common.document_display import ensure_document_numbers

        # Lazily assign per-series document numbers (HR-0001, INSP-0001, ...)
        # to any submissions that don't have one yet. Existing numbers never change.
        try:
            ensure_document_numbers()
        except Exception as num_err:
            db.session.rollback()
            current_app.logger.warning(f"Document numbering skipped: {num_err}")

        # Get all submissions with user info
        submissions = Submission.query.order_by(Submission.created_at.desc()).all()
        
        documents = []
        for submission in submissions:
            mt = submission.module_type or ''
            if mt in STOREKEEPING_MODULE_TYPES:
                continue
            # Get user info
            user = User.query.get(submission.user_id) if submission.user_id else None
            
            # Get completed jobs for this submission
            jobs = Job.query.filter_by(
                submission_id=submission.id,
                status='completed'
            ).all()
            
            # Extract document URLs from jobs
            excel_url = None
            pdf_url = None
            for job in jobs:
                if job.result_data:
                    excel_url = job.result_data.get('excel_url') or job.result_data.get('excel')
                    pdf_url = job.result_data.get('pdf_url') or job.result_data.get('pdf')
                    if excel_url or pdf_url:
                        break
            
            # If no URLs in jobs, check files table
            if not excel_url and not pdf_url:
                report_files = File.query.filter_by(
                    submission_id=submission.id
                ).filter(
                    File.file_type.in_(['report_excel', 'report_pdf'])
                ).all()
                
                for file in report_files:
                    if file.file_type == 'report_excel' and file.cloud_url:
                        excel_url = file.cloud_url
                    elif file.file_type == 'report_pdf' and file.cloud_url:
                        pdf_url = file.cloud_url
            
            # Format module type for display (HR uses same labels as HR module)
            module_display = get_module_display_name(mt)

            doc_labels = build_document_labels(submission, module_display)
            
            documents.append({
                'id': submission.id,
                'doc_number': submission.doc_number or '',
                'submission_id': submission.submission_id,
                'module_type': submission.module_type,
                'module_display': module_display,
                'document_title': doc_labels['document_title'],
                'document_subtitle': doc_labels['document_subtitle'],
                'document_ref': doc_labels['document_ref'],
                'site_name': submission.site_name or 'N/A',
                'visit_date': submission.visit_date.isoformat() if submission.visit_date else None,
                'status': submission.status,
                'created_at': submission.created_at.isoformat() + 'Z' if submission.created_at else None,  # Add 'Z' to indicate UTC
                'created_at_timestamp': submission.created_at.timestamp() if submission.created_at else None,
                'user': {
                    'id': user.id if user else None,
                    'username': user.username if user else 'Unknown',
                    'full_name': user.full_name if user else None,
                    'email': user.email if user else 'N/A',
                    'designation': user.designation if user and hasattr(user, 'designation') else None
                },
                'workflow_status': submission.workflow_status if hasattr(submission, 'workflow_status') else 'submitted',
                'supervisor_id': submission.supervisor_id if hasattr(submission, 'supervisor_id') else None,
                'manager_id': submission.manager_id if hasattr(submission, 'manager_id') else None,
                'supervisor_notified_at': submission.supervisor_notified_at.isoformat() + 'Z' if hasattr(submission, 'supervisor_notified_at') and submission.supervisor_notified_at else None,
                'supervisor_reviewed_at': submission.supervisor_reviewed_at.isoformat() + 'Z' if hasattr(submission, 'supervisor_reviewed_at') and submission.supervisor_reviewed_at else None,
                'manager_notified_at': submission.manager_notified_at.isoformat() + 'Z' if hasattr(submission, 'manager_notified_at') and submission.manager_notified_at else None,
                'manager_reviewed_at': submission.manager_reviewed_at.isoformat() + 'Z' if hasattr(submission, 'manager_reviewed_at') and submission.manager_reviewed_at else None,
                'excel_url': excel_url,
                'pdf_url': pdf_url,
                'has_documents': bool(excel_url or pdf_url)
            })
        
        return success_response({
            'documents': documents,
            'count': len(documents)
        })
    except Exception as e:
        current_app.logger.error(f"Error listing documents: {str(e)}", exc_info=True)
        return error_response('Failed to fetch documents', status_code=500, error_code='DATABASE_ERROR')


def log_audit(user_id, action, resource_type=None, resource_id=None, details=None):
    """Create audit log entry"""
    try:
        log = AuditLog(
            user_id=user_id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent'),
            details=details
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        current_app.logger.error(f"Failed to create audit log: {str(e)}")
        db.session.rollback()


@admin_bp.route('/users/<int:user_id>/designation', methods=['PUT'])
@jwt_required()
@admin_required
def set_user_designation(user_id):
    """Set user designation for new 5-stage workflow"""
    try:
        admin_id = get_jwt_identity()
        data = request.get_json()
        designation = data.get('designation')
        
        # New valid designations for 5-stage workflow
        valid_designations = [
            'supervisor',           # Stage 1: Creates and submits forms
            'operations_manager',   # Stage 2: First approval
            'business_development', # Stage 3: Parallel review
            'procurement',          # Stage 3: Parallel review
            'general_manager',      # Stage 4: Final approval
            'hr_manager',           # HR module lead
            'employee',             # Staff (organizational label)
            'technician',           # Field technician — HR routing: supervisor → OM → GM → HR HO
            'admin',                # Organizational admin label (distinct from role=admin)
            None                    # No designation (regular user)
        ]
        
        if designation not in valid_designations:
            return error_response(
                'Invalid designation. Must be one of: supervisor, operations_manager, business_development, procurement, general_manager, hr_manager, employee, technician, admin, or null', 
                status_code=400, 
                error_code='VALIDATION_ERROR'
            )
        
        user = User.query.get_or_404(user_id)
        old_designation = user.designation
        user.designation = designation
        
        db.session.commit()
        
        log_audit(admin_id, 'set_designation', 'user', str(user_id), {
            'target_user': user.username,
            'old_designation': old_designation,
            'new_designation': designation,
            'set_by': User.query.get(admin_id).username if admin_id else 'system'
        })
        
        return success_response({
            'user': _admin_user_dict(user),
            'message': f'Designation updated to {designation or "None"}'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error setting designation: {str(e)}", exc_info=True)
        return error_response('Failed to set designation', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/submissions/<submission_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_submission(submission_id):
    """Update a submitted form (admin can modify any field) and regenerate documents"""
    try:
        admin_id = get_jwt_identity()
        data = request.get_json()
        
        from app.models import Submission, Job
        from datetime import datetime
        import os
        
        submission = Submission.query.filter_by(submission_id=submission_id).first_or_404()
        
        # Update form_data
        if 'form_data' in data:
            submission.form_data = data['form_data']
        
        # Update other fields if provided
        if 'site_name' in data:
            submission.site_name = data['site_name']
        if 'visit_date' in data:
            try:
                submission.visit_date = datetime.strptime(data['visit_date'], '%Y-%m-%d').date()
            except ValueError:
                return error_response('Invalid date format. Use YYYY-MM-DD', status_code=400, error_code='VALIDATION_ERROR')
        
        submission.updated_at = utc_now_naive()
        db.session.commit()
        
        # Delete old jobs and their associated files to force regeneration
        old_jobs = Job.query.filter_by(submission_id=submission.id).all()
        for old_job in old_jobs:
            # Delete old generated files if they exist locally
            if old_job.result_data:
                results = old_job.result_data if isinstance(old_job.result_data, dict) else {}
                excel_filename = results.get('excel_filename') or results.get('excel')
                pdf_filename = results.get('pdf_filename') or results.get('pdf')
                
                GENERATED_DIR = current_app.config.get('GENERATED_DIR')
                if GENERATED_DIR:
                    if excel_filename and isinstance(excel_filename, str):
                        excel_path = os.path.join(GENERATED_DIR, excel_filename)
                        if os.path.exists(excel_path):
                            try:
                                os.remove(excel_path)
                                current_app.logger.info(f"Deleted old Excel file: {excel_filename}")
                            except Exception as e:
                                current_app.logger.warning(f"Could not delete old Excel file: {e}")
                    
                    if pdf_filename and isinstance(pdf_filename, str):
                        pdf_path = os.path.join(GENERATED_DIR, pdf_filename)
                        if os.path.exists(pdf_path):
                            try:
                                os.remove(pdf_path)
                                current_app.logger.info(f"Deleted old PDF file: {pdf_filename}")
                            except Exception as e:
                                current_app.logger.warning(f"Could not delete old PDF file: {e}")
            
            db.session.delete(old_job)
        
        db.session.commit()
        
        # Trigger document regeneration for unified inspection (+ legacy types)
        job_id = None
        if submission.module_type in ('inspection', 'hvac_mep', 'civil', 'cleaning'):
            from common.db_utils import create_job_db
            from module_inspection.routes import get_paths, process_job

            new_job = create_job_db(submission)
            job_id = new_job.job_id

            GENERATED_DIR, UPLOADS_DIR, JOBS_DIR, EXECUTOR = get_paths()

            if EXECUTOR:
                EXECUTOR.submit(
                    process_job,
                    submission.submission_id,
                    job_id,
                    current_app.config,
                    current_app._get_current_object()
                )
                current_app.logger.info(f"✅ Regeneration job {job_id} queued for submission {submission_id}")
            else:
                current_app.logger.error("ThreadPoolExecutor not available for document regeneration")

        log_audit(admin_id, 'update_submission', 'submission', submission_id, {
            'site_name': submission.site_name,
            'module_type': submission.module_type,
            'updated_by': User.query.get(admin_id).username if admin_id else 'system',
            'regeneration_job_id': job_id
        })
        
        return success_response({
            'submission': submission.to_dict(),
            'message': 'Submission updated successfully. Documents are being regenerated.',
            'job_id': job_id,
            'regenerating': True
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating submission: {str(e)}", exc_info=True)
        return error_response('Failed to update submission', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/submissions/<submission_id>/close', methods=['POST'])
@jwt_required()
@admin_required
def close_submission(submission_id):
    """Close a submission from admin side (read-only for all users)"""
    try:
        admin_id = get_jwt_identity()
        data = request.get_json() or {}
        reason = (data.get('reason') or '').strip()

        from app.models import Submission

        submission = Submission.query.filter_by(submission_id=submission_id).first_or_404()
        if submission.workflow_status == 'closed_by_admin':
            return success_response({'submission_id': submission_id}, message='Submission already closed')

        # Update workflow/status fields
        submission.workflow_status = 'closed_by_admin'
        submission.status = 'closed'

        # Store close metadata in form_data (no schema change required).
        raw_fd = submission.form_data
        if isinstance(raw_fd, str):
            try:
                form_data = dict(json.loads(raw_fd))
            except Exception:
                form_data = {}
        else:
            form_data = shallow_copy_form_data(submission)

        form_data['_admin_closed'] = {
            'closed_at': utc_now_naive().isoformat() + 'Z',
            'closed_by': admin_id,
            'reason': reason or 'Closed by admin'
        }
        submission.form_data = form_data

        db.session.commit()

        log_audit(admin_id, 'close_submission', 'submission', submission_id, {
            'reason': reason or 'Closed by admin'
        })

        return success_response({'submission_id': submission_id}, message='Submission closed by admin')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error closing submission: {str(e)}", exc_info=True)
        return error_response('Failed to close submission', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/submissions/<submission_id>', methods=['GET'])
@jwt_required()
@admin_required
def get_submission(submission_id):
    """Get submission details for editing"""
    try:
        from app.models import Submission
        submission = Submission.query.filter_by(submission_id=submission_id).first_or_404()
        
        # Get user info
        user = User.query.get(submission.user_id) if submission.user_id else None
        supervisor = User.query.get(submission.supervisor_id) if hasattr(submission, 'supervisor_id') and submission.supervisor_id else None
        manager = User.query.get(submission.manager_id) if hasattr(submission, 'manager_id') and submission.manager_id else None
        
        submission_dict = submission.to_dict()
        submission_dict['user'] = user.to_dict() if user else None
        submission_dict['supervisor'] = supervisor.to_dict() if supervisor else None
        submission_dict['manager'] = manager.to_dict() if manager else None
        
        return success_response({'submission': submission_dict})
    except Exception as e:
        current_app.logger.error(f"Error getting submission: {str(e)}", exc_info=True)
        return error_response('Failed to get submission', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/users/by-designation/<designation>', methods=['GET'])
@jwt_required()
@admin_required
def get_users_by_designation(designation):
    """Get all users with a specific designation"""
    try:
        valid_designations = [
            'supervisor',
            'operations_manager',
            'business_development',
            'procurement',
            'general_manager',
            'hr_manager',
            'employee',
            'technician',
            'admin',
        ]
        
        if designation not in valid_designations:
            return error_response(
                'Invalid designation',
                status_code=400,
                error_code='VALIDATION_ERROR'
            )
        
        users = User.query.filter_by(designation=designation, is_active=True).all()
        
        return success_response({
            'users': [user.to_dict() for user in users],
            'count': len(users),
            'designation': designation
        })
    except Exception as e:
        current_app.logger.error(f"Error getting users by designation: {str(e)}", exc_info=True)
        return error_response('Failed to get users', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/workflow/stats', methods=['GET'])
@jwt_required()
@admin_required
def get_workflow_stats():
    """Get workflow statistics for dashboard"""
    try:
        from app.models import Submission
        
        # Count submissions by workflow status
        stats = {
            'total_submissions': Submission.query.count(),
            'by_status': {},
            'by_designation': {}
        }
        
        # Count by workflow status
        statuses = [
            'submitted',
            'operations_manager_review',
            'operations_manager_approved',
            'bd_procurement_review',
            'general_manager_review',
            'completed',
            'rejected'
        ]
        
        for status in statuses:
            count = Submission.query.filter_by(workflow_status=status).count()
            stats['by_status'][status] = count
        
        # Count users by designation
        designations = [
            'supervisor',
            'operations_manager',
            'business_development',
            'procurement',
            'general_manager',
            'hr_manager',
            'employee',
            'technician',
            'admin',
        ]
        
        for designation in designations:
            count = User.query.filter_by(designation=designation, is_active=True).count()
            stats['by_designation'][designation] = count
        
        return success_response(stats)
    except Exception as e:
        current_app.logger.error(f"Error getting workflow stats: {str(e)}", exc_info=True)
        return error_response('Failed to get statistics', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/dashboard-overview', methods=['GET'])
@jwt_required()
@admin_required
def dashboard_overview():
    """Single aggregated payload for the admin dashboard (users, submissions, devices, BD, DocHub, audit)."""
    try:
        from app.models import Submission
        from sqlalchemy import func, or_, and_

        # --- Users ---
        vis = ~User.username.in_(DIRECTORY_HIDDEN_USERNAMES)
        user_total = User.query.filter(vis).count()
        user_active = User.query.filter(vis, User.is_active.is_(True)).count()
        user_inactive = max(0, user_total - user_active)
        role_rows = db.session.query(User.role, func.count(User.id)).group_by(User.role).all()
        by_role = {str(r or 'unknown'): c for r, c in role_rows}
        default_password_count = User.query.filter_by(password_changed=False).count()
        by_designation_active = db.session.query(
            User.designation, func.count(User.id)
        ).filter(
            User.is_active.is_(True),
            User.designation.isnot(None),
            User.designation != ''
        ).group_by(User.designation).all()
        designation_breakdown = {str(d or 'unknown'): c for d, c in by_designation_active}

        # --- Submissions (inspection documents) ---
        sub_total = Submission.query.count()
        mod_rows = db.session.query(Submission.module_type, func.count(Submission.id)).group_by(
            Submission.module_type
        ).all()
        by_module = {str(m or 'unknown'): c for m, c in mod_rows}

        pipeline_open = Submission.query.filter(
            and_(
                or_(
                    Submission.workflow_status.is_(None),
                    ~Submission.workflow_status.in_(['completed', 'closed_by_admin', 'rejected'])
                ),
                or_(
                    Submission.status.is_(None),
                    ~Submission.status.in_(['completed', 'closed'])
                )
            )
        ).count()

        completed_closed = Submission.query.filter(
            or_(
                Submission.workflow_status.in_(['completed', 'closed_by_admin']),
                Submission.status.in_(['completed', 'closed'])
            )
        ).count()

        rejected_count = Submission.query.filter(Submission.workflow_status == 'rejected').count()

        in_review_notified = Submission.query.filter(
            or_(
                Submission.workflow_status.like('%reviewing%'),
                Submission.workflow_status.like('%notified%')
            )
        ).count()

        ws_rows = db.session.query(Submission.workflow_status, func.count(Submission.id)).group_by(
            Submission.workflow_status
        ).all()
        workflow_status_breakdown = {str(ws or 'unknown'): c for ws, c in ws_rows}

        # --- Devices ---
        dev_total = Device.query.count()
        dev_online = Device.query.filter_by(status='online').count()
        dev_offline = Device.query.filter_by(status='offline').count()
        dev_pending = Device.query.filter_by(status='update').count()

        employee_active = User.query.filter(
            User.is_active.is_(True),
            User.designation == 'employee',
        ).count()

        hr_form_count = Submission.query.filter(Submission.module_type.startswith('hr_')).count()
        inspection_module_types = ('hvac_mep', 'civil', 'cleaning')
        inspection_form_count = Submission.query.filter(
            Submission.module_type.in_(inspection_module_types)
        ).count()

        # --- Business development (light aggregates) ---
        bd_projects = BDProject.query.count()
        bd_pipeline_value = float(db.session.query(func.coalesce(func.sum(BDProject.value_amount), 0)).scalar() or 0)
        bd_active = BDProject.query.filter(BDProject.status.in_(['active', 'proposal', 'prospect'])).count()
        bd_contacts = BDContact.query.count()
        now = utc_now_naive()
        bd_overdue_fu = BDFollowUp.query.filter(
            BDFollowUp.status != 'done',
            BDFollowUp.due_at.isnot(None),
            BDFollowUp.due_at < now
        ).count()
        bd_open_fu = BDFollowUp.query.filter(BDFollowUp.status != 'done').count()
        won = BDProject.query.filter_by(status='won').count()
        lost = BDProject.query.filter_by(status='lost').count()
        bd_win_rate = int(round((won / (won + lost)) * 100)) if (won + lost) > 0 else 0

        # --- DocHub & admin tools ---
        dochub_access_grants = DocHubAccess.query.filter_by(can_access=True).count()
        pp_projects = AdminPersonalProject.query.count()

        # --- Recent audit (security / visibility) ---
        AUDIT_RECENT_LIMIT = 100
        audit_total = AuditLog.query.count()
        recent_logs = (
            AuditLog.query
            .order_by(AuditLog.created_at.desc())
            .limit(AUDIT_RECENT_LIMIT)
            .all()
        )
        user_ids = {log.user_id for log in recent_logs if log.user_id}
        users_by_id = {}
        if user_ids:
            users_by_id = {
                u.id: u.username
                for u in User.query.filter(User.id.in_(user_ids)).all()
            }
        audit_recent = []
        for log in recent_logs:
            audit_recent.append({
                'id': log.id,
                'action': log.action,
                'resource_type': log.resource_type,
                'resource_id': log.resource_id,
                'created_at': log.created_at.isoformat() if log.created_at else None,
                'username': users_by_id.get(log.user_id) or '—',
            })

        # --- Service tickets / work orders (global admin view) ---
        ticketing_snapshot = {'available': False}
        try:
            from app.models import Ticket as _TkModel

            TERMINAL = ('closed', 'resolved', 'cancelled')
            total_tk_all = db.session.query(func.count(_TkModel.id)).scalar() or 0
            active_pipeline = db.session.query(func.count(_TkModel.id)).filter(
                ~_TkModel.status.in_(TERMINAL),
            ).scalar() or 0
            supervisor_queue = db.session.query(func.count(_TkModel.id)).filter(
                _TkModel.status.in_(('open', 'pending_supervisor')),
            ).scalar() or 0
            field_active = db.session.query(func.count(_TkModel.id)).filter(
                _TkModel.status.in_((
                    'assigned', 'site_attended', 'work_started', 'in_progress',
                    'pending_parts',
                )),
            ).scalar() or 0
            verification_lane = db.session.query(func.count(_TkModel.id)).filter(
                _TkModel.status.in_(('pending_verification', 'verification', 'work_completed')),
            ).scalar() or 0
            urgent_ct = db.session.query(func.count(_TkModel.id)).filter(
                ~_TkModel.status.in_(TERMINAL),
                _TkModel.priority.in_(('critical', 'high')),
            ).scalar() or 0
            on_hold_ct = db.session.query(func.count(_TkModel.id)).filter_by(status='on_hold').scalar() or 0

            recent_rows = _TkModel.query.order_by(_TkModel.updated_at.desc()).limit(7).all()
            recent_tickets_admin = []
            for t in recent_rows:
                title = (t.title or '').strip()
                if len(title) > 92:
                    title = title[:90] + '…'
                recent_tickets_admin.append({
                    'ticket_id': t.ticket_id,
                    'title': title,
                    'status': t.status,
                    'priority': t.priority,
                    'project': (t.project or '')[:52],
                    'updated_at': t.updated_at.isoformat() if t.updated_at else None,
                })

            ticketing_snapshot = {
                'available': True,
                'total': int(total_tk_all),
                'active_pipeline': int(active_pipeline),
                'supervisor_queue': int(supervisor_queue),
                'field_active': int(field_active),
                'awaiting_verification': int(verification_lane),
                'urgent_priority': int(urgent_ct),
                'on_hold': int(on_hold_ct),
                'recent': recent_tickets_admin,
            }
        except Exception as tick_exc:
            current_app.logger.warning('dashboard ticketing aggregate skipped: %s', tick_exc)

        return success_response({
            'generated_at': utc_now_naive().isoformat() + 'Z',
            'users': {
                'total': user_total,
                'active': user_active,
                'inactive': user_inactive,
                'by_role': by_role,
                'designation_active': designation_breakdown,
                'default_password_count': default_password_count,
            },
            'submissions': {
                'total': sub_total,
                'by_module': by_module,
                'pipeline_open': pipeline_open,
                'completed_or_closed': completed_closed,
                'rejected': rejected_count,
                'in_review_or_notified': in_review_notified,
                'workflow_status_breakdown': workflow_status_breakdown,
            },
            'devices': {
                'total': dev_total,
                'online': dev_online,
                'offline': dev_offline,
                'pending_updates': dev_pending,
            },
            'side_summaries': {
                'users_total': user_total,
                'employees_active': employee_active,
                'hr_forms_total': hr_form_count,
                'inspection_forms_total': inspection_form_count,
                'devices_total': dev_total,
                'pipeline_open': pipeline_open,
                'tickets_active': (
                    ticketing_snapshot.get('active_pipeline')
                    if ticketing_snapshot.get('available') else None
                ),
                'bd_projects': bd_projects,
                'bd_followups_open': bd_open_fu,
                'audit_total': audit_total,
            },
            'bd': {
                'projects_total': bd_projects,
                'pipeline_value': bd_pipeline_value,
                'active_deals': bd_active,
                'contacts': bd_contacts,
                'followups_open': bd_open_fu,
                'followups_overdue': bd_overdue_fu,
                'win_rate': bd_win_rate,
            },
            'tools': {
                'dochub_access_grants': dochub_access_grants,
                'personal_progress_projects': pp_projects,
            },
            'audit_recent': audit_recent,
            'audit_total': audit_total,
            'ticketing': ticketing_snapshot,
        })
    except Exception as e:
        current_app.logger.error(f"Error building dashboard overview: {str(e)}", exc_info=True)
        return error_response('Failed to load dashboard overview', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/designations', methods=['GET'])
@jwt_required()
@admin_required
def get_valid_designations():
    """Get list of valid designations with descriptions"""
    try:
        designations = [
            {
                'value': 'supervisor',
                'label': 'Supervisor/Inspector',
                'description': 'Stage 1: Creates and submits forms',
                'stage': 1
            },
            {
                'value': 'operations_manager',
                'label': 'Operations Manager',
                'description': 'Stage 2: First approval level',
                'stage': 2
            },
            {
                'value': 'business_development',
                'label': 'Business Development',
                'description': 'Stage 3: Parallel review with Procurement',
                'stage': 3
            },
            {
                'value': 'procurement',
                'label': 'Procurement',
                'description': 'Stage 3: Parallel review with Business Development',
                'stage': 3
            },
            {
                'value': 'general_manager',
                'label': 'General Manager',
                'description': 'Stage 4: Final approval',
                'stage': 4
            },
            {
                'value': 'hr_manager',
                'label': 'HR Manager',
                'description': 'HR module workflow lead',
                'stage': 0
            },
            {
                'value': 'employee',
                'label': 'Employee',
                'description': 'Staff — no document workflow reviewer role',
                'stage': 0
            },
            {
                'value': 'technician',
                'label': 'Technician',
                'description': 'Field technician — HR forms route via supervisor chain on the PDF trail',
                'stage': 0
            },
            {
                'value': 'admin',
                'label': 'Admin',
                'description': 'Organizational admin label (separate from system Administrator role)',
                'stage': 0
            }
        ]
        
        return success_response({
            'designations': designations,
            'count': len(designations)
        })
    except Exception as e:
        current_app.logger.error(f"Error getting designations: {str(e)}", exc_info=True)
        return error_response('Failed to get designations', status_code=500, error_code='DATABASE_ERROR')


# ============== Device Management (Admin only) ==============


@admin_bp.route('/dochub/access-users', methods=['GET'])
@jwt_required()
@admin_required
def list_dochub_access_users():
    """List users with DocHub access flags (admin control)."""
    try:
        users = User.query.order_by(User.full_name.asc(), User.username.asc()).all()
        access_map = {
            row.user_id: row.can_access
            for row in DocHubAccess.query.all()
        }

        data = []
        for u in users:
            data.append({
                'id': u.id,
                'username': u.username,
                'full_name': u.full_name,
                'email': u.email,
                'role': u.role,
                'is_active': u.is_active,
                'can_access_dochub': True if u.role == 'admin' else access_map.get(u.id, True)
            })
        return success_response({'users': data, 'count': len(data)})
    except Exception as e:
        current_app.logger.error(f"Error listing DocHub access users: {str(e)}", exc_info=True)
        return error_response('Failed to fetch users', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/dochub/access-users/<int:user_id>', methods=['POST'])
@jwt_required()
@admin_required
def set_dochub_user_access(user_id):
    """Grant/revoke DocHub access for a user."""
    try:
        admin_id = get_jwt_identity()
        user = User.query.get_or_404(user_id)
        data = request.get_json() or {}
        can_access = bool(data.get('can_access', True))

        if user.role == 'admin':
            return error_response('Admin access cannot be revoked', status_code=400, error_code='VALIDATION_ERROR')

        row = DocHubAccess.query.filter_by(user_id=user.id).first()
        if row:
            row.can_access = can_access
            row.updated_by = admin_id
        else:
            db.session.add(DocHubAccess(
                user_id=user.id,
                can_access=can_access,
                updated_by=admin_id
            ))

        db.session.commit()
        return success_response({
            'user_id': user.id,
            'can_access_dochub': can_access
        }, message='DocHub access updated')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error setting DocHub user access: {str(e)}", exc_info=True)
        return error_response('Failed to update access', status_code=500, error_code='DATABASE_ERROR')


def _normalize_device_excel_column_name(value):
    raw = str(value or '').strip().lower()
    cleaned = ''.join(ch if ch.isalnum() else ' ' for ch in raw)
    return ' '.join(cleaned.split())

@admin_bp.route('/devices', methods=['GET'])
@jwt_required()
@admin_required
def list_devices():
    """Get all registered devices"""
    try:
        devices = Device.query.order_by(Device.created_at.desc()).all()
        devices_data = [d.to_dict() for d in devices]
        return success_response({
            'devices': devices_data,
            'count': len(devices_data)
        })
    except Exception as e:
        current_app.logger.error(f"Error listing devices: {str(e)}", exc_info=True)
        return error_response('Failed to fetch devices', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/devices', methods=['POST'])
@jwt_required()
@admin_required
def create_device():
    """Enroll a new device"""
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        device_type = (data.get('device_type') or 'Laptop').strip()
        os = (data.get('os') or 'Windows 11').strip()
        user_email = (data.get('assigned_user_email') or '').strip()
        serial = (data.get('serial_or_asset_tag') or '').strip()

        if not name:
            return error_response('Device name is required', status_code=400, error_code='VALIDATION_ERROR')

        assigned_user_id = None
        if user_email:
            user = User.query.filter_by(email=user_email).first()
            if user:
                assigned_user_id = user.id

        # Generate unique device_id
        import random
        existing_ids = {d.device_id for d in Device.query.with_entities(Device.device_id).all()}
        for _ in range(50):
            dev_id = 'DEV-' + str(random.randint(1000, 9999))
            if dev_id not in existing_ids:
                break
        else:
            dev_id = 'DEV-' + str(random.randint(10000, 99999))

        device = Device(
            device_id=dev_id,
            name=name,
            device_type=device_type,
            os=os,
            status='idle',
            health=random.randint(80, 100),
            assigned_user_id=assigned_user_id,
            serial_or_asset_tag=serial or None,
            last_active_at=utc_now_naive()
        )
        db.session.add(device)
        db.session.commit()

        return success_response({
            'device': device.to_dict(),
            'message': f'Device "{name}" enrolled successfully'
        }, status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating device: {str(e)}", exc_info=True)
        return error_response('Failed to enroll device', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/devices/<int:id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_device(id):
    """Remove a device"""
    try:
        device = Device.query.get_or_404(id)
        name = device.name
        db.session.delete(device)
        db.session.commit()
        return success_response({'message': f'Device "{name}" removed successfully'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting device: {str(e)}", exc_info=True)
        return error_response('Failed to remove device', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/devices/stats', methods=['GET'])
@jwt_required()
@admin_required
def device_stats():
    """Get device statistics for dashboard"""
    try:
        total = Device.query.count()
        online = Device.query.filter_by(status='online').count()
        offline = Device.query.filter_by(status='offline').count()
        pending_updates = Device.query.filter_by(status='update').count()
        return success_response({
            'total': total,
            'online': online,
            'offline': offline,
            'pending_updates': pending_updates
        })
    except Exception as e:
        current_app.logger.error(f"Error getting device stats: {str(e)}", exc_info=True)
        return error_response('Failed to get statistics', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/devices/import-excel', methods=['POST'])
@jwt_required()
@admin_required
def import_devices_excel():
    """Bulk import devices from Excel file."""
    try:
        if 'file' not in request.files:
            return error_response('No file provided', status_code=400, error_code='VALIDATION_ERROR')

        file = request.files['file']
        if not file or not file.filename:
            return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')

        filename = file.filename.lower()
        if not filename.endswith(('.xlsx', '.xls')):
            return error_response('Invalid file format. Upload .xlsx or .xls', status_code=400, error_code='VALIDATION_ERROR')

        try:
            import pandas as pd
        except ImportError:
            return error_response(
                'Excel import requires pandas/openpyxl/xlrd dependencies',
                status_code=500,
                error_code='DEPENDENCY_ERROR'
            )

        # Read file robustly (.xlsx, .xls, or HTML-based .xls)
        try:
            from common.kynvera_excel_brand import read_import_dataframe

            df = read_import_dataframe(file, preferred_sheets=('Devices',))
        except Exception as read_error:
            current_app.logger.error(f"Device Excel import read error: {read_error}", exc_info=True)
            return error_response(f'Could not parse Excel file: {read_error}', status_code=400, error_code='VALIDATION_ERROR')

        if df is None or df.empty:
            return error_response('Excel file is empty', status_code=400, error_code='VALIDATION_ERROR')

        # Flatten multi-level headers if present
        if hasattr(df.columns, 'levels'):
            flat_cols = []
            for col in df.columns:
                if isinstance(col, tuple):
                    flat_cols.append(' '.join([str(part).strip() for part in col if str(part).strip() and str(part).strip().lower() != 'nan']))
                else:
                    flat_cols.append(str(col))
            df.columns = flat_cols

        normalized_cols = {_normalize_device_excel_column_name(c): c for c in df.columns}
        alias_map = {
            'name': 'name',
            'device': 'name',
            'device name': 'name',
            'device type': 'device_type',
            'type': 'device_type',
            'os': 'os',
            'operating system': 'os',
            'status': 'status',
            'health': 'health',
            'health percent': 'health',
            'assigned user email': 'assigned_user_email',
            'user email': 'assigned_user_email',
            'email': 'assigned_user_email',
            'serial': 'serial_or_asset_tag',
            'serial number': 'serial_or_asset_tag',
            'asset tag': 'serial_or_asset_tag',
            'serial or asset tag': 'serial_or_asset_tag',
        }

        canonical_to_original = {}
        for normalized_name, original_name in normalized_cols.items():
            canonical_name = alias_map.get(normalized_name)
            if canonical_name:
                canonical_to_original[canonical_name] = original_name

        if 'name' not in canonical_to_original:
            return error_response(
                'Excel must include a "Device Name" (or "Name") column',
                status_code=400,
                error_code='VALIDATION_ERROR'
            )

        # Existing keys to avoid duplicates
        existing_devices = Device.query.with_entities(Device.name, Device.serial_or_asset_tag).all()
        existing_keys = {
            f"{(name or '').strip().lower()}|{(serial or '').strip().lower()}"
            for name, serial in existing_devices
        }
        existing_ids = {d.device_id for d in Device.query.with_entities(Device.device_id).all()}

        def cell(row, canonical_name, default=''):
            col = canonical_to_original.get(canonical_name)
            if not col:
                return default
            return row.get(col, default)

        def safe_int(value, default=0):
            try:
                if pd.isna(value):
                    return int(default)
                return int(float(value))
            except Exception:
                return int(default)

        def normalize_status(value):
            s = str(value or '').strip().lower()
            if s in ['online', 'offline', 'update', 'idle']:
                return s
            if 'warn' in s or 'update' in s:
                return 'update'
            if 'off' in s:
                return 'offline'
            if 'on' in s:
                return 'online'
            return 'idle'

        imported = 0
        skipped_duplicates = 0
        skipped_empty = 0
        errors = []

        import random
        for idx, row in df.iterrows():
            try:
                name = str(cell(row, 'name', '')).strip()
                if not name or name.lower() == 'nan':
                    skipped_empty += 1
                    continue

                device_type = str(cell(row, 'device_type', 'Laptop')).strip() or 'Laptop'
                os = str(cell(row, 'os', 'Windows 11')).strip() or 'Windows 11'
                status = normalize_status(cell(row, 'status', 'idle'))
                health = max(0, min(100, safe_int(cell(row, 'health', random.randint(75, 100)), random.randint(75, 100))))
                user_email = str(cell(row, 'assigned_user_email', '')).strip()
                serial = str(cell(row, 'serial_or_asset_tag', '')).strip()
                if serial.lower() == 'nan':
                    serial = ''

                dup_key = f"{name.lower()}|{serial.lower()}"
                if dup_key in existing_keys:
                    skipped_duplicates += 1
                    continue

                assigned_user_id = None
                if user_email and user_email.lower() != 'nan':
                    matched_user = User.query.filter_by(email=user_email).first()
                    if matched_user:
                        assigned_user_id = matched_user.id

                # Generate unique device_id
                for _ in range(80):
                    dev_id = 'DEV-' + str(random.randint(1000, 9999))
                    if dev_id not in existing_ids:
                        break
                else:
                    dev_id = 'DEV-' + str(random.randint(10000, 99999))
                existing_ids.add(dev_id)

                device = Device(
                    device_id=dev_id,
                    name=name,
                    device_type=device_type,
                    os=os,
                    status=status,
                    health=health,
                    assigned_user_id=assigned_user_id,
                    serial_or_asset_tag=serial or None,
                    last_active_at=utc_now_naive()
                )
                db.session.add(device)
                existing_keys.add(dup_key)
                imported += 1
            except Exception as row_error:
                errors.append(f"Row {idx + 2}: {row_error}")

        db.session.commit()
        return success_response({
            'imported': imported,
            'skipped_duplicates': skipped_duplicates,
            'skipped_empty': skipped_empty,
            'total_rows': int(len(df)),
            'errors': errors[:15]
        }, message=f'Imported {imported} device(s)')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importing devices from Excel: {str(e)}", exc_info=True)
        return error_response('Failed to import devices', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/devices/sample-excel', methods=['GET'])
@jwt_required()
@admin_required
def download_devices_sample_excel():
    """Download a sample Excel file with multiple device rows."""
    try:
        from app.admin.excel_templates import build_devices_sample_bytes

        output = BytesIO(build_devices_sample_bytes())
        filename = f"device_import_sample_{utc_now_naive().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        current_app.logger.error(f"Error generating sample device Excel: {str(e)}", exc_info=True)
        return error_response('Failed to generate sample Excel', status_code=500, error_code='DATABASE_ERROR')


def _parse_iso_date(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value)).date()
    except Exception:
        return None


def _parse_iso_datetime(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace('Z', '+00:00')).replace(tzinfo=None)
    except Exception:
        return None


def _bd_activity(icon, title, description='', badge='', bg='#fff4ef', event_time=None, user_id=None):
    activity = BDActivity(
        icon=icon,
        bg=bg,
        title=title,
        description=description,
        badge=badge,
        event_time=event_time or utc_now_naive(),
        created_by=user_id
    )
    db.session.add(activity)


def _bd_dt_24(dt=None):
    """Format a datetime as YYYY-MM-DD HH:MM (24h) for timeline / quote logs."""
    value = dt or utc_now_naive()
    try:
        return value.strftime('%Y-%m-%d %H:%M')
    except Exception:
        return ''


def _normalize_excel_column_name(value):
    raw = str(value or '').strip().lower()
    cleaned = ''.join(ch if ch.isalnum() else ' ' for ch in raw)
    return ' '.join(cleaned.split())


def _status_stage_progress_from_contract_status(status_text):
    s = (status_text or '').strip().lower()
    if 'active' in s:
        return 'active', 'negotiation', 70
    if 'expired' in s:
        return 'lost', 'closing', 100
    if 'draft' in s:
        return 'prospect', 'prospecting', 15
    if 'won' in s:
        return 'won', 'closing', 100
    if 'proposal' in s:
        return 'proposal', 'proposal', 55
    return 'prospect', 'qualifying', 25


def _parse_excel_date(value):
    try:
        import pandas as pd
        dt = pd.to_datetime(value, errors='coerce')
        if pd.isna(dt):
            return None
        return dt.date()
    except Exception:
        return None


def _parse_excel_float(value, default=0.0):
    try:
        import pandas as pd
        if pd.isna(value):
            return float(default)
        if isinstance(value, str):
            value = value.replace(',', '').strip()
        return float(value)
    except Exception:
        return float(default)


def _seed_bd_data_if_empty(user_id):
    """Top up the BD pipeline with sample deals, quotes, follow-ups, and contacts."""
    from app.bd.sample_data import ensure_bd_sample_pipeline
    ensure_bd_sample_pipeline(user_id)


@admin_bp.route('/bd/dashboard-data', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_dashboard_data():
    """Get all BD dashboard data (scoped to salesperson unless elevated)."""
    try:
        user = _bd_current_user()
        user_id = user.id if user else get_jwt_identity()
        _seed_bd_data_if_empty(user_id)

        view = (request.args.get('view') or 'all').strip().lower()
        owner_id = request.args.get('owner_id', type=int)
        projects_q = _bd_scoped_projects_query(user).order_by(BDProject.updated_at.desc())
        if user_sees_all_bd_deals(user) and owner_id:
            projects_q = projects_q.filter(BDProject.owner_user_id == owner_id)
        projects = projects_q.all()

        today = date.today()
        renewal_cutoff = today + timedelta(days=45)
        for p in projects:
            if (p.status or '') in ('lost',):
                continue
            linked = TicketProject.query.filter_by(bd_project_id=p.id).all()
            for tp in linked:
                rd = tp.renewal_date or tp.project_end_date
                if rd and today <= rd <= renewal_cutoff and (p.status or '') != 'under_renewal':
                    p.status = 'under_renewal'
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

        if view == 'active':
            projects = [p for p in projects if (p.status or '') in ('active', 'proposal', 'prospect')]
        elif view == 'lost':
            projects = [p for p in projects if (p.status or '') == 'lost']
        elif view in ('renewal', 'under_renewal'):
            projects = [p for p in projects if (p.status or '') == 'under_renewal']

        project_ids = {p.id for p in projects}
        if user_sees_all_bd_deals(user):
            followups = BDFollowUp.query.order_by(BDFollowUp.created_at.desc()).all()
            contacts = BDContact.query.order_by(BDContact.updated_at.desc()).all()
            activities = BDActivity.query.order_by(BDActivity.event_time.desc()).limit(50).all()
        else:
            followups = BDFollowUp.query.filter(
                db.or_(BDFollowUp.created_by == user.id, BDFollowUp.project_id.in_(project_ids or [-1]))
            ).order_by(BDFollowUp.created_at.desc()).all()
            contacts = BDContact.query.filter(BDContact.created_by == user.id).order_by(BDContact.updated_at.desc()).all()
            activities = BDActivity.query.filter(BDActivity.created_by == user.id).order_by(
                BDActivity.event_time.desc()
            ).limit(50).all()

        total_value = sum(float(p.value_amount or 0) for p in projects)
        active_deals = len([p for p in projects if p.status in ['active', 'proposal', 'prospect', 'under_renewal']])
        won = len([p for p in projects if p.status == 'won'])
        lost = len([p for p in projects if p.status == 'lost'])
        under_renewal = len([p for p in projects if p.status == 'under_renewal'])
        win_rate = int(round((won / (won + lost)) * 100)) if (won + lost) > 0 else 0
        avg_deal_size = int(round(total_value / len(projects))) if projects else 0
        overdue_followups = len([
            f for f in followups
            if f.status != 'done' and f.due_at and f.due_at < utc_now_naive()
        ])

        stage_order = ['prospecting', 'qualifying', 'proposal', 'negotiation', 'closing']
        stage_stats = []
        for stage in stage_order:
            items = [p for p in projects if (p.stage or '').lower() == stage]
            stage_value = sum(float(p.value_amount or 0) for p in items)
            stage_stats.append({
                'stage': stage,
                'count': len(items),
                'value': stage_value
            })

        from app.bd import analytics as BDA
        ticket_projects = TicketProject.query.all()
        forecast = BDA.weighted_forecast(projects)
        funnel = BDA.conversion_funnel(projects)
        stalled = BDA.stalled_deals(projects)
        outcome = BDA.outcome_loop(ticket_projects)

        owners = []
        if user_sees_all_bd_deals(user):
            for u in User.query.filter(User.is_active.is_(True)).order_by(User.full_name).all():
                if user_has_bd_access(u) or u.role == 'admin':
                    owners.append({
                        'id': u.id,
                        'name': u.full_name or u.username,
                    })

        quotes_q = Quotation.query.order_by(Quotation.updated_at.desc())
        if not user_sees_all_bd_deals(user):
            quotes_q = quotes_q.filter(db.or_(
                Quotation.owner_user_id == user.id,
                Quotation.created_by_id == user.id,
            ))
        if user_sees_all_bd_deals(user) and owner_id:
            quotes_q = quotes_q.filter(Quotation.owner_user_id == owner_id)
        quotations = quotes_q.limit(200).all()

        return success_response({
            'projects': [p.to_dict() for p in projects],
            'followups': [f.to_dict() for f in followups],
            'contacts': [c.to_dict() for c in contacts],
            'activities': [a.to_dict() for a in activities],
            'owners': owners,
            'quotations': [q.to_dict(include_items=False) for q in quotations],
            'can_prepare_quotations': user_can_prepare_quotations(user),
            'scope': {
                'sees_all': user_sees_all_bd_deals(user),
                'view': view,
            },
            'stats': {
                'total_pipeline': total_value,
                'active_deals': active_deals,
                'under_renewal': under_renewal,
                'win_rate': win_rate,
                'avg_deal_size': avg_deal_size,
                'overdue_followups': overdue_followups,
                'stage_stats': stage_stats,
                'forecast': forecast,
                'funnel': funnel,
                'stalled': stalled,
                'outcome': outcome,
            }
        })
    except Exception as e:
        current_app.logger.error(f"Error fetching BD dashboard data: {str(e)}", exc_info=True)
        return error_response('Failed to fetch BD dashboard data', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/projects', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_create_project():
    """Create a BD project/deal."""
    try:
        user = _bd_current_user()
        user_id = user.id
        data = request.get_json() or {}

        name = (data.get('name') or '').strip()
        company = (data.get('company') or '').strip()
        if not name or not company:
            return error_response('Project name and company are required', status_code=400, error_code='VALIDATION_ERROR')

        owner_uid = _resolve_owner_user_id(data, default_user_id=user_id)
        if not user_sees_all_bd_deals(user):
            owner_uid = user_id

        project = BDProject(
            name=name,
            company=company,
            stage=(data.get('stage') or 'prospecting').strip().lower(),
            status=(data.get('status') or 'active').strip().lower(),
            priority=(data.get('priority') or 'med').strip().lower(),
            value_amount=float(data.get('value_amount') or 0),
            progress=max(0, min(100, int(data.get('progress') or 0))),
            owner=(data.get('owner') or '').strip() or None,
            owner_user_id=owner_uid,
            next_action=(data.get('next_action') or '').strip() or None,
            expected_close_date=_parse_iso_date(data.get('expected_close_date')),
            notes=(data.get('notes') or '').strip() or None,
            primary_contact_name=(data.get('primary_contact_name') or '').strip() or None,
            primary_contact_email=(data.get('primary_contact_email') or '').strip() or None,
            created_by=user_id
        )
        _sync_owner_display(project)
        db.session.add(project)

        if project.primary_contact_name:
            existing_contact = BDContact.query.filter_by(
                name=project.primary_contact_name,
                company=project.company
            ).first()
            if not existing_contact:
                db.session.add(BDContact(
                    name=project.primary_contact_name,
                    title='Primary Contact',
                    company=project.company,
                    email=project.primary_contact_email,
                    tags=['Project Contact'],
                    created_by=user_id
                ))

        if project.next_action:
            due_dt = None
            if project.expected_close_date:
                due_dt = datetime.combine(project.expected_close_date, datetime.min.time())
            db.session.add(BDFollowUp(
                title=project.next_action,
                company=project.company,
                followup_type='note',
                due_at=due_dt,
                status='open',
                project=project,
                created_by=user_id
            ))

        _bd_activity(
            icon='📁',
            title=f'Project created — {project.name}',
            description=f'Added deal for {project.company}',
            badge=project.company,
            bg='#fff4ef',
            user_id=user_id
        )

        db.session.commit()
        return success_response({'project': project.to_dict()}, message='Project created successfully', status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating BD project: {str(e)}", exc_info=True)
        return error_response('Failed to create project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/projects/<int:project_id>', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_get_project(project_id):
    """Project workspace payload: deal + related quotes, follow-ups, activity."""
    try:
        user = _bd_current_user()
        project = db.session.get(BDProject, project_id)
        if not project:
            return error_response('Project not found', status_code=404, error_code='NOT_FOUND')
        if not _can_access_bd_project(user, project):
            return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')

        quotations = (
            Quotation.query.filter_by(bd_project_id=project.id)
            .order_by(Quotation.created_at.desc())
            .limit(100)
            .all()
        )
        followups = (
            BDFollowUp.query.filter(
                db.or_(
                    BDFollowUp.project_id == project.id,
                    db.and_(
                        BDFollowUp.company.isnot(None),
                        BDFollowUp.company == project.company,
                    ),
                )
            )
            .order_by(BDFollowUp.created_at.desc())
            .limit(50)
            .all()
        )
        company = (project.company or '').strip()
        activities_q = BDActivity.query
        if company:
            activities_q = activities_q.filter(
                db.or_(
                    BDActivity.badge == company,
                    BDActivity.title.ilike(f'%{project.name}%'),
                    BDActivity.description.ilike(f'%{company}%'),
                )
            )
        else:
            activities_q = activities_q.filter(BDActivity.title.ilike(f'%{project.name}%'))
        activities = activities_q.order_by(BDActivity.event_time.desc()).limit(40).all()

        contacts = []
        if company:
            contacts = (
                BDContact.query.filter(BDContact.company == company)
                .order_by(BDContact.name.asc())
                .limit(20)
                .all()
            )

        quoted_total = round(sum(float(q.grand_total or 0) for q in quotations), 2)
        latest = quotations[0] if quotations else None
        owners = []
        if user_sees_all_bd_deals(user):
            for u in User.query.filter(User.is_active.is_(True)).order_by(User.full_name).all():
                if user_has_bd_access(u) or u.role == 'admin':
                    owners.append({
                        'id': u.id,
                        'name': u.full_name or u.username,
                    })

        return success_response({
            'project': project.to_dict(),
            'quotations': [q.to_dict() for q in quotations],
            'followups': [f.to_dict() for f in followups],
            'activities': [a.to_dict() for a in activities],
            'contacts': [c.to_dict() for c in contacts],
            'owners': owners,
            'can_prepare_quotations': user_can_prepare_quotations(user),
            'summary': {
                'quote_count': len(quotations),
                'quoted_total': quoted_total,
                'latest_quote_no': latest.quote_no if latest else None,
                'latest_quote_status': latest.status if latest else None,
                'open_followups': sum(1 for f in followups if (f.status or 'open') != 'done'),
            },
        })
    except Exception as e:
        current_app.logger.error(f"Error fetching BD project {project_id}: {str(e)}", exc_info=True)
        return error_response('Failed to fetch project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/projects/<int:project_id>', methods=['PUT'])
@jwt_required()
@bd_access_required
def bd_update_project(project_id):
    """Update a BD project/deal."""
    try:
        user = _bd_current_user()
        user_id = user.id
        data = request.get_json() or {}
        project = BDProject.query.get_or_404(project_id)
        if not _can_access_bd_project(user, project):
            return error_response('You can only edit your own deals', status_code=403, error_code='OWNERSHIP_REQUIRED')

        name = (data.get('name') or project.name or '').strip()
        company = (data.get('company') or project.company or '').strip()
        if not name or not company:
            return error_response('Project name and company are required', status_code=400, error_code='VALIDATION_ERROR')

        old_name = project.name
        old_company = project.company

        project.name = name
        project.company = company
        project.stage = (data.get('stage') or project.stage or 'prospecting').strip().lower()
        project.status = (data.get('status') or project.status or 'active').strip().lower()
        project.priority = (data.get('priority') or project.priority or 'med').strip().lower()
        project.value_amount = float(data.get('value_amount') if data.get('value_amount') is not None else (project.value_amount or 0))
        project.progress = max(0, min(100, int(data.get('progress') if data.get('progress') is not None else (project.progress or 0))))
        if data.get('owner') is not None:
            project.owner = (data.get('owner') or '').strip() or None
        if 'owner_user_id' in data or 'ownerUserId' in data:
            if user_sees_all_bd_deals(user):
                project.owner_user_id = _resolve_owner_user_id(data, default_user_id=project.owner_user_id)
            else:
                project.owner_user_id = user.id
            _sync_owner_display(project)
        project.next_action = (data.get('next_action') if data.get('next_action') is not None else project.next_action or '')
        project.next_action = project.next_action.strip() or None
        project.expected_close_date = _parse_iso_date(data.get('expected_close_date')) if 'expected_close_date' in data else project.expected_close_date
        project.notes = (data.get('notes') if data.get('notes') is not None else project.notes or '')
        project.notes = project.notes.strip() or None
        project.primary_contact_name = (data.get('primary_contact_name') if data.get('primary_contact_name') is not None else project.primary_contact_name or '')
        project.primary_contact_name = project.primary_contact_name.strip() or None
        project.primary_contact_email = (data.get('primary_contact_email') if data.get('primary_contact_email') is not None else project.primary_contact_email or '')
        project.primary_contact_email = project.primary_contact_email.strip() or None

        if project.primary_contact_name:
            existing_contact = BDContact.query.filter_by(
                name=project.primary_contact_name,
                company=project.company
            ).first()
            if not existing_contact:
                db.session.add(BDContact(
                    name=project.primary_contact_name,
                    title='Primary Contact',
                    company=project.company,
                    email=project.primary_contact_email,
                    tags=['Project Contact'],
                    created_by=user_id
                ))

        _bd_activity(
            icon='✏️',
            title=f'Project updated — {project.name}',
            description=f'Updated project details ({old_name} / {old_company})',
            badge=project.company,
            bg='#fef6e4',
            user_id=user_id
        )

        db.session.commit()
        return success_response({'project': project.to_dict()}, message='Project updated successfully')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating BD project: {str(e)}", exc_info=True)
        return error_response('Failed to update project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/projects/<int:project_id>', methods=['DELETE'])
@jwt_required()
@bd_access_required
def bd_delete_project(project_id):
    """Delete a BD deal. Quotes and ticketing projects are unlinked, not removed."""
    try:
        user = _bd_current_user()
        project = db.session.get(BDProject, project_id)
        if not project:
            return error_response('Project not found', status_code=404, error_code='NOT_FOUND')
        if not _can_access_bd_project(user, project):
            return error_response('You can only delete your own deals', status_code=403, error_code='OWNERSHIP_REQUIRED')

        name = project.name
        company = project.company

        BDFollowUp.query.filter_by(project_id=project.id).delete(synchronize_session=False)
        Quotation.query.filter_by(bd_project_id=project.id).update(
            {Quotation.bd_project_id: None}, synchronize_session=False
        )
        TicketProject.query.filter_by(bd_project_id=project.id).update(
            {TicketProject.bd_project_id: None}, synchronize_session=False
        )
        try:
            from app.models import QhsiTraining
            QhsiTraining.query.filter_by(bd_project_id=project.id).update(
                {QhsiTraining.bd_project_id: None}, synchronize_session=False
            )
        except Exception:
            pass

        db.session.delete(project)
        _bd_activity(
            icon='🗑️',
            title=f'Project deleted — {name}',
            description=f'Removed deal for {company}',
            badge=company or 'Deleted',
            bg='#fdf0ee',
            user_id=user.id,
        )
        db.session.commit()
        return success_response({'deleted_id': project_id}, message='Project deleted')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting BD project {project_id}: {str(e)}", exc_info=True)
        return error_response('Failed to delete project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/projects/import-excel', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_import_projects_excel():
    """Bulk import BD projects from client contract Excel."""
    try:
        user_id = get_jwt_identity()
        user = User.query.get(user_id)

        if 'file' not in request.files:
            return error_response('No file provided', status_code=400, error_code='VALIDATION_ERROR')

        file = request.files['file']
        if not file or not file.filename:
            return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')

        filename = file.filename.lower()
        if not filename.endswith(('.xlsx', '.xls')):
            return error_response('Invalid file format. Upload .xlsx or .xls', status_code=400, error_code='VALIDATION_ERROR')

        try:
            import pandas as pd
            from io import StringIO
        except ImportError:
            return error_response(
                'Excel import requires pandas/openpyxl/xlrd dependencies',
                status_code=500,
                error_code='DEPENDENCY_ERROR'
            )

        # Read file robustly (.xlsx, legacy .xls, or HTML-based .xls)
        try:
            if filename.endswith('.xlsx'):
                df = pd.read_excel(file)
            else:
                try:
                    df = pd.read_excel(file, engine='xlrd')
                except Exception:
                    file.stream.seek(0)
                    html_text = file.stream.read().decode('utf-8', errors='ignore')
                    tables = pd.read_html(StringIO(html_text))
                    if not tables:
                        return error_response('Could not read any table from uploaded Excel file', status_code=400, error_code='VALIDATION_ERROR')
                    df = max(tables, key=lambda t: t.shape[0])
        except Exception as read_error:
            current_app.logger.error(f"BD Excel import read error: {read_error}", exc_info=True)
            return error_response(f'Could not parse Excel file: {read_error}', status_code=400, error_code='VALIDATION_ERROR')

        if df is None or df.empty:
            return error_response('Excel file is empty', status_code=400, error_code='VALIDATION_ERROR')

        # Flatten multi-level headers if present
        if hasattr(df.columns, 'levels'):
            flat_cols = []
            for col in df.columns:
                if isinstance(col, tuple):
                    flat_cols.append(' '.join([str(part).strip() for part in col if str(part).strip() and str(part).strip().lower() != 'nan']))
                else:
                    flat_cols.append(str(col))
            df.columns = flat_cols

        normalized_cols = {_normalize_excel_column_name(c): c for c in df.columns}
        alias_map = {
            'code': 'code',
            'contract': 'contract',
            'contract name': 'contract',
            'contract reference': 'contract',
            'reference code': 'reference_code',
            'client': 'client',
            'customer': 'client',
            'start date': 'start_date',
            'end date': 'end_date',
            'renewal date': 'renewal_date',
            'status': 'status',
            'contract amount': 'contract_amount',
            'amount': 'contract_amount',
            'payment type': 'payment_type',
            'invoicing schedule': 'invoicing_schedule'
        }

        canonical_to_original = {}
        for normalized_name, original_name in normalized_cols.items():
            canonical_name = alias_map.get(normalized_name)
            if canonical_name:
                canonical_to_original[canonical_name] = original_name

        if 'contract' not in canonical_to_original and 'client' not in canonical_to_original:
            return error_response(
                'Excel must include at least "Contract" or "Client" column',
                status_code=400,
                error_code='VALIDATION_ERROR'
            )

        existing_projects = BDProject.query.with_entities(BDProject.name, BDProject.company).all()
        existing_keys = {
            f"{(name or '').strip().lower()}|{(company or '').strip().lower()}"
            for name, company in existing_projects
        }
        existing_contacts = {
            f"{(c.name or '').strip().lower()}|{(c.company or '').strip().lower()}"
            for c in BDContact.query.with_entities(BDContact.name, BDContact.company).all()
        }

        imported = 0
        skipped_duplicates = 0
        skipped_empty = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                def cell(canonical_name, default=''):
                    col = canonical_to_original.get(canonical_name)
                    if not col:
                        return default
                    return row.get(col, default)

                contract_name = str(cell('contract', '')).strip()
                client_name = str(cell('client', '')).strip()
                if contract_name.lower() == 'nan':
                    contract_name = ''
                if client_name.lower() == 'nan':
                    client_name = ''

                if not contract_name and not client_name:
                    skipped_empty += 1
                    continue

                # Use contract title as project name and client as company
                project_name = contract_name or client_name
                company_name = client_name or contract_name

                duplicate_key = f"{project_name.lower()}|{company_name.lower()}"
                if duplicate_key in existing_keys:
                    skipped_duplicates += 1
                    continue

                status_text = str(cell('status', '')).strip()
                status, stage, progress = _status_stage_progress_from_contract_status(status_text)
                contract_amount = _parse_excel_float(cell('contract_amount', 0), default=0.0)
                renewal_date = _parse_excel_date(cell('renewal_date', None))
                end_date = _parse_excel_date(cell('end_date', None))
                start_date = _parse_excel_date(cell('start_date', None))
                expected_close = renewal_date or end_date

                if contract_amount >= 300000:
                    priority = 'high'
                elif contract_amount >= 100000:
                    priority = 'med'
                else:
                    priority = 'low'

                if status == 'active':
                    next_action = 'Relationship review with client'
                elif status == 'lost':
                    next_action = 'Renewal/revival outreach'
                else:
                    next_action = 'Qualification follow-up'

                code = str(cell('code', '')).strip()
                ref_code = str(cell('reference_code', '')).strip()
                payment_type = str(cell('payment_type', '')).strip()
                invoicing_schedule = str(cell('invoicing_schedule', '')).strip()

                notes_parts = []
                if code and code.lower() != 'nan':
                    notes_parts.append(f"Code: {code}")
                if ref_code and ref_code.lower() != 'nan':
                    notes_parts.append(f"Reference: {ref_code}")
                if start_date:
                    notes_parts.append(f"Start Date: {start_date.isoformat()}")
                if end_date:
                    notes_parts.append(f"End Date: {end_date.isoformat()}")
                if payment_type and payment_type.lower() != 'nan':
                    notes_parts.append(f"Payment Type: {payment_type}")
                if invoicing_schedule and invoicing_schedule.lower() != 'nan':
                    notes_parts.append(f"Invoicing: {invoicing_schedule}")

                project = BDProject(
                    name=project_name,
                    company=company_name,
                    stage=stage,
                    status=status,
                    priority=priority,
                    value_amount=contract_amount,
                    progress=progress,
                    owner=(user.full_name if user and user.full_name else (user.username if user else 'Admin')),
                    owner_user_id=user_id,
                    next_action=next_action,
                    expected_close_date=expected_close,
                    notes=' | '.join(notes_parts) if notes_parts else None,
                    created_by=user_id
                )
                db.session.add(project)
                existing_keys.add(duplicate_key)

                contact_key = f"{company_name.lower()}|{company_name.lower()}"
                if company_name and contact_key not in existing_contacts:
                    db.session.add(BDContact(
                        name=company_name,
                        title='Client',
                        company=company_name,
                        tags=['Imported Client'],
                        created_by=user_id
                    ))
                    existing_contacts.add(contact_key)

                imported += 1
            except Exception as row_error:
                errors.append(f"Row {idx + 2}: {row_error}")

        if imported > 0:
            _bd_activity(
                icon='📥',
                title='Projects imported from Excel',
                description=f'Imported {imported} projects',
                badge='Excel Import',
                bg='#e8f0fb',
                user_id=user_id
            )

        db.session.commit()
        return success_response({
            'imported': imported,
            'skipped_duplicates': skipped_duplicates,
            'skipped_empty': skipped_empty,
            'total_rows': int(len(df)),
            'errors': errors[:15]
        }, message=f'Imported {imported} project(s) from Excel')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error importing BD projects from Excel: {str(e)}", exc_info=True)
        return error_response('Failed to import projects from Excel', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/followups', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_create_followup():
    """Create a BD follow-up."""
    try:
        user_id = get_jwt_identity()
        data = request.get_json() or {}

        title = (data.get('title') or '').strip()
        if not title:
            return error_response('Follow-up title is required', status_code=400, error_code='VALIDATION_ERROR')

        followup = BDFollowUp(
            title=title,
            company=(data.get('company') or '').strip() or None,
            followup_type=(data.get('followup_type') or 'note').strip().lower(),
            due_at=_parse_iso_datetime(data.get('due_at')),
            status='open',
            details=(data.get('details') or '').strip() or None,
            project_id=data.get('project_id'),
            created_by=user_id
        )
        db.session.add(followup)
        _bd_activity(
            icon='🔔',
            title=f'Follow-up added — {title}',
            description=(followup.details or 'No extra details'),
            badge=(followup.company or 'Follow-up'),
            bg='#fef6e4',
            user_id=user_id
        )
        db.session.commit()
        return success_response({'followup': followup.to_dict()}, message='Follow-up created successfully', status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating BD follow-up: {str(e)}", exc_info=True)
        return error_response('Failed to create follow-up', status_code=500, error_code='DATABASE_ERROR')


def _can_access_followup(user, followup):
    if user_sees_all_bd_deals(user):
        return True
    if followup.created_by == user.id:
        return True
    if followup.project_id:
        project = BDProject.query.get(followup.project_id)
        if project and _can_access_bd_project(user, project):
            return True
    return False


@admin_bp.route('/bd/followups/<int:followup_id>', methods=['PUT'])
@jwt_required()
@bd_access_required
def bd_update_followup(followup_id):
    """Update a follow-up, including marking it done or reopening it."""
    try:
        user = _bd_current_user()
        followup = db.session.get(BDFollowUp, followup_id)
        if not followup:
            return error_response('Follow-up not found', status_code=404, error_code='NOT_FOUND')
        if not _can_access_followup(user, followup):
            return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')

        data = request.get_json(silent=True) or {}
        if 'title' in data:
            title = (data.get('title') or '').strip()
            if not title:
                return error_response('Follow-up title is required', status_code=400, error_code='VALIDATION_ERROR')
            followup.title = title
        if 'company' in data:
            followup.company = (data.get('company') or '').strip() or None
        if 'followup_type' in data:
            followup.followup_type = (data.get('followup_type') or followup.followup_type or 'note').strip().lower()
        if 'due_at' in data:
            followup.due_at = _parse_iso_datetime(data.get('due_at'))
        if 'details' in data:
            followup.details = (data.get('details') or '').strip() or None
        if 'project_id' in data:
            followup.project_id = data.get('project_id') or None
        if 'status' in data:
            status = (data.get('status') or '').strip().lower()
            if status not in ('open', 'done'):
                return error_response('Status must be open or done', status_code=400, error_code='VALIDATION_ERROR')
            followup.status = status

        done_now = followup.status == 'done'
        _bd_activity(
            icon='✅' if done_now else '🔔',
            title=('Follow-up completed — ' if done_now else 'Follow-up updated — ') + followup.title,
            description=(followup.details or followup.company or ''),
            badge=(followup.company or 'Follow-up'),
            bg='#e8f5ee' if done_now else '#fef6e4',
            user_id=user.id,
        )
        db.session.commit()
        return success_response({'followup': followup.to_dict()}, message='Follow-up updated')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating BD follow-up {followup_id}: {str(e)}", exc_info=True)
        return error_response('Failed to update follow-up', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/bd/contacts', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_create_contact():
    """Create a BD contact."""
    try:
        user_id = get_jwt_identity()
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        if not name:
            return error_response('Contact name is required', status_code=400, error_code='VALIDATION_ERROR')

        tags = data.get('tags') or []
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(',') if t.strip()]

        contact = BDContact(
            name=name,
            title=(data.get('title') or '').strip() or None,
            company=(data.get('company') or '').strip() or None,
            email=(data.get('email') or '').strip() or None,
            phone=(data.get('phone') or '').strip() or None,
            tags=tags if isinstance(tags, list) else [],
            created_by=user_id
        )
        db.session.add(contact)
        _bd_activity(
            icon='👥',
            title=f'Contact added — {contact.name}',
            description=f'{contact.title or "Contact"} at {contact.company or "Unknown Company"}',
            badge=(contact.company or 'Contacts'),
            bg='#e8f0fb',
            user_id=user_id
        )
        db.session.commit()
        return success_response({'contact': contact.to_dict()}, message='Contact created successfully', status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating BD contact: {str(e)}", exc_info=True)
        return error_response('Failed to create contact', status_code=500, error_code='DATABASE_ERROR')


def _can_access_contact(user, contact):
    if user_sees_all_bd_deals(user):
        return True
    return contact.created_by == user.id


@admin_bp.route('/bd/contacts/<int:contact_id>', methods=['DELETE'])
@jwt_required()
@bd_access_required
def bd_delete_contact(contact_id):
    try:
        user = _bd_current_user()
        contact = db.session.get(BDContact, contact_id)
        if not contact:
            return error_response('Contact not found', status_code=404, error_code='NOT_FOUND')
        if not _can_access_contact(user, contact):
            return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
        name = contact.name
        company = contact.company
        db.session.delete(contact)
        _bd_activity(
            icon='🗑️',
            title=f'Contact deleted — {name}',
            description=f'{company or "No company"}',
            badge=(company or 'Contacts'),
            bg='#fdf0ee',
            user_id=user.id,
        )
        db.session.commit()
        return success_response({'deleted_id': contact_id}, message='Contact deleted')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting BD contact {contact_id}: {str(e)}", exc_info=True)
        return error_response('Failed to delete contact', status_code=500, error_code='DATABASE_ERROR')


def _gen_quote_no():
    return f'QT-{uuid.uuid4().hex[:8].upper()}'


def _gen_quote_ref_no(quote_no=None):
    yy = date.today().strftime('%y')
    yyyy = date.today().strftime('%Y')
    token = (quote_no or uuid.uuid4().hex).replace('QT-', '')[:6].upper()
    return f'KYQ/{yy}RR{token}Rev1/{yyyy}'


def _normalize_quote_signature(raw):
    if raw is None:
        return None
    sig = str(raw).strip()
    if not sig:
        return None
    if not sig.startswith('data:image'):
        return None
    if len(sig) > 1_500_000:
        return None
    return sig


def _recompute_quote_totals(quote, tax_pct=None, discount_amount=None):
    from app.bd.quotation_builder import amount_to_aed_words

    subtotal = sum(float(it.total_price or 0) for it in quote.items)
    quote.subtotal = round(subtotal, 2)
    if discount_amount is not None:
        try:
            quote.discount_amount = max(0.0, float(discount_amount))
        except (TypeError, ValueError):
            pass
    discount = max(0.0, float(quote.discount_amount or 0))
    if discount > subtotal:
        discount = subtotal
        quote.discount_amount = discount
    net = round(max(subtotal - discount, 0), 2)
    if tax_pct is not None:
        try:
            quote.tax_pct = float(tax_pct)
        except (TypeError, ValueError):
            pass
    pct = float(quote.tax_pct or 0)
    quote.tax_amount = round(net * pct / 100.0, 2)
    quote.grand_total = net
    quote.amount_in_words = amount_to_aed_words(net)


def _apply_quote_letter_fields(quote, data, proj=None):
    from app.models import (
        QUOTATION_DEFAULT_EXCLUSIONS,
        QUOTATION_DEFAULT_INTRO,
        QUOTATION_DEFAULT_NOTES,
        QUOTATION_DEFAULT_SIGNATORY_EMAIL,
        QUOTATION_DEFAULT_SIGNATORY_NAME,
        QUOTATION_DEFAULT_SIGNATORY_PHONE,
        QUOTATION_DEFAULT_SIGNOFF_LABEL,
        QUOTATION_DEFAULT_TERMS,
    )

    str_fields = (
        'ref_no', 'kind_attn', 'client_tel', 'subject', 'project_name',
        'intro_text', 'notes_text', 'exclusions_text', 'terms_text',
        'signatory_name', 'signatory_email', 'signatory_phone', 'signoff_label',
        'notes',
    )
    for field in str_fields:
        if field in data:
            setattr(quote, field, (data.get(field) or '').strip() or None)

    if 'prepared_signature' in data:
        raw_sig = data.get('prepared_signature')
        if raw_sig in (None, ''):
            quote.prepared_signature = None
        else:
            norm = _normalize_quote_signature(raw_sig)
            if norm:
                quote.prepared_signature = norm

    if not quote.intro_text:
        quote.intro_text = QUOTATION_DEFAULT_INTRO
    if not quote.notes_text:
        quote.notes_text = QUOTATION_DEFAULT_NOTES
    if not quote.exclusions_text:
        quote.exclusions_text = QUOTATION_DEFAULT_EXCLUSIONS
    if not quote.terms_text:
        quote.terms_text = QUOTATION_DEFAULT_TERMS
    if not quote.signatory_name:
        quote.signatory_name = QUOTATION_DEFAULT_SIGNATORY_NAME
    if not quote.signatory_email:
        quote.signatory_email = QUOTATION_DEFAULT_SIGNATORY_EMAIL
    if not quote.signatory_phone:
        quote.signatory_phone = QUOTATION_DEFAULT_SIGNATORY_PHONE
    if not quote.signoff_label:
        quote.signoff_label = QUOTATION_DEFAULT_SIGNOFF_LABEL
    if not quote.project_name and proj is not None:
        quote.project_name = proj.name
    if not quote.kind_attn and quote.contact_name:
        quote.kind_attn = quote.contact_name


def _build_quote_items(quote, items_data):
    quote.items.clear()
    if not items_data:
        return 'At least one line item is required'
    for raw in items_data:
        desc = (raw.get('description') or raw.get('material_name') or '').strip()
        if not desc:
            continue
        try:
            qty = float(raw.get('quantity') or 1)
            price = float(raw.get('unit_price') or 0)
        except (TypeError, ValueError):
            return 'Invalid quantity or unit price'
        quote.items.append(QuotationItem(
            description=desc,
            details=(raw.get('details') or raw.get('description_detail') or '').strip() or None,
            quantity=qty,
            unit=(raw.get('unit') or '').strip() or None,
            unit_price=price,
            total_price=round(qty * price, 2),
        ))
    if not quote.items:
        return 'At least one line item is required'
    return None


def _can_access_quote(user, quote):
    if user_sees_all_bd_deals(user):
        return True
    return quote.owner_user_id in (None, user.id) or quote.created_by_id == user.id


@admin_bp.route('/bd/quotations', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_list_quotations():
    user = _bd_current_user()
    q = Quotation.query.order_by(Quotation.created_at.desc())
    if not user_sees_all_bd_deals(user):
        q = q.filter(db.or_(
            Quotation.owner_user_id == user.id,
            Quotation.created_by_id == user.id,
        ))
    project_id = request.args.get('bd_project_id', type=int)
    if project_id:
        q = q.filter(Quotation.bd_project_id == project_id)
    return success_response({
        'quotations': [x.to_dict(include_items=False) for x in q.limit(200).all()],
        'can_prepare_quotations': user_can_prepare_quotations(user),
    })


@admin_bp.route('/bd/quotations', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_create_quotation():
    user = _bd_current_user()
    if not user_can_prepare_quotations(user):
        return error_response(
            'You do not have permission to prepare quotations',
            status_code=403,
            error_code='ACCESS_DENIED',
        )
    data = request.get_json(silent=True) or {}

    bd_id = data.get('bd_project_id')
    if not bd_id:
        return error_response(
            'Quotations must be created from a project (bd_project_id is required)',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    try:
        bd_id = int(bd_id)
    except (TypeError, ValueError):
        return error_response('Invalid bd_project_id', status_code=400, error_code='VALIDATION_ERROR')

    proj = BDProject.query.get(bd_id)
    if not proj:
        return error_response('BD project not found', status_code=404, error_code='NOT_FOUND')
    if not _can_access_bd_project(user, proj):
        return error_response('Access denied to that deal', status_code=403, error_code='OWNERSHIP_REQUIRED')

    company = (data.get('company_name') or data.get('company') or proj.company or '').strip()
    if not company:
        return error_response('Company name is required', status_code=400, error_code='VALIDATION_ERROR')

    owner_uid = _resolve_owner_user_id(data, default_user_id=user.id)
    if not user_sees_all_bd_deals(user):
        owner_uid = user.id
    if data.get('owner_user_id') is None and data.get('ownerUserId') is None and proj.owner_user_id:
        owner_uid = proj.owner_user_id if user_sees_all_bd_deals(user) else user.id

    contact_name = (data.get('contact_name') or '').strip() or proj.primary_contact_name
    contact_email = (data.get('contact_email') or '').strip() or proj.primary_contact_email
    quote_no = _gen_quote_no()
    ref_no = (data.get('ref_no') or '').strip() or _gen_quote_ref_no(quote_no)

    quote = Quotation(
        quote_no=quote_no,
        ref_no=ref_no,
        bd_project_id=bd_id,
        company_name=company,
        contact_name=contact_name or None,
        contact_email=contact_email or None,
        quote_date=_parse_iso_date(data.get('quote_date')) or date.today(),
        valid_until=_parse_iso_date(data.get('valid_until')),
        status='draft',
        notes=(data.get('notes') or '').strip() or None,
        owner_user_id=owner_uid,
        created_by_id=user.id,
        tax_pct=float(data.get('tax_pct') if data.get('tax_pct') is not None else 5),
        discount_amount=float(data.get('discount_amount') or 0),
        project_name=(data.get('project_name') or proj.name or '').strip() or None,
    )
    _apply_quote_letter_fields(quote, data, proj=proj)
    err = _build_quote_items(quote, data.get('items') or [])
    if err:
        return error_response(err, status_code=400, error_code='VALIDATION_ERROR')
    _recompute_quote_totals(quote, data.get('tax_pct'), data.get('discount_amount'))
    db.session.add(quote)
    created_stamp = _bd_dt_24(quote.created_at or utc_now_naive())
    _bd_activity(
        '📄',
        f'Quotation created — {quote.quote_no} for {proj.name}',
        f'{company} · Created {created_stamp}',
        badge=proj.name,
        user_id=user.id,
    )
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Quotation created', status_code=201)


@admin_bp.route('/bd/quotations/<int:quote_id>', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_get_quotation(quote_id):
    user = _bd_current_user()
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    return success_response({'quotation': quote.to_dict()})


@admin_bp.route('/bd/quotations/<int:quote_id>', methods=['PUT'])
@jwt_required()
@bd_access_required
def bd_update_quotation(quote_id):
    user = _bd_current_user()
    if not user_can_prepare_quotations(user):
        return error_response(
            'You do not have permission to prepare quotations',
            status_code=403,
            error_code='ACCESS_DENIED',
        )
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    if quote.status not in ('draft', 'sent', 'rejected', 'pending_approval'):
        return error_response('Only draft/sent/rejected quotations can be edited', status_code=400, error_code='VALIDATION_ERROR')
    data = request.get_json(silent=True) or {}
    if 'company_name' in data or 'company' in data:
        quote.company_name = (data.get('company_name') or data.get('company') or quote.company_name).strip()
    for field in ('contact_name', 'contact_email', 'notes'):
        if field in data:
            setattr(quote, field, (data.get(field) or '').strip() or None)
    if 'valid_until' in data:
        quote.valid_until = _parse_iso_date(data.get('valid_until'))
    if 'quote_date' in data:
        parsed = _parse_iso_date(data.get('quote_date'))
        if parsed:
            quote.quote_date = parsed
    _apply_quote_letter_fields(quote, data, proj=quote.bd_project)
    if 'items' in data:
        err = _build_quote_items(quote, data.get('items') or [])
        if err:
            return error_response(err, status_code=400, error_code='VALIDATION_ERROR')
    _recompute_quote_totals(quote, data.get('tax_pct'), data.get('discount_amount'))
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Quotation updated')


@admin_bp.route('/bd/quotations/<int:quote_id>/submit', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_submit_quotation(quote_id):
    user = _bd_current_user()
    if not user_can_prepare_quotations(user):
        return error_response(
            'You do not have permission to prepare quotations',
            status_code=403,
            error_code='ACCESS_DENIED',
        )
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    if quote.status not in ('draft', 'sent', 'rejected', 'pending_approval'):
        return error_response('Quotation cannot be submitted from current status', status_code=400, error_code='VALIDATION_ERROR')
    data = request.get_json(silent=True) or {}
    if data.get('confirm_thanks_remarks') not in (True, 'true', '1', 1, None):
        pass
    from app.models import (
        QUOTATION_DEFAULT_SIGNATORY_EMAIL,
        QUOTATION_DEFAULT_SIGNATORY_NAME,
        QUOTATION_DEFAULT_SIGNATORY_PHONE,
        QUOTATION_DEFAULT_SIGNOFF_LABEL,
    )
    for field, default in (
        ('signoff_label', QUOTATION_DEFAULT_SIGNOFF_LABEL),
        ('signatory_name', QUOTATION_DEFAULT_SIGNATORY_NAME),
        ('signatory_email', QUOTATION_DEFAULT_SIGNATORY_EMAIL),
        ('signatory_phone', QUOTATION_DEFAULT_SIGNATORY_PHONE),
    ):
        if field in data:
            setattr(quote, field, (data.get(field) or '').strip() or default)
    sig = _normalize_quote_signature(data.get('signature') or data.get('prepared_signature'))
    if not sig:
        return error_response(
            'E-signature is required to submit (draw, type, or upload via the signature pad)',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    quote.prepared_signature = sig
    if not quote.signatory_name:
        quote.signatory_name = getattr(user, 'full_name', None) or user.username
    if not quote.signoff_label:
        quote.signoff_label = QUOTATION_DEFAULT_SIGNOFF_LABEL
    now = utc_now_naive()
    quote.status = 'pending_approval'
    quote.submitted_at = now
    quote.approved_at = None
    quote.approved_by_id = None
    quote.rejected_at = None
    quote.rejection_notes = None
    proj_name = quote.bd_project.name if quote.bd_project else ''
    submitted_stamp = _bd_dt_24(quote.submitted_at)
    created_stamp = _bd_dt_24(quote.created_at) if quote.created_at else ''
    desc_bits = [quote.company_name or '']
    if created_stamp:
        desc_bits.append(f'Created {created_stamp}')
    desc_bits.append(f'Submitted {submitted_stamp}')
    _bd_activity(
        '📤',
        f'Quotation submitted for approval — {quote.quote_no}' + (f' for {proj_name}' if proj_name else ''),
        ' · '.join(b for b in desc_bits if b),
        badge=proj_name or quote.company_name or 'Pending approval',
        event_time=quote.submitted_at,
        user_id=user.id,
    )
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Quotation submitted for approval')


@admin_bp.route('/bd/quotations/<int:quote_id>/approve', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_approve_quotation(quote_id):
    user = _bd_current_user()
    des = (getattr(user, 'designation', None) or '').strip().lower()
    if user.role != 'admin' and des not in ('general_manager', 'operations_manager'):
        return error_response('Only GM / Ops Manager / Admin may approve proposals', status_code=403, error_code='ACCESS_DENIED')
    quote = Quotation.query.get_or_404(quote_id)
    if quote.status != 'pending_approval':
        return error_response('Quotation is not pending approval', status_code=400, error_code='VALIDATION_ERROR')
    data = request.get_json(silent=True) or {}
    sig = _normalize_quote_signature(data.get('signature') or data.get('approval_signature'))
    if not sig:
        return error_response(
            'E-signature image is required for approval (use the signature pad)',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    quote.status = 'approved'
    quote.approved_at = utc_now_naive()
    quote.approved_by_id = user.id
    quote.approval_signature = sig
    quote.approval_notes = (data.get('notes') or '').strip() or None
    if quote.bd_project:
        quote.bd_project.status = 'won'
        quote.bd_project.stage = 'closing'
        quote.bd_project.progress = 100
    approved_stamp = _bd_dt_24(quote.approved_at)
    _bd_activity(
        '✅',
        f'Quotation approved — {quote.quote_no}',
        f'{quote.company_name} · Approved {approved_stamp}',
        badge='Approved',
        bg='#fff4ef',
        event_time=quote.approved_at,
        user_id=user.id,
    )
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Proposal approved')


@admin_bp.route('/bd/quotations/<int:quote_id>/reject', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_reject_quotation(quote_id):
    user = _bd_current_user()
    des = (getattr(user, 'designation', None) or '').strip().lower()
    if user.role != 'admin' and des not in ('general_manager', 'operations_manager'):
        return error_response('Only GM / Ops Manager / Admin may reject proposals', status_code=403, error_code='ACCESS_DENIED')
    quote = Quotation.query.get_or_404(quote_id)
    if quote.status != 'pending_approval':
        return error_response('Only pending quotations can be rejected', status_code=400, error_code='VALIDATION_ERROR')
    data = request.get_json(silent=True) or {}
    quote.status = 'rejected'
    quote.rejected_at = utc_now_naive()
    quote.rejection_notes = (data.get('notes') or '').strip() or None
    rejected_stamp = _bd_dt_24(quote.rejected_at)
    _bd_activity(
        '❌',
        f'Quotation rejected — {quote.quote_no}',
        f'{quote.company_name} · Rejected {rejected_stamp}',
        badge='Rejected',
        bg='#fdf0ee',
        event_time=quote.rejected_at,
        user_id=user.id,
    )
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Proposal rejected')


@admin_bp.route('/bd/quotations/<int:quote_id>/cancel', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_cancel_quotation(quote_id):
    user = _bd_current_user()
    if not user_can_prepare_quotations(user):
        return error_response(
            'You do not have permission to cancel quotations',
            status_code=403,
            error_code='ACCESS_DENIED',
        )
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    if quote.status not in ('draft', 'sent', 'pending_approval', 'rejected'):
        return error_response(
            'Only draft, pending, or rejected quotations can be cancelled',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    data = request.get_json(silent=True) or {}
    notes = (data.get('notes') or data.get('cancellation_notes') or '').strip()
    if not notes:
        return error_response(
            'Cancellation notes are required',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    quote.status = 'cancelled'
    quote.rejected_at = utc_now_naive()
    quote.rejection_notes = notes
    cancelled_stamp = _bd_dt_24(quote.rejected_at)
    _bd_activity(
        '🚫',
        f'Quotation cancelled — {quote.quote_no}',
        f'{quote.company_name} · Cancelled {cancelled_stamp} · {notes[:120]}',
        badge='Cancelled',
        bg='#f4f4f5',
        event_time=quote.rejected_at,
        user_id=user.id,
    )
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='Quotation cancelled')


@admin_bp.route('/bd/quotations/<int:quote_id>/lpo', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_quote_upload_lpo(quote_id):
    user = _bd_current_user()
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    if 'file' not in request.files:
        return error_response('No file provided', status_code=400, error_code='VALIDATION_ERROR')
    f = request.files['file']
    if not f or not f.filename:
        return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')
    if quote.status != 'approved':
        return error_response(
            'Attach an LPO after the quotation is approved',
            status_code=400,
            error_code='VALIDATION_ERROR',
        )
    upload_root = os.path.join(current_app.root_path, '..', 'uploads', 'quotations')
    upload_root = os.path.abspath(upload_root)
    os.makedirs(upload_root, exist_ok=True)
    safe = f"{quote.quote_no}_LPO_{uuid.uuid4().hex[:6]}_{f.filename.replace('/', '_')}"
    path = os.path.join(upload_root, safe)
    f.save(path)
    quote.lpo_filename = f.filename
    quote.lpo_path = path
    db.session.commit()
    return success_response({'quotation': quote.to_dict()}, message='LPO attached')


@admin_bp.route('/bd/quotations/<int:quote_id>/lpo', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_quote_download_lpo(quote_id):
    user = _bd_current_user()
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    if quote.lpo_cloud_url:
        from flask import redirect
        return redirect(quote.lpo_cloud_url)
    if not quote.lpo_path or not os.path.isfile(quote.lpo_path):
        return error_response('LPO not found', status_code=404, error_code='NOT_FOUND')
    return send_file(quote.lpo_path, as_attachment=True, download_name=quote.lpo_filename or 'lpo.pdf')


@admin_bp.route('/bd/quotations/<int:quote_id>/pdf', methods=['GET'])
@jwt_required()
@bd_access_required
def bd_quote_pdf(quote_id):
    user = _bd_current_user()
    quote = Quotation.query.get_or_404(quote_id)
    if not _can_access_quote(user, quote):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    try:
        from app.bd.quotation_builder import build_quotation_pdf
    except Exception as e:
        current_app.logger.error('quotation_builder: %s', e, exc_info=True)
        return error_response('PDF generator unavailable', status_code=500, error_code='DEPENDENCY_ERROR')
    out_dir = os.path.join(current_app.root_path, '..', 'generated', 'quotations')
    os.makedirs(os.path.abspath(out_dir), exist_ok=True)
    out_path = os.path.abspath(os.path.join(out_dir, f'{quote.quote_no}.pdf'))
    build_quotation_pdf(quote, out_path)
    return send_file(out_path, as_attachment=True, download_name=f'{quote.quote_no}.pdf', mimetype='application/pdf')


@admin_bp.route('/bd/projects/<int:project_id>/promote', methods=['POST'])
@jwt_required()
@bd_access_required
def bd_promote_won_deal(project_id):
    """On win: create/link a ticketing project for delivery."""
    user = _bd_current_user()
    project = BDProject.query.get_or_404(project_id)
    if not _can_access_bd_project(user, project):
        return error_response('Access denied', status_code=403, error_code='OWNERSHIP_REQUIRED')
    data = request.get_json(silent=True) or {}
    project.status = 'won'
    project.stage = 'closing'
    project.progress = 100

    created = {}
    if data.get('create_ticket_project', True):
        existing = TicketProject.query.filter_by(bd_project_id=project.id).first()
        if not existing:
            tp = TicketProject(
                name=project.name[:160],
                client_name=project.company,
                description=project.notes,
                bd_project_id=project.id,
                project_value=project.value_amount,
                is_active=True,
            )
            db.session.add(tp)
            db.session.flush()
            created['ticket_project_id'] = tp.id
        else:
            created['ticket_project_id'] = existing.id

    _bd_activity('🏆', f'Deal won & promoted — {project.name}', project.company, badge='Won', bg='#fff4ef', user_id=user.id)
    db.session.commit()
    return success_response({'project': project.to_dict(), 'created': created}, message='Deal promoted')


# ============== Personal progress (admin workspace) ==============

_PERSONAL_PROJECT_STATUSES = frozenset({'planning', 'active', 'on_hold', 'done', 'archived'})
_PERSONAL_STEP_STATUSES = frozenset({'pending', 'in_progress', 'done', 'blocked', 'skipped'})


def _pp_parse_date(val):
    if not val:
        return None
    try:
        return datetime.fromisoformat(str(val).replace('Z', '+00:00')).date()
    except Exception:
        return _parse_iso_date(val)


def _pp_clear_other_focus(user_id, except_project_id=None):
    q = AdminPersonalProject.query.filter_by(user_id=user_id, is_current_focus=True)
    if except_project_id is not None:
        q = q.filter(AdminPersonalProject.id != except_project_id)
    q.update({'is_current_focus': False}, synchronize_session=False)


def _pp_apply_step_status(step, status):
    st = (status or 'pending').strip().lower()
    if st not in _PERSONAL_STEP_STATUSES:
        st = 'pending'
    step.status = st
    if st == 'done':
        if not step.completed_at:
            step.completed_at = utc_now_naive()
    else:
        step.completed_at = None


def _pp_sync_steps(project, steps_payload, max_steps=120):
    """Replace/update steps from ordered list; preserves ids when possible."""
    if steps_payload is None:
        return
    if not isinstance(steps_payload, list):
        raise ValueError('steps must be a list')
    if len(steps_payload) > max_steps:
        raise ValueError(f'At most {max_steps} steps per project')

    incoming_ids = set()
    for item in steps_payload:
        if not isinstance(item, dict):
            continue
        sid = item.get('id')
        if sid is not None:
            incoming_ids.add(int(sid))

    for row in list(project.steps.all()):
        if row.id not in incoming_ids:
            db.session.delete(row)

    db.session.flush()
    existing = {s.id: s for s in project.steps.all()}

    for order, raw in enumerate(steps_payload):
        if not isinstance(raw, dict):
            continue
        title = (raw.get('title') or '').strip()
        if not title:
            continue
        sid = raw.get('id')
        if sid is not None:
            step = existing.get(int(sid))
            if step and step.project_id == project.id:
                step.title = title[:255]
                step.description = (raw.get('description') or '').strip() or None
                _pp_apply_step_status(step, raw.get('status'))
                step.sort_order = order
                step.due_date = _pp_parse_date(raw.get('dueDate'))
                step.notes = (raw.get('notes') or '').strip() or None
                continue
        step = AdminPersonalProgressStep(
            project_id=project.id,
            title=title[:255],
            description=(raw.get('description') or '').strip() or None,
            sort_order=order,
            due_date=_pp_parse_date(raw.get('dueDate')),
            notes=(raw.get('notes') or '').strip() or None,
        )
        _pp_apply_step_status(step, raw.get('status'))
        db.session.add(step)


@admin_bp.route('/personal-progress', methods=['GET'])
@jwt_required()
@admin_required
def personal_progress_list():
    """List current admin user's personal projects with steps and rollups."""
    try:
        user_id = int(get_jwt_identity())
        status_filter = (request.args.get('status') or '').strip().lower()
        focus_only = request.args.get('focus') in ('1', 'true', 'yes')

        q = AdminPersonalProject.query.filter_by(user_id=user_id)
        if status_filter and status_filter != 'all':
            q = q.filter(AdminPersonalProject.status == status_filter)
        if focus_only:
            q = q.filter_by(is_current_focus=True)

        projects = q.order_by(
            AdminPersonalProject.is_current_focus.desc(),
            AdminPersonalProject.sort_order.asc(),
            AdminPersonalProject.updated_at.desc(),
        ).all()

        data = [p.to_dict(include_steps=True) for p in projects]
        summary = {
            'total': len(data),
            'active': sum(1 for p in projects if (p.status or '') == 'active'),
            'withFocus': sum(1 for p in projects if p.is_current_focus),
        }
        return success_response({'projects': data, 'summary': summary})
    except Exception as e:
        current_app.logger.error(f'personal_progress_list: {e}', exc_info=True)
        return error_response('Failed to load personal progress', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/personal-progress/projects', methods=['POST'])
@jwt_required()
@admin_required
def personal_progress_create_project():
    try:
        user_id = int(get_jwt_identity())
        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        if not title:
            return error_response('Title is required', status_code=400, error_code='VALIDATION_ERROR')

        st = (data.get('status') or 'active').strip().lower()
        if st not in _PERSONAL_PROJECT_STATUSES:
            st = 'active'
        pr = (data.get('priority') or 'med').strip().lower()
        if pr not in ('low', 'med', 'high'):
            pr = 'med'

        focus = bool(data.get('isCurrentFocus') or data.get('is_current_focus'))
        if focus:
            _pp_clear_other_focus(user_id)

        tags = data.get('tags') or []
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(',') if t.strip()]
        if not isinstance(tags, list):
            tags = []

        project = AdminPersonalProject(
            user_id=user_id,
            title=title[:255],
            summary=(data.get('summary') or '').strip() or None,
            status=st,
            priority=pr,
            category=(data.get('category') or '').strip()[:80] or None,
            start_date=_pp_parse_date(data.get('startDate') or data.get('start_date')),
            target_date=_pp_parse_date(data.get('targetDate') or data.get('target_date')),
            link_url=(data.get('linkUrl') or data.get('link_url') or '').strip()[:500] or None,
            tags=tags,
            notes=(data.get('notes') or '').strip() or None,
            is_current_focus=focus,
            sort_order=int(data.get('sortOrder') or data.get('sort_order') or 0),
        )
        db.session.add(project)
        db.session.flush()

        steps_in = data.get('steps')
        if steps_in is not None:
            _pp_sync_steps(project, steps_in)

        db.session.commit()
        project = AdminPersonalProject.query.get(project.id)
        return success_response({'project': project.to_dict(include_steps=True)}, message='Project created', status_code=201)
    except ValueError as ve:
        db.session.rollback()
        return error_response(str(ve), status_code=400, error_code='VALIDATION_ERROR')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'personal_progress_create_project: {e}', exc_info=True)
        return error_response('Failed to create project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/personal-progress/projects/<int:project_id>', methods=['PUT'])
@jwt_required()
@admin_required
def personal_progress_update_project(project_id):
    try:
        user_id = int(get_jwt_identity())
        project = AdminPersonalProject.query.filter_by(id=project_id, user_id=user_id).first_or_404()
        data = request.get_json() or {}

        if 'title' in data:
            t = (data.get('title') or '').strip()
            if not t:
                return error_response('Title cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
            project.title = t[:255]

        if 'summary' in data:
            project.summary = (data.get('summary') or '').strip() or None
        if 'status' in data:
            st = (data.get('status') or project.status or 'active').strip().lower()
            project.status = st if st in _PERSONAL_PROJECT_STATUSES else project.status
        if 'priority' in data:
            pr = (data.get('priority') or 'med').strip().lower()
            project.priority = pr if pr in ('low', 'med', 'high') else project.priority
        if 'category' in data:
            project.category = (data.get('category') or '').strip()[:80] or None
        if 'startDate' in data or 'start_date' in data:
            project.start_date = _pp_parse_date(data.get('startDate', data.get('start_date')))
        if 'targetDate' in data or 'target_date' in data:
            project.target_date = _pp_parse_date(data.get('targetDate', data.get('target_date')))
        if 'linkUrl' in data or 'link_url' in data:
            project.link_url = (data.get('linkUrl') or data.get('link_url') or '').strip()[:500] or None
        if 'notes' in data:
            project.notes = (data.get('notes') or '').strip() or None
        if 'sortOrder' in data or 'sort_order' in data:
            project.sort_order = int(data.get('sortOrder', data.get('sort_order') or 0))

        if 'tags' in data:
            tags = data.get('tags') or []
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(',') if t.strip()]
            project.tags = tags if isinstance(tags, list) else []

        if 'isCurrentFocus' in data or 'is_current_focus' in data:
            focus = bool(data.get('isCurrentFocus', data.get('is_current_focus')))
            if focus:
                _pp_clear_other_focus(user_id, except_project_id=project.id)
            project.is_current_focus = focus

        if 'steps' in data:
            _pp_sync_steps(project, data.get('steps'))

        project.updated_at = utc_now_naive()
        db.session.commit()
        project = AdminPersonalProject.query.get(project_id)
        return success_response({'project': project.to_dict(include_steps=True)}, message='Saved')
    except ValueError as ve:
        db.session.rollback()
        return error_response(str(ve), status_code=400, error_code='VALIDATION_ERROR')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'personal_progress_update_project: {e}', exc_info=True)
        return error_response('Failed to update project', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/personal-progress/projects/<int:project_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def personal_progress_delete_project(project_id):
    try:
        user_id = int(get_jwt_identity())
        project = AdminPersonalProject.query.filter_by(id=project_id, user_id=user_id).first_or_404()
        db.session.delete(project)
        db.session.commit()
        return success_response({}, message='Project deleted')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'personal_progress_delete_project: {e}', exc_info=True)
        return error_response('Failed to delete project', status_code=500, error_code='DATABASE_ERROR')


# ──────────────────────────────────────────────────────────────────────────────
# Technicians CRUD + Excel import/export
# ──────────────────────────────────────────────────────────────────────────────

def _tech_parse_date(raw):
    if not raw:
        return None
    if hasattr(raw, 'date'):
        return raw.date()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(raw).strip(), fmt).date()
        except ValueError:
            pass
    return None


def _technician_login_matches_roster(user, roster_by_email, roster_by_name):
    """Skip login rows that already have a technicians-table record."""
    email = (user.email or '').strip().lower()
    if email and email in roster_by_email:
        return True
    for key in (
        (user.full_name or '').strip().lower(),
        (user.username or '').strip().lower(),
    ):
        if key and key in roster_by_name:
            return True
    return False


def _technician_dict_from_login_user(user):
    """Virtual roster row for accounts whose designation is technician."""
    rm = user.reporting_manager
    return {
        'id': -int(user.id),
        'user_id': user.id,
        'source': 'login',
        'employee_id': f'@{user.username}',
        'full_name': user.full_name or user.username,
        'designation': user.job_designation or 'Technician',
        'department': None,
        'specialization': None,
        'phone': None,
        'email': user.email,
        'salary': None,
        'joining_date': user.employment_start_date.isoformat() if user.employment_start_date else None,
        'status': 'active' if user.is_active else 'inactive',
        'notes': None,
        'created_at': user.created_at.isoformat() if user.created_at else None,
        'supervisor_user_id': user.reporting_manager_id,
        'supervisor_name': (rm.full_name or rm.username) if rm else None,
        'supervisor_username': rm.username if rm else None,
    }


@admin_bp.route('/technicians', methods=['GET'])
@jwt_required()
@admin_required
def list_technicians():
    _ensure_technicians_supervisor_column()
    status_filter = (request.args.get('status') or '').strip().lower()
    q = Technician.query.options(joinedload(Technician.supervisor_user))
    if status_filter and status_filter != 'all':
        q = q.filter(Technician.status == status_filter)
    techs = q.order_by(Technician.full_name).all()
    roster = [t.to_dict() for t in techs]
    for row in roster:
        row['source'] = 'roster'

    roster_by_email = {
        (r.get('email') or '').strip().lower()
        for r in roster
        if (r.get('email') or '').strip()
    }
    roster_by_name = set()
    for r in roster:
        for key in (r.get('full_name'),):
            if key:
                roster_by_name.add(str(key).strip().lower())

    login_users = (
        User.query.options(joinedload(User.reporting_manager))
        .filter(User.designation == 'technician')
        .order_by(User.full_name, User.username)
        .all()
    )
    for row in roster:
        matched = next(
            (u for u in login_users if _technician_login_matches_roster(
                u,
                {(row.get('email') or '').strip().lower()} - {''},
                {(row.get('full_name') or '').strip().lower()} - {''},
            )),
            None,
        )
        if matched:
            row['has_login'] = True
            row['user_id'] = matched.id

    for user in login_users:
        if _technician_login_matches_roster(user, roster_by_email, roster_by_name):
            continue
        login_row = _technician_dict_from_login_user(user)
        if status_filter and status_filter != 'all' and login_row.get('status') != status_filter:
            continue
        roster.append(login_row)

    roster.sort(key=lambda r: (r.get('full_name') or '').lower())
    return success_response({'technicians': roster})


def _create_linked_technician_login(data, *, full_name, email, phone, joining_date, supervisor_user_id, job_title):
    """Create a User(designation=technician) alongside a roster row. Raises ValueError on validation."""
    from app.auth.routes import validate_email, validate_password

    username = (data.get('username') or data.get('login_username') or '').strip()
    if not username:
        raise ValueError('Username is required to create an app login')
    email = (email or '').strip().lower()
    if not email:
        raise ValueError('Email is required to create an app login')
    if not validate_email(email):
        raise ValueError('Invalid email format')
    if User.query.filter_by(username=username).first():
        raise ValueError(f'Username "{username}" already exists')
    if User.query.filter_by(email=email).first():
        raise ValueError('Email already in use')

    raw_pw = (data.get('password') or data.get('login_password') or '').strip()
    temp_password = raw_pw or (os.environ.get('ADMIN_RESET_PASSWORD') or '').strip() or DEFAULT_ADMIN_RESET_PASSWORD
    if raw_pw:
        ok, msg = validate_password(raw_pw)
        if not ok:
            raise ValueError(msg)
    else:
        ok, msg = validate_password(temp_password)
        if not ok:
            temp_password = DEFAULT_ADMIN_RESET_PASSWORD

    user = User(
        username=username,
        email=email,
        full_name=full_name,
        role='user',
        designation='technician',
    )
    user.set_password(temp_password)
    user.password_changed = bool(raw_pw)
    user.job_designation = _truncate_job_designation(job_title) or 'Technician'
    user.employment_start_date = joining_date
    if phone:
        user.phone = phone
    if supervisor_user_id:
        user.reporting_manager_id = supervisor_user_id
    user.access_ticketing = True
    user.access_submitted_forms = True
    db.session.add(user)
    db.session.flush()
    return user, (None if raw_pw else temp_password)


def _suggest_staff_login_username(employee_id, full_name):
    emp = ''.join(ch for ch in (employee_id or '').lower() if ch.isalnum())
    if emp:
        return emp[:80]
    name = '.'.join(part for part in (full_name or '').lower().split() if part)
    name = ''.join(ch if ch.isalnum() or ch == '.' else '' for ch in name).strip('.')
    return (name or 'staff')[:80]


def _user_with_email(email):
    email = (email or '').strip().lower()
    if not email:
        return None
    return User.query.filter(func.lower(User.email) == email).first()


@admin_bp.route('/technicians', methods=['POST'])
@jwt_required()
@admin_required
def create_technician():
    data = request.get_json() or {}
    emp_id = (data.get('employee_id') or '').strip()
    name = (data.get('full_name') or '').strip()
    _ensure_technicians_supervisor_column()
    try:
        sup_coerced = _coerce_optional_supervisor_user_id(data)
    except ValueError as e:
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    if not emp_id or not name:
        return error_response('employee_id and full_name are required', status_code=400, error_code='VALIDATION_ERROR')
    if Technician.query.filter_by(employee_id=emp_id).first():
        return error_response(f'Employee ID "{emp_id}" already exists', status_code=409, error_code='DUPLICATE')
    create_login = bool(data.get('create_login'))
    supervisor_id = sup_coerced if sup_coerced != '__unset__' else None
    phone = (data.get('phone') or '').strip() or None
    email = (data.get('email') or '').strip() or None
    joining_date = _tech_parse_date(data.get('joining_date'))
    job_title = (data.get('designation') or '').strip() or None
    try:
        login_user = None
        temp_password = None
        if create_login:
            login_user, temp_password = _create_linked_technician_login(
                data,
                full_name=name,
                email=email,
                phone=phone,
                joining_date=joining_date,
                supervisor_user_id=supervisor_id,
                job_title=job_title,
            )
        t = Technician(
            employee_id=emp_id,
            full_name=name,
            designation=job_title,
            department=(data.get('department') or '').strip() or None,
            specialization=(data.get('specialization') or '').strip() or None,
            phone=phone,
            email=email,
            salary=float(data['salary']) if data.get('salary') not in (None, '') else None,
            joining_date=joining_date,
            status=(data.get('status') or 'active').strip().lower(),
            notes=(data.get('notes') or '').strip() or None,
            supervisor_user_id=supervisor_id,
        )
        db.session.add(t)
        db.session.commit()
        payload = {'technician': t.to_dict(), 'login_created': bool(login_user)}
        if login_user:
            payload['user'] = _admin_user_dict(login_user)
            try:
                from common.email_service import send_account_created_email
                payload['email_sent'] = send_account_created_email(
                    user_email=login_user.email,
                    username=login_user.username,
                    full_name=login_user.full_name,
                    temp_password=temp_password,
                )
            except Exception as email_error:
                current_app.logger.warning('Account created email failed: %s', email_error)
                payload['email_sent'] = False
        if temp_password:
            payload['temp_password'] = temp_password
        message = 'Technician created'
        if login_user and payload.get('email_sent'):
            message += '. A welcome email with the username was sent.'
        return success_response(payload, message=message, status_code=201)
    except ValueError as e:
        db.session.rollback()
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'create_technician: {e}', exc_info=True)
        return error_response('Failed to create technician', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/technicians/<int:tech_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_technician(tech_id):
    _ensure_technicians_supervisor_column()
    t = Technician.query.get_or_404(tech_id)
    data = request.get_json() or {}
    try:
        sup_coerced = _coerce_optional_supervisor_user_id(data)
    except ValueError as e:
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    if sup_coerced != '__unset__':
        t.supervisor_user_id = sup_coerced
    if 'employee_id' in data:
        new_eid = (data['employee_id'] or '').strip()
        if not new_eid:
            return error_response('employee_id cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
        conflict = Technician.query.filter(Technician.employee_id == new_eid, Technician.id != tech_id).first()
        if conflict:
            return error_response(f'Employee ID "{new_eid}" already in use', status_code=409, error_code='DUPLICATE')
        t.employee_id = new_eid
    if 'full_name' in data:
        n = (data['full_name'] or '').strip()
        if not n:
            return error_response('full_name cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
        t.full_name = n
    for field in ('designation', 'department', 'specialization', 'phone', 'email', 'notes'):
        if field in data:
            setattr(t, field, (data[field] or '').strip() or None)
    if 'salary' in data:
        t.salary = float(data['salary']) if data['salary'] not in (None, '') else None
    if 'joining_date' in data:
        t.joining_date = _tech_parse_date(data['joining_date'])
    if 'status' in data:
        s = (data['status'] or 'active').strip().lower()
        if s in ('active', 'inactive', 'on_leave'):
            t.status = s
    try:
        db.session.commit()
        return success_response({'technician': t.to_dict()})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'update_technician: {e}', exc_info=True)
        return error_response('Failed to update technician', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/technicians/<int:tech_id>/create-login', methods=['POST'])
@jwt_required()
@admin_required
def convert_technician_to_system_user(tech_id):
    """Create a system login from an existing staff (roster) record, keeping their details."""
    _ensure_technicians_supervisor_column()
    t = Technician.query.get_or_404(tech_id)
    data = request.get_json() or {}

    email_override = (data.get('email') or '').strip() or None
    if email_override:
        t.email = email_override
    phone_override = data.get('phone')
    if phone_override is not None:
        t.phone = (phone_override or '').strip() or None
    if 'full_name' in data and (data.get('full_name') or '').strip():
        t.full_name = (data.get('full_name') or '').strip()
    if 'designation' in data:
        t.designation = (data.get('designation') or '').strip() or None
    if 'joining_date' in data:
        t.joining_date = _tech_parse_date(data.get('joining_date'))
    try:
        sup_coerced = _coerce_optional_supervisor_user_id(data)
    except ValueError as e:
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    if sup_coerced != '__unset__':
        t.supervisor_user_id = sup_coerced

    existing = _user_with_email(t.email)
    if existing:
        return error_response(
            'This staff member already has a system account (same email).',
            status_code=409,
            error_code='DUPLICATE_EMAIL',
        )

    login_data = dict(data)
    if not (login_data.get('username') or '').strip():
        login_data['username'] = _suggest_staff_login_username(t.employee_id, t.full_name)

    try:
        login_user, temp_password = _create_linked_technician_login(
            login_data,
            full_name=t.full_name,
            email=t.email,
            phone=t.phone,
            joining_date=t.joining_date,
            supervisor_user_id=t.supervisor_user_id,
            job_title=t.designation,
        )
        db.session.commit()
        payload = {
            'technician': t.to_dict(),
            'login_created': True,
            'user': _admin_user_dict(login_user),
        }
        if temp_password:
            payload['temp_password'] = temp_password
        try:
            from common.email_service import send_account_created_email
            payload['email_sent'] = send_account_created_email(
                user_email=login_user.email,
                username=login_user.username,
                full_name=login_user.full_name,
                temp_password=temp_password,
            )
        except Exception as email_error:
            current_app.logger.warning('Account created email failed: %s', email_error)
            payload['email_sent'] = False
        message = 'Staff converted to a system user'
        if payload.get('email_sent'):
            message += '. A welcome email with the username was sent.'
        return success_response(payload, message=message)
    except ValueError as e:
        db.session.rollback()
        return error_response(str(e), status_code=400, error_code='VALIDATION_ERROR')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'convert_technician_to_system_user: {e}', exc_info=True)
        return error_response('Failed to convert staff to a system user', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/technicians/<int:tech_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_technician(tech_id):
    t = Technician.query.get_or_404(tech_id)
    try:
        db.session.delete(t)
        db.session.commit()
        return success_response({}, message='Technician deleted')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'delete_technician: {e}', exc_info=True)
        return error_response('Failed to delete technician', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/technicians/import-excel', methods=['POST'])
@jwt_required()
@admin_required
def import_technicians_excel():
    """Import technicians from an uploaded Excel file."""
    try:
        import openpyxl
    except ImportError:
        return error_response('openpyxl is required for Excel import (pip install openpyxl)', status_code=500, error_code='DEPENDENCY_ERROR')

    if 'file' not in request.files:
        return error_response('No file uploaded', status_code=400, error_code='VALIDATION_ERROR')
    f = request.files['file']
    if not f.filename:
        return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')

    try:
        wb = openpyxl.load_workbook(BytesIO(f.read()), read_only=True, data_only=True)
        from common.kynvera_excel_brand import resolve_import_sheet_name

        sheet_name = resolve_import_sheet_name(wb.sheetnames, ('Technicians',))
        ws = wb[sheet_name] if sheet_name else wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        return error_response(f'Could not read Excel file: {e}', status_code=400, error_code='PARSE_ERROR')

    if not rows:
        return error_response('Excel file is empty', status_code=400, error_code='VALIDATION_ERROR')

    header_row = [str(c or '').strip().lower().replace(' ', '_') for c in rows[0]]
    COL = {h: i for i, h in enumerate(header_row)}

    def _get(row, key, alt=None):
        idx = COL.get(key, COL.get(alt))
        if idx is None:
            return None
        v = row[idx] if idx < len(row) else None
        return str(v).strip() if v is not None and str(v).strip() else None

    created, skipped, errors = 0, 0, []
    _ensure_technicians_supervisor_column()
    for row_num, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            continue
        emp_id = _get(row, 'employee_id', 'employee_id')
        name = _get(row, 'full_name', 'name')
        if not emp_id or not name:
            errors.append(f'Row {row_num}: employee_id and full_name are required — skipped')
            continue
        if Technician.query.filter_by(employee_id=emp_id).first():
            skipped += 1
            continue
        salary_raw = _get(row, 'salary')
        try:
            salary = float(salary_raw) if salary_raw else None
        except ValueError:
            salary = None
        sup_raw = _get(row, 'supervisor_user_id', 'supervisor_id')
        supervisor_uid = None
        if sup_raw:
            try:
                supervisor_uid = int(str(sup_raw).strip().replace(',', '').split('.', 1)[0])
            except ValueError:
                errors.append(f'Row {row_num}: invalid supervisor_user_id — skipped')
                continue
            if not _valid_roster_supervisor_user(supervisor_uid):
                errors.append(f'Row {row_num}: supervisor_user_id not an active supervisor/manager — skipped')
                continue
        t = Technician(
            employee_id=emp_id,
            full_name=name,
            designation=_get(row, 'designation'),
            department=_get(row, 'department'),
            specialization=_get(row, 'specialization'),
            phone=_get(row, 'phone'),
            email=_get(row, 'email'),
            salary=salary,
            joining_date=_tech_parse_date(_get(row, 'joining_date')),
            status=(_get(row, 'status') or 'active').lower(),
            notes=_get(row, 'notes'),
            supervisor_user_id=supervisor_uid,
        )
        db.session.add(t)
        created += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'import_technicians_excel commit: {e}', exc_info=True)
        return error_response('Database error during import', status_code=500, error_code='DATABASE_ERROR')

    return success_response({
        'created': created,
        'skipped': skipped,
        'errors': errors,
    }, message=f'{created} technician(s) imported, {skipped} skipped (duplicate ID).')


@admin_bp.route('/technicians/export-template', methods=['GET'])
@jwt_required()
@admin_required
def export_technicians_template():
    """Return a blank Excel template for technician bulk import."""
    try:
        from app.admin.excel_templates import build_technicians_template_bytes
    except ImportError:
        return error_response('openpyxl is required (pip install openpyxl)', status_code=500, error_code='DEPENDENCY_ERROR')

    buf = BytesIO(build_technicians_template_bytes())
    return send_file(
        buf,
        as_attachment=True,
        download_name='technicians_import_template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# ---------------------------------------------------------------------------
# Knowledge Base (feeds the Kynvera assistant)
# ---------------------------------------------------------------------------

KB_ALLOWED_EXTENSIONS = {'pdf', 'docx', 'txt', 'md'}
KB_CATEGORIES = ['General', 'HR', 'Inspection', 'Procurement', 'Ticketing',
                 'Reports', 'Business Development', 'Policy', 'IT', 'Workflow']


def _kb_invalidate_cache():
    try:
        from module_assistant.knowledge import invalidate_cache
        invalidate_cache()
    except Exception as e:
        current_app.logger.warning(f"KB cache invalidate failed: {e}")


def _kb_clean_category(value):
    val = (value or 'General').strip()
    return val if val in KB_CATEGORIES else 'General'


@admin_bp.route('/knowledge-base', methods=['GET'])
@jwt_required()
@admin_required
def kb_list_entries():
    """List knowledge base entries with optional search and category filter."""
    try:
        q = (request.args.get('q') or '').strip().lower()
        category = (request.args.get('category') or '').strip()

        query = KnowledgeBaseEntry.query
        if category and category in KB_CATEGORIES:
            query = query.filter(KnowledgeBaseEntry.category == category)
        entries = query.order_by(KnowledgeBaseEntry.updated_at.desc()).all()

        if q:
            entries = [
                e for e in entries
                if q in (e.title or '').lower()
                or q in (e.content or '').lower()
                or q in (e.keywords or '').lower()
            ]

        data = [e.to_dict(include_content=False) for e in entries]
        return success_response({
            'entries': data,
            'count': len(data),
            'categories': KB_CATEGORIES,
        })
    except Exception as e:
        current_app.logger.error(f"Error listing knowledge base: {str(e)}", exc_info=True)
        return error_response('Failed to fetch knowledge base', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/<int:entry_id>', methods=['GET'])
@jwt_required()
@admin_required
def kb_get_entry(entry_id):
    """Get a single knowledge base entry with full content."""
    try:
        entry = KnowledgeBaseEntry.query.get_or_404(entry_id)
        return success_response({'entry': entry.to_dict(include_content=True)})
    except Exception as e:
        current_app.logger.error(f"Error getting KB entry: {str(e)}", exc_info=True)
        return error_response('Failed to fetch entry', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base', methods=['POST'])
@jwt_required()
@admin_required
def kb_create_entry():
    """Create a text knowledge record."""
    try:
        data = request.get_json() or {}
        title = (data.get('title') or '').strip()
        content = (data.get('content') or '').strip()
        if not title:
            return error_response('Title is required', status_code=400, error_code='VALIDATION_ERROR')
        if not content:
            return error_response('Content is required for a text record', status_code=400, error_code='VALIDATION_ERROR')

        keywords = (data.get('keywords') or '').strip()
        if isinstance(data.get('keywords'), list):
            keywords = ', '.join(str(k).strip() for k in data['keywords'] if str(k).strip())

        entry = KnowledgeBaseEntry(
            title=title[:255],
            content=content,
            keywords=keywords or None,
            category=_kb_clean_category(data.get('category')),
            answer_link=(data.get('answer_link') or '').strip() or None,
            source_type='text',
            is_active=bool(data.get('is_active', True)),
            created_by=get_jwt_identity(),
        )
        db.session.add(entry)
        db.session.commit()
        _kb_invalidate_cache()
        return success_response({'entry': entry.to_dict()}, message='Knowledge record created', status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error creating KB entry: {str(e)}", exc_info=True)
        return error_response('Failed to create knowledge record', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/upload', methods=['POST'])
@jwt_required()
@admin_required
def kb_upload_document():
    """Upload a document; extract its text into a knowledge record."""
    try:
        from werkzeug.utils import secure_filename
        from module_assistant.extract import extract_text

        f = request.files.get('file')
        if not f or not f.filename:
            return error_response('No file selected', status_code=400, error_code='VALIDATION_ERROR')

        ext = (f.filename.rsplit('.', 1)[1].lower() if '.' in f.filename else '')
        if ext not in KB_ALLOWED_EXTENSIONS:
            return error_response(
                'Unsupported file type. Allowed: PDF, DOCX, TXT, MD',
                status_code=400, error_code='VALIDATION_ERROR'
            )

        generated_root = current_app.config.get('GENERATED_DIR')
        if not generated_root:
            return error_response('Generated directory not configured', status_code=500, error_code='CONFIG_ERROR')

        kb_dir = os.path.join(generated_root, 'knowledge')
        os.makedirs(kb_dir, exist_ok=True)

        original_name = secure_filename(f.filename)
        import uuid
        unique_name = f"{uuid.uuid4().hex[:12]}_{original_name}"
        stored_path = os.path.join(kb_dir, unique_name)
        f.save(stored_path)

        extracted = extract_text(stored_path, ext)

        title = (request.form.get('title') or '').strip()
        if not title:
            title = os.path.splitext(original_name)[0].replace('_', ' ').strip() or 'Untitled Document'

        keywords = (request.form.get('keywords') or '').strip()

        entry = KnowledgeBaseEntry(
            title=title[:255],
            content=extracted,
            keywords=keywords or None,
            category=_kb_clean_category(request.form.get('category')),
            answer_link=(request.form.get('answer_link') or '').strip() or None,
            source_type='upload',
            file_name=original_name,
            stored_path=stored_path,
            file_type=ext.upper(),
            is_active=str(request.form.get('is_active', 'true')).lower() != 'false',
            created_by=get_jwt_identity(),
        )
        db.session.add(entry)
        db.session.commit()
        _kb_invalidate_cache()

        msg = 'Document uploaded and indexed'
        if not extracted:
            msg = 'Document uploaded, but no text could be extracted. You can add content manually.'
        return success_response({'entry': entry.to_dict(include_content=False)}, message=msg, status_code=201)
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error uploading KB document: {str(e)}", exc_info=True)
        return error_response('Failed to upload document', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/<int:entry_id>', methods=['PUT'])
@jwt_required()
@admin_required
def kb_update_entry(entry_id):
    """Update a knowledge record or toggle its active state."""
    try:
        entry = KnowledgeBaseEntry.query.get_or_404(entry_id)
        data = request.get_json() or {}

        if 'title' in data:
            t = (data.get('title') or '').strip()
            if not t:
                return error_response('Title cannot be empty', status_code=400, error_code='VALIDATION_ERROR')
            entry.title = t[:255]
        if 'content' in data:
            entry.content = (data.get('content') or '').strip()
        if 'keywords' in data:
            kw = data.get('keywords')
            if isinstance(kw, list):
                kw = ', '.join(str(k).strip() for k in kw if str(k).strip())
            entry.keywords = (kw or '').strip() or None
        if 'category' in data:
            entry.category = _kb_clean_category(data.get('category'))
        if 'answer_link' in data:
            entry.answer_link = (data.get('answer_link') or '').strip() or None
        if 'is_active' in data:
            entry.is_active = bool(data.get('is_active'))

        db.session.commit()
        _kb_invalidate_cache()
        return success_response({'entry': entry.to_dict()}, message='Knowledge record updated')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error updating KB entry: {str(e)}", exc_info=True)
        return error_response('Failed to update knowledge record', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/<int:entry_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def kb_delete_entry(entry_id):
    """Delete a knowledge record (and its stored file if any)."""
    try:
        entry = KnowledgeBaseEntry.query.get_or_404(entry_id)
        title = entry.title
        stored_path = entry.stored_path

        db.session.delete(entry)
        db.session.commit()
        _kb_invalidate_cache()

        if stored_path and os.path.isfile(stored_path):
            try:
                os.remove(stored_path)
            except OSError:
                pass

        return success_response({'message': f'Knowledge record "{title}" deleted'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting KB entry: {str(e)}", exc_info=True)
        return error_response('Failed to delete knowledge record', status_code=500, error_code='DATABASE_ERROR')


# Marks the start of the full page text within a 'link' record's content,
# so the assistant excerpt surfaces the summary while search covers everything.
_KB_LINK_SEPARATOR = '\n\n--- Full page content ---\n\n'


def _kb_build_link_content(summary, full_text):
    """Store summary first (for excerpts) + full text (for keyword search)."""
    summary = (summary or '').strip()
    full_text = (full_text or '').strip()
    if summary and full_text:
        return f"{summary}{_KB_LINK_SEPARATOR}{full_text}"
    return summary or full_text


@admin_bp.route('/knowledge-base/link', methods=['POST'])
@jwt_required()
@admin_required
def kb_add_link():
    """Fetch a URL, summarize it, and store it as a knowledge record."""
    try:
        from module_assistant.fetch_url import fetch_url_text, summarize_extractive, FetchError

        data = request.get_json() or {}
        url = (data.get('url') or '').strip()
        if not url:
            return error_response('A URL is required', status_code=400, error_code='VALIDATION_ERROR')

        try:
            page_title, full_text = fetch_url_text(url)
        except FetchError as fe:
            return error_response(str(fe), status_code=400, error_code='FETCH_ERROR')

        summary = summarize_extractive(full_text)

        title = (data.get('title') or '').strip() or (page_title or '').strip() or url
        keywords = (data.get('keywords') or '').strip()
        if isinstance(data.get('keywords'), list):
            keywords = ', '.join(str(k).strip() for k in data['keywords'] if str(k).strip())

        entry = KnowledgeBaseEntry(
            title=title[:255],
            content=_kb_build_link_content(summary, full_text),
            keywords=keywords or None,
            category=_kb_clean_category(data.get('category')),
            answer_link=(data.get('answer_link') or '').strip() or url,
            source_type='link',
            source_url=url[:1000],
            fetched_at=utc_now_naive(),
            is_active=bool(data.get('is_active', True)),
            created_by=get_jwt_identity(),
        )
        db.session.add(entry)
        db.session.commit()
        _kb_invalidate_cache()
        return success_response(
            {'entry': entry.to_dict(include_content=False), 'summary': summary},
            message='Link fetched, summarized and indexed',
            status_code=201,
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error adding KB link: {str(e)}", exc_info=True)
        return error_response('Failed to add link', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/<int:entry_id>/refetch', methods=['POST'])
@jwt_required()
@admin_required
def kb_refetch_link(entry_id):
    """Re-fetch and re-summarize an existing link record."""
    try:
        from module_assistant.fetch_url import fetch_url_text, summarize_extractive, FetchError

        entry = KnowledgeBaseEntry.query.get_or_404(entry_id)
        if entry.source_type != 'link' or not entry.source_url:
            return error_response('This record is not a link', status_code=400, error_code='VALIDATION_ERROR')

        try:
            page_title, full_text = fetch_url_text(entry.source_url)
        except FetchError as fe:
            return error_response(str(fe), status_code=400, error_code='FETCH_ERROR')

        summary = summarize_extractive(full_text)
        entry.content = _kb_build_link_content(summary, full_text)
        entry.fetched_at = utc_now_naive()
        db.session.commit()
        _kb_invalidate_cache()
        return success_response(
            {'entry': entry.to_dict(include_content=False), 'summary': summary},
            message='Link content refreshed',
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error refetching KB link: {str(e)}", exc_info=True)
        return error_response('Failed to refresh link', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/knowledge-base/refresh-app', methods=['POST'])
@jwt_required()
@admin_required
def kb_refresh_app_data():
    """Auto-generate / update KB entries from live application data."""
    from datetime import datetime, timezone
    from app.models import Submission, Ticket

    def _utcnow_naive():
        return datetime.now(timezone.utc).replace(tzinfo=None)

    def _upsert(title, content, keywords, category, answer_link=None):
        entry = KnowledgeBaseEntry.query.filter_by(title=title, source_type='auto').first()
        if entry:
            entry.content = content
            entry.keywords = keywords
            entry.updated_at = _utcnow_naive()
            return 'updated'
        else:
            entry = KnowledgeBaseEntry(
                title=title,
                content=content,
                keywords=keywords,
                category=category,
                answer_link=answer_link,
                source_type='auto',
                is_active=True,
                created_by=get_jwt_identity(),
            )
            db.session.add(entry)
            return 'created'

    try:
        results = {}

        # ── Procurement snapshot ─────────────────────────────────────────────
        try:
            proc_materials = Submission.query.filter(
                Submission.module_type.in_(['procurement_material', 'catalog_material'])
            ).all()
            proc_props = Submission.query.filter_by(module_type='procurement_property').all()

            import json as _json
            mat_names = []
            for s in proc_materials[:10]:
                fd = s.form_data
                if isinstance(fd, str):
                    try:
                        fd = _json.loads(fd)
                    except Exception:
                        fd = {}
                name = (fd or {}).get('material_name') or (fd or {}).get('name') or ''
                if name:
                    mat_names.append(name)

            prop_names = []
            for s in proc_props[:10]:
                fd = s.form_data
                if isinstance(fd, str):
                    try:
                        fd = _json.loads(fd)
                    except Exception:
                        fd = {}
                pn = (fd or {}).get('property_name') or s.site_name or ''
                if pn and pn not in prop_names:
                    prop_names.append(pn)

            mat_line = f"There are {len(proc_materials)} material record(s) in the catalog."
            if mat_names:
                mat_line += f" Examples: {', '.join(mat_names[:8])}."
            prop_line = f"There are {len(proc_props)} registered property/properties."
            if prop_names:
                prop_line += f" Properties: {', '.join(prop_names[:8])}."

            results['procurement'] = _upsert(
                title='[Auto] Procurement — Live Snapshot',
                content=(
                    f"Procurement module live data snapshot (refreshed {_utcnow_naive().strftime('%d %b %Y')}).\n\n"
                    f"{mat_line}\n{prop_line}\n\n"
                    "The Procurement module manages material catalogs, registered properties, and pricing lists. "
                    "Users with procurement access can add materials, assign them to properties, and export reports."
                ),
                keywords='procurement, materials, catalog, properties, pricing, material list',
                category='Procurement',
                answer_link='/procurement/',
            )
        except Exception as e:
            current_app.logger.warning(f"KB refresh: procurement error: {e}")
            results['procurement'] = 'error'

        # ── Ticketing snapshot ───────────────────────────────────────────────
        try:
            OPEN_S = {'open', 'pending_supervisor'}
            IN_PROG_S = {'in_progress', 'pending_parts', 'pending_verification'}
            all_tickets = Ticket.query.all()
            open_c = sum(1 for t in all_tickets if t.status in OPEN_S)
            in_prog_c = sum(1 for t in all_tickets if t.status in IN_PROG_S)
            closed_c = sum(1 for t in all_tickets if t.status == 'closed')

            recent_titles = [t.title for t in sorted(all_tickets, key=lambda x: x.created_at or _utcnow_naive(), reverse=True)[:5]]

            results['ticketing'] = _upsert(
                title='[Auto] Ticketing — Live Snapshot',
                content=(
                    f"Ticketing module live data snapshot (refreshed {_utcnow_naive().strftime('%d %b %Y')}).\n\n"
                    f"Total tickets: {len(all_tickets)}. Open: {open_c}. In progress: {in_prog_c}. Closed: {closed_c}.\n"
                    + (f"Recent ticket titles: {', '.join(recent_titles)}.\n" if recent_titles else '')
                    + "\nThe Ticketing module handles work orders and service requests. "
                    "Ticket statuses flow: open → pending_supervisor → in_progress → pending_parts → "
                    "pending_verification → closed. Reporters can track status; supervisors assign technicians."
                ),
                keywords='tickets, work orders, open tickets, service requests, ticketing, status',
                category='Ticketing',
                answer_link='/tickets/',
            )
        except Exception as e:
            current_app.logger.warning(f"KB refresh: ticketing error: {e}")
            results['ticketing'] = 'error'

        # ── Inspection snapshot ──────────────────────────────────────────────
        try:
            INSP_TYPES = ('hvac_mep', 'civil', 'cleaning')
            from datetime import date
            from sqlalchemy import extract
            now = _utcnow_naive()
            insp_all = Submission.query.filter(
                Submission.module_type.in_(INSP_TYPES)
            ).all()
            by_type = {t: sum(1 for s in insp_all if s.module_type == t) for t in INSP_TYPES}
            this_month = sum(
                1 for s in insp_all
                if (s.created_at or now).month == now.month and (s.created_at or now).year == now.year
            )
            results['inspection'] = _upsert(
                title='[Auto] Inspection — Live Snapshot',
                content=(
                    f"Inspection module live data snapshot (refreshed {now.strftime('%d %b %Y')}).\n\n"
                    f"Total inspections submitted: {len(insp_all)}. This month: {this_month}.\n"
                    f"HVAC & MEP: {by_type['hvac_mep']}. Civil Works: {by_type['civil']}. Cleaning: {by_type['cleaning']}.\n\n"
                    "Inspection forms cover HVAC & MEP, Civil Works, and Cleaning Services. "
                    "Each form includes site details, checklist items, photos, and signatures. "
                    "Submitted forms go through supervisor → operations manager → GM approval workflow."
                ),
                keywords='inspection, hvac, mep, civil, cleaning, site visit, inspection form, checklist',
                category='Inspection',
                answer_link='/inspection/',
            )
        except Exception as e:
            current_app.logger.warning(f"KB refresh: inspection error: {e}")
            results['inspection'] = 'error'

        # ── QHSI is hidden from production — drop auto snapshots ───────────────
        try:
            KnowledgeBaseEntry.query.filter(
                KnowledgeBaseEntry.source_type == 'auto',
                or_(
                    KnowledgeBaseEntry.title.like('%QHSI%'),
                    KnowledgeBaseEntry.category == 'QHSI',
                ),
            ).delete(synchronize_session=False)
            results['qhsi'] = 'hidden'
        except Exception as e:
            current_app.logger.warning(f"KB refresh: qhsi hide error: {e}")
            results['qhsi'] = 'error'

        # ── Modules overview ─────────────────────────────────────────────────
        results['overview'] = _upsert(
            title='[Auto] Kynvera Modules — Overview',
            content=(
                "Kynvera modules overview (auto-generated).\n\n"
                "Available modules:\n"
                "- Inspection Hub: HVAC & MEP, Civil Works, and Cleaning Services inspection forms with photo capture, signatures, and PDF reports.\n"
                "- Procurement: Material catalog management, registered properties, pricing and Excel import/export.\n"
                "- Ticketing: Work order and service request management with supervisor/technician workflow.\n"
                "- HR: Leave applications, commencement, duty resumption, termination, asset tracking, visa/passport processing, grievances, appraisals.\n"
                "- DocHub: Company document library with policies, manuals and downloadable files.\n"
                "- MMR: Monthly Maintenance Reports with chargeable configuration and automated scheduling.\n"
                "- Business Development: BD email module for project pipeline and client communications.\n"
                "- Workflow Hub: Central dashboard for pending reviews and submitted forms.\n"
                "- Dashboard: Quick stats, recent activity, and navigation to all modules."
            ),
            keywords='modules, features, hvac, procurement, ticketing, hr, dochub, mmr, bd, workflow, inspection',
            category='General',
            answer_link='/dashboard',
        )

        db.session.commit()
        _kb_invalidate_cache()

        created = sum(1 for v in results.values() if v == 'created')
        updated = sum(1 for v in results.values() if v == 'updated')
        return success_response(
            {'results': results, 'created': created, 'updated': updated},
            message=f'App data refreshed: {created} created, {updated} updated.',
        )

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"KB refresh-app error: {e}", exc_info=True)
        return error_response('Failed to refresh app data', status_code=500, error_code='INTERNAL_ERROR')


@admin_bp.route('/database/status', methods=['GET'])
@jwt_required()
@admin_required
def database_status():
    """Plain-language status of the database this running app is using."""
    try:
        from app.database_admin import describe_current_database, module_table_counts

        payload = describe_current_database()
        payload['modules'] = module_table_counts()
        return success_response(payload)
    except Exception as e:
        current_app.logger.error(f'Database status: {e}', exc_info=True)
        return error_response('Could not read database status', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/database/backups', methods=['GET'])
@jwt_required()
@admin_required
def database_backups_list():
    """Recent backup downloads recorded on this instance."""
    try:
        from app.database_admin import list_backups

        return success_response({'backups': list_backups(40)})
    except Exception as e:
        current_app.logger.error(f'Database backups list: {e}', exc_info=True)
        return error_response('Could not load backup history', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/database/tables/<table_name>', methods=['GET'])
@jwt_required()
@admin_required
def database_browse_table(table_name):
    """Read-only page of rows for one table on THIS instance's database."""
    try:
        from app.database_admin import UnknownTable, browse_table

        page = request.args.get('page', 1)
        per_page = request.args.get('per_page', 50)
        q = request.args.get('q', '') or ''
        payload = browse_table(table_name, page=page, per_page=per_page, q=q)
        return success_response(payload)
    except UnknownTable as e:
        return error_response(str(e), status_code=404, error_code='NOT_FOUND')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Database browse {table_name}: {e}', exc_info=True)
        return error_response('Could not read that table', status_code=500, error_code='DATABASE_ERROR')


@admin_bp.route('/database/backup', methods=['POST'])
@jwt_required()
@admin_required
def database_backup_download():
    """Create a copy of THIS instance's database and stream it to the admin."""
    from flask import after_this_request
    from app.database_admin import (
        BackupError, BackupTooLarge, create_backup_file,
        current_engine_kind, current_environment,
    )

    admin_id = get_jwt_identity()
    try:
        admin_id = int(admin_id)
    except (TypeError, ValueError):
        pass

    env = current_environment()
    engine = current_engine_kind()
    artifact = None
    try:
        artifact = create_backup_file()
        row = DatabaseBackup(
            created_by_user_id=admin_id if isinstance(admin_id, int) else None,
            environment=env,
            engine=engine,
            filename=artifact['filename'],
            size_bytes=artifact['size_bytes'],
            status='ok',
            kind='download',
        )
        db.session.add(row)
        db.session.commit()
        log_audit(admin_id, 'database_backup', 'database', env, {
            'filename': artifact['filename'],
            'size_bytes': artifact['size_bytes'],
            'engine': engine,
        })

        path = artifact['path']
        keep = artifact.get('keep')

        if not keep:
            @after_this_request
            def _cleanup_backup(response):
                try:
                    if path and os.path.isfile(path):
                        os.remove(path)
                except OSError:
                    pass
                return response

        return send_file(
            path,
            as_attachment=True,
            download_name=artifact['filename'],
            mimetype=artifact.get('mimetype') or 'application/octet-stream',
        )
    except BackupTooLarge as e:
        db.session.rollback()
        _record_failed_backup(admin_id, env, engine, str(e))
        return error_response(str(e), status_code=413, error_code='BACKUP_TOO_LARGE')
    except BackupError as e:
        db.session.rollback()
        _record_failed_backup(admin_id, env, engine, str(e))
        return error_response(str(e), status_code=500, error_code='BACKUP_FAILED')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Database backup: {e}', exc_info=True)
        _record_failed_backup(admin_id, env, engine, 'Backup failed')
        if artifact and artifact.get('path') and not artifact.get('keep'):
            try:
                os.remove(artifact['path'])
            except OSError:
                pass
        return error_response('Could not create a database copy', status_code=500, error_code='BACKUP_FAILED')


def _record_failed_backup(admin_id, env, engine, message):
    try:
        db.session.rollback()
        row = DatabaseBackup(
            created_by_user_id=admin_id if isinstance(admin_id, int) else None,
            environment=env,
            engine=engine,
            filename='(failed)',
            size_bytes=0,
            status='failed',
            error_message=(message or '')[:500],
            kind='download',
        )
        db.session.add(row)
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.warning('Could not record failed database backup', exc_info=True)
