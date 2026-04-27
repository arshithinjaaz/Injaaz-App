"""
Civil report generators for Excel and PDF.
"""
import os
import logging
import sys
from datetime import datetime, timedelta

from common.datetime_utils import utc_now_naive

# Dubai timezone offset (Gulf Standard Time, UTC+4)
DUBAI_OFFSET = timedelta(hours=4)

def get_dubai_time():
    """Get current time in Dubai timezone (GST - Gulf Standard Time, UTC+4)"""
    utc_now = utc_now_naive()
    return utc_now + DUBAI_OFFSET

def format_dubai_datetime(dt=None, format_str='%Y-%m-%d %H:%M:%S'):
    """Format datetime in Dubai timezone (GST, UTC+4)"""
    if dt is None:
        dt = get_dubai_time()
    elif isinstance(dt, datetime):
        # Assume UTC if naive, add Dubai offset
        if dt.tzinfo is None:
            dt = dt + DUBAI_OFFSET
        else:
            # Convert to UTC first, then add Dubai offset
            from datetime import timezone as dt_timezone
            utc_dt = dt.astimezone(dt_timezone.utc).replace(tzinfo=None)
            dt = utc_dt + DUBAI_OFFSET
    else:
        dt = get_dubai_time()
    return dt.strftime(format_str)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from common.utils import get_image_for_pdf

logger = logging.getLogger(__name__)

# Try importing PIL for better image handling
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    logger.warning("PIL/Pillow not available - signature aspect ratio may not be perfectly preserved")

# Try importing professional PDF service, fall back if unavailable
try:
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from app.services.professional_pdf_service import (
        create_professional_pdf,
        create_header_with_logo,
        create_info_table,
        create_data_table,
        add_photo_grid,
        add_signatures_section,
        add_section_heading,
        add_item_heading,
        add_paragraph,
        get_professional_styles
    )
    USE_PROFESSIONAL_PDF = True
    logger.info("✅ Professional PDF service loaded successfully")
except Exception as e:
    logger.warning(f"⚠️ Professional PDF service not available: {e}. Using basic PDF generation.")
    USE_PROFESSIONAL_PDF = False

def _write_materials_sheet(ws, materials):
    """Write a 'Materials Used' sheet - matches reference format."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill('solid', fgColor='125435')
    alt_fill     = PatternFill('solid', fgColor='E3F2FD')  # Light blue for zebra striping
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    body_font    = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right',  vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='BBDEFB')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Material Name', 'Brand', 'Department', 'UOM', 'Quantity', 'Unit Price (AED)', 'Line Total (AED)']
    col_widths = [6, 42, 16, 14, 10, 10, 18, 20]

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "MATERIALS & COST BREAKDOWN"
    title_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    title_cell.alignment = center_align
    title_cell.fill = header_fill
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = "Selected inspection materials with pricing and cost totals"
    sub_cell.font = Font(name='Calibri', bold=False, color='125435', size=10)
    sub_cell.alignment = center_align
    sub_cell.fill = PatternFill('solid', fgColor='E8F5E9')
    ws.row_dimensions[2].height = 22

    header_row = 4
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    grand_total = 0.0
    data_start_row = header_row + 1
    for idx, m in enumerate(materials, 1):
        row_idx = data_start_row + idx - 1
        row_fill = PatternFill('solid', fgColor='FFFFFF') if row_idx % 2 == 0 else alt_fill
        qty = float(m.get('quantity', 1) or 0)
        unit_price = float(m.get('unit_price', 0) or 0)
        line_total = qty * unit_price
        grand_total += line_total
        row_data = [
            idx,
            str(m.get('name', '')),
            str(m.get('brand', '') or ''),
            str(m.get('department', '') or ''),
            str(m.get('uom', '') or ''),
            qty,
            unit_price,
            line_total,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = center_align if col_idx in (1, 6) else (right_align if col_idx in (7, 8) else left_align)
            if col_idx in (7, 8):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    total_row = data_start_row + len(materials)
    for col_idx in range(1, 9):
        cell = ws.cell(row=total_row, column=col_idx, value='' if col_idx < 7 else None)
        cell.font = Font(name='Calibri', bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='E8F5E9')
        cell.border = border
        cell.alignment = right_align if col_idx in (7, 8) else center_align
    grand_total_cell = ws.cell(row=total_row, column=7, value='Grand Total (AED)')
    grand_total_cell.alignment = right_align
    total_value_cell = ws.cell(row=total_row, column=8, value=grand_total)
    total_value_cell.alignment = right_align
    total_value_cell.number_format = '#,##0.00'
    ws.row_dimensions[total_row].height = 24

    data_end_row = total_row - 1
    if data_end_row >= data_start_row:
        ws.auto_filter.ref = f"A{header_row}:H{data_end_row}"
    ws.freeze_panes = f"A{data_start_row}"


def _add_summary_kpi_cards(ws, start_row, cards):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row + 1].height = 28

    groups = [(1, 2), (3, 4), (5, 6), (7, 8)]
    for idx, (c1, c2) in enumerate(groups):
        if idx >= len(cards):
            break
        label, value = cards[idx]
        ws.merge_cells(start_row=start_row, start_column=c1, end_row=start_row, end_column=c2)
        ws.merge_cells(start_row=start_row + 1, start_column=c1, end_row=start_row + 1, end_column=c2)

        lbl_cell = ws.cell(row=start_row, column=c1, value=label)
        lbl_cell.font = Font(name='Calibri', size=9, bold=True, color='125435')
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')
        lbl_cell.fill = PatternFill('solid', fgColor='E8F5E9')
        lbl_cell.border = border

        val_cell = ws.cell(row=start_row + 1, column=c1, value=value)
        val_cell.font = Font(name='Calibri', size=13, bold=True, color='0F172A')
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        val_cell.fill = PatternFill('solid', fgColor='FFFFFF')
        val_cell.border = border

        ws.cell(row=start_row, column=c2).fill = PatternFill('solid', fgColor='E8F5E9')
        ws.cell(row=start_row, column=c2).border = border
        ws.cell(row=start_row + 1, column=c2).fill = PatternFill('solid', fgColor='FFFFFF')
        ws.cell(row=start_row + 1, column=c2).border = border

    return start_row + 3


def create_excel_report(data, output_dir):
    """Generate Civil Works Excel report with professional formatting."""
    try:
        # Import professional Excel service
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app.services.professional_excel_service import (
            create_professional_excel_workbook,
            add_logo_and_title,
            add_info_section,
            add_data_table,
            add_section_header,
            finalize_workbook,
            recenter_logo
        )
        
        logger.info(f"Creating professional Civil Excel report in {output_dir}")
        
        # Extract data - data is already in the correct format from submit_with_urls
        project_name = data.get('project_name', 'Unknown_Project')
        project_name = project_name.replace(' ', '_') if project_name else 'Unknown_Project'

        # Collect all photos from work items
        work_items = data.get('work_items', [])
        logger.info(f"📸 Processing {len(work_items)} work items for Excel")
        all_photos = []
        for idx, item in enumerate(work_items):
            photos = item.get('photos', [])
            logger.info(f"  Work item {idx + 1}: {len(photos)} photos")
            all_photos.extend(photos)
        logger.info(f"📸 Total photos collected: {len(all_photos)}")
        
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"Civil_{project_name}_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        materials = data.get('materials_required', [])

        materials_total = 0.0
        for m in materials if isinstance(materials, list) else []:
            try:
                materials_total += float(m.get('quantity', 1) or 0) * float(m.get('unit_price', 0) or 0)
            except Exception:
                pass
        priced_materials = sum(
            1 for m in materials if isinstance(materials, list)
            if (str(m.get('unit_price', '')).strip() not in ('', '0', '0.0', '0.00'))
        )

        # Sheet 1: Summary
        wb, ws_summary = create_professional_excel_workbook(
            title="Civil Works Inspection Report",
            sheet_name="Summary"
        )
        current_row = add_logo_and_title(
            ws_summary,
            title="CIVIL WORKS INSPECTION REPORT",
            subtitle=f"Project: {project_name.replace('_', ' ')}",
            max_columns=8
        )
        current_row = add_info_section(
            ws_summary,
            [
                ('Project Name', data.get('project_name', 'N/A')),
                ('Visit Date', data.get('visit_date', 'N/A')),
                ('Location', data.get('location', 'N/A')),
                ('Inspector', data.get('inspector_name', 'N/A')),
                ('Report Generated', format_dubai_datetime() + ' (GST)'),
                ('Work Items', str(len(work_items))),
                ('Selected Materials', str(len(materials) if isinstance(materials, list) else 0)),
            ],
            current_row,
            title="Executive Overview",
            max_columns=8
        )
        current_row = _add_summary_kpi_cards(ws_summary, current_row, [
            ("Total Work Items", len(work_items)),
            ("Total Materials", len(materials) if isinstance(materials, list) else 0),
            ("Priced Materials", priced_materials),
            ("Materials Value (AED)", f"{materials_total:,.2f}")
        ])
        ws_summary.freeze_panes = "A8"
        finalize_workbook(ws_summary)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [23, 32, 17, 10, 18, 10, 23, 10]):
            ws_summary.column_dimensions[_cl].width = _w
        recenter_logo(ws_summary)

        # Sheet 2: Work Items
        ws_items = wb.create_sheet(title="Work Items")
        items_row = add_logo_and_title(
            ws_items,
            title="CIVIL WORKS - ITEM REGISTER",
            subtitle=f"Project: {project_name.replace('_', ' ')}",
            max_columns=8
        )
        items_row = add_info_section(
            ws_items,
            [
                ('Work Description', data.get('description_of_work', 'N/A')),
                ('Area', data.get('area', 'N/A')),
                ('Total Photos', str(len(all_photos))),
            ],
            items_row,
            title="Execution Scope",
            max_columns=8
        )
        headers = ['#', 'Description', 'Quantity', 'Material', 'Material Qty', 'Price', 'Labour', 'Comments']
        table_data = []
        for idx, item in enumerate(work_items, 1):
            table_data.append([
                idx,
                item.get('description', 'N/A'),
                float(item.get('quantity', 0) or 0),
                item.get('material', 'N/A'),
                item.get('material_qty', 'N/A'),
                item.get('price', 'N/A'),
                item.get('labour', 'N/A'),
                item.get('comments', 'N/A')
            ])
        col_widths = {
            'A': 6,
            'B': 32,
            'C': 10,
            'D': 18,
            'E': 14,
            'F': 12,
            'G': 14,
            'H': 26
        }
        table_start = add_section_header(ws_items, "Work Item Details", items_row, span_columns=8)
        add_data_table(ws_items, headers, table_data, table_start, col_widths=col_widths)
        if table_data:
            header_row = table_start
            last_row = header_row + len(table_data)
            ws_items.auto_filter.ref = f"A{header_row}:H{last_row}"
            ws_items.freeze_panes = f"A{header_row + 1}"
        finalize_workbook(ws_items)
        # Lock column widths appropriate for Work Items content
        for _cl, _w in zip('ABCDEFGH', [16, 40, 12, 25, 16, 12, 20, 32]):
            ws_items.column_dimensions[_cl].width = _w
        recenter_logo(ws_items)

        # Sheet 3: Materials & Cost
        ws_mat = wb.create_sheet(title="Materials & Cost")
        if materials and isinstance(materials, list):
            _write_materials_sheet(ws_mat, materials)
        else:
            ws_mat.merge_cells('A1:H1')
            ws_mat['A1'] = "MATERIALS & COST BREAKDOWN"
            ws_mat.merge_cells('A3:H3')
            ws_mat['A3'] = "No materials were selected for this inspection."
        finalize_workbook(ws_mat)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [13, 25, 15, 17, 11, 15, 24, 25]):
            ws_mat.column_dimensions[_cl].width = _w

        # Save workbook
        wb.save(excel_path)
        
        if not os.path.exists(excel_path):
            raise Exception(f"Excel file not created at {excel_path}")
        
        logger.info(f"✅ Professional Civil Excel report created: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"❌ Civil Excel generation error: {str(e)}")
        raise

def create_pdf_report(data, output_dir):
    """Generate professional Civil Works PDF report with branding."""
    try:
        logger.info(f"Creating professional Civil PDF report in {output_dir}")
        
        # Handle data structure - could be direct form_data or wrapped in submission dict
        if 'form_data' in data:
            # Data from database (wrapped in submission dict)
            form_data = data.get('form_data', {})
            project_name = form_data.get('project_name') or data.get('site_name', 'Unknown_Project')
            work_items = form_data.get('work_items', [])
            # Extract other fields from form_data
            inspector_name = form_data.get('inspector_name', 'N/A')
            description_of_work = form_data.get('description_of_work', 'N/A')
            location = form_data.get('location', 'N/A')
            visit_date = form_data.get('visit_date') or data.get('visit_date', 'N/A')
            inspector_signature = form_data.get('supervisor_signature', '') or form_data.get('inspector_signature', {})
            supervisor_comments = form_data.get('supervisor_comments', '')
        else:
            # Direct form_data (from submit_with_urls)
            form_data = data
            project_name = data.get('project_name', 'Unknown_Project')
            work_items = data.get('work_items', [])
            inspector_name = data.get('inspector_name', 'N/A')
            description_of_work = data.get('description_of_work', 'N/A')
            location = data.get('location', 'N/A')
            visit_date = data.get('visit_date', 'N/A')
            inspector_signature = data.get('supervisor_signature', '') or data.get('inspector_signature', {})
            supervisor_comments = data.get('supervisor_comments', '')
        
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"Civil_{project_name.replace(' ', '_')}_{timestamp}.pdf"
        pdf_path = os.path.join(output_dir, pdf_filename)
        
        # Collect all photos from work items
        logger.info(f"📸 Processing {len(work_items)} work items for PDF")
        logger.info(f"📸 Full data structure keys: {list(data.keys())}")
        all_photos = []
        for idx, item in enumerate(work_items):
            photos = item.get('photos', [])
            logger.info(f"  Work item {idx + 1}: {len(photos)} photos")
            logger.info(f"    Work item keys: {list(item.keys())}")
            for pidx, photo in enumerate(photos):
                if isinstance(photo, dict):
                    photo_url = photo.get('url', 'NO URL KEY')
                    logger.info(f"    Photo {pidx + 1} (dict): url={photo_url[:80] if photo_url else 'NONE'}..., is_cloud={photo.get('is_cloud', 'N/A')}")
                else:
                    logger.info(f"    Photo {pidx + 1} (not dict): {str(photo)[:80]}...")
            all_photos.extend(photos)
        logger.info(f"📸 Total photos collected: {len(all_photos)}")
        if len(all_photos) == 0:
            logger.warning("⚠️ NO PHOTOS FOUND! Checking data structure...")
            logger.warning(f"   Work items: {work_items}")
            logger.warning(f"   First work item structure: {work_items[0] if work_items else 'NO ITEMS'}")
        
        # Container for PDF elements
        story = []
        styles = get_professional_styles()
        
        # HEADER WITH LOGO
        create_header_with_logo(
            story,
            "CIVIL WORKS INSPECTION REPORT",
            f"Project: {project_name}"
        )
        
        # Compact separator (match HVAC formatting)
        story.append(Spacer(1, 0.04*inch))
        
        # PROJECT INFORMATION
        add_section_heading(story, "Project Information")
        
        project_data = [
            ['Project Name:', data.get('project_name', 'N/A')],
            ['Location:', data.get('location', 'N/A')],
            ['Visit Date:', data.get('visit_date', 'N/A')],
            ['Inspector:', data.get('inspector_name', 'N/A')],
            ['Description of Work:', data.get('description_of_work', 'N/A')],
            ['Report Generated:', format_dubai_datetime() + ' (GST)'],
            ['Total Photos:', str(len(all_photos))]
        ]
        
        project_table = create_info_table(project_data, col_widths=[2.35*inch, 4.65*inch])
        story.append(project_table)
        story.append(Spacer(1, 0.1*inch))
        
        # WORK ITEMS - Use work_items array directly
        add_section_heading(story, "Work Items")
        
        if work_items:
            for idx, item in enumerate(work_items, 1):
                add_item_heading(story, f"Work Item {idx}")
                
                item_data = [
                    ['Description:', item.get('description', 'N/A')],
                    ['Quantity:', item.get('quantity', 'N/A')],
                    ['Material:', item.get('material', 'N/A')],
                    ['Material Quantity:', item.get('material_qty', 'N/A')],
                    ['Price:', item.get('price', 'N/A')],
                    ['Labour:', item.get('labour', 'N/A')],
                    ['Photos:', str(len(item.get('photos', [])))]
                ]
                
                item_table = create_info_table(item_data, col_widths=[2.35*inch, 4.65*inch])
                story.append(item_table)
                story.append(Spacer(1, 0.06*inch))
                
                # Add photos for this work item
                photos = item.get('photos', [])
                logger.info(f"    Adding photos for work item {idx}: {len(photos)} photos")
                if photos:
                    add_paragraph(story, f"<b>Attached Photos ({len(photos)} total):</b>")
                    story.append(Spacer(1, 0.04*inch))
                    logger.info(f"    Calling add_photo_grid with {len(photos)} photos")
                    add_photo_grid(story, photos)
                else:
                    logger.warning(f"    ⚠️ No photos found for work item {idx}")
                
                # Small spacer between items (no page break - avoids huge gaps)
                if idx < len(work_items):
                    story.append(Spacer(1, 0.15*inch))
        else:
            add_paragraph(story, "No work items recorded.")
        
        # PHOTOS SECTION - Show all photos if not already shown per item
        if all_photos and not work_items:
            story.append(Spacer(1, 0.2*inch))
            add_section_heading(story, f"Attached Photos ({len(all_photos)} total)")
            add_photo_grid(story, all_photos)
        elif not all_photos:
            story.append(Spacer(1, 0.2*inch))
            add_paragraph(story, "No photos attached.")
        
        # MATERIALS REQUIRED SECTION
        materials = data.get('materials_required', [])
        if materials and isinstance(materials, list) and len(materials) > 0:
            add_section_heading(story, "Materials Required")
            mat_headers = ['#', 'Material Name', 'Brand', 'Department', 'UOM', 'Qty', 'Unit Price (AED)', 'Line Total (AED)']
            mat_data = []
            materials_total = 0.0
            for idx, m in enumerate(materials, 1):
                qty = float(m.get('quantity', 1) or 0)
                unit_price = float(m.get('unit_price', 0) or 0)
                line_total = qty * unit_price
                materials_total += line_total
                mat_data.append([
                    str(idx),
                    str(m.get('name', 'N/A')),
                    str(m.get('brand', '') or '—'),
                    str(m.get('department', '') or '—'),
                    str(m.get('uom', '') or '—'),
                    f"{qty:g}",
                    f"{unit_price:,.2f}",
                    f"{line_total:,.2f}",
                ])
            mat_col_widths = [0.3*inch, 1.8*inch, 0.9*inch, 0.8*inch, 0.5*inch, 0.45*inch, 0.9*inch, 0.95*inch]
            mat_table_data = [mat_headers] + mat_data
            mat_table = Table(mat_table_data, colWidths=mat_col_widths, repeatRows=1)
            mat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#125435')),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0, 0), (-1, 0), 8),
                ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE',   (0, 1), (-1, -1), 8),
                ('ALIGN',      (0, 1), (0, -1), 'CENTER'),
                ('ALIGN',      (5, 1), (5, -1), 'CENTER'),
                ('ALIGN',      (6, 1), (7, -1), 'RIGHT'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FAF5')]),
                ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(mat_table)
            add_paragraph(story, f"<b>Materials Grand Total (AED):</b> {materials_total:,.2f}")
            story.append(Spacer(1, 0.15*inch))

        # SIGNATURES PAGE - Professional format with all reviewer signatures
        # Extract all reviewer data similar to HVAC module
        signatures = {}
        
        # Get nested data dict if it exists
        nested_data = data.get('data') if isinstance(data.get('data'), dict) else {}
        
        # Extract supervisor signature - try multiple paths
        supervisor_sig = None
        if inspector_signature:
            supervisor_sig = inspector_signature
        else:
            supervisor_sig_raw = data.get('supervisor_signature')
            if supervisor_sig_raw is not None and supervisor_sig_raw != '' and supervisor_sig_raw != 'None':
                supervisor_sig = supervisor_sig_raw
            elif nested_data and nested_data.get('supervisor_signature'):
                supervisor_sig = nested_data.get('supervisor_signature')
            elif isinstance(data.get('form_data'), dict):
                form_data_dict = data.get('form_data', {})
                if form_data_dict.get('supervisor_signature'):
                    supervisor_sig = form_data_dict.get('supervisor_signature')
        
        # Handle supervisor signature format
        if supervisor_sig:
            if isinstance(supervisor_sig, dict) and supervisor_sig.get('url'):
                signatures['Supervisor'] = supervisor_sig
            elif isinstance(supervisor_sig, str) and (supervisor_sig.startswith('data:image') or supervisor_sig.startswith('http') or supervisor_sig.startswith('/')):
                signatures['Supervisor'] = supervisor_sig
        
        # Extract Operations Manager data
        operations_manager_comments = None
        operations_manager_comments_raw = data.get('operations_manager_comments')
        if operations_manager_comments_raw is not None and operations_manager_comments_raw != 'None' and operations_manager_comments_raw != '':
            operations_manager_comments = operations_manager_comments_raw
        elif nested_data and nested_data.get('operations_manager_comments'):
            operations_manager_comments = nested_data.get('operations_manager_comments')
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            if form_data_dict.get('operations_manager_comments'):
                operations_manager_comments = form_data_dict.get('operations_manager_comments')
        
        if operations_manager_comments is None:
            operations_manager_comments = ''
        
        # Extract Operations Manager signature
        opman_sig = None
        opman_sig_raw = data.get('operations_manager_signature') or data.get('opMan_signature')
        if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
            opman_sig = opman_sig_raw
        elif nested_data:
            opman_sig_raw = nested_data.get('operations_manager_signature') or nested_data.get('opMan_signature')
            if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
                opman_sig = opman_sig_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            opman_sig_raw = form_data_dict.get('operations_manager_signature') or form_data_dict.get('opMan_signature')
            if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
                opman_sig = opman_sig_raw
        
        if opman_sig:
            if isinstance(opman_sig, dict) and opman_sig.get('url'):
                signatures['Operations Manager'] = opman_sig
            elif isinstance(opman_sig, str) and (opman_sig.startswith('data:image') or opman_sig.startswith('http') or opman_sig.startswith('/')):
                signatures['Operations Manager'] = opman_sig
        
        # Extract Business Development data
        business_dev_comments = None
        supervisor_comments_for_validation = supervisor_comments
        business_dev_comments_raw = data.get('business_dev_comments') or data.get('business_development_comments')
        if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
            if business_dev_comments_raw != supervisor_comments_for_validation:
                business_dev_comments = business_dev_comments_raw
        elif nested_data:
            business_dev_comments_raw = nested_data.get('business_dev_comments') or nested_data.get('business_development_comments')
            if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
                if business_dev_comments_raw != supervisor_comments_for_validation:
                    business_dev_comments = business_dev_comments_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            business_dev_comments_raw = form_data_dict.get('business_dev_comments') or form_data_dict.get('business_development_comments')
            if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
                if business_dev_comments_raw != supervisor_comments_for_validation:
                    business_dev_comments = business_dev_comments_raw
        
        if business_dev_comments is None:
            business_dev_comments = ''
        
        # Extract Business Development signature
        business_dev_sig = None
        business_dev_sig_raw = data.get('business_dev_signature') or data.get('businessDevSignature')
        if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
            business_dev_sig = business_dev_sig_raw
        elif nested_data:
            business_dev_sig_raw = nested_data.get('business_dev_signature') or nested_data.get('businessDevSignature')
            if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
                business_dev_sig = business_dev_sig_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            business_dev_sig_raw = form_data_dict.get('business_dev_signature') or form_data_dict.get('businessDevSignature')
            if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
                business_dev_sig = business_dev_sig_raw
        
        if business_dev_sig:
            if isinstance(business_dev_sig, dict) and business_dev_sig.get('url'):
                signatures['Business Development'] = business_dev_sig
            elif isinstance(business_dev_sig, str) and (business_dev_sig.startswith('data:image') or business_dev_sig.startswith('http') or business_dev_sig.startswith('/')):
                signatures['Business Development'] = business_dev_sig
        
        # Extract Procurement data
        procurement_comments = None
        procurement_comments_raw = data.get('procurement_comments')
        if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
            procurement_comments = procurement_comments_raw
        elif nested_data:
            procurement_comments_raw = nested_data.get('procurement_comments')
            if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
                procurement_comments = procurement_comments_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            procurement_comments_raw = form_data_dict.get('procurement_comments')
            if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
                procurement_comments = procurement_comments_raw
        
        if procurement_comments is None:
            procurement_comments = ''
        
        # Extract Procurement signature
        procurement_sig = None
        procurement_sig_raw = data.get('procurement_signature') or data.get('procurementSignature')
        if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
            procurement_sig = procurement_sig_raw
        elif nested_data:
            procurement_sig_raw = nested_data.get('procurement_signature') or nested_data.get('procurementSignature')
            if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
                procurement_sig = procurement_sig_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            procurement_sig_raw = form_data_dict.get('procurement_signature') or form_data_dict.get('procurementSignature')
            if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
                procurement_sig = procurement_sig_raw
        
        if procurement_sig:
            if isinstance(procurement_sig, dict) and procurement_sig.get('url'):
                signatures['Procurement'] = procurement_sig
            elif isinstance(procurement_sig, str) and (procurement_sig.startswith('data:image') or procurement_sig.startswith('http') or procurement_sig.startswith('/')):
                signatures['Procurement'] = procurement_sig
        
        # Extract General Manager data
        general_manager_comments = data.get('general_manager_comments', '') or (nested_data.get('general_manager_comments') if nested_data else '') or ''
        general_manager_sig = data.get('general_manager_signature', '') or data.get('generalManagerSignature', '')
        if general_manager_sig:
            if isinstance(general_manager_sig, dict) and general_manager_sig.get('url'):
                signatures['General Manager'] = general_manager_sig
            elif isinstance(general_manager_sig, str) and (general_manager_sig.startswith('data:image') or general_manager_sig.startswith('http')):
                signatures['General Manager'] = general_manager_sig
        
        # Helper function to add comment and signature together for a reviewer (same as HVAC)
        def add_reviewer_section(role_name, comments, signature_data, always_show_signature=False):
            """Add comments and signature together for a reviewer with aspect-ratio-preserved signatures"""
            has_content = False
            
            if comments and comments.strip():
                add_section_heading(story, f"{role_name} Comments")
                add_paragraph(story, comments)
                story.append(Spacer(1, 0.1*inch))
                has_content = True
            
            if signature_data or always_show_signature:
                styles = get_professional_styles()
                sig_rows = []
                
                if signature_data:
                    try:
                        from common.utils import get_image_for_pdf
                        from PIL import Image as PILImage
                        
                        img_data, is_url = get_image_for_pdf(signature_data)
                        if img_data:
                            max_width = 2.5 * inch
                            max_height = 1.2 * inch
                            
                            try:
                                if is_url:
                                    img_data.seek(0)
                                    pil_img = PILImage.open(img_data)
                                else:
                                    pil_img = PILImage.open(img_data)
                                
                                orig_width, orig_height = pil_img.size
                                
                                width_ratio = max_width / orig_width
                                height_ratio = max_height / orig_height
                                scale_ratio = min(width_ratio, height_ratio)
                                
                                final_width = orig_width * scale_ratio
                                final_height = orig_height * scale_ratio
                                
                                original_ratio = orig_width / orig_height if orig_height > 0 else 1
                                final_ratio = final_width / final_height if final_height > 0 else 1
                                
                                if is_url:
                                    img_data.seek(0)
                                    sig_img = Image(img_data, width=final_width, height=final_height)
                                else:
                                    sig_img = Image(img_data, width=final_width, height=final_height)
                                
                                logger.info(f"✅ {role_name} signature aspect ratio: Original={orig_width}x{orig_height} (ratio={original_ratio:.3f}), Final={final_width:.2f}x{final_height:.2f} (ratio={final_ratio:.3f}), Scale={scale_ratio:.3f}")
                                
                                if abs(original_ratio - final_ratio) > 0.01:
                                    logger.warning(f"⚠️ {role_name} signature aspect ratio mismatch! Original={original_ratio:.3f}, Final={final_ratio:.3f}")
                            except Exception as pil_error:
                                logger.warning(f"PIL image processing failed for {role_name}, using fallback: {pil_error}")
                                if is_url:
                                    img_data.seek(0)
                                    sig_img = Image(img_data)
                                else:
                                    sig_img = Image(img_data)
                                
                                if hasattr(sig_img, 'imageWidth') and hasattr(sig_img, 'imageHeight'):
                                    orig_width = sig_img.imageWidth
                                    orig_height = sig_img.imageHeight
                                    if orig_width > 0 and orig_height > 0:
                                        width_ratio = max_width / orig_width
                                        height_ratio = max_height / orig_height
                                        scale_ratio = min(width_ratio, height_ratio)
                                        final_width = orig_width * scale_ratio
                                        final_height = orig_height * scale_ratio
                                        sig_img.drawWidth = final_width
                                        sig_img.drawHeight = final_height
                                    else:
                                        sig_img.drawWidth = max_width
                                else:
                                    sig_img.drawWidth = max_width
                            
                            sig_rows.append([
                                Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                                sig_img
                            ])
                        else:
                            sig_rows.append([
                                Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                                Paragraph("Signature not available", styles['Small'])
                            ])
                    except Exception as e:
                        logger.error(f"Error processing {role_name} signature: {str(e)}")
                        sig_rows.append([
                            Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                            Paragraph("Error loading signature", styles['Small'])
                        ])
                else:
                    sig_rows.append([
                        Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                        Paragraph("<i>Not signed</i>", styles['Small'])
                    ])
                
                if sig_rows:
                    sig_table = Table(sig_rows, colWidths=[2*inch, 3.5*inch])
                    sig_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#125435')),
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    story.append(sig_table)
                    story.append(Spacer(1, 0.15*inch))
                    has_content = True
            
            return has_content
        
        # Check if Operations Manager has approved
        om_has_approved = False
        if data.get('operations_manager_approved_at') or data.get('operations_manager_id'):
            om_has_approved = True
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'operations_manager_approved' in workflow_status or 'bd_procurement' in workflow_status:
                om_has_approved = True
        
        # Check if BD has approved
        bd_has_approved = False
        if data.get('business_dev_approved_at') or data.get('business_dev_id'):
            bd_has_approved = True
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'bd_procurement' in workflow_status or 'general_manager' in workflow_status:
                bd_has_approved = True
        
        # Check if Procurement has approved
        procurement_has_approved = False
        if data.get('procurement_approved_at') or data.get('procurement_id'):
            procurement_has_approved = True
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'bd_procurement' in workflow_status or 'general_manager' in workflow_status:
                procurement_has_approved = True
        
        # Add reviewer sections in workflow order
        # 1. Supervisor - ALWAYS show
        supervisor_comments_display = supervisor_comments if supervisor_comments and supervisor_comments.strip() else None
        supervisor_sig_display = signatures.get('Supervisor')
        add_reviewer_section("Supervisor", supervisor_comments_display, supervisor_sig_display, always_show_signature=True)
        
        # 2. Operations Manager - show if approved or has data
        if operations_manager_comments or signatures.get('Operations Manager') or om_has_approved:
            add_reviewer_section("Operations Manager", operations_manager_comments, signatures.get('Operations Manager'), always_show_signature=True)
        
        # 3. Business Development - show if approved or has data
        if business_dev_comments or signatures.get('Business Development') or bd_has_approved:
            add_reviewer_section("Business Development", business_dev_comments, signatures.get('Business Development'), always_show_signature=True)
        
        # 4. Procurement - show if approved or has data
        if procurement_comments or signatures.get('Procurement') or procurement_has_approved:
            add_reviewer_section("Procurement", procurement_comments, signatures.get('Procurement'), always_show_signature=True)
        
        # 5. General Manager - show if has data
        if general_manager_comments or signatures.get('General Manager'):
            add_reviewer_section("General Manager", general_manager_comments, signatures.get('General Manager'), always_show_signature=True)
        
        # Build professional PDF with logo and branding
        create_professional_pdf(
            pdf_path, 
            story, 
            report_title=f"Civil Works - {project_name}"
        )
        
        logger.info(f"✅ Professional Civil PDF created successfully: {pdf_path}")
        return os.path.basename(pdf_path)
        
    except Exception as e:
        logger.error(f"❌ Civil PDF generation error: {str(e)}")
        raise
