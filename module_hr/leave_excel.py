"""
Leave Tracker — Excel export / import with Remaining formulas and sick alerts.
"""
from __future__ import annotations

import logging
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.models import (
    LEAVE_COMPANIES,
    LEAVE_SICK_ALERT_CRITICAL,
    LEAVE_SICK_ALERT_WARNING,
    LEAVE_SICK_ENTITLEMENT,
    LEAVE_TRACKER_MONTH_LABELS,
    LEAVE_TRACKER_MONTHS,
    LEAVE_TRACKER_YEAR,
    LEAVE_WINDOW_END,
    LEAVE_WINDOW_START,
    normalize_leave_company,
    parse_employee_company,
    LEAVE_TYPES,
    LeaveEmployee,
    LeaveLog,
    LeavePlan,
    db,
    recompute_monthly_usage,
)
from common.datetime_utils import utc_now_naive
from common.kynvera_excel_brand import (
    HINT_FONT,
    THIN,
    TITLE_FONT,
    InstructionSpec,
    style_header_row,
    write_instructions_sheet,
)

logger = logging.getLogger(__name__)

WARN_FILL = PatternFill('solid', fgColor='FEF3C7')
CRIT_FILL = PatternFill('solid', fgColor='FFEDD5')
EXHAUST_FILL = PatternFill('solid', fgColor='FECACA')
INPUT_FILL = PatternFill('solid', fgColor='FFF2CC')

MONTH_HEADERS = [LEAVE_TRACKER_MONTH_LABELS[m] for m in LEAVE_TRACKER_MONTHS]

SICK_HEADERS = (
    'Emp ID', 'Name', 'Designation', 'Company',
    *MONTH_HEADERS, 'Used', 'Remaining', 'Alert',
)
ANNUAL_HEADERS = (
    'Emp ID', 'Name', 'Designation', 'Company', 'Entitlement',
    *MONTH_HEADERS, 'Used', 'Remaining',
)
STAFF_HEADERS = ('Emp ID', 'Name', 'Designation', 'Company', 'Annual Entitlement', 'Active')
PLANS_HEADERS = ('Emp ID', 'Name', 'Company', 'Start Date', 'End Date', 'Days', 'Notes')
# Matches Injaaz Leave Tracker workbook → "Leave Log" sheet
LOGS_HEADERS = (
    'SN',
    'Emp ID',
    'Employee Name',
    'Designation',
    'Company',
    'Project',
    'Leave Type',
    'Start Date',
    'End Date',
    'No. of Days',
    'Reason / Notes',
    'Approved',
    'Month',
)
# Yellow input columns in the official Leave Log template (1-based)
LOG_INPUT_COLS = (2, 7, 8, 9, 11, 12)  # Emp ID, Type, Start, End, Reason, Approved
LOG_SHEET_NAMES = ('Leave Log', 'Leave Logs')


def _style_header(ws, ncols: int) -> None:
    style_header_row(ws, 1, ncols)


def _autosize(ws, min_w=10, max_w=36) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_w
        for cell in col:
            val = cell.value
            if val is None:
                continue
            width = max(width, min(max_w, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


def _parse_days(raw) -> Optional[float]:
    if raw is None or raw == '':
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_int(raw) -> Optional[int]:
    d = _parse_days(raw)
    if d is None:
        return None
    return int(round(d))


def _normalize_emp_id(raw) -> str:
    if raw is None:
        return ''
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    if isinstance(raw, int):
        return str(raw)
    s = str(raw).strip()
    if re.fullmatch(r'\d+\.0', s):
        return s[:-2]
    return s


def _as_date(raw) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    return date.fromisoformat(str(raw).strip()[:10])


def _normalize_leave_type(raw) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if s in LEAVE_TYPES:
        return s
    if s.startswith('sick'):
        return 'sick'
    if s.startswith('annual'):
        return 'annual'
    return None


def _find_header_row(ws, max_scan: int = 15) -> tuple[int, list[str]]:
    """Return (1-based row index, headers) for the first row containing Emp ID + leave type."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        headers = [str(c or '').strip() for c in row]
        upper = {h.upper() for h in headers if h}
        if 'EMP ID' in upper and any(
            h in upper for h in ('LEAVE TYPE', 'TYPE', 'TYPE OF LEAVE')
        ):
            return i, headers
    # Fallback: first row
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return 1, [str(c or '').strip() for c in first]


def _map_log_columns(headers: list[str]) -> dict[str, int]:
    """Map canonical keys → column index from Leave Log header aliases."""
    col: dict[str, int] = {}
    for i, h in enumerate(headers):
        key = (h or '').strip().upper().replace('\n', ' ')
        key = ' '.join(key.split())
        if key in ('EMP ID', 'EMPID', 'EMPLOYEE ID'):
            col['emp_id'] = i
        elif key in ('LEAVE TYPE', 'TYPE', 'TYPE OF LEAVE'):
            col['leave_type'] = i
        elif key in (
            'START DATE', 'FROM', 'DATE', 'LEAVE START DATE', 'LEAVE START',
        ):
            col.setdefault('start', i)
        elif key in ('END DATE', 'TO', 'LEAVE END DATE', 'LEAVE END'):
            col['end'] = i
        elif key in (
            'NO. OF DAYS', 'NO OF DAYS', 'DAYS', 'NUMBER OF DAYS', 'NO.OF DAYS',
        ):
            col['days'] = i
        elif key in (
            'REASON / NOTES', 'REASON/NOTES', 'REASON', 'NOTES',
            'ACTUAL LEAVE RESUMPTION',
        ):
            col.setdefault('notes', i)
        elif key == 'APPROVED':
            col['approved'] = i
        elif key in ('EMPLOYEE NAME', 'NAME', 'EMP NAME'):
            col['name'] = i
        elif key == 'COMPANY':
            col['company'] = i
        elif key == 'DESIGNATION':
            col['designation'] = i
    return col


def _leave_log_sheet(wb) -> Optional[Any]:
    for name in LOG_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    # Case-insensitive fallback
    lower = {n.lower(): n for n in wb.sheetnames}
    for name in LOG_SHEET_NAMES:
        if name.lower() in lower:
            return wb[lower[name.lower()]]
    # First sheet that looks like a leave log (covers Sheet1 HR exports)
    for name in wb.sheetnames:
        ws = wb[name]
        _hr, headers = _find_header_row(ws)
        col = _map_log_columns(headers)
        if 'emp_id' in col and 'start' in col and 'leave_type' in col:
            return ws
    return None


def _alert_label(level: str) -> str:
    return {
        'warning': 'Approaching limit',
        'critical': 'Nearly exhausted',
        'exhausted': 'Exhausted',
    }.get(level or '', '')


def _set_usage(employee: LeaveEmployee, leave_type: str, month: int, days: Optional[float]) -> None:
    """Replace month logs with a synthetic monthly log (or clear), then recompute."""
    year = LEAVE_TRACKER_YEAR
    month_start = date(year, month, 1)
    month_end = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
    existing = LeaveLog.query.filter(
        LeaveLog.employee_id == employee.id,
        LeaveLog.leave_type == leave_type,
        LeaveLog.leave_date >= month_start,
        LeaveLog.leave_date < month_end,
    ).all()
    for lg in existing:
        db.session.delete(lg)
    db.session.flush()
    if days is not None and days > 0:
        db.session.add(LeaveLog(
            employee_id=employee.id,
            leave_type=leave_type,
            leave_date=month_start,
            days=float(days),
            notes='Imported from monthly sheet',
        ))
        db.session.flush()
    recompute_monthly_usage(employee.id, leave_type, year, month)


def build_leave_workbook(
    employees: list[LeaveEmployee],
    plans: Optional[list[LeavePlan]] = None,
    logs: Optional[list[LeaveLog]] = None,
) -> BytesIO:
    """Export Sick, Annual, Staff, Plans, and Leave Log sheets with formulas + CF."""
    wb = Workbook()

    # ── Sick Leave ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Sick Leave'
    ws.append(list(SICK_HEADERS))
    _style_header(ws, len(SICK_HEADERS))

    first_month_col = 5
    used_col = first_month_col + len(LEAVE_TRACKER_MONTHS)
    rem_col = used_col + 1
    alert_col = rem_col + 1

    for i, emp in enumerate(employees, start=2):
        sick_map = emp.usage_map('sick')
        row_vals = [
            emp.emp_id,
            emp.full_name,
            emp.designation or '',
            emp.company or '',
        ]
        for m in LEAVE_TRACKER_MONTHS:
            d = sick_map.get(m)
            row_vals.append(d if d is not None else None)
        ws.append(row_vals + [None, None, None])

        used_letter = get_column_letter(used_col)
        first_m = get_column_letter(first_month_col)
        last_m = get_column_letter(used_col - 1)
        ws.cell(i, used_col).value = f'=SUM({first_m}{i}:{last_m}{i})'
        ws.cell(i, rem_col).value = f'={LEAVE_SICK_ENTITLEMENT}-{used_letter}{i}'
        ws.cell(i, alert_col).value = (
            f'=IF({used_letter}{i}>={LEAVE_SICK_ENTITLEMENT},"Exhausted",'
            f'IF({used_letter}{i}>={LEAVE_SICK_ALERT_CRITICAL},"Nearly exhausted",'
            f'IF({used_letter}{i}>={LEAVE_SICK_ALERT_WARNING},"Approaching limit","")))'
        )
        for c in range(1, alert_col + 1):
            ws.cell(i, c).border = THIN

    last_row = max(2, len(employees) + 1)
    used_range = f'{get_column_letter(used_col)}2:{get_column_letter(used_col)}{last_row}'
    row_range = f'A2:L{last_row}'
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[f'${get_column_letter(used_col)}2>={LEAVE_SICK_ENTITLEMENT}'],
            fill=EXHAUST_FILL,
        ),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[
                f'AND(${get_column_letter(used_col)}2>={LEAVE_SICK_ALERT_CRITICAL},'
                f'${get_column_letter(used_col)}2<{LEAVE_SICK_ENTITLEMENT})'
            ],
            fill=CRIT_FILL,
        ),
    )
    ws.conditional_formatting.add(
        row_range,
        FormulaRule(
            formula=[
                f'AND(${get_column_letter(used_col)}2>={LEAVE_SICK_ALERT_WARNING},'
                f'${get_column_letter(used_col)}2<{LEAVE_SICK_ALERT_CRITICAL})'
            ],
            fill=WARN_FILL,
        ),
    )
    _ = used_range
    rem_letter_hdr = get_column_letter(rem_col)
    ws.conditional_formatting.add(
        f'{rem_letter_hdr}2:{rem_letter_hdr}{last_row}',
        CellIsRule(operator='lessThanOrEqual', formula=['0'], fill=EXHAUST_FILL),
    )
    _autosize(ws)

    # ── Annual Leave ────────────────────────────────────────────────────────
    ws_a = wb.create_sheet('Annual Leave')
    ws_a.append(list(ANNUAL_HEADERS))
    _style_header(ws_a, len(ANNUAL_HEADERS))
    a_first_month = 6
    a_used_col = a_first_month + len(LEAVE_TRACKER_MONTHS)
    a_rem_col = a_used_col + 1

    for i, emp in enumerate(employees, start=2):
        annual_map = emp.usage_map('annual')
        row_vals = [
            emp.emp_id,
            emp.full_name,
            emp.designation or '',
            emp.company or '',
            emp.annual_entitlement if emp.annual_entitlement is not None else None,
        ]
        for m in LEAVE_TRACKER_MONTHS:
            d = annual_map.get(m)
            row_vals.append(d if d is not None else None)
        ws_a.append(row_vals + [None, None])

        first_m = get_column_letter(a_first_month)
        last_m = get_column_letter(a_used_col - 1)
        used_letter = get_column_letter(a_used_col)
        ent_letter = get_column_letter(5)
        ws_a.cell(i, a_used_col).value = f'=SUM({first_m}{i}:{last_m}{i})'
        ws_a.cell(i, a_rem_col).value = f'=IF({ent_letter}{i}="","",{ent_letter}{i}-{used_letter}{i})'
        for c in range(1, a_rem_col + 1):
            ws_a.cell(i, c).border = THIN
    _autosize(ws_a)

    # ── Staff ───────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet('Staff')
    ws_s.append(list(STAFF_HEADERS))
    _style_header(ws_s, len(STAFF_HEADERS))
    for emp in employees:
        ws_s.append([
            emp.emp_id,
            emp.full_name,
            emp.designation or '',
            emp.company or '',
            emp.annual_entitlement if emp.annual_entitlement is not None else None,
            'Y' if emp.active else 'N',
        ])
    _autosize(ws_s)

    # ── Plans ───────────────────────────────────────────────────────────────
    ws_p = wb.create_sheet('Plans')
    ws_p.append(list(PLANS_HEADERS))
    _style_header(ws_p, len(PLANS_HEADERS))
    for plan in plans or []:
        d = plan.to_dict()
        ws_p.append([
            d.get('emp_id'),
            d.get('full_name'),
            d.get('company'),
            d.get('start_date'),
            d.get('end_date'),
            d.get('days'),
            d.get('notes') or '',
        ])
    _autosize(ws_p)

    # ── Leave Log (official column layout) ──────────────────────────────────
    ws_l = wb.create_sheet('Leave Log')
    ws_l.append(list(LOGS_HEADERS))
    _style_header(ws_l, len(LOGS_HEADERS))
    for sn, log in enumerate(logs or [], start=1):
        d = log.to_dict()
        leave_type = (d.get('leave_type') or '').capitalize()
        start = d.get('leave_date')
        end = d.get('end_date') or start
        month_label = ''
        if start:
            try:
                month_label = date.fromisoformat(str(start)[:10]).strftime('%B %Y')
            except Exception:
                month_label = d.get('month_label') or ''
        ws_l.append([
            sn,
            d.get('emp_id'),
            d.get('full_name'),
            d.get('designation') or '',
            d.get('company') or '',
            '',  # Project (optional HR field)
            leave_type,
            start,
            end,
            d.get('days'),
            d.get('notes') or '',
            'Yes',
            month_label,
        ])
    for col_i in (8, 9):  # Start / End date columns
        for row_i in range(2, 2 + len(logs or [])):
            ws_l.cell(row_i, col_i).number_format = 'YYYY-MM-DD'
    _autosize(ws_l)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_leave_log_template_bytes() -> BytesIO:
    """
    Blank Leave Log template matching the Injaaz Leave Tracker workbook layout
    (title + hint + header row + yellow input cells + dropdowns).
    """
    wb = Workbook()

    ws = wb.active
    ws.title = 'Leave Log'
    ws.merge_cells('A1:M1')
    ws['A1'] = (
        f'Leave Log — Enter every Sick / Annual leave instance here '
        f'(from {LEAVE_WINDOW_START.isoformat()})'
    )
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:M2')
    ws['A2'] = (
        'Fill yellow cells only: Emp ID, Leave Type, Start/End Date, Reason / Notes, Approved. '
        'No. of Days and Month calculate automatically. Upload this file via Import on Leave Tracker. '
        f'Dates must fall on or after {LEAVE_WINDOW_START.isoformat()}.'
    )
    ws['A2'].font = HINT_FONT
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 36

    header_row = 3
    for col_idx, header in enumerate(LOGS_HEADERS, start=1):
        ws.cell(header_row, col_idx, header)
    style_header_row(ws, header_row, len(LOGS_HEADERS))

    blank_rows = 50
    first_data = header_row + 1
    last_data = header_row + blank_rows
    for r in range(first_data, last_data + 1):
        # SN
        ws.cell(r, 1).value = f'=IF(B{r}="","",ROW()-{header_row})'
        # No. of Days
        ws.cell(r, 10).value = f'=IF(AND(H{r}<>"",I{r}<>""),I{r}-H{r}+1,"")'
        # Month label
        ws.cell(r, 13).value = f'=IF(H{r}="","",TEXT(H{r},"mmmm yyyy"))'
        ws.cell(r, 8).number_format = 'YYYY-MM-DD'
        ws.cell(r, 9).number_format = 'YYYY-MM-DD'
        for col_i in LOG_INPUT_COLS:
            ws.cell(r, col_i).fill = INPUT_FILL
            ws.cell(r, col_i).border = THIN

    # Example row
    ws.cell(first_data, 2).value = 170
    ws.cell(first_data, 3).value = 'Example — replace Emp ID / dates before import'
    ws.cell(first_data, 7).value = 'Sick'
    ws.cell(first_data, 8).value = date(LEAVE_TRACKER_YEAR, 8, 3)
    ws.cell(first_data, 9).value = date(LEAVE_TRACKER_YEAR, 8, 3)
    ws.cell(first_data, 11).value = 'Example sick leave'
    ws.cell(first_data, 12).value = 'Yes'

    type_dv = DataValidation(
        type='list',
        formula1='"Sick,Annual"',
        allow_blank=True,
        showErrorMessage=True,
        showDropDown=False,
    )
    type_dv.add(f'G{first_data}:G{last_data}')
    ws.add_data_validation(type_dv)

    appr_dv = DataValidation(
        type='list',
        formula1='"Yes,No,Pending"',
        allow_blank=True,
        showDropDown=False,
    )
    appr_dv.add(f'L{first_data}:L{last_data}')
    ws.add_data_validation(appr_dv)

    widths = {
        'A': 6, 'B': 10, 'C': 28, 'D': 20, 'E': 12, 'F': 14,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 22, 'L': 10, 'M': 14,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w
    ws.freeze_panes = 'A4'

    write_instructions_sheet(wb, InstructionSpec(
        title='Leave log template',
        module_label='HR / Leave Tracker',
        about=(
            f'Blank Leave Log for Sick and Annual leave in the current tracker window '
            f'(from {LEAVE_WINDOW_START.isoformat()}).',
            'Fill yellow cells only. No. of Days and Month calculate automatically from the dates.',
            'Emp ID must already exist in Leave Tracker (seed staff first if needed).',
        ),
        how_to=(
            'Open the Leave Log sheet. Keep the coral header on row 3.',
            'Enter Emp ID (must match an existing employee).',
            'Choose Leave Type: Sick or Annual (dropdown).',
            'Enter Start Date and End Date (inclusive). No. of Days fills in automatically.',
            'Optionally fill Reason / Notes and Approved (Yes / No / Pending).',
            'Save as .xlsx and click Import on the Leave Tracker page.',
        ),
        columns=(
            ('SN', 'Auto. Leave blank; formula fills when Emp ID is set.'),
            ('Emp ID', 'Required. Must already exist in Leave Tracker.'),
            ('Employee Name', 'Optional on import (looked up from Emp ID).'),
            ('Designation', 'Optional on import.'),
            ('Company', 'Optional on import.'),
            ('Project', 'Optional.'),
            ('Leave Type', 'Required. Sick or Annual.'),
            ('Start Date', f'Required. Inclusive. Must fall in {LEAVE_WINDOW_START.isoformat()} – {LEAVE_WINDOW_END.isoformat()}.'),
            ('End Date', 'Required. Inclusive. No. of Days = End − Start + 1.'),
            ('No. of Days', 'Auto from dates. Do not type over the formula.'),
            ('Reason / Notes', 'Optional.'),
            ('Approved', 'Optional. Yes / No / Pending.'),
            ('Month', 'Auto from Start Date. Do not type over the formula.'),
        ),
        example_headers=LOGS_HEADERS,
        example_rows=((
            1,
            170,
            'Example — replace Emp ID / dates before import',
            '',
            '',
            '',
            'Sick',
            date(LEAVE_TRACKER_YEAR, 8, 3).isoformat(),
            date(LEAVE_TRACKER_YEAR, 8, 3).isoformat(),
            1,
            'Example sick leave',
            'Yes',
            f'August {LEAVE_TRACKER_YEAR}',
        ),),
        import_rules=(
            f'Only dates from {LEAVE_WINDOW_START.isoformat()} to {LEAVE_WINDOW_END.isoformat()} are imported.',
            'Unknown Emp IDs are reported and skipped; other rows still import.',
            'You can also upload the full Leave Tracker workbook — the Leave Log sheet is detected automatically.',
            'Employee Name / Designation / Company / Project are optional (looked up from Emp ID).',
        ),
    ))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


STAFF_SHEET_NAMES = ('Staff', 'Employee List', 'Employees')
STAFF_EXAMPLE_IDS = frozenset({'INJ-0000', 'EXAMPLE', 'SAMPLE'})


def _is_staff_example_row(emp_id: str, name: str) -> bool:
    key = (emp_id or '').strip().upper()
    label = (name or '').strip().lower()
    if key in STAFF_EXAMPLE_IDS:
        return True
    return label.startswith('example') or label.startswith('[sample]')


def _staff_sheet(wb):
    for name in STAFF_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    return None


def build_staff_workbook(employees: Optional[list] = None, *, template_only: bool = False) -> BytesIO:
    """Employee List template (example row) or export of the current staff roster."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Staff'
    ws.append(list(STAFF_HEADERS))
    _style_header(ws, len(STAFF_HEADERS))

    rows: list[list[Any]] = []
    if template_only:
        rows.append(['INJ-0000', 'Example Name', 'Facility Supervisor', 'Kynvera', 30, 'Y'])
    else:
        for emp in employees or []:
            rows.append([
                emp.emp_id,
                emp.full_name,
                emp.designation or '',
                emp.company or '',
                emp.annual_entitlement if emp.annual_entitlement is not None else None,
                'Y' if emp.active else 'N',
            ])

    for i, values in enumerate(rows, start=2):
        for col, val in enumerate(values, start=1):
            cell = ws.cell(i, col, val)
            cell.border = THIN
            if template_only:
                cell.fill = INPUT_FILL

    last_row = max(200, 2 + len(rows))
    active_dv = DataValidation(
        type='list',
        formula1='"Y,N"',
        allow_blank=True,
        showDropDown=False,
    )
    active_dv.add(f'F2:F{last_row}')
    ws.add_data_validation(active_dv)

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10
    ws.freeze_panes = 'A2'

    write_instructions_sheet(wb, InstructionSpec(
        title='Employee List template',
        module_label='HR / Employee List',
        about=(
            'Staff roster used by Employee List and Leave Tracker.',
            'One row per person. Emp ID must be unique.',
        ),
        how_to=(
            'Open the Staff sheet. Keep the coral header on row 1.',
            'Fill Emp ID, Name, Designation, Company, and optional Annual Entitlement.',
            'Active is Y or N. Leave blank for Y.',
            'Replace the example row before importing.',
            'Save as .xlsx and click Import on Employee List.',
        ),
        columns=(
            ('Emp ID', 'Required. Unique staff number, e.g. INJ-0042.'),
            ('Name', 'Required. Full name.'),
            ('Designation', 'Optional job title.'),
            ('Company', 'Kynvera, Tourism, or L&P (dropdown).'),
            ('Annual Entitlement', 'Optional annual leave days, e.g. 30.'),
            ('Active', 'Y (default) or N to deactivate.'),
        ),
        example_headers=STAFF_HEADERS,
        example_rows=(('INJ-0000', 'Example Name', 'Facility Supervisor', 'Kynvera', 30, 'Y'),),
        import_rules=(
            'Existing Emp IDs are updated; new IDs are added.',
            'The example row (INJ-0000 / Example Name) is skipped.',
            'The full Leave Tracker workbook also works if it has a Staff sheet.',
        ),
    ))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_staff_workbook(file_storage) -> dict[str, Any]:
    """Import / update staff from a Staff (or Employee List) sheet."""
    wb = load_workbook(file_storage, data_only=False)
    ws = _staff_sheet(wb)
    if ws is None:
        sheet_hint = ', '.join(wb.sheetnames[:8]) or '(empty workbook)'
        raise ValueError(
            f'No "Staff" sheet found (sheets: {sheet_hint}). '
            'Download Template from Employee List, or upload a Leave Tracker workbook that includes Staff.'
        )

    created = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    by_emp = {e.emp_id.strip().upper(): e for e in LeaveEmployee.query.all()}

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    emp_i = col.get('Emp ID')
    if emp_i is None:
        raise ValueError('Staff sheet is missing an Emp ID column.')

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or emp_i >= len(row) or not row[emp_i]:
            continue
        emp_id = _normalize_emp_id(row[emp_i])
        if not emp_id:
            continue
        name = str(row[col['Name']]).strip() if 'Name' in col and row[col['Name']] else ''
        if not name and 'Employee Name' in col and row[col['Employee Name']]:
            name = str(row[col['Employee Name']]).strip()
        if _is_staff_example_row(emp_id, name):
            skipped += 1
            continue
        desig = ''
        if 'Designation' in col and row[col['Designation']]:
            desig = str(row[col['Designation']]).strip()
        company = 'Kynvera'
        if 'Company' in col and row[col['Company']]:
            company = parse_employee_company(row[col['Company']])
        ent = None
        if 'Annual Entitlement' in col:
            ent = _parse_int(row[col['Annual Entitlement']])
        active = True
        if 'Active' in col and row[col['Active']] is not None:
            active = str(row[col['Active']]).strip().upper() not in ('N', 'NO', '0', 'FALSE')

        key = emp_id.upper()
        emp = by_emp.get(key)
        if not emp:
            if not name:
                errors.append(f'Staff: skip {emp_id} — missing name')
                continue
            emp = LeaveEmployee(
                emp_id=emp_id,
                full_name=name,
                designation=desig,
                company=company,
                annual_entitlement=ent,
                active=active,
            )
            db.session.add(emp)
            by_emp[key] = emp
            created += 1
        else:
            if name:
                emp.full_name = name
            if desig:
                emp.designation = desig
            if company:
                emp.company = company
            if ent is not None:
                emp.annual_entitlement = ent
            emp.active = active
            emp.updated_at = utc_now_naive()
            updated += 1

    db.session.commit()
    return {
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors[:20],
    }


def import_leave_workbook(file_storage) -> dict[str, Any]:
    """
    Import Staff + Leave Log (preferred) and/or monthly Sick/Annual sheets.
    Accepts the official Injaaz "Leave Log" sheet (or "Leave Logs") and
    creates dated entries that recompute staff master months.
    """
    wb = load_workbook(file_storage, data_only=False)
    created = 0
    updated = 0
    usage_updates = 0
    logs_created = 0
    errors: list[str] = []

    by_emp = {e.emp_id.strip().upper(): e for e in LeaveEmployee.query.all()}

    def _get_or_note(emp_id_raw: str) -> Optional[LeaveEmployee]:
        if not emp_id_raw:
            return None
        key = str(emp_id_raw).strip().upper()
        return by_emp.get(key)

    if 'Staff' in wb.sheetnames:
        ws = wb['Staff']
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        emp_i = col.get('Emp ID')
        if emp_i is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or emp_i >= len(row) or not row[emp_i]:
                    continue
                emp_id = _normalize_emp_id(row[emp_i])
                if not emp_id:
                    continue
                key = emp_id.upper()
                name = str(row[col['Name']]).strip() if 'Name' in col and row[col['Name']] else ''
                if not name and 'Employee Name' in col and row[col['Employee Name']]:
                    name = str(row[col['Employee Name']]).strip()
                desig = ''
                if 'Designation' in col and row[col['Designation']]:
                    desig = str(row[col['Designation']]).strip()
                company = 'Kynvera'
                if 'Company' in col and row[col['Company']]:
                    company = parse_employee_company(row[col['Company']])
                ent = None
                if 'Annual Entitlement' in col:
                    ent = _parse_int(row[col['Annual Entitlement']])
                active = True
                if 'Active' in col and row[col['Active']] is not None:
                    active = str(row[col['Active']]).strip().upper() not in ('N', 'NO', '0', 'FALSE')

                emp = by_emp.get(key)
                if not emp:
                    if not name:
                        errors.append(f'Staff: skip {emp_id} — missing name')
                        continue
                    emp = LeaveEmployee(
                        emp_id=emp_id,
                        full_name=name,
                        designation=desig,
                        company=company,
                        annual_entitlement=ent,
                        active=active,
                    )
                    db.session.add(emp)
                    by_emp[key] = emp
                    created += 1
                else:
                    if name:
                        emp.full_name = name
                    if desig:
                        emp.designation = desig
                    if company:
                        emp.company = company
                    if ent is not None:
                        emp.annual_entitlement = ent
                    emp.active = active
                    emp.updated_at = utc_now_naive()
                    updated += 1

        db.session.flush()

    has_logs_sheet = False
    log_ws = _leave_log_sheet(wb)
    if log_ws is not None:
        has_logs_sheet = True
        from app.models import months_touched_by_range

        header_row, headers = _find_header_row(log_ws)
        col = _map_log_columns(headers)
        if 'emp_id' in col and 'start' in col and 'leave_type' in col:
            for row in log_ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or col['emp_id'] >= len(row):
                    continue
                emp_raw = row[col['emp_id']]
                if emp_raw is None or str(emp_raw).strip() == '':
                    continue
                emp_id = _normalize_emp_id(emp_raw)
                emp = _get_or_note(emp_id)
                if not emp:
                    errors.append(f'Leave Log: unknown Emp ID {emp_id}')
                    continue
                lt = _normalize_leave_type(row[col['leave_type']] if col['leave_type'] < len(row) else None)
                if not lt:
                    errors.append(f'Leave Log: bad type for {emp.emp_id}')
                    continue
                try:
                    leave_date = _as_date(row[col['start']])
                    end_date = leave_date
                    if 'end' in col and col['end'] < len(row) and row[col['end']]:
                        end_date = _as_date(row[col['end']])
                except Exception:
                    errors.append(f'Leave Log: bad date for {emp.emp_id}')
                    continue
                if end_date < leave_date:
                    errors.append(f'Leave Log: end before start for {emp.emp_id}')
                    continue
                if leave_date < LEAVE_WINDOW_START or end_date > LEAVE_WINDOW_END:
                    errors.append(
                        f'Leave Log: date out of window for {emp.emp_id} '
                        f'({leave_date}–{end_date})'
                    )
                    continue
                cal_days = (end_date - leave_date).days + 1
                days = None
                if 'days' in col and col['days'] < len(row):
                    days = _parse_days(row[col['days']])
                if days is None or days <= 0:
                    days = float(cal_days)
                notes = ''
                if 'notes' in col and col['notes'] < len(row) and row[col['notes']]:
                    notes = str(row[col['notes']]).strip()
                if notes.lower().startswith('example'):
                    continue
                if 'approved' in col and col['approved'] < len(row) and row[col['approved']]:
                    appr = str(row[col['approved']]).strip()
                    if appr and appr.lower() not in ('yes', 'y', 'true', '1'):
                        notes = (notes + f' [Approved: {appr}]').strip() if notes else f'Approved: {appr}'
                db.session.add(LeaveLog(
                    employee_id=emp.id,
                    leave_type=lt,
                    leave_date=leave_date,
                    end_date=end_date if end_date != leave_date else None,
                    days=float(days),
                    notes=notes,
                ))
                db.session.flush()
                for y, m in months_touched_by_range(leave_date, end_date):
                    recompute_monthly_usage(emp.id, lt, y, m)
                logs_created += 1
        else:
            errors.append(
                'Leave Log: missing Emp ID / Leave Type / Start Date columns '
                '(expected the official Leave Log header row)'
            )
    else:
        # Helpful when file only has Leave Log / wrong sheet name and no monthly sheets
        sheet_hint = ', '.join(wb.sheetnames[:8]) or '(empty workbook)'
        if not any(n in wb.sheetnames for n in ('Sick Leave', 'Annual Leave', 'Staff')):
            errors.append(
                f'No "Leave Log" sheet found (sheets: {sheet_hint}). '
                'Use Template download or rename the sheet to Leave Log.'
            )

    def _import_usage_sheet(sheet_name: str, leave_type: str, has_entitlement: bool) -> None:
        nonlocal usage_updates, updated, created
        if sheet_name not in wb.sheetnames:
            return
        # Prefer Leave Logs as source of truth — skip monthly if logs sheet present
        if has_logs_sheet:
            # Still allow entitlement updates from Annual sheet
            if not has_entitlement:
                return
        ws = wb[sheet_name]
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        if 'Emp ID' not in col:
            errors.append(f'{sheet_name}: missing Emp ID column')
            return
        month_cols = {}
        for m, label in LEAVE_TRACKER_MONTH_LABELS.items():
            if label in col:
                month_cols[m] = col[label]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            emp_raw = row[col['Emp ID']]
            if not emp_raw:
                continue
            emp = _get_or_note(str(emp_raw))
            if not emp:
                name = str(row[col['Name']]).strip() if 'Name' in col and row[col['Name']] else str(emp_raw)
                desig = str(row[col['Designation']]).strip() if 'Designation' in col and row[col['Designation']] else ''
                company = normalize_leave_company(
                    row[col['Company']] if 'Company' in col and row[col['Company']] else None
                ) or 'Kynvera'
                emp = LeaveEmployee(
                    emp_id=str(emp_raw).strip(),
                    full_name=name,
                    designation=desig,
                    company=company if company in LEAVE_COMPANIES else 'Kynvera',
                )
                db.session.add(emp)
                db.session.flush()
                by_emp[emp.emp_id.strip().upper()] = emp
                created += 1

            if has_entitlement and 'Entitlement' in col:
                ent = _parse_int(row[col['Entitlement']])
                if ent is not None:
                    emp.annual_entitlement = ent
                    emp.updated_at = utc_now_naive()
                    updated += 1

            if has_logs_sheet:
                continue

            for m, ci in month_cols.items():
                if ci >= len(row):
                    continue
                days = _parse_days(row[ci])
                if days is None and (row[ci] is None or str(row[ci]).strip() == ''):
                    _set_usage(emp, leave_type, m, None)
                    usage_updates += 1
                    continue
                if days is not None:
                    _set_usage(emp, leave_type, m, days)
                    usage_updates += 1

    _import_usage_sheet('Sick Leave', 'sick', False)
    _import_usage_sheet('Annual Leave', 'annual', True)

    plans_created = 0
    if 'Plans' in wb.sheetnames:
        ws = wb['Plans']
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = {h: i for i, h in enumerate(headers)}
        if 'Emp ID' in col and 'Start Date' in col and 'End Date' in col:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col['Emp ID']]:
                    continue
                emp = _get_or_note(str(row[col['Emp ID']]))
                if not emp:
                    errors.append(f'Plans: unknown Emp ID {row[col["Emp ID"]]}')
                    continue
                start_raw = row[col['Start Date']]
                end_raw = row[col['End Date']]
                try:
                    if isinstance(start_raw, datetime):
                        start = start_raw.date()
                    elif isinstance(start_raw, date):
                        start = start_raw
                    else:
                        start = date.fromisoformat(str(start_raw)[:10])
                    if isinstance(end_raw, datetime):
                        end = end_raw.date()
                    elif isinstance(end_raw, date):
                        end = end_raw
                    else:
                        end = date.fromisoformat(str(end_raw)[:10])
                except Exception:
                    errors.append(f'Plans: bad dates for {emp.emp_id}')
                    continue
                if end < start:
                    errors.append(f'Plans: end before start for {emp.emp_id}')
                    continue
                notes = ''
                if 'Notes' in col and row[col['Notes']]:
                    notes = str(row[col['Notes']]).strip()
                days = LeavePlan.calendar_days(start, end)
                db.session.add(LeavePlan(
                    employee_id=emp.id,
                    start_date=start,
                    end_date=end,
                    days=days,
                    notes=notes,
                ))
                plans_created += 1

    db.session.commit()
    return {
        'created': created,
        'updated': updated,
        'usage_updates': usage_updates,
        'logs_created': logs_created,
        'plans_created': plans_created,
        'errors': errors[:50],
    }


def parse_staff_list_workbook(path_or_file) -> list[dict[str, str]]:
    """
    Parse staff list Excel (sheets Kynvera / Tourism / L&P; legacy INJAAZ still accepted).
    Returns list of {emp_id, full_name, designation, company}.
    """
    wb = load_workbook(path_or_file, data_only=True)
    rows: list[dict[str, str]] = []
    seen: set[str] = set()

    for sheet_name in wb.sheetnames:
        company = normalize_leave_company(sheet_name.strip(), default=None)
        if not company:
            continue

        ws = wb[sheet_name]
        header_row = None
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(c).strip() if c is not None else '' for c in row]
            if 'EMP ID' in [v.upper() for v in vals] or 'Employee Name' in vals:
                header_row = i
                headers = vals
                break
        if header_row is None:
            continue

        # normalize header indices
        hmap = {}
        for idx, h in enumerate(headers):
            key = h.strip().upper()
            if key in ('EMP ID', 'EMPID', 'EMPLOYEE ID'):
                hmap['emp_id'] = idx
            elif key in ('EMPLOYEE NAME', 'NAME'):
                hmap['name'] = idx
            elif key in ('DESIGNATION', 'ROLE', 'POSITION'):
                hmap['designation'] = idx

        if 'emp_id' not in hmap or 'name' not in hmap:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row:
                continue
            emp_raw = row[hmap['emp_id']] if hmap['emp_id'] < len(row) else None
            name_raw = row[hmap['name']] if hmap['name'] < len(row) else None
            if emp_raw is None or name_raw is None:
                continue
            emp_id = str(emp_raw).strip()
            # openpyxl may give ints
            if re.fullmatch(r'\d+\.0', emp_id):
                emp_id = emp_id[:-2]
            name = str(name_raw).strip()
            if not emp_id or not name:
                continue
            key = emp_id.upper()
            if key in seen:
                continue
            seen.add(key)
            desig = ''
            if 'designation' in hmap and hmap['designation'] < len(row) and row[hmap['designation']]:
                desig = str(row[hmap['designation']]).strip()
            rows.append({
                'emp_id': emp_id,
                'full_name': name,
                'designation': desig,
                'company': company,
            })
    return rows


def seed_employees_from_staff_list(path_or_file, replace_inactive: bool = False) -> dict[str, Any]:
    """Insert staff list rows into leave_employees. Skips existing emp_ids."""
    parsed = parse_staff_list_workbook(path_or_file)
    existing = {e.emp_id.strip().upper(): e for e in LeaveEmployee.query.all()}
    created = 0
    skipped = 0
    for item in parsed:
        key = item['emp_id'].strip().upper()
        if key in existing:
            skipped += 1
            continue
        emp = LeaveEmployee(
            emp_id=item['emp_id'],
            full_name=item['full_name'],
            designation=item.get('designation') or '',
            company=normalize_leave_company(item.get('company')) or 'Kynvera',
            active=True,
        )
        db.session.add(emp)
        existing[key] = emp
        created += 1
    db.session.commit()
    return {
        'created': created,
        'skipped': skipped,
        'total_parsed': len(parsed),
    }


def split_plan_days_by_month(start: date, end: date) -> dict[int, int]:
    """Split inclusive calendar days across months (year ignored beyond tracker year)."""
    if end < start:
        return {}
    buckets: dict[int, int] = {}
    cur = start
    while cur <= end:
        if (
            LEAVE_WINDOW_START <= cur <= LEAVE_WINDOW_END
            and cur.year == LEAVE_TRACKER_YEAR
        ):
            buckets[cur.month] = buckets.get(cur.month, 0) + 1
        cur += timedelta(days=1)
    return buckets
