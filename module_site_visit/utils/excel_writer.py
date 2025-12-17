"""
Excel report generator for HVAC/MEP site visits
"""
import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)

def create_report_workbook(output_dir, visit_info, processed_items):
    """Generate Excel workbook for HVAC/MEP site visit."""
    try:
        os.makedirs(output_dir, exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "HVAC MEP Report"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Title
        ws['A1'] = "HVAC/MEP Site Visit Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        # Visit Information
        row = 3
        ws[f'A{row}'] = "Building Name:"
        ws[f'B{row}'] = visit_info.get('building_name', '')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Address:"
        ws[f'B{row}'] = visit_info.get('building_address', '')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Visit Date:"
        ws[f'B{row}'] = visit_info.get('visit_date', '')
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Technician:"
        ws[f'B{row}'] = visit_info.get('technician_name', '')
        ws[f'A{row}'].font = Font(bold=True)
        
        # Items table header
        row += 2
        headers = ['Asset', 'System', 'Description', 'Quantity', 'Brand', 'Comments']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Items data
        for item in processed_items:
            row += 1
            ws.cell(row=row, column=1, value=item.get('asset', ''))
            ws.cell(row=row, column=2, value=item.get('system', ''))
            ws.cell(row=row, column=3, value=item.get('description', ''))
            ws.cell(row=row, column=4, value=item.get('quantity', ''))
            ws.cell(row=row, column=5, value=item.get('brand', ''))
            ws.cell(row=row, column=6, value=item.get('comments', ''))
            
            image_paths = item.get('image_paths', [])
            if image_paths:
                ws.cell(row=row, column=7, value=f"{len(image_paths)} photo(s)")
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20
        
        # Save file
        timestamp = int(time.time())
        filename = f"hvac_mep_report_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        logger.info(f"Excel report created: {filepath}")
        
        return filepath, filename
        
    except Exception as e:
        logger.error(f"Failed to create Excel report: {str(e)}")
        raise
