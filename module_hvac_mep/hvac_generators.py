import logging
import os
import json
import traceback
from datetime import datetime, timedelta, timezone as dt_timezone
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import base64
from common.utils import get_image_for_pdf

# Try importing PIL for better image handling
try:
    from PIL import Image as PILImage
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

logger = logging.getLogger(__name__)

# Dubai timezone offset (Gulf Standard Time, UTC+4)
DUBAI_OFFSET = timedelta(hours=4)

def get_dubai_time():
    """Get current time in Dubai timezone (GST - Gulf Standard Time, UTC+4)"""
    # Get UTC time and add 4 hours for Dubai time
    utc_now = datetime.utcnow()
    dubai_time = utc_now + DUBAI_OFFSET
    return dubai_time

def format_dubai_datetime(dt=None, format_str='%Y-%m-%d %H:%M:%S'):
    """Format datetime in Dubai timezone (GST, UTC+4)"""
    if dt is None:
        dt = get_dubai_time()
    elif isinstance(dt, datetime):
        # If datetime has timezone info, convert to UTC first, then add Dubai offset
        if dt.tzinfo is not None:
            # Convert to UTC
            utc_dt = dt.astimezone(dt_timezone.utc).replace(tzinfo=None)
        else:
            # Assume UTC if naive
            utc_dt = dt
        # Add Dubai offset (UTC+4)
        dt = utc_dt + DUBAI_OFFSET
    else:
        # If not datetime, get current Dubai time
        dt = get_dubai_time()
    return dt.strftime(format_str)

# Try importing professional PDF service, fall back if unavailable
try:
    import sys
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from app.services.professional_pdf_service import (
        create_professional_pdf,
        create_header_with_logo,
        create_info_table,
        create_data_table,
        add_photo_grid,
        add_signatures_section,
        add_section_heading,
        add_item_heading,
        add_paragraph,
        get_professional_styles
    )
    USE_PROFESSIONAL_PDF = True
    logger.info("✅ Professional PDF service loaded successfully")
except Exception as e:
    logger.warning(f"⚠️ Professional PDF service not available: {e}. Using basic PDF generation.")
    USE_PROFESSIONAL_PDF = False

def _write_materials_sheet(ws, materials):
    """Write a 'Materials Used' sheet - matches reference format HVAC_MEP_Injaaz_*.xlsx."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    header_fill  = PatternFill('solid', fgColor='125435')
    alt_fill     = PatternFill('solid', fgColor='E3F2FD')  # Light blue for zebra striping
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    body_font    = Font(name='Calibri', size=10)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    right_align  = Alignment(horizontal='right',  vertical='center', wrap_text=True)
    thin         = Side(style='thin', color='BBDEFB')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['#', 'Material Name', 'Brand', 'Department', 'UOM', 'Quantity', 'Unit Price (AED)', 'Line Total (AED)']
    # Column widths: A narrow, B widest, G/H wide, C-F medium
    col_widths = [6, 42, 16, 14, 10, 10, 18, 20]

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "MATERIALS & COST BREAKDOWN"
    title_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=13)
    title_cell.alignment = center_align
    title_cell.fill = header_fill
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:H2')
    sub_cell = ws['A2']
    sub_cell.value = "Selected inspection materials with pricing and cost totals"
    sub_cell.font = Font(name='Calibri', bold=False, color='125435', size=10)
    sub_cell.alignment = center_align
    sub_cell.fill = PatternFill('solid', fgColor='E8F5E9')
    ws.row_dimensions[2].height = 22

    header_row = 4
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    grand_total = 0.0
    data_start_row = header_row + 1
    for idx, m in enumerate(materials, 1):
        row_idx = data_start_row + idx - 1
        row_fill = PatternFill('solid', fgColor='FFFFFF') if row_idx % 2 == 0 else alt_fill
        qty = float(m.get('quantity', 1) or 0)
        unit_price = float(m.get('unit_price', 0) or 0)
        line_total = qty * unit_price
        grand_total += line_total
        row_data = [
            idx,
            str(m.get('name', '')),
            str(m.get('brand', '') or ''),
            str(m.get('department', '') or ''),
            str(m.get('uom', '') or ''),
            qty,
            unit_price,
            line_total,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = body_font
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = center_align if col_idx in (1, 6) else (right_align if col_idx in (7, 8) else left_align)
            if col_idx in (7, 8):
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row_idx].height = 22

    total_row = data_start_row + len(materials)
    for col_idx in range(1, 9):
        cell = ws.cell(row=total_row, column=col_idx, value='' if col_idx < 7 else None)
        cell.font = Font(name='Calibri', bold=True, size=10)
        cell.fill = PatternFill('solid', fgColor='E8F5E9')
        cell.border = border
        cell.alignment = right_align if col_idx in (7, 8) else center_align
    grand_total_cell = ws.cell(row=total_row, column=7, value='Grand Total (AED)')
    grand_total_cell.alignment = right_align
    total_value_cell = ws.cell(row=total_row, column=8, value=grand_total)
    total_value_cell.alignment = right_align
    total_value_cell.number_format = '#,##0.00'
    ws.row_dimensions[total_row].height = 24

    data_end_row = total_row - 1
    if data_end_row >= data_start_row:
        ws.auto_filter.ref = f"A{header_row}:H{data_end_row}"
    ws.freeze_panes = f"A{data_start_row}"


def _add_summary_kpi_cards(ws, start_row, cards):
    """Add KPI card band (4 cards, 2 columns each) to summary sheet."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[start_row].height = 20
    ws.row_dimensions[start_row + 1].height = 28

    groups = [(1, 2), (3, 4), (5, 6), (7, 8)]
    for idx, (c1, c2) in enumerate(groups):
        if idx >= len(cards):
            break
        label, value = cards[idx]
        ws.merge_cells(start_row=start_row, start_column=c1, end_row=start_row, end_column=c2)
        ws.merge_cells(start_row=start_row + 1, start_column=c1, end_row=start_row + 1, end_column=c2)

        lbl_cell = ws.cell(row=start_row, column=c1, value=label)
        lbl_cell.font = Font(name='Calibri', size=9, bold=True, color='125435')
        lbl_cell.alignment = Alignment(horizontal='center', vertical='center')
        lbl_cell.fill = PatternFill('solid', fgColor='E8F5E9')
        lbl_cell.border = border

        val_cell = ws.cell(row=start_row + 1, column=c1, value=value)
        val_cell.font = Font(name='Calibri', size=13, bold=True, color='0F172A')
        val_cell.alignment = Alignment(horizontal='center', vertical='center')
        val_cell.fill = PatternFill('solid', fgColor='FFFFFF')
        val_cell.border = border

        ws.cell(row=start_row, column=c2).fill = PatternFill('solid', fgColor='E8F5E9')
        ws.cell(row=start_row, column=c2).border = border
        ws.cell(row=start_row + 1, column=c2).fill = PatternFill('solid', fgColor='FFFFFF')
        ws.cell(row=start_row + 1, column=c2).border = border

    return start_row + 3


def create_excel_report(data, output_dir):
    """Generate HVAC/MEP Excel report with professional formatting."""
    try:
        # Import professional Excel service
        sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        from app.services.professional_excel_service import (
            create_professional_excel_workbook,
            add_logo_and_title,
            add_info_section,
            add_data_table,
            add_section_header,
            finalize_workbook,
            recenter_logo
        )
        
        logger.info(f"Creating professional Excel report in {output_dir}")
        
        # Generate filename
        site_name = data.get('site_name', 'Unknown_Site').replace(' ', '_')
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"HVAC_MEP_{site_name}_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Core data
        items = data.get('items', [])
        materials = data.get('materials_required', []) if isinstance(data.get('materials_required', []), list) else []
        materials_total = 0.0
        for m in materials:
            try:
                materials_total += float(m.get('quantity', 1) or 0) * float(m.get('unit_price', 0) or 0)
            except Exception:
                pass
        priced_materials = sum(
            1 for m in materials
            if (str(m.get('unit_price', '')).strip() not in ('', '0', '0.0', '0.00'))
        )

        # Sheet 1: Executive Summary
        wb, ws_summary = create_professional_excel_workbook(
            title="HVAC & MEP Inspection Report",
            sheet_name="Summary"
        )
        current_row = add_logo_and_title(
            ws_summary,
            title="HVAC & MEP INSPECTION REPORT",
            subtitle=f"Site: {data.get('site_name', 'N/A')}",
            max_columns=8
        )
        summary_info = [
            ('Site Name', data.get('site_name', 'N/A')),
            ('Visit Date', data.get('visit_date', 'N/A')),
            ('Report Generated', format_dubai_datetime() + ' (GST)'),
            ('Inspection Items', str(len(items))),
            ('Selected Materials', str(len(materials))),
            ('Materials Total (AED)', f"{materials_total:,.2f}")
        ]
        current_row = add_info_section(ws_summary, summary_info, current_row, title="Executive Overview", max_columns=8)
        current_row = _add_summary_kpi_cards(ws_summary, current_row, [
            ("Total Items", len(items)),
            ("Total Materials", len(materials)),
            ("Priced Materials", priced_materials),
            ("Materials Value (AED)", f"{materials_total:,.2f}")
        ])
        ws_summary.freeze_panes = "A8"
        finalize_workbook(ws_summary)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [23, 32, 17, 10, 18, 10, 23, 10]):
            ws_summary.column_dimensions[_cl].width = _w
        recenter_logo(ws_summary)

        # Sheet 2: Inspection Items
        ws_items = wb.create_sheet(title="Inspection Items")
        items_row = add_logo_and_title(
            ws_items,
            title="HVAC & MEP - INSPECTION ITEMS",
            subtitle=f"Site: {data.get('site_name', 'N/A')}",
            max_columns=8
        )
        items_row = add_info_section(
            ws_items,
            [
                ('Inspector Scope', 'Detailed asset-wise inspection checklist'),
                ('Items Included', str(len(items))),
            ],
            items_row,
            title="Inspection Register",
            max_columns=8
        )
        headers = ['#', 'Asset', 'System', 'Description', 'Quantity', 'Brand', 'Specification', 'Comments']
        table_data = []
        for idx, item in enumerate(items, 1):
            table_data.append([
                idx,
                item.get('asset', 'N/A'),
                item.get('system', 'N/A'),
                item.get('description', 'N/A'),
                float(item.get('quantity', 0) or 0),
                item.get('brand', 'N/A'),
                item.get('specification', 'N/A'),
                item.get('comments', 'N/A')
            ])
        col_widths = {
            'A': 6,   # #
            'B': 18,  # Asset
            'C': 18,  # System
            'D': 30,  # Description
            'E': 10,  # Quantity
            'F': 16,  # Brand
            'G': 22,  # Specification
            'H': 30   # Comments
        }
        table_start_row = add_section_header(ws_items, "Item-wise Findings", items_row, span_columns=8)
        add_data_table(ws_items, headers, table_data, table_start_row, title=None, col_widths=col_widths)
        if table_data:
            header_row = table_start_row
            last_row = header_row + len(table_data)
            ws_items.auto_filter.ref = f"A{header_row}:H{last_row}"
            ws_items.freeze_panes = f"A{header_row + 1}"
        finalize_workbook(ws_items)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [16, 47, 13, 47, 15, 14, 20, 31]):
            ws_items.column_dimensions[_cl].width = _w
        recenter_logo(ws_items)

        # Sheet 3: Materials & Cost
        ws_mat = wb.create_sheet(title="Materials & Cost")
        if materials:
            _write_materials_sheet(ws_mat, materials)
        else:
            ws_mat.merge_cells('A1:H1')
            ws_mat['A1'] = "MATERIALS & COST BREAKDOWN"
            ws_mat['A1'].font = ws_summary['A1'].font
            ws_mat['A1'].fill = ws_summary['A1'].fill
            ws_mat['A1'].alignment = ws_summary['A1'].alignment
            ws_mat.merge_cells('A3:H3')
            ws_mat['A3'] = "No materials were selected for this inspection."
        finalize_workbook(ws_mat)
        # Lock column widths to match reference format exactly
        for _cl, _w in zip('ABCDEFGH', [13, 25, 15, 17, 11, 15, 24, 25]):
            ws_mat.column_dimensions[_cl].width = _w

        # Save workbook
        wb.save(excel_path)
        
        if not os.path.exists(excel_path):
            raise Exception(f"Excel file not created at {excel_path}")
        
        logger.info(f"✅ Professional Excel report created: {excel_path}")
        return excel_path
        
    except Exception as e:
        logger.error(f"❌ Excel generation error: {str(e)}")
        raise

def create_pdf_report(data, output_dir):
    """Generate comprehensive HVAC/MEP PDF report with professional branding."""
    try:
        logger.info(f"Creating professional HVAC/MEP PDF report in {output_dir}")
        
        # Generate filename
        site_name = data.get('site_name', 'Unknown_Site').replace(' ', '_')
        timestamp = get_dubai_time().strftime('%Y%m%d_%H%M%S')
        pdf_filename = f"HVAC_MEP_{site_name}_{timestamp}.pdf"
        pdf_path = os.path.join(output_dir, pdf_filename)
        
        # Container for PDF elements
        story = []
        styles = get_professional_styles()
        
        # HEADER WITH LOGO
        create_header_with_logo(
            story,
            "HVAC & MEP INSPECTION REPORT",
            f"Site: {data.get('site_name', 'N/A')}"
        )
        
        # Compact separator
        story.append(Spacer(1, 0.04*inch))
        
        # SITE INFORMATION
        add_section_heading(story, "Site Information")
        
        site_info_data = [
            ['Site Name:', data.get('site_name', 'N/A')],
            ['Visit Date:', data.get('visit_date', 'N/A')],
            ['Report Generated:', format_dubai_datetime() + ' (GST)'],
            ['Total Items Inspected:', str(len(data.get('items', [])))]
        ]
        
        site_table = create_info_table(site_info_data, col_widths=[2.35*inch, 4.65*inch])
        story.append(site_table)
        story.append(Spacer(1, 0.1*inch))
        
        # INSPECTION ITEMS
        items = data.get('items', [])
        
        if items:
            add_section_heading(story, "Inspection Items")
            
            for idx, item in enumerate(items, 1):
                # Item header
                add_item_heading(story, f"Item {idx}: {item.get('asset', 'N/A')}")
                
                # Item details table - All fields
                item_details = [
                    ['Asset Name:', item.get('asset', 'N/A')],
                    ['System Type:', item.get('system', 'N/A')],
                    ['Description:', item.get('description', 'N/A')],
                    ['Quantity:', str(item.get('quantity', 'N/A'))],
                    ['Brand:', item.get('brand', 'N/A')],
                    ['Specification:', item.get('specification', 'N/A')],
                    ['Comments:', item.get('comments', 'N/A')],
                    ['Photos Attached:', str(len(item.get('photos', [])))]
                ]
                
                item_table = create_info_table(item_details, col_widths=[2.35*inch, 4.65*inch])
                story.append(item_table)
                story.append(Spacer(1, 0.06*inch))
                
                # PHOTOS - Support both cloud URLs and local paths
                photos = item.get('photos', [])
                
                if photos:
                    add_paragraph(story, f"<b>Attached Photos ({len(photos)} total):</b>")
                    story.append(Spacer(1, 0.04*inch))
                    add_photo_grid(story, photos)
                
                # Small spacer between items (no page break - avoids huge gaps)
                if idx < len(items):
                    story.append(Spacer(1, 0.15*inch))
        
        else:
            add_paragraph(story, "No inspection items recorded.")

        # MATERIALS REQUIRED SECTION
        materials = data.get('materials_required', [])
        if materials and isinstance(materials, list) and len(materials) > 0:
            add_section_heading(story, "Materials Required")
            mat_headers = ['#', 'Material Name', 'Brand', 'Department', 'UOM', 'Qty', 'Unit Price (AED)', 'Line Total (AED)']
            mat_data = []
            materials_total = 0.0
            for idx, m in enumerate(materials, 1):
                qty = float(m.get('quantity', 1) or 0)
                unit_price = float(m.get('unit_price', 0) or 0)
                line_total = qty * unit_price
                materials_total += line_total
                mat_data.append([
                    str(idx),
                    str(m.get('name', 'N/A')),
                    str(m.get('brand', '') or '—'),
                    str(m.get('department', '') or '—'),
                    str(m.get('uom', '') or '—'),
                    f"{qty:g}",
                    f"{unit_price:,.2f}",
                    f"{line_total:,.2f}",
                ])
            mat_col_widths = [0.3*inch, 1.8*inch, 0.9*inch, 0.8*inch, 0.5*inch, 0.45*inch, 0.9*inch, 0.95*inch]
            mat_table_data = [mat_headers] + mat_data
            mat_table = Table(mat_table_data, colWidths=mat_col_widths, repeatRows=1)
            mat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#125435')),
                ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
                ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0, 0), (-1, 0), 8),
                ('ALIGN',      (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME',   (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE',   (0, 1), (-1, -1), 8),
                ('ALIGN',      (0, 1), (0, -1), 'CENTER'),
                ('ALIGN',      (5, 1), (5, -1), 'CENTER'),
                ('ALIGN',      (6, 1), (7, -1), 'RIGHT'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0FAF5')]),
                ('GRID',       (0, 0), (-1, -1), 0.5, colors.HexColor('#BBDEFB')),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(mat_table)
            add_paragraph(story, f"<b>Materials Grand Total (AED):</b> {materials_total:,.2f}")
            story.append(Spacer(1, 0.15*inch))

        # SIGNATURES PAGE - Professional format with all signatures
        signatures = {}
        
        # Get nested data dict if it exists (extract once to avoid f-string issues)
        nested_data = data.get('data') if isinstance(data.get('data'), dict) else {}
        
        # Check for supervisor signature (new workflow field)
        # Try multiple paths: direct key, nested in data, form_data, etc.
        supervisor_sig = None
        supervisor_sig_path = None
        
        # Check direct path first - handle None/null explicitly
        supervisor_sig_raw = data.get('supervisor_signature')
        if supervisor_sig_raw is not None and supervisor_sig_raw != '' and supervisor_sig_raw != 'None':
            supervisor_sig = supervisor_sig_raw
            supervisor_sig_path = 'direct (supervisor_signature)'
        # Check nested in data
        elif nested_data:
            supervisor_sig_raw = nested_data.get('supervisor_signature')
            if supervisor_sig_raw is not None and supervisor_sig_raw != '' and supervisor_sig_raw != 'None':
                supervisor_sig = supervisor_sig_raw
                supervisor_sig_path = 'nested (data.supervisor_signature)'
        # Check in form_data if it exists
        if supervisor_sig is None and isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            supervisor_sig_raw = form_data_dict.get('supervisor_signature')
            if supervisor_sig_raw is not None and supervisor_sig_raw != '' and supervisor_sig_raw != 'None':
                supervisor_sig = supervisor_sig_raw
                supervisor_sig_path = 'form_data (supervisor_signature)'
            # Also check nested form_data['data']
            elif isinstance(form_data_dict.get('data'), dict):
                nested_form_data = form_data_dict.get('data', {})
                supervisor_sig_raw = nested_form_data.get('supervisor_signature')
                if supervisor_sig_raw is not None and supervisor_sig_raw != '' and supervisor_sig_raw != 'None':
                    supervisor_sig = supervisor_sig_raw
                    supervisor_sig_path = 'form_data.data (supervisor_signature)'
        
        # Also check if supervisor_signature is an object with url property
        if supervisor_sig and isinstance(supervisor_sig, dict):
            if supervisor_sig.get('url'):
                supervisor_sig = supervisor_sig.get('url')
                logger.info(f"✅ Extracted supervisor signature URL from object format")
            else:
                # Object without url - might be invalid, set to None
                logger.warning(f"⚠️ Supervisor signature is object but has no 'url' property: {supervisor_sig}")
                supervisor_sig = None
        
        # Convert empty strings to None
        if supervisor_sig == '' or supervisor_sig == 'None':
            supervisor_sig = None
        
        # Check for supervisor comments - try all possible paths - handle None/null explicitly
        supervisor_comments = None
        supervisor_comments_raw = data.get('supervisor_comments')
        if supervisor_comments_raw is not None and supervisor_comments_raw != 'None':
            supervisor_comments = supervisor_comments_raw
        elif nested_data:
            supervisor_comments_raw = nested_data.get('supervisor_comments')
            if supervisor_comments_raw is not None and supervisor_comments_raw != 'None':
                supervisor_comments = supervisor_comments_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            supervisor_comments_raw = form_data_dict.get('supervisor_comments')
            if supervisor_comments_raw is not None and supervisor_comments_raw != 'None':
                supervisor_comments = supervisor_comments_raw
            elif isinstance(form_data_dict.get('data'), dict):
                nested_form_data = form_data_dict.get('data', {})
                supervisor_comments_raw = nested_form_data.get('supervisor_comments')
                if supervisor_comments_raw is not None and supervisor_comments_raw != 'None':
                    supervisor_comments = supervisor_comments_raw
        
        # Convert None to empty string (so we can show placeholder)
        if supervisor_comments is None:
            supervisor_comments = ''
        elif supervisor_comments == 'None':
            supervisor_comments = ''
        
        # Check for Operations Manager comments (try multiple paths) - handle None/null explicitly
        operations_manager_comments = None
        operations_manager_comments_raw = data.get('operations_manager_comments')
        if operations_manager_comments_raw is not None and operations_manager_comments_raw != 'None':
            operations_manager_comments = operations_manager_comments_raw
        elif data.get('opMan_comments'):
            operations_manager_comments = data.get('opMan_comments')
        elif data.get('opman_comments'):
            operations_manager_comments = data.get('opman_comments')
        elif nested_data:
            operations_manager_comments_raw = nested_data.get('operations_manager_comments')
            if operations_manager_comments_raw is not None and operations_manager_comments_raw != 'None':
                operations_manager_comments = operations_manager_comments_raw
            elif nested_data.get('opMan_comments'):
                operations_manager_comments = nested_data.get('opMan_comments')
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            operations_manager_comments_raw = form_data_dict.get('operations_manager_comments')
            if operations_manager_comments_raw is not None and operations_manager_comments_raw != 'None':
                operations_manager_comments = operations_manager_comments_raw
            elif form_data_dict.get('opMan_comments'):
                operations_manager_comments = form_data_dict.get('opMan_comments')
            elif isinstance(form_data_dict.get('data'), dict):
                nested_form_data = form_data_dict.get('data', {})
                operations_manager_comments_raw = nested_form_data.get('operations_manager_comments')
                if operations_manager_comments_raw is not None and operations_manager_comments_raw != 'None':
                    operations_manager_comments = operations_manager_comments_raw
        
        # Convert None to empty string
        if operations_manager_comments is None:
            operations_manager_comments = ''
        elif operations_manager_comments == 'None':
            operations_manager_comments = ''
        
        # Log Operations Manager comments detection for debugging
        logger.info(f"🔍 Checking Operations Manager comments in PDF generation:")
        logger.info(f"  - Direct operations_manager_comments: {bool(data.get('operations_manager_comments'))} (value: {str(data.get('operations_manager_comments', ''))[:50] if data.get('operations_manager_comments') else 'None'})")
        logger.info(f"  - Nested in data: {bool(nested_data.get('operations_manager_comments') if nested_data else False)}")
        logger.info(f"  - Final operations_manager_comments length: {len(str(operations_manager_comments)) if operations_manager_comments else 0}")
        if operations_manager_comments:
            logger.info(f"  - Operations Manager comments found: {str(operations_manager_comments)[:100]}...")
        else:
            logger.warning(f"  - ⚠️ No Operations Manager comments found in data")
        
        # Check for Business Development comments - handle None/null explicitly
        # CRITICAL: Only use actual BD comments, never fall back to supervisor comments
        business_dev_comments = None
        supervisor_comments = data.get('supervisor_comments') or (nested_data.get('supervisor_comments') if nested_data else None) or (data.get('form_data', {}).get('supervisor_comments') if isinstance(data.get('form_data'), dict) else None)
        
        business_dev_comments_raw = data.get('business_dev_comments') or data.get('business_development_comments')
        if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
            # Validate that BD comments are not supervisor comments
            if business_dev_comments_raw != supervisor_comments:
                business_dev_comments = business_dev_comments_raw
            else:
                logger.warning(f"⚠️ PDF: BD comments appear to be supervisor comments, ignoring (value: {business_dev_comments_raw[:50]}...)")
        elif nested_data:
            business_dev_comments_raw = nested_data.get('business_dev_comments') or nested_data.get('business_development_comments')
            if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
                # Validate that BD comments are not supervisor comments
                if business_dev_comments_raw != supervisor_comments:
                    business_dev_comments = business_dev_comments_raw
                else:
                    logger.warning(f"⚠️ PDF: BD comments in nested_data appear to be supervisor comments, ignoring (value: {business_dev_comments_raw[:50]}...)")
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            business_dev_comments_raw = form_data_dict.get('business_dev_comments') or form_data_dict.get('business_development_comments')
            if business_dev_comments_raw is not None and business_dev_comments_raw != 'None' and business_dev_comments_raw != '':
                # Validate that BD comments are not supervisor comments
                if business_dev_comments_raw != supervisor_comments:
                    business_dev_comments = business_dev_comments_raw
                else:
                    logger.warning(f"⚠️ PDF: BD comments in form_data appear to be supervisor comments, ignoring (value: {business_dev_comments_raw[:50]}...)")
        
        # Convert None to empty string
        if business_dev_comments is None:
            business_dev_comments = ''
        
        # Log BD comments extraction
        logger.info(f"🔍 Business Development comments extraction:")
        logger.info(f"  - business_dev_comments: {bool(business_dev_comments)} (value: {business_dev_comments[:50] if business_dev_comments else 'None'}...)")
        if supervisor_comments:
            logger.info(f"  - supervisor_comments (for comparison): {supervisor_comments[:50]}...")
            if business_dev_comments == supervisor_comments:
                logger.warning(f"  - ⚠️ WARNING: BD comments match supervisor comments - this should not happen!")
        
        # Check for Procurement comments - handle None/null explicitly and check multiple paths
        procurement_comments = None
        procurement_comments_raw = data.get('procurement_comments')
        if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
            procurement_comments = procurement_comments_raw
        elif nested_data:
            procurement_comments_raw = nested_data.get('procurement_comments')
            if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
                procurement_comments = procurement_comments_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            procurement_comments_raw = form_data_dict.get('procurement_comments')
            if procurement_comments_raw is not None and procurement_comments_raw != 'None' and procurement_comments_raw != '':
                procurement_comments = procurement_comments_raw
        
        # Convert None to empty string
        if procurement_comments is None:
            procurement_comments = ''
        
        # Log Procurement comments extraction
        logger.info(f"🔍 Procurement comments extraction:")
        logger.info(f"  - procurement_comments: {bool(procurement_comments)} (value: {procurement_comments[:50] if procurement_comments else 'None'}...)")
        logger.info(f"  - supervisor_comments (for comparison): {bool(supervisor_comments)} (value: {supervisor_comments[:50] if supervisor_comments else 'None'}...)")
        
        # Check for General Manager comments
        general_manager_comments = (
            data.get('general_manager_comments') or
            (nested_data.get('general_manager_comments') if nested_data else '') or
            ''
        )
        
        # Log signature detection for debugging
        logger.info(f"🔍 Checking supervisor signature in PDF generation:")
        logger.info(f"  - Direct supervisor_signature: {bool(data.get('supervisor_signature'))} (value type: {type(data.get('supervisor_signature'))})")
        logger.info(f"  - Nested in data: {bool(nested_data.get('supervisor_signature') if nested_data else False)}")
        logger.info(f"  - Found supervisor_sig via: {supervisor_sig_path}")
        logger.info(f"  - Final supervisor_sig type: {type(supervisor_sig)}, length: {len(str(supervisor_sig)) if supervisor_sig else 0}")
        if supervisor_sig:
            logger.info(f"  - Signature preview: {str(supervisor_sig)[:100]}...")
        
        if supervisor_sig:
            # Handle different signature formats
            if isinstance(supervisor_sig, dict):
                # Dictionary format with 'url' key
                if supervisor_sig.get('url'):
                    signatures['Supervisor'] = supervisor_sig
                    logger.debug("✅ Found supervisor signature in dict format with URL")
            elif isinstance(supervisor_sig, str):
                # String format - check if it's a valid image data URL or HTTP URL
                sig_str = supervisor_sig.strip()
                if sig_str and (sig_str.startswith('data:image') or sig_str.startswith('http://') or sig_str.startswith('https://') or sig_str.startswith('/')):
                    signatures['Supervisor'] = sig_str
                    logger.debug(f"✅ Found supervisor signature as string (starts with: {sig_str[:50]}...)")
                elif sig_str:
                    # Might be a base64 string without data: prefix, try it anyway
                    logger.warning(f"⚠️ Supervisor signature doesn't match expected format: {sig_str[:50]}...")
                    signatures['Supervisor'] = sig_str
            else:
                logger.warning(f"⚠️ Supervisor signature has unexpected type: {type(supervisor_sig)}")
        
        # Check for Operations Manager signature (try multiple key variations)
        logger.info(f"🔍 Checking Operations Manager signature in PDF generation:")
        logger.info(f"  - data.get('operations_manager_signature'): {bool(data.get('operations_manager_signature'))}")
        logger.info(f"  - data.get('opMan_signature'): {bool(data.get('opMan_signature'))}")
        logger.info(f"  - data.get('opman_signature'): {bool(data.get('opman_signature'))}")
        logger.info(f"  - data.get('data') is dict: {isinstance(data.get('data'), dict)}")
        if nested_data:
            logger.info(f"  - nested_data.get('operations_manager_signature'): {bool(nested_data.get('operations_manager_signature'))}")
            logger.info(f"  - nested_data.get('opMan_signature'): {bool(nested_data.get('opMan_signature'))}")
            logger.info(f"  - nested_data.get('opman_signature'): {bool(nested_data.get('opman_signature'))}")
        
        # Try all possible paths for Operations Manager signature - check ALL locations including form_data
        opman_sig = None
        opman_sig_path = None
        
        # Check direct paths first
        opman_sig_raw = data.get('operations_manager_signature')
        if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
            opman_sig = opman_sig_raw
            opman_sig_path = 'direct (operations_manager_signature)'
        elif data.get('opMan_signature'):
            opman_sig = data.get('opMan_signature')
            opman_sig_path = 'direct (opMan_signature)'
        elif data.get('opman_signature'):
            opman_sig = data.get('opman_signature')
            opman_sig_path = 'direct (opman_signature)'
        # Check nested paths
        elif nested_data:
            opman_sig_raw = nested_data.get('operations_manager_signature')
            if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
                opman_sig = opman_sig_raw
                opman_sig_path = 'nested (data.operations_manager_signature)'
            elif nested_data.get('opMan_signature'):
                opman_sig = nested_data.get('opMan_signature')
                opman_sig_path = 'nested (data.opMan_signature)'
        # Check form_data path (CRITICAL - this is where it's saved when OM approves)
        if opman_sig is None and isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            opman_sig_raw = form_data_dict.get('operations_manager_signature')
            if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
                opman_sig = opman_sig_raw
                opman_sig_path = 'form_data (operations_manager_signature)'
                logger.info(f"✅ Found Operations Manager signature in form_data")
            elif form_data_dict.get('opMan_signature'):
                opman_sig = form_data_dict.get('opMan_signature')
                opman_sig_path = 'form_data (opMan_signature)'
            # Also check nested form_data['data']
            elif isinstance(form_data_dict.get('data'), dict):
                nested_form_data = form_data_dict.get('data', {})
                opman_sig_raw = nested_form_data.get('operations_manager_signature')
                if opman_sig_raw is not None and opman_sig_raw != '' and opman_sig_raw != 'None':
                    opman_sig = opman_sig_raw
                    opman_sig_path = 'form_data.data (operations_manager_signature)'
        
        if opman_sig:
            logger.info(f"✅ Found Operations Manager signature via: {opman_sig_path}")
            logger.info(f"  - Signature type: {type(opman_sig)}")
            if isinstance(opman_sig, str):
                logger.info(f"  - Signature length: {len(opman_sig)}")
                logger.info(f"  - Signature preview: {opman_sig[:100]}...")
            elif isinstance(opman_sig, dict):
                logger.info(f"  - Signature dict keys: {list(opman_sig.keys())}")
                logger.info(f"  - Signature URL: {opman_sig.get('url', 'N/A')[:100] if opman_sig.get('url') else 'N/A'}")
            
            # Handle object format with url property
            if isinstance(opman_sig, dict) and opman_sig.get('url'):
                signatures['Operations Manager'] = opman_sig
                logger.info("✅ Added Operations Manager signature to signatures dict (dict format with URL)")
            elif isinstance(opman_sig, str) and (opman_sig.startswith('data:image') or opman_sig.startswith('http') or opman_sig.startswith('/')):
                signatures['Operations Manager'] = opman_sig
                logger.info(f"✅ Added Operations Manager signature to signatures dict (string format, length: {len(opman_sig)})")
            else:
                logger.warning(f"⚠️ Operations Manager signature found but format unexpected: {type(opman_sig)}")
                # Try to add it anyway if it's not empty
                if opman_sig and str(opman_sig).strip() and opman_sig != 'None':
                    signatures['Operations Manager'] = opman_sig
                    logger.info("⚠️ Added Operations Manager signature despite unexpected format")
        else:
            logger.warning("⚠️ Operations Manager signature not found in any path")
            logger.warning(f"  - Available keys in data: {list(data.keys())[:20]}")
            if isinstance(data.get('form_data'), dict):
                logger.warning(f"  - Available keys in form_data: {list(data.get('form_data').keys())[:20]}")
            if nested_data:
                logger.warning(f"  - Available keys in nested_data: {list(nested_data.keys())[:20]}")
        
        # Check for Business Development signature (try multiple key variations)
        logger.info(f"🔍 Checking Business Development signature in PDF generation:")
        logger.info(f"  - data.get('business_dev_signature'): {bool(data.get('business_dev_signature'))}")
        logger.info(f"  - data.get('businessDevSignature'): {bool(data.get('businessDevSignature'))}")
        if nested_data:
            logger.info(f"  - data.get('data').get('business_dev_signature'): {bool(nested_data.get('business_dev_signature'))}")
        
        # Check for Business Development signature - handle None/null explicitly and check multiple paths
        business_dev_sig = None
        business_dev_sig_raw = data.get('business_dev_signature') or data.get('businessDevSignature')
        if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
            business_dev_sig = business_dev_sig_raw
        elif nested_data:
            business_dev_sig_raw = nested_data.get('business_dev_signature') or nested_data.get('businessDevSignature')
            if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
                business_dev_sig = business_dev_sig_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            business_dev_sig_raw = form_data_dict.get('business_dev_signature') or form_data_dict.get('businessDevSignature')
            if business_dev_sig_raw is not None and business_dev_sig_raw != 'None' and business_dev_sig_raw != '':
                business_dev_sig = business_dev_sig_raw
        
        if business_dev_sig:
            if isinstance(business_dev_sig, dict) and business_dev_sig.get('url'):
                signatures['Business Development'] = business_dev_sig
                logger.info("✅ Found Business Development signature in dict format with URL")
            elif isinstance(business_dev_sig, str) and (business_dev_sig.startswith('data:image') or business_dev_sig.startswith('http') or business_dev_sig.startswith('/')):
                signatures['Business Development'] = business_dev_sig
                logger.info(f"✅ Found Business Development signature as string (length: {len(business_dev_sig)})")
            else:
                logger.warning(f"⚠️ Business Development signature found but format unexpected: {type(business_dev_sig)}")
        else:
            logger.debug("ℹ️ Business Development signature not found (may not be approved yet)")
        
        # Check for Procurement signature - handle None/null explicitly and check multiple paths
        procurement_sig = None
        procurement_sig_raw = data.get('procurement_signature') or data.get('procurementSignature')
        if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
            procurement_sig = procurement_sig_raw
        elif nested_data:
            procurement_sig_raw = nested_data.get('procurement_signature') or nested_data.get('procurementSignature')
            if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
                procurement_sig = procurement_sig_raw
        elif isinstance(data.get('form_data'), dict):
            form_data_dict = data.get('form_data', {})
            procurement_sig_raw = form_data_dict.get('procurement_signature') or form_data_dict.get('procurementSignature')
            if procurement_sig_raw is not None and procurement_sig_raw != 'None' and procurement_sig_raw != '':
                procurement_sig = procurement_sig_raw
        
        if procurement_sig:
            if isinstance(procurement_sig, dict) and procurement_sig.get('url'):
                signatures['Procurement'] = procurement_sig
                logger.info("✅ Found Procurement signature in dict format with URL")
            elif isinstance(procurement_sig, str) and (procurement_sig.startswith('data:image') or procurement_sig.startswith('http') or procurement_sig.startswith('/')):
                signatures['Procurement'] = procurement_sig
                logger.info(f"✅ Found Procurement signature as string (length: {len(procurement_sig)})")
            else:
                logger.warning(f"⚠️ Procurement signature found but format unexpected: {type(procurement_sig)}")
        else:
            logger.debug("ℹ️ Procurement signature not found (may not be approved yet)")
        
        # Check for General Manager signature
        general_manager_sig = data.get('general_manager_signature', '') or data.get('generalManagerSignature', '')
        if general_manager_sig:
            if isinstance(general_manager_sig, dict) and general_manager_sig.get('url'):
                signatures['General Manager'] = general_manager_sig
            elif isinstance(general_manager_sig, str) and (general_manager_sig.startswith('data:image') or general_manager_sig.startswith('http')):
                signatures['General Manager'] = general_manager_sig
        
        # Helper function to add comment and signature together for a reviewer
        def add_reviewer_section(role_name, comments, signature_data, always_show_signature=False):
            """Add comments and signature together for a reviewer
            
            Args:
                role_name: Name of the reviewer role
                comments: Comments text (can be None/empty)
                signature_data: Signature data (can be None/empty)
                always_show_signature: If True, always show signature section even if missing (default: False)
            """
            has_content = False
            
            if comments and comments.strip():
                add_section_heading(story, f"{role_name} Comments")
                add_paragraph(story, comments)
                story.append(Spacer(1, 0.1*inch))
                has_content = True
            
            # Always show signature section if signature_data exists OR if always_show_signature is True
            if signature_data or always_show_signature:
                # Add signature section for this reviewer
                styles = get_professional_styles()
                sig_rows = []
                
                if signature_data:
                    try:
                        from common.utils import get_image_for_pdf
                        from PIL import Image as PILImage
                        
                        img_data, is_url = get_image_for_pdf(signature_data)
                        if img_data:
                            # Calculate size maintaining aspect ratio
                            max_width = 2.5 * inch
                            max_height = 1.2 * inch
                            
                            if HAS_PIL:
                                # Use PIL to get actual image dimensions for proper aspect ratio
                                try:
                                    if is_url:
                                        img_data.seek(0)
                                        pil_img = PILImage.open(img_data)
                                    else:
                                        pil_img = PILImage.open(img_data)
                                    
                                    orig_width, orig_height = pil_img.size
                                    
                                    # Calculate scaling factor to fit within max dimensions while maintaining aspect ratio
                                    width_ratio = max_width / orig_width
                                    height_ratio = max_height / orig_height
                                    scale_ratio = min(width_ratio, height_ratio)  # Use min to ensure it fits within bounds
                                    
                                    final_width = orig_width * scale_ratio
                                    final_height = orig_height * scale_ratio
                                    
                                    # Verify aspect ratio is maintained
                                    original_ratio = orig_width / orig_height if orig_height > 0 else 1
                                    final_ratio = final_width / final_height if final_height > 0 else 1
                                    
                                    # Create ReportLab Image with calculated dimensions
                                    # By calculating both dimensions from the same scale_ratio, aspect ratio is preserved
                                    if is_url:
                                        img_data.seek(0)
                                        sig_img = Image(img_data, width=final_width, height=final_height)
                                    else:
                                        sig_img = Image(img_data, width=final_width, height=final_height)
                                    
                                    # Log dimensions for verification
                                    logger.info(f"✅ {role_name} signature aspect ratio: Original={orig_width}x{orig_height} (ratio={original_ratio:.3f}), Final={final_width:.2f}x{final_height:.2f} (ratio={final_ratio:.3f}), Scale={scale_ratio:.3f}")
                                    
                                    # Double-check: aspect ratios should match (within rounding error)
                                    if abs(original_ratio - final_ratio) > 0.01:
                                        logger.warning(f"⚠️ {role_name} signature aspect ratio mismatch! Original={original_ratio:.3f}, Final={final_ratio:.3f}")
                                except Exception as pil_error:
                                    logger.warning(f"PIL image processing failed, using fallback: {pil_error}")
                                    if is_url:
                                        img_data.seek(0)
                                        sig_img = Image(img_data)
                                    else:
                                        sig_img = Image(img_data)
                                    
                                    # Get image dimensions and calculate aspect-ratio-preserving size
                                    if hasattr(sig_img, 'imageWidth') and hasattr(sig_img, 'imageHeight'):
                                        orig_width = sig_img.imageWidth
                                        orig_height = sig_img.imageHeight
                                        if orig_width > 0 and orig_height > 0:
                                            # Calculate scaling factor to fit within max dimensions
                                            width_ratio = max_width / orig_width
                                            height_ratio = max_height / orig_height
                                            scale_ratio = min(width_ratio, height_ratio)
                                            
                                            # Set dimensions maintaining aspect ratio
                                            final_width = orig_width * scale_ratio
                                            final_height = orig_height * scale_ratio
                                            sig_img.drawWidth = final_width
                                            sig_img.drawHeight = final_height
                                            logger.debug(f"✅ {role_name} signature (fallback): Original={orig_width}x{orig_height}, Final={final_width:.2f}x{final_height:.2f}, Ratio={scale_ratio:.3f}")
                                        else:
                                            # If dimensions unknown, set max width only and let height adjust automatically
                                            sig_img.drawWidth = max_width
                                            logger.debug(f"⚠️ {role_name} signature: Unknown dimensions, using max width only")
                                    else:
                                        # Fallback: set max width only and let height adjust automatically
                                        sig_img.drawWidth = max_width
                                        logger.debug(f"⚠️ {role_name} signature: No dimension attributes, using max width only")
                            else:
                                # Fallback without PIL
                                if is_url:
                                    img_data.seek(0)
                                    sig_img = Image(img_data)
                                else:
                                    sig_img = Image(img_data)
                                
                                # Get image dimensions and calculate aspect-ratio-preserving size
                                if hasattr(sig_img, 'imageWidth') and hasattr(sig_img, 'imageHeight'):
                                    orig_width = sig_img.imageWidth
                                    orig_height = sig_img.imageHeight
                                    if orig_width > 0 and orig_height > 0:
                                        # Calculate scaling factor to fit within max dimensions
                                        width_ratio = max_width / orig_width
                                        height_ratio = max_height / orig_height
                                        scale_ratio = min(width_ratio, height_ratio)
                                        
                                        # Set dimensions maintaining aspect ratio
                                        final_width = orig_width * scale_ratio
                                        final_height = orig_height * scale_ratio
                                        sig_img.drawWidth = final_width
                                        sig_img.drawHeight = final_height
                                        logger.debug(f"✅ {role_name} signature (no PIL): Original={orig_width}x{orig_height}, Final={final_width:.2f}x{final_height:.2f}, Ratio={scale_ratio:.3f}")
                                    else:
                                        # If dimensions unknown, set max width only and let height adjust automatically
                                        sig_img.drawWidth = max_width
                                        logger.debug(f"⚠️ {role_name} signature: Unknown dimensions, using max width only")
                                else:
                                    # Fallback: set max width only and let height adjust automatically
                                    sig_img.drawWidth = max_width
                                    logger.debug(f"⚠️ {role_name} signature: No dimension attributes, using max width only")
                            
                            sig_rows.append([
                                Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                                sig_img
                            ])
                        else:
                            sig_rows.append([
                                Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                                Paragraph("Signature not available", styles['Small'])
                            ])
                    except Exception as e:
                        logger.error(f"Error processing {role_name} signature: {str(e)}")
                        logger.error(traceback.format_exc())
                        sig_rows.append([
                            Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                            Paragraph("Error loading signature", styles['Small'])
                        ])
                else:
                    # No signature data - show placeholder if always_show_signature is True
                    sig_rows.append([
                        Paragraph(f"<b>{role_name} Signature:</b>", styles['Normal']),
                        Paragraph("<i>Not signed</i>", styles['Small'])
                    ])
                
                if sig_rows:
                    sig_table = Table(sig_rows, colWidths=[2*inch, 3.5*inch])
                    sig_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#125435')),
                        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ]))
                    story.append(sig_table)
                    story.append(Spacer(1, 0.15*inch))
                    has_content = True
            
            return has_content
        
        # Add reviewer sections in workflow order: comments + signature together for each
        # 1. Supervisor - ALWAYS show this section (required field)
        # Always add supervisor section, even if empty (show placeholders)
        supervisor_comments_display = supervisor_comments if supervisor_comments and supervisor_comments.strip() else None
        supervisor_sig_display = signatures.get('Supervisor')
        
        # Log supervisor data for debugging
        logger.info(f"🔍 Supervisor section in PDF generation:")
        logger.info(f"  - Raw supervisor_comments: {repr(supervisor_comments)}")
        logger.info(f"  - Comments present: {bool(supervisor_comments_display)}")
        logger.info(f"  - Comments length: {len(supervisor_comments_display) if supervisor_comments_display else 0}")
        logger.info(f"  - Raw supervisor_sig: {type(supervisor_sig) if supervisor_sig else 'None'}")
        logger.info(f"  - Supervisor sig path: {supervisor_sig_path}")
        logger.info(f"  - Signature present: {bool(supervisor_sig_display)}")
        logger.info(f"  - Signature type: {type(supervisor_sig_display) if supervisor_sig_display else 'None'}")
        if supervisor_sig_display:
            if isinstance(supervisor_sig_display, str):
                logger.info(f"  - Signature length: {len(supervisor_sig_display)}")
                logger.info(f"  - Signature preview: {supervisor_sig_display[:100]}...")
            elif isinstance(supervisor_sig_display, dict):
                logger.info(f"  - Signature dict keys: {list(supervisor_sig_display.keys())}")
                logger.info(f"  - Signature URL: {supervisor_sig_display.get('url', 'N/A')[:100] if supervisor_sig_display.get('url') else 'N/A'}")
        
        # Log data structure for debugging
        logger.info(f"  - Data keys: {list(data.keys())[:20]}")
        if isinstance(data.get('form_data'), dict):
            logger.info(f"  - form_data keys: {list(data.get('form_data').keys())[:20]}")
        
        # Always show supervisor section
        add_section_heading(story, "Supervisor Review")
        styles = get_professional_styles()
        # Add comments with more spacing and bold
        if supervisor_comments_display:
            story.append(Paragraph(supervisor_comments_display, styles['CommentLead']))
        else:
            story.append(Paragraph("<i>No comments provided</i>", styles['CommentLead']))
        story.append(Spacer(1, 0.1*inch))
        
        # Add signature section
        styles = get_professional_styles()
        sig_rows = []
        
        if supervisor_sig_display:
            try:
                from common.utils import get_image_for_pdf
                
                img_data, is_url = get_image_for_pdf(supervisor_sig_display)
                if img_data:
                    # Calculate size maintaining aspect ratio
                    max_width = 2.5 * inch
                    max_height = 1.2 * inch
                    
                    if HAS_PIL:
                        # Use PIL to get actual image dimensions for proper aspect ratio
                        try:
                            if is_url:
                                # BytesIO stream
                                img_data.seek(0)
                                pil_img = PILImage.open(img_data)
                            else:
                                # File path
                                pil_img = PILImage.open(img_data)
                            
                            # Get original dimensions
                            orig_width, orig_height = pil_img.size
                            
                            # Calculate scaling factor to fit within max dimensions while maintaining aspect ratio
                            width_ratio = max_width / orig_width
                            height_ratio = max_height / orig_height
                            scale_ratio = min(width_ratio, height_ratio)  # Use min to ensure it fits within bounds
                            
                            # Calculate final dimensions
                            final_width = orig_width * scale_ratio
                            final_height = orig_height * scale_ratio
                            
                            # Verify aspect ratio is maintained
                            original_ratio = orig_width / orig_height if orig_height > 0 else 1
                            final_ratio = final_width / final_height if final_height > 0 else 1
                            
                            # Create ReportLab Image with calculated dimensions
                            # By calculating both dimensions from the same scale_ratio, aspect ratio is preserved
                            if is_url:
                                img_data.seek(0)  # Reset stream
                                sig_img = Image(img_data, width=final_width, height=final_height)
                            else:
                                sig_img = Image(img_data, width=final_width, height=final_height)
                            
                            # Log dimensions for verification
                            logger.info(f"✅ Supervisor signature aspect ratio: Original={orig_width}x{orig_height} (ratio={original_ratio:.3f}), Final={final_width:.2f}x{final_height:.2f} (ratio={final_ratio:.3f}), Scale={scale_ratio:.3f}")
                            
                            # Double-check: aspect ratios should match (within rounding error)
                            if abs(original_ratio - final_ratio) > 0.01:
                                logger.warning(f"⚠️ Supervisor signature aspect ratio mismatch! Original={original_ratio:.3f}, Final={final_ratio:.3f}")
                        except Exception as pil_error:
                            logger.warning(f"PIL image processing failed, using fallback: {pil_error}")
                            # Fallback: use max dimensions but let ReportLab maintain aspect ratio
                            if is_url:
                                img_data.seek(0)
                                sig_img = Image(img_data)
                            else:
                                sig_img = Image(img_data)
                            
                            # Get image dimensions and calculate aspect-ratio-preserving size
                            if hasattr(sig_img, 'imageWidth') and hasattr(sig_img, 'imageHeight'):
                                orig_width = sig_img.imageWidth
                                orig_height = sig_img.imageHeight
                                if orig_width > 0 and orig_height > 0:
                                    # Calculate scaling factor to fit within max dimensions
                                    width_ratio = max_width / orig_width
                                    height_ratio = max_height / orig_height
                                    scale_ratio = min(width_ratio, height_ratio)
                                    
                                    # Set dimensions maintaining aspect ratio
                                    final_width = orig_width * scale_ratio
                                    final_height = orig_height * scale_ratio
                                    sig_img.drawWidth = final_width
                                    sig_img.drawHeight = final_height
                                    logger.debug(f"✅ Supervisor signature (fallback): Original={orig_width}x{orig_height}, Final={final_width:.2f}x{final_height:.2f}, Ratio={scale_ratio:.3f}")
                                else:
                                    # If dimensions unknown, set max width only and let height adjust automatically
                                    sig_img.drawWidth = max_width
                                    logger.debug(f"⚠️ Supervisor signature: Unknown dimensions, using max width only")
                            else:
                                # Fallback: set max width only and let height adjust automatically
                                sig_img.drawWidth = max_width
                                logger.debug(f"⚠️ Supervisor signature: No dimension attributes, using max width only")
                    else:
                        # Fallback without PIL: load image and calculate dimensions manually
                        if is_url:
                            img_data.seek(0)
                            sig_img = Image(img_data)
                        else:
                            sig_img = Image(img_data)
                        
                        # Get image dimensions from ReportLab Image object
                        orig_width = sig_img.imageWidth
                        orig_height = sig_img.imageHeight
                        
                        if orig_width > 0 and orig_height > 0:
                            # Calculate scaling factor to fit within max dimensions
                            width_ratio = max_width / orig_width
                            height_ratio = max_height / orig_height
                            scale_ratio = min(width_ratio, height_ratio)
                            
                            # Set dimensions maintaining aspect ratio
                            final_width = orig_width * scale_ratio
                            final_height = orig_height * scale_ratio
                            sig_img.drawWidth = final_width
                            sig_img.drawHeight = final_height
                            logger.debug(f"✅ Supervisor signature (no PIL): Original={orig_width}x{orig_height}, Final={final_width:.2f}x{final_height:.2f}, Ratio={scale_ratio:.3f}")
                        else:
                            # If dimensions unknown, use max width only and let height adjust automatically
                            sig_img.drawWidth = max_width
                            logger.debug(f"⚠️ Supervisor signature: Unknown dimensions, using max width only")
                    
                    sig_rows.append([
                        Paragraph(f"<b>Supervisor Signature:</b>", styles['Normal']),
                        sig_img
                    ])
                else:
                    sig_rows.append([
                        Paragraph(f"<b>Supervisor Signature:</b>", styles['Normal']),
                        Paragraph("Signature not available", styles['Small'])
                    ])
            except Exception as e:
                logger.error(f"Error processing Supervisor signature: {str(e)}")
                logger.error(traceback.format_exc())
                sig_rows.append([
                    Paragraph(f"<b>Supervisor Signature:</b>", styles['Normal']),
                    Paragraph("Error loading signature", styles['Small'])
                ])
        else:
            # Show placeholder for missing signature
            sig_rows.append([
                Paragraph(f"<b>Supervisor Signature:</b>", styles['Normal']),
                Paragraph("<i>Not signed</i>", styles['Small'])
            ])
        
        if sig_rows:
            sig_table = Table(sig_rows, colWidths=[2*inch, 3.5*inch])
            sig_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.75, colors.HexColor('#125435')),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F5E9')),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(sig_table)
            story.append(Spacer(1, 0.15*inch))
        
        # 2. Operations Manager - ALWAYS show if comments exist (even without signature)
        ops_mgr_comments = operations_manager_comments.strip() if operations_manager_comments and operations_manager_comments.strip() else None
        ops_mgr_sig = signatures.get('Operations Manager')
        logger.info(f"🔍 Operations Manager section check:")
        logger.info(f"  - Comments present: {bool(ops_mgr_comments)}")
        logger.info(f"  - Comments value: {str(ops_mgr_comments)[:100] if ops_mgr_comments else 'None'}")
        logger.info(f"  - Signature present: {bool(ops_mgr_sig)}")
        logger.info(f"  - Signature type: {type(ops_mgr_sig) if ops_mgr_sig else 'None'}")
        if ops_mgr_sig and isinstance(ops_mgr_sig, str):
            logger.info(f"  - Signature length: {len(ops_mgr_sig)}")
            logger.info(f"  - Signature preview: {ops_mgr_sig[:100]}...")
        elif ops_mgr_sig and isinstance(ops_mgr_sig, dict):
            logger.info(f"  - Signature dict keys: {list(ops_mgr_sig.keys())}")
            logger.info(f"  - Signature URL: {ops_mgr_sig.get('url', 'N/A')[:100] if ops_mgr_sig.get('url') else 'N/A'}")
        
        # Check if Operations Manager has approved (even if data is missing, we should show the section)
        om_has_approved = False
        if data.get('operations_manager_approved_at') or data.get('operations_manager_id'):
            om_has_approved = True
            logger.info(f"✅ Operations Manager has approved (operations_manager_approved_at or operations_manager_id present)")
        # Also check workflow status
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'operations_manager_approved' in workflow_status or 'bd_procurement' in workflow_status:
                om_has_approved = True
                logger.info(f"✅ Operations Manager has approved (based on workflow_status: {workflow_status})")
        
        logger.info(f"🔍 Operations Manager section check:")
        logger.info(f"  - Comments present: {bool(ops_mgr_comments)}")
        logger.info(f"  - Comments value: {ops_mgr_comments[:50] if ops_mgr_comments else 'None'}")
        logger.info(f"  - Signature present: {bool(ops_mgr_sig)}")
        logger.info(f"  - Signature type: {type(ops_mgr_sig)}")
        logger.info(f"  - OM has approved (based on workflow/approval fields): {om_has_approved}")
        
        # Always show Operations Manager section if OM has approved OR if we have comments/signature
        # Use always_show_signature=True to ensure signature section appears even if missing
        if ops_mgr_comments or ops_mgr_sig or om_has_approved:
            logger.info(f"✅ Adding Operations Manager section to PDF (comments: {bool(ops_mgr_comments)}, signature: {bool(ops_mgr_sig)}, approved: {om_has_approved})")
            # Always show signature section for Operations Manager (even if missing, show "Not signed")
            add_reviewer_section("Operations Manager", ops_mgr_comments, ops_mgr_sig, always_show_signature=True)
        else:
            logger.warning("⚠️ Skipping Operations Manager section - no comments, signature, or approval found")
            # Log available keys for debugging
            logger.warning(f"  - Available keys in data: {list(data.keys())[:30]}")
            if isinstance(data.get('form_data'), dict):
                logger.warning(f"  - Available keys in form_data: {list(data.get('form_data').keys())[:30]}")
        
        # 3. Business Development
        # Only show BD section if they have actually signed (don't show "Not signed" placeholder)
        # Check if BD has signed (has signature data)
        bd_has_signed = bool(signatures.get('Business Development'))
        bd_has_approved = False
        if data.get('business_dev_approved_at') or data.get('business_dev_id'):
            bd_has_approved = True
        # Also check workflow status
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'bd_procurement' in workflow_status or 'general_manager' in workflow_status:
                bd_has_approved = True
        
        # Only show BD section if BD has actually signed (has signature) OR has comments
        # Don't show "Not signed" placeholder - only show when they've signed
        if bd_has_signed or (business_dev_comments and business_dev_comments.strip()):
            logger.info(f"✅ Adding Business Development section to PDF (comments: {bool(business_dev_comments)}, signature: {bool(bd_has_signed)}, approved: {bd_has_approved})")
            # Use always_show_signature=False so "Not signed" is not shown if signature is missing
            add_reviewer_section("Business Development", business_dev_comments, signatures.get('Business Development'), always_show_signature=False)
        
        # 4. Procurement
        # Only show Procurement section if they have actually signed (don't show "Not signed" placeholder)
        # Check if Procurement has signed (has signature data)
        procurement_has_signed = bool(signatures.get('Procurement'))
        procurement_has_approved = False
        if data.get('procurement_approved_at') or data.get('procurement_id'):
            procurement_has_approved = True
        # Also check workflow status
        if data.get('workflow_status'):
            workflow_status = str(data.get('workflow_status'))
            if 'bd_procurement' in workflow_status or 'general_manager' in workflow_status:
                procurement_has_approved = True
        
        # Only show Procurement section if Procurement has actually signed (has signature) OR has comments
        # Don't show "Not signed" placeholder - only show when they've signed
        if procurement_has_signed or (procurement_comments and procurement_comments.strip()):
            logger.info(f"✅ Adding Procurement section to PDF (comments: {bool(procurement_comments)}, signature: {bool(procurement_has_signed)}, approved: {procurement_has_approved})")
            # Use always_show_signature=False so "Not signed" is not shown if signature is missing
            add_reviewer_section("Procurement", procurement_comments, signatures.get('Procurement'), always_show_signature=False)
        
        # 5. General Manager
        if general_manager_comments or signatures.get('General Manager'):
            add_reviewer_section("General Manager", general_manager_comments, signatures.get('General Manager'))
        
        # Add document signed timestamp if any signatures exist
        if any(signatures.values()):
            styles = get_professional_styles()
            story.append(Spacer(1, 0.15*inch))
            story.append(Paragraph(
                f"<i>Document signed on: {format_dubai_datetime(format_str='%B %d, %Y at %H:%M')} (GST)</i>",
                styles['Small']
            ))
        
        # Build professional PDF with logo and branding
        create_professional_pdf(
            pdf_path, 
            story, 
            report_title=f"HVAC/MEP Inspection - {data.get('site_name', 'N/A')}"
        )
        
        logger.info(f"✅ Professional HVAC/MEP PDF created successfully: {pdf_path}")
        return pdf_path
        
    except Exception as e:
        logger.error(f"❌ PDF generation error: {str(e)}")
        raise