"""
Shared Kynvera branding for openpyxl import templates.

Colors and naming come from common.kynvera_brand (docs/KYNVERA_DESIGN.md).
Calibri is used because Excel cannot embed the in-app SF Pro stack.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from io import BytesIO, StringIO
from typing import Any, Iterable, Optional, Sequence

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common.kynvera_brand import (
    BRAND,
    BRAND_DARK,
    BRAND_WASH,
    CHROME,
    COMPANY_NAME,
    FOOTER_CONFIDENTIAL,
    TAGLINE,
    WHITE,
    hex_rgb,
)

logger = logging.getLogger(__name__)

FONT_NAME = "Calibri"

# Hex without '#' — openpyxl PatternFill fgColor
HEADER_FILL_HEX = hex_rgb(BRAND)
PRIMARY_DARK_HEX = hex_rgb(BRAND_DARK)
SOFT_WASH_HEX = hex_rgb(BRAND_WASH)
TEXT_DARK_HEX = hex_rgb(CHROME["text"])
TEXT_MID_HEX = hex_rgb(CHROME["text_mid"])
TEXT_MUTED_HEX = hex_rgb(CHROME["text_muted"])
HAIRLINE_HEX = hex_rgb(CHROME["hairline"])
SURFACE_ALT_HEX = hex_rgb(CHROME["well"])
WHITE_HEX = hex_rgb(WHITE)
# Font rgb must be 8-char AARRGGBB. "FFFFFF" is stored as 00FFFFFF (invisible on Excel for Mac).
WHITE_FONT_HEX = "FFFFFFFF"

HEADER_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)
SOFT_FILL = PatternFill("solid", fgColor=SOFT_WASH_HEX)
SURFACE_FILL = PatternFill("solid", fgColor=SURFACE_ALT_HEX)
EXAMPLE_FILL = SOFT_FILL

HEADER_FONT = Font(name=FONT_NAME, bold=True, color=WHITE_FONT_HEX, size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color=TEXT_DARK_HEX)
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color=WHITE_FONT_HEX)
BODY_FONT = Font(name=FONT_NAME, size=11, color=TEXT_DARK_HEX)
HINT_FONT = Font(name=FONT_NAME, size=10, color=TEXT_MID_HEX)
MUTED_FONT = Font(name=FONT_NAME, size=10, color=TEXT_MUTED_HEX)
BANNER_FONT = Font(name=FONT_NAME, bold=True, size=18, color=WHITE_FONT_HEX)
BANNER_TAG_FONT = Font(name=FONT_NAME, size=11, color=WHITE_FONT_HEX, bold=False)
COL_NAME_FONT = Font(name=FONT_NAME, bold=True, size=11, color=TEXT_DARK_HEX)
FOOTER_FONT = Font(name=FONT_NAME, italic=True, size=9, color=TEXT_MUTED_HEX)
EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, size=10, color=TEXT_MID_HEX)

THIN = Border(
    left=Side(style="thin", color=HAIRLINE_HEX),
    right=Side(style="thin", color=HAIRLINE_HEX),
    top=Side(style="thin", color=HAIRLINE_HEX),
    bottom=Side(style="thin", color=HAIRLINE_HEX),
)

WRAP = Alignment(wrap_text=True, vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)

INSTRUCTIONS_SHEET_NAMES = frozenset({"instructions", "instruction", "readme", "how to use"})

DEFAULT_DO_NOT = (
    "Do not rename or delete the coral header row on the data sheet(s).",
    "Do not rename the data sheet — importers look it up by name.",
    "Do not upload the Instructions sheet as data; it is ignored on import.",
    "Replace or delete example rows before importing if you do not want them created.",
    "Keep the file as .xlsx (Excel Workbook). CSV is only accepted where the importer says so.",
)

def paint_tab(ws: Worksheet) -> None:
    ws.sheet_properties.tabColor = HEADER_FILL_HEX


def style_header_row(ws: Worksheet, row: int, ncols: int, *, freeze: bool = True) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = THIN
    ws.row_dimensions[row].height = 28
    if freeze:
        ws.freeze_panes = f"A{row + 1}"
    paint_tab(ws)


def write_header_row(ws: Worksheet, headers: Sequence[str], row: int = 1, *, freeze: bool = True) -> None:
    for col, header in enumerate(headers, start=1):
        ws.cell(row, col, header)
    style_header_row(ws, row, len(headers), freeze=freeze)


def style_data_cell(cell, *, example: bool = False) -> None:
    cell.border = THIN
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.font = EXAMPLE_FONT if example else BODY_FONT
    if example:
        cell.fill = EXAMPLE_FILL


def style_example_row(ws: Worksheet, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        style_data_cell(ws.cell(row, col), example=True)


def write_data_row(ws: Worksheet, row: int, values: Sequence[Any], *, example: bool = False) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row, col, val)
        style_data_cell(cell, example=example)


def apply_column_widths(ws: Worksheet, widths: Sequence[float] | dict[str, float]) -> None:
    if isinstance(widths, dict):
        for letter, width in widths.items():
            ws.column_dimensions[letter].width = width
        return
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def autosize_columns(ws: Worksheet, ncols: int, min_w: float = 12, max_w: float = 36) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = min_w
        for cell in ws[letter]:
            val = cell.value
            if val is None:
                continue
            width = max(width, min(max_w, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


@dataclass
class InstructionSpec:
    title: str
    module_label: str
    about: Sequence[str]
    how_to: Sequence[str]
    columns: Sequence[tuple[str, str]]
    example_headers: Sequence[str] = field(default_factory=tuple)
    example_rows: Sequence[Sequence[Any]] = field(default_factory=tuple)
    import_rules: Sequence[str] = field(default_factory=tuple)
    do_not: Sequence[str] = field(default_factory=lambda: DEFAULT_DO_NOT)
    extra_sections: Sequence[tuple[str, Sequence[str]]] = field(default_factory=tuple)


def _merge(ws: Worksheet, row: int, start: int, end: int) -> None:
    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)


def _paint_banner_row(ws: Worksheet, row: int, value: str, font: Font, span: int, height: float) -> None:
    """Fill + font first, then merge. Excel for Mac keeps the pre-merge font of A1."""
    for col in range(1, span + 1):
        cell = ws.cell(row, col, value if col == 1 else None)
        cell.fill = HEADER_FILL
        cell.font = font
        cell.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        cell.border = Border()
    if span > 1:
        _merge(ws, row, 1, span)
    top = ws.cell(row, 1)
    top.value = value
    top.font = font
    top.fill = HEADER_FILL
    top.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = height


def _section_heading(ws: Worksheet, row: int, title: str, span: int = 2) -> int:
    cell = ws.cell(row, 1, title)
    cell.font = SECTION_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(vertical="center", indent=1)
    for col in range(1, span + 1):
        extra = ws.cell(row, col)
        extra.fill = HEADER_FILL
        extra.font = SECTION_FONT
        extra.border = THIN
    if span > 1:
        _merge(ws, row, 1, span)
    ws.cell(row, 1).font = SECTION_FONT
    ws.cell(row, 1).fill = HEADER_FILL
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_paragraphs(ws: Worksheet, row: int, lines: Sequence[str], *, numbered: bool = False, span: int = 2) -> int:
    for i, line in enumerate(lines, start=1):
        text = f"{i}. {line}" if numbered else line
        cell = ws.cell(row, 1, text)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        if span > 1:
            _merge(ws, row, 1, span)
        ws.row_dimensions[row].height = max(18, 16 + 12 * (len(text) // 90))
        row += 1
    return row


def write_instructions_sheet(wb, spec: InstructionSpec, *, index: int = 0) -> Worksheet:
    """Create or move the Instructions sheet so it is always first."""
    if "Instructions" in wb.sheetnames:
        ws = wb["Instructions"]
        current = wb.sheetnames.index("Instructions")
        if current != index:
            wb.move_sheet(ws, offset=index - current)
    else:
        ws = wb.create_sheet("Instructions", index)

    span = 2
    example_cols = max(len(spec.example_headers), 2)
    banner_span = max(span, min(example_cols, 8), 4)

    ws.sheet_view.showGridLines = False
    paint_tab(ws)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    # Drop any leftover drawings if this sheet was reused.
    ws._images = []

    _paint_banner_row(ws, 1, COMPANY_NAME, BANNER_FONT, banner_span, 32)
    _paint_banner_row(ws, 2, TAGLINE, BANNER_TAG_FONT, banner_span, 20)

    row = 4
    title_cell = ws.cell(row, 1, spec.title)
    title_cell.font = TITLE_FONT
    _merge(ws, row, 1, banner_span)
    ws.row_dimensions[row].height = 24
    row += 1
    mod = ws.cell(row, 1, f"{spec.module_label}  ·  Import template")
    mod.font = HINT_FONT
    _merge(ws, row, 1, banner_span)
    row += 2

    row = _section_heading(ws, row, "About this template", span=banner_span)
    row = _write_paragraphs(ws, row, spec.about, span=banner_span)
    row += 1

    row = _section_heading(ws, row, "How to fill the data", span=banner_span)
    row = _write_paragraphs(ws, row, spec.how_to, numbered=True, span=banner_span)
    row += 1

    row = _section_heading(ws, row, "Column guide", span=2)
    ws.cell(row, 1, "Column").font = HEADER_FONT
    ws.cell(row, 2, "Required / format / allowed values").font = HEADER_FONT
    ws.cell(row, 1).fill = HEADER_FILL
    ws.cell(row, 2).fill = HEADER_FILL
    ws.cell(row, 1).alignment = CENTER_WRAP
    ws.cell(row, 2).alignment = CENTER_WRAP
    ws.cell(row, 1).border = THIN
    ws.cell(row, 2).border = THIN
    ws.row_dimensions[row].height = 22
    row += 1
    for i, (name, desc) in enumerate(spec.columns):
        name_cell = ws.cell(row, 1, name)
        desc_cell = ws.cell(row, 2, desc)
        name_cell.font = COL_NAME_FONT
        desc_cell.font = BODY_FONT
        desc_cell.alignment = WRAP
        name_cell.alignment = Alignment(vertical="center")
        name_cell.border = THIN
        desc_cell.border = THIN
        if i % 2:
            name_cell.fill = SURFACE_FILL
            desc_cell.fill = SURFACE_FILL
        ws.row_dimensions[row].height = max(20, 16 + 10 * (len(desc) // 70))
        row += 1
    row += 1

    if spec.example_headers:
        row = _section_heading(ws, row, "Worked example", span=max(span, len(spec.example_headers)))
        hint = ws.cell(row, 1, "Copy this shape onto the data sheet. Replace or delete example rows before import.")
        hint.font = HINT_FONT
        _merge(ws, row, 1, max(span, len(spec.example_headers)))
        row += 1
        write_header_row(ws, spec.example_headers, row=row, freeze=False)
        row += 1
        for values in spec.example_rows:
            write_data_row(ws, row, values, example=True)
            row += 1
        row += 1

    if spec.import_rules:
        row = _section_heading(ws, row, "Import rules", span=banner_span)
        row = _write_paragraphs(ws, row, spec.import_rules, span=banner_span)
        row += 1

    for heading, lines in spec.extra_sections:
        row = _section_heading(ws, row, heading, span=banner_span)
        row = _write_paragraphs(ws, row, lines, span=banner_span)
        row += 1

    if spec.do_not:
        row = _section_heading(ws, row, "Do not", span=banner_span)
        row = _write_paragraphs(ws, row, spec.do_not, span=banner_span)
        row += 1

    foot = ws.cell(row + 1, 1, FOOTER_CONFIDENTIAL)
    foot.font = FOOTER_FONT

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72
    for col in range(3, max(banner_span, example_cols) + 1):
        letter = get_column_letter(col)
        if ws.column_dimensions[letter].width is None or ws.column_dimensions[letter].width < 14:
            ws.column_dimensions[letter].width = 16

    return ws


def resolve_import_sheet_name(
    sheetnames: Sequence[str],
    preferred: Iterable[str] = (),
) -> str | None:
    """Pick a data sheet, skipping Instructions. Prefer named sheets when present."""
    names = [n for n in sheetnames if n]
    if not names:
        return None
    preferred_lower = [p.strip().lower() for p in preferred if p]
    by_lower = {n.strip().lower(): n for n in names}
    for key in preferred_lower:
        if key in by_lower:
            return by_lower[key]
    for name in names:
        if name.strip().lower() not in INSTRUCTIONS_SHEET_NAMES:
            return name
    return names[0]


def _seek(file_storage) -> None:
    for attr in ("seek",):
        if hasattr(file_storage, attr):
            try:
                file_storage.seek(0)
            except Exception:
                pass
    stream = getattr(file_storage, "stream", None)
    if stream is not None and hasattr(stream, "seek"):
        try:
            stream.seek(0)
        except Exception:
            pass


def _file_bytes(file_storage) -> bytes:
    if isinstance(file_storage, (bytes, bytearray)):
        return bytes(file_storage)
    if hasattr(file_storage, "read"):
        _seek(file_storage)
        data = file_storage.read()
        _seek(file_storage)
        if isinstance(data, str):
            return data.encode("utf-8")
        return data or b""
    raise TypeError("Expected a file-like object or bytes")


def read_import_dataframe(file_storage, preferred_sheets: Sequence[str] = ()):
    """Read the data sheet of an import workbook, skipping Instructions.

    Accepts .xlsx / .xls / .csv (and HTML-disguised .xls). Preferred sheet
    names are tried first; otherwise the first non-Instructions sheet is used.
    """
    import pandas as pd

    filename = (getattr(file_storage, "filename", None) or "").lower()
    if filename.endswith(".csv"):
        _seek(file_storage)
        return pd.read_csv(file_storage)

    data = _file_bytes(file_storage)

    def _from_excel(engine: Optional[str] = None):
        kwargs = {"engine": engine} if engine else {}
        xl = pd.ExcelFile(BytesIO(data), **kwargs)
        name = resolve_import_sheet_name(list(xl.sheet_names), preferred_sheets)
        return pd.read_excel(xl, sheet_name=name)

    last_err: Optional[Exception] = None

    if filename.endswith(".xlsx") or not filename.endswith(".xls"):
        try:
            return _from_excel("openpyxl")
        except ImportError:
            raise
        except Exception as e:
            last_err = e
            logger.warning("openpyxl could not read workbook: %s", e)
            if filename.endswith(".xlsx"):
                raise ValueError(f"Could not read Excel: {e}") from e

    try:
        return _from_excel("xlrd")
    except Exception as e:
        last_err = last_err or e

    try:
        html_text = data.decode("utf-8", errors="ignore")
        tables = pd.read_html(StringIO(html_text))
        if not tables:
            raise ValueError("Could not read any table from the file")
        return max(tables, key=lambda t: t.shape[0])
    except Exception as e:
        if last_err:
            raise ValueError(f"Could not read Excel: {last_err}") from last_err
        raise ValueError(f"Could not read Excel: {e}") from e
