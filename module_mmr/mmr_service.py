"""
MMR Service – Excel parsing and professional report generation.
Handles: Reactive Workorder Details sheet with columns:
  WorkOrder No, Reported Date, Priority, Work Description, Client,
  Reported By, Service Group, Assigned Date, Work Start Date, Closed By,
  Status, Contract, BaseUnit, Space
"""
import os
import logging
import zipfile
from io import BytesIO, StringIO
from datetime import datetime

import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import pixels_to_EMU

logger = logging.getLogger(__name__)

# Built-in rules (admin can disable each; defaults match original CAFM logic).
DEFAULT_BUILTIN_RULES = {
    'facade_cleaning_non_chargeable': True,
    'elevator_non_chargeable': True,
    'roof_top_non_chargeable': True,
    'garden_city_ac_non_chargeable': True,
    'apt_number_chargeable': True,
    'residential_hvac_non_chargeable': True,
    'cafm_reception_outside_exit_non_chargeable': True,
    'floor_word_non_chargeable': True,
}

# Defaults for MMR chargeable admin settings (see admin UI + MmrChargeableConfig model)
DEFAULT_MMR_CHARGEABLE_CONFIG = {
    # When True (recommended for CAFM exports): BaseUnit text that is not "Apt No + number"
    # and not reception/floor/exit-style labels resolves to Non-Chargeable.
    # When False: legacy behaviour — those locations resolve to Chargeable.
    'non_apartment_baseunit_non_chargeable': True,
    # Substring overrides (longest match wins). Applied after core rules except where
    # facade/elevator/roof/Garden AC or apartment numbers lock the outcome.
    'baseunit_overrides': [],  # [{ "pattern": "lobby", "chargeable": true }, ...]
    'builtin_rules': dict(DEFAULT_BUILTIN_RULES),
    # Last saved Location Register (Excel/HTML): rows with chargeable flags for UI + overrides merge
    'location_register_state': None,
}

_mmr_chargeable_config_cache: dict | None = None

EXPECTED_COLS = [
    'WorkOrder No', 'Reported Date', 'Priority', 'Work Description',
    'Client', 'Reported By', 'Service Group', 'Assigned Date',
    'Work Start Date', 'Closed By', 'Status', 'Contract', 'BaseUnit', 'Space'
]

PRIMARY   = "125435"
SECONDARY = "2E7D32"
ACCENT    = "E8F5E9"
WHITE     = "FFFFFF"
ALT_ROW   = "F9FAFB"
BORDER    = "D1D5DB"

LOGO_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'static', 'logo.png'
)

# ──────────────────────────────────────────────────────────────────────────────
# Report date extraction (from parmToDate or Reported Date)
# ──────────────────────────────────────────────────────────────────────────────

def get_report_date_from_excel(file_path: str):
    """Extract report date from CAFM Excel. Tries parmToDate first, else max(Reported Date).
    Returns datetime or None."""
    try:
        df_raw = pd.read_excel(file_path, sheet_name='Reactive Workorder Details', header=None)
        # parmToDate is in col 0, value in next row
        for i in range(len(df_raw) - 1):
            val = df_raw.iloc[i, 0]
            if pd.notna(val) and str(val).strip().lower() == 'parmtodate':
                next_val = df_raw.iloc[i + 1, 0]
                if pd.notna(next_val):
                    dt = pd.to_datetime(next_val, errors='coerce')
                    if pd.notna(dt):
                        return dt.to_pydatetime()
                break
        # Fallback: max Reported Date from parsed data
        df = pd.read_excel(file_path, sheet_name='Reactive Workorder Details')
        if 'Reported Date' in df.columns:
            df['Reported Date'] = pd.to_datetime(df['Reported Date'], errors='coerce')
            max_d = df['Reported Date'].max()
            if pd.notna(max_d):
                return max_d.to_pydatetime()
    except Exception:
        pass
    return None


def get_report_date_from_df(df: pd.DataFrame):
    """Get report date from DataFrame (max of Reported Date). Returns datetime or None."""
    r = get_report_date_range_from_df(df)
    return r[1] if r else None  # max date


def get_report_date_range_from_df(df: pd.DataFrame, exclude_today: bool = True) -> tuple | None:
    """Get (min, max, n_unique) Reported Date from DataFrame. Excludes today (save date) by default.
    Returns (min_dt, max_dt, count_of_unique_dates) or None."""
    if df is None or df.empty or 'Reported Date' not in df.columns:
        return None
    try:
        from datetime import date
        work = df.copy()
        work['Reported Date'] = pd.to_datetime(work['Reported Date'], errors='coerce')
        work = work.dropna(subset=['Reported Date'])
        if work.empty:
            return None
        if exclude_today:
            today = date.today()
            work = work[work['Reported Date'].dt.date < today]
            if work.empty:
                return None
        min_d = work['Reported Date'].min()
        max_d = work['Reported Date'].max()
        n_unique = work['Reported Date'].dt.date.nunique()
        if pd.notna(min_d) and pd.notna(max_d):
            return (min_d.to_pydatetime(), max_d.to_pydatetime(), int(n_unique))
    except Exception:
        pass
    return None


def split_df_by_month(df: pd.DataFrame) -> list[tuple[str, pd.DataFrame]]:
    """Split a DataFrame into (month_label, sub_df) chunks by Reported Date year-month.

    Returns a list sorted chronologically, e.g.:
      [('January 2026', <df>), ('February 2026', <df>), ...]
    Rows with no parseable Reported Date are grouped under 'Unknown'.
    """
    if df is None or df.empty:
        return []

    work = df.copy()
    work['Reported Date'] = pd.to_datetime(work['Reported Date'], errors='coerce')

    known = work.dropna(subset=['Reported Date']).copy()
    unknown = work[work['Reported Date'].isna()].copy()

    chunks: list[tuple[str, pd.DataFrame]] = []
    if not known.empty:
        known['_ym'] = known['Reported Date'].dt.to_period('M')
        for period, grp in known.groupby('_ym', sort=True):
            label = period.strftime('%B %Y')
            grp = grp.drop(columns=['_ym'])
            chunks.append((label, grp.reset_index(drop=True)))
    if not unknown.empty:
        chunks.append(('Unknown', unknown.reset_index(drop=True)))
    return chunks


def generate_monthly_zip(df: pd.DataFrame, report_format: str = 'daily') -> tuple[bytes, list[str]]:
    """Generate one Excel report per calendar month and return a ZIP archive.

    report_format: 'daily' | 'monthly' — prefix for each workbook inside the ZIP.

    Returns:
        (zip_bytes, list_of_filenames_inside_zip)
    """
    months = split_df_by_month(df)
    if not months:
        raise ValueError('No data to generate monthly reports from.')

    kind = (report_format or 'daily').strip().lower()
    prefix = 'Monthly Report' if kind == 'monthly' else 'Daily Report'

    buf = BytesIO()
    filenames: list[str] = []

    with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        for label, month_df in months:
            excel_bytes = generate_report_excel(month_df)
            fname = f'{prefix} – {label}.xlsx'
            zf.writestr(fname, excel_bytes)
            filenames.append(fname)

    buf.seek(0)
    return buf.getvalue(), filenames


# ──────────────────────────────────────────────────────────────────────────────
# Parsing
# ──────────────────────────────────────────────────────────────────────────────

def parse_excel(file_path: str) -> pd.DataFrame:
    """Read the 'Reactive Workorder Details' sheet and normalise values."""
    try:
        df = pd.read_excel(file_path, sheet_name='Reactive Workorder Details')
    except Exception:
        # Fall back to active sheet if named sheet not found
        df = pd.read_excel(file_path)

    return _normalise_parsed_df(df)


def parse_saved_report_excel(file_path: str) -> pd.DataFrame:
    """Parse a saved report Excel (from report folder). Reads 'All Work Orders' sheet.
    Header is at row 5 (1-based), data starts at row 6. Pandas header=4 = 0-based row 4 = 5th row."""
    try:
        df = pd.read_excel(file_path, sheet_name='All Work Orders', header=4)
    except Exception:
        return pd.DataFrame()
    return _normalise_parsed_df(df)


def _clean_cafm_text(val) -> str:
    """Strip Excel/CSV artefacts (e.g. _x000d_) and collapse whitespace."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ''
    t = str(val).strip()
    if t.lower() == 'nan':
        return ''
    t = t.replace('_x000d_', ' ').replace('_x000D_', ' ').replace('\r', ' ')
    return ' '.join(t.split())


def _normalise_parsed_df(df: pd.DataFrame) -> pd.DataFrame:
    """Common normalisation for parsed Excel data."""
    df.columns = [str(c).strip() for c in df.columns]

    # Drop completely empty rows (copy avoids SettingWithCopyWarning on chained assignment)
    df = df.dropna(how='all').copy()

    # Normalise string columns (BaseUnit cleaned again below for CAFM artefacts)
    for col in ['Client', 'Service Group', 'Status', 'Contract',
                'Space', 'Priority', 'Reported By', 'Closed By']:
        if col in df.columns:
            df[col] = df[col].fillna('').astype(str).str.strip()

    for col in ['Work Description', 'Specific Area', 'BaseUnit']:
        if col in df.columns:
            df[col] = df[col].apply(_clean_cafm_text)

    # Normalise work order number – keep rows with valid WO formats:
    # - Pure numeric: 10017690
    # - Manual prefix: Manual - 10017703
    # (trailing metadata rows like parmFromDate / parmToDate are discarded)
    if 'WorkOrder No' in df.columns:
        df = df[df['WorkOrder No'].notna()]
        df['WorkOrder No'] = df['WorkOrder No'].astype(str).str.strip()
        # Match digits-only OR "Manual - " + digits
        wo_valid = df['WorkOrder No'].str.match(r'^(?:\d+|Manual\s*-\s*\d+)$', case=False, na=False)
        df = df[wo_valid]

    # Parse dates
    for col in ['Reported Date', 'Assigned Date', 'Work Start Date']:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')

    return df.reset_index(drop=True)


def df_to_rows(df: pd.DataFrame) -> list:
    """Convert DataFrame to JSON-serialisable list of dicts."""
    rows = []
    for _, row in df.iterrows():
        r = {}
        for col in df.columns:
            val = row[col]
            if pd.isna(val):
                r[col] = ''
            elif isinstance(val, pd.Timestamp):
                r[col] = val.strftime('%Y-%m-%d')
            else:
                r[col] = str(val).strip()
        rows.append(r)
    return rows


def rows_to_df(rows: list) -> pd.DataFrame:
    """Convert list of dicts (from frontend) back to DataFrame. Handles date strings."""
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    date_cols = ['Reported Date', 'Assigned Date', 'Work Start Date']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df


# ──────────────────────────────────────────────────────────────────────────────
# Dashboard aggregation
# ──────────────────────────────────────────────────────────────────────────────

def compute_dashboard(df: pd.DataFrame) -> dict:
    """Return all aggregated stats needed by the front-end dashboard."""

    def count_by(col: str) -> dict:
        if col not in df.columns:
            return {}
        series = df[col].replace('', None).dropna()
        return {str(k): int(v) for k, v in series.value_counts().items()}

    def uniq(col: str) -> list:
        if col not in df.columns:
            return []
        vals = df[col].replace('', None).dropna().unique().tolist()
        return sorted([str(v) for v in vals])

    # Chargeable/Non-Chargeable: use resolved logic (Space + BaseUnit when Space empty)
    if 'Space' in df.columns and 'BaseUnit' in df.columns:
        resolved = _get_resolved_chargeable_series(df)
        by_space = {str(k): int(v) for k, v in resolved.value_counts().items()}
        spaces_filter = sorted(resolved.unique().astype(str).tolist())
    else:
        by_space = {str(k): int(v) for k, v in
                    (df['Space'].apply(_normalise_space).value_counts().items())} if 'Space' in df.columns else {}
        spaces_filter = uniq('Space')

    return {
        'total': len(df),
        'by_client':        count_by('Client'),
        'by_contract':      count_by('Contract'),
        'by_service_group': count_by('Service Group'),
        'by_space':         by_space,
        'by_status':        count_by('Status'),
        'by_priority':      count_by('Priority'),
        'filters': {
            'clients':        uniq('Client'),
            'contracts':      uniq('Contract'),
            'service_groups': uniq('Service Group'),
            'spaces':         spaces_filter,
            'statuses':       uniq('Status'),
            'priorities':     uniq('Priority'),
        }
    }


# ──────────────────────────────────────────────────────────────────────────────
# Excel report generation
# ──────────────────────────────────────────────────────────────────────────────

def _normalise_space(val: str) -> str:
    """Map raw Space values (including typos) to clean labels."""
    v = val.strip().lower()
    if v in ('chargebale', 'chargeable'):
        return 'Chargeable'
    if v in ('non-chargebale', 'non-chargeable', 'non chargebale', 'non chargeable'):
        return 'Non-Chargeable'
    return val.strip() or 'Unknown'


# AC/HVAC in Service Group / complaint text (word-boundary "ac" avoids matching "space", etc.)
_AC_HVAC_TEXT_RE = re.compile(
    r'(?:\b(?:hvac|a/c|a-c|chiller|vrf|cooling|ventilation|duct(?:ing)?|thermostat|split\s*ac|'
    r'ac\s+system|mechanical\s+ac|air\s*condition(?:ing|er)?|airconditioning)\b|\bac\b)',
    re.IGNORECASE,
)

# Clients/contracts where rows with empty BaseUnit default to Chargeable (office / internal sites).
_ALL_BASEUNITS_CHARGEABLE_CLIENTS = ('askaan', 'ajman holding', 'injaaz')

# CAFM often uses "Elevator system" or the typo "Elevater system". Use elevat(or|er)
# so we do not match unrelated words like "elevation".
_ELEVATOR_SERVICE_GROUP_RE = re.compile(r'elevat(or|er)', re.IGNORECASE)

# Roof / rooftop mentions (incl. "roof toop", "rooftop", "roof top") → Non-Chargeable
_ROOF_TOP_RE = re.compile(r'roof\s*to+p', re.IGNORECASE)

# BaseUnit like "Apt No 911", "AptNo 12" → always Chargeable (checked before reception/floor rules)
_APT_NO_WITH_NUMBER_RE = re.compile(r'apt\s*no\s*\d+', re.IGNORECASE)


def _extra_project_blob_for_chargeable(df: pd.DataFrame, i: int) -> str:
    """Concatenate optional CAFM location columns so Garden City can be detected when not on Client/Contract."""
    parts: list[str] = []
    for col in ('Property', 'Site', 'Project', 'Location', 'Building', 'Tower'):
        if col not in df.columns:
            continue
        v = df[col].iat[i]
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return ' '.join(parts)


def _complaint_text_for_chargeable(df: pd.DataFrame, i: int) -> str:
    for col in ('Complaint', 'Complaint Description', 'Complaint Details'):
        if col not in df.columns:
            continue
        v = df[col].iat[i]
        if pd.isna(v):
            continue
        s = str(v).strip()
        if s:
            return s
    return ''


def _project_blob_from_row_dict(r: dict) -> str:
    """Same fields as _extra_project_blob_for_chargeable for dashboard row dicts."""
    parts: list[str] = []
    if not isinstance(r, dict):
        return ''
    for k in (
        'property_display',
        'Property display',
        'property',
        'Property',
        'Site',
        'site',
        'Project',
        'project',
        'Location',
        'location',
        'Building',
        'building',
        'Tower',
        'tower',
    ):
        v = r.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s:
            parts.append(s)
    return ' '.join(parts)


def _complaint_from_row_dict(r: dict) -> str:
    for k in ('Complaint', 'Complaint Description', 'Complaint Details'):
        v = r.get(k) if isinstance(r, dict) else None
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ''


# CAFM property code for **Garden City** estate (PY0002). PY0003 and other codes are not rolled up here.
# Extend this set only when another code should alias to Garden City.
_GARDEN_CITY_PROPERTY_CODES = frozenset({'py0002'})


def _alnum_lower(s: str) -> str:
    return re.sub(r'[^a-z0-9]', '', (s or '').lower())


def _blob_contains_garden_city_property_code(blob: str) -> bool:
    compact = _alnum_lower(blob)
    if not compact:
        return False
    for code in _GARDEN_CITY_PROPERTY_CODES:
        if _alnum_lower(code) in compact:
            return True
    return False


def _text_indicates_garden_city_project(
    client_val: str,
    contract_val: str,
    property_val: str = '',
    site_val: str = '',
) -> bool:
    """
    True when the row is for the Garden City estate (CAFM may put this in Property/Site,
    not only Client/Contract). Also matches known Garden City property codes (e.g. PY0003).
    """
    blob = ' '.join(
        p for p in (client_val, contract_val, property_val, site_val) if p and str(p).strip()
    ).strip().lower()
    if not blob:
        return False
    if 'garden' in blob and 'city' in blob:
        return True
    return _blob_contains_garden_city_property_code(blob)


def _text_indicates_ac_hvac_complaint(
    service_group_val: str,
    work_description_val: str,
    specific_area_val: str,
    complaint_val: str = '',
) -> bool:
    """Service Group or complaint/description text indicates an AC / HVAC-type job."""
    sg = (service_group_val or '').strip()
    if sg and _AC_HVAC_TEXT_RE.search(sg):
        return True
    blob = ' '.join(
        p
        for p in (work_description_val, specific_area_val, complaint_val)
        if p and str(p).strip()
    ).strip()
    if not blob:
        return False
    return bool(_AC_HVAC_TEXT_RE.search(blob))


def _text_indicates_roof_top(*parts: str) -> bool:
    """True when any combined text mentions roof top / rooftop (common CAFM typos included)."""
    blob = ' '.join(p for p in parts if p and str(p).strip())
    if not blob:
        return False
    if _ROOF_TOP_RE.search(blob):
        return True
    return False


def _baseunit_is_non_chargeable_cafm_labels(bu_lower: str) -> bool:
    """True when BaseUnit matches CAFM labels that bill as Non-Chargeable (reception, outside, exit/entry)."""
    if not bu_lower:
        return False
    if 'reception' in bu_lower:
        return True
    if 'outside' in bu_lower or 'out side' in bu_lower:
        return True
    if 'exit' in bu_lower and 'entry' in bu_lower:
        return True
    if 'exit/' in bu_lower or 'exit /' in bu_lower:
        return True
    return False


def merge_builtin_rules_payload(raw: dict | None) -> dict:
    """Merge admin JSON into DEFAULT_BUILTIN_RULES (unknown keys ignored)."""
    br = dict(DEFAULT_BUILTIN_RULES)
    if isinstance(raw, dict):
        for k in DEFAULT_BUILTIN_RULES:
            if k in raw:
                br[k] = bool(raw[k])
    return br


def _sanitize_location_register_state(lrs: dict | None) -> dict | None:
    """Trim and cap location_register_state for JSON storage."""
    if lrs is None:
        return None
    if not isinstance(lrs, dict):
        return None
    rows_in = lrs.get('rows') or []
    if not isinstance(rows_in, list):
        rows_in = []
    clean: list[dict] = []
    for r in rows_in[:50000]:
        if not isinstance(r, dict):
            continue
        bu = (r.get('base_unit') or '').strip()
        if not bu:
            continue
        rec: dict = {
            'base_unit': bu[:500],
            'funct_type': (str(r.get('funct_type') or ''))[:200],
            'property': (str(r.get('property') or ''))[:300],
            'zone': (str(r.get('zone') or ''))[:300],
            'is_apt': bool(r.get('is_apt')),
            'chargeable': bool(r.get('chargeable')),
        }
        pd = (str(r.get('property_display') or '')).strip()
        if pd:
            rec['property_display'] = pd[:300]
        sz = (str(r.get('sub_zone') or ''))[:300].strip()
        if sz:
            rec['sub_zone'] = sz
        for k, lim in (
            ('city', 120),
            ('area', 120),
            ('area_group', 120),
            ('sub_zone_code', 120),
            ('base_unit_code', 120),
            ('funct_sub_type', 200),
        ):
            v = r.get(k)
            if v is not None and str(v).strip():
                rec[k] = (str(v))[:lim]
        for k in ('permit_required', 'tenantable'):
            if k in r and r.get(k) is not None:
                rec[k] = r.get(k)
        kid = (str(r.get('key_id') or '')).strip()
        if kid:
            rec['key_id'] = kid[:120]
        clean.append(rec)
    cd = lrs.get('columns_detected')
    if not isinstance(cd, dict):
        cd = {}
    return {
        'source_filename': (str(lrs.get('source_filename') or ''))[:255],
        'columns_detected': {str(k)[:80]: str(v)[:200] for k, v in cd.items()},
        'rows': clean,
    }


def _merge_mmr_chargeable_config(raw: dict | None) -> dict:
    out = {
        'non_apartment_baseunit_non_chargeable': DEFAULT_MMR_CHARGEABLE_CONFIG[
            'non_apartment_baseunit_non_chargeable'
        ],
        'baseunit_overrides': [],
        'builtin_rules': dict(DEFAULT_BUILTIN_RULES),
        'location_register_state': None,
    }
    if isinstance(raw, dict):
        if 'non_apartment_baseunit_non_chargeable' in raw:
            out['non_apartment_baseunit_non_chargeable'] = bool(
                raw['non_apartment_baseunit_non_chargeable']
            )
        if raw.get('baseunit_overrides') is not None:
            out['baseunit_overrides'] = list(raw['baseunit_overrides'])
        out['builtin_rules'] = merge_builtin_rules_payload(raw.get('builtin_rules'))
        if 'location_register_state' in raw:
            out['location_register_state'] = _sanitize_location_register_state(
                raw.get('location_register_state')
            )
    return out


def _br(cfg: dict, key: str) -> bool:
    """True if built-in rule `key` is enabled (default: DEFAULT_BUILTIN_RULES)."""
    br = cfg.get('builtin_rules') or {}
    if key in br:
        return bool(br[key])
    return bool(DEFAULT_BUILTIN_RULES.get(key, True))


def _norm_header(val) -> str:
    return re.sub(r'\s+', ' ', str(val).strip().lower())


def _find_register_column(df: pd.DataFrame, required_phrases: list[str]) -> str | None:
    """Pick column whose header matches the best-scoring phrase (substring match)."""
    best_col = None
    best = 0
    for col in df.columns:
        nh = _norm_header(col)
        for phrase in required_phrases:
            if phrase in nh or nh == phrase:
                sc = len(phrase)
                if sc > best:
                    best = sc
                    best_col = str(col).strip()
    return best_col


def _column_if_exact(df: pd.DataFrame, want: str) -> str | None:
    """Pick column whose normalized header equals `want` (single best match)."""
    want_n = _norm_header(want)
    for col in df.columns:
        if _norm_header(col) == want_n:
            return str(col).strip()
    return None


def _find_property_display_column(df: pd.DataFrame, prop_col: str | None) -> str | None:
    """
    Optional column with human-readable property / estate name when the main Property column
    holds a code (e.g. PRC-0036). Used for register headlines instead of the code.
    """
    if df is None or df.empty:
        return None
    for col in df.columns:
        c = str(col).strip()
        if prop_col and c == prop_col:
            continue
        nh = _norm_header(col)
        if nh in (
            'property name',
            'estate name',
            'project name',
            'building name',
            'site name',
            'property description',
            'project title',
            'property title',
            'estate',
        ):
            return c
        if ('property' in nh and 'name' in nh) or nh.endswith('property title'):
            if 'code' not in nh and 'id' not in nh and 'key' not in nh:
                return c
    return None


def _property_headline_source_for_row(r: dict) -> str:
    """Prefer display name from register; fall back to Property code field."""
    disp = (r.get('property_display') or '').strip()
    code = (r.get('property') or '').strip()
    return disp if disp else code


def _map_register_columns(df: pd.DataFrame) -> dict[str, str | None]:
    """Map logical fields to CAFM-style columns (Location Register .xlsx).

    Prefer a plain **Zone** column over **Sub Zone** so both can coexist (CAFM exports both).
    Optional columns (city, sub zone, codes, etc.) are used for grouping and display only.
    """
    bu = _find_register_column(
        df,
        [
            'base unit name',
            'base unit',
            'bu name',
            'unit name',
            'location name',
            'location',
        ],
    )
    if not bu:
        bu = _find_register_column(df, ['name'])
    bu_code = _find_register_column(df, ['base unit code', 'bu code', 'unit code'])
    ft = _find_register_column(
        df,
        ['bu funct type', 'funct type', 'function type', 'functional type', 'bu function'],
    )
    ft_sub = _find_register_column(
        df,
        ['bu funct sub type', 'bu function sub type', 'funct sub type', 'function sub type'],
    )
    prop_code = _column_if_exact(df, 'Property Code') or _find_register_column(
        df,
        ['property code', 'property id', 'project code', 'prc', 'prc code', 'property ref'],
    )
    prop_name = _column_if_exact(df, 'Property Name') or _find_register_column(
        df,
        [
            'property name',
            'property description',
            'project name',
            'building name',
            'estate name',
            'site name',
        ],
    )
    prop_generic = None
    if not prop_code and not prop_name:
        prop_generic = _find_register_column(df, ['property', 'project'])

    prop_display_col = None
    if prop_code and prop_name and prop_code != prop_name:
        prop = prop_code
        prop_display_col = prop_name
    elif prop_code and not prop_name:
        prop = prop_code
        prop_display_col = _find_property_display_column(df, prop)
    elif prop_name and not prop_code:
        prop = prop_name
        prop_display_col = _find_property_display_column(df, prop)
    elif prop_generic:
        prop = prop_generic
        prop_display_col = _find_property_display_column(df, prop)
    else:
        prop = None
        prop_display_col = None
    # Main zone: exact "Zone" first, then names — never pick "Sub Zone" as the main zone column.
    zone = _column_if_exact(df, 'Zone')
    if not zone:
        zone = _find_register_column(
            df,
            [
                'zone name',
                'area zone name',
                'location zone',
            ],
        )
    if not zone:
        for col in df.columns:
            nh = _norm_header(col)
            if nh == 'zone' or nh.endswith(' zone') and 'sub' not in nh:
                zone = str(col).strip()
                break
    if not zone:
        zone = _find_register_column(
            df,
            [
                'zone code',
                'bu zone',
                'area zone',
                'zone',
            ],
        )
    # Sub zone (finer grouping under Zone) — "Sub Zone", not "Sub Zone Code"
    sub_zone = _column_if_exact(df, 'Sub Zone')
    if not sub_zone:
        sub_zone = _find_register_column(
            df,
            ['sub zone name', 'subzone name', 'sub-zone name'],
        )
    if not sub_zone:
        for col in df.columns:
            nh = _norm_header(col)
            if nh in ('sub zone', 'subzone') and 'code' not in nh:
                sub_zone = str(col).strip()
                break
    sub_zone_code = _find_register_column(
        df,
        ['sub zone code', 'subzone code', 'sub-zone code'],
    )
    city = _column_if_exact(df, 'City') or _find_register_column(df, ['city'])
    area = _column_if_exact(df, 'Area') or _find_register_column(df, ['area'])
    area_group = _find_register_column(df, ['area group', 'areagroup'])
    permit_req = _find_register_column(df, ['permit required', 'permit'])
    tenantable = _column_if_exact(df, 'Tenantable') or _find_register_column(df, ['tenantable'])
    # Stable CAFM / register identifier (highlighted “Key ID” in many exports) — dedupe & UI row key
    key_id = _find_register_column(
        df,
        [
            'key id',
            'keyid',
            'location key id',
            'location id',
            'cafm id',
            'cafm key',
            'record id',
            'register id',
            'base unit id',
            'bu id',
        ],
    )

    return {
        'base_unit': bu,
        'base_unit_code': bu_code,
        'key_id': key_id,
        'funct_type': ft,
        'funct_sub_type': ft_sub,
        'property': prop,
        'property_display': prop_display_col,
        'zone': zone,
        'sub_zone': sub_zone,
        'sub_zone_code': sub_zone_code,
        'city': city,
        'area': area,
        'area_group': area_group,
        'permit_required': permit_req,
        'tenantable': tenantable,
    }


def _strip_utf8_bom(data: bytes) -> bytes:
    if data.startswith(b'\xef\xbb\xbf'):
        return data[3:]
    return data


def _decode_html_bytes(data: bytes) -> str:
    data = _strip_utf8_bom(data)
    for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode('utf-8', errors='replace')


_MSO_COND_OPEN = re.compile(r'<!--\s*\[if[^\]]*\]\s*>', re.IGNORECASE | re.DOTALL)
_MSO_COND_CLOSE = re.compile(r'<!\[endif\]\s*-->', re.IGNORECASE)


def _unwrap_mso_conditional_comments(html: str) -> str:
    """Strip Office conditional-comment wrappers so <table> inside is visible to parsers."""
    s = _MSO_COND_OPEN.sub('', html)
    return _MSO_COND_CLOSE.sub('', s)


def _inline_html_comments_with_table_markup(html: str) -> str:
    """Unwrap <!-- ... --> blocks that contain table rows so BeautifulSoup/pandas can see them."""
    out: list[str] = []
    i = 0
    while True:
        start = html.find('<!--', i)
        if start < 0:
            out.append(html[i:])
            break
        out.append(html[i:start])
        end = html.find('-->', start + 4)
        if end < 0:
            out.append(html[start:])
            break
        body = html[start + 4 : end]
        low = body.lower()
        if '<table' in low or '<tr' in low or '<td' in low:
            out.append(body)
        else:
            out.append(html[start : end + 3])
        i = end + 3
    return ''.join(out)


def _dedupe_column_names(names: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for raw in names:
        h = str(raw).strip() or 'col'
        if h in seen:
            seen[h] += 1
            h = f'{h}_{seen[h]}'
        else:
            seen[h] = 0
        out.append(h)
    return out


def _grid_to_dataframe(grid: list[list[str]]) -> pd.DataFrame | None:
    if not grid:
        return None
    width = max(len(r) for r in grid)
    norm = [list(r) + [''] * (width - len(r)) for r in grid]
    if len(norm) == 1:
        return pd.DataFrame([norm[0]], columns=_dedupe_column_names([str(i) for i in range(width)]))
    header = [str(c).strip() for c in norm[0]]
    cols = _dedupe_column_names([h if h else f'col_{i}' for i, h in enumerate(header)])
    body = norm[1:]
    return pd.DataFrame(body, columns=cols)


def _dataframes_from_tr_tags(html: str) -> list[pd.DataFrame]:
    """Build DataFrames from <tr>/<td> when pandas.read_html finds no <table> (or MSO quirks)."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []
    soup = BeautifulSoup(html, 'lxml')
    out: list[pd.DataFrame] = []

    def append_from_grid(grid: list[list[str]]) -> None:
        df = _grid_to_dataframe(grid)
        if df is not None and not df.empty:
            out.append(df)

    for table in soup.find_all('table'):
        grid: list[list[str]] = []
        for tr in table.find_all('tr'):
            cells = tr.find_all(['td', 'th'])
            if not cells:
                continue
            grid.append([c.get_text(' ', strip=True) for c in cells])
        append_from_grid(grid)

    if not out:
        orphan_rows: list[list[str]] = []
        for tr in soup.find_all('tr'):
            if tr.find_parent('table') is not None:
                continue
            cells = tr.find_all(['td', 'th'])
            if not cells:
                continue
            orphan_rows.append([c.get_text(' ', strip=True) for c in cells])
        append_from_grid(orphan_rows)

    return out


def _read_html_dataframes(data: bytes) -> list[pd.DataFrame]:
    """
    Extract tables from HTML/Office "web page" exports.

    Order: unwrap MSO conditional comments → pandas read_html (whole doc, cleaned doc,
    per-<table>) → manual <tr>/<td> extraction (when pandas reports no tables).
    """
    dfs: list[pd.DataFrame] = []
    text = _unwrap_mso_conditional_comments(_decode_html_bytes(data))
    text = _inline_html_comments_with_table_markup(text)
    buf = BytesIO(text.encode('utf-8'))

    for flavor in ('lxml', 'html5lib', None):
        try:
            kw = {'displayed_only': False}
            if flavor is not None:
                kw['flavor'] = flavor
            buf.seek(0)
            tbs = pd.read_html(buf, **kw)
            for t in tbs or []:
                if t is not None and not t.empty:
                    dfs.append(t)
            if dfs:
                return dfs
        except Exception:
            pass

    try:
        from bs4 import BeautifulSoup
    except ImportError as e:
        raise ValueError(
            'Could not read HTML table export: beautifulsoup4 is required. '
            'pip install beautifulsoup4 lxml html5lib'
        ) from e

    soup = BeautifulSoup(text, 'lxml')

    if not dfs:
        soup_clean = BeautifulSoup(text, 'lxml')
        for tag in soup_clean(['script', 'style', 'noscript']):
            tag.decompose()
        for flavor in ('lxml', 'html5lib', None):
            try:
                kw = {'displayed_only': False}
                if flavor is not None:
                    kw['flavor'] = flavor
                tbs = pd.read_html(StringIO(str(soup_clean)), **kw)
                for t in tbs or []:
                    if t is not None and not t.empty:
                        dfs.append(t)
                if dfs:
                    return dfs
            except Exception:
                pass

    for table in soup.find_all('table'):
        fragment = str(table)
        for flavor in ('lxml', 'html5lib', None):
            try:
                kw = {'displayed_only': False}
                if flavor is not None:
                    kw['flavor'] = flavor
                tbs = pd.read_html(StringIO(fragment), **kw)
                for t in tbs or []:
                    if t is not None and not t.empty:
                        dfs.append(t)
                break
            except Exception:
                continue

    if not dfs:
        dfs.extend(_dataframes_from_tr_tags(text))

    return dfs


def _pick_largest_table(dfs: list[pd.DataFrame]) -> pd.DataFrame | None:
    if not dfs:
        return None

    def score(df: pd.DataFrame) -> int:
        if df is None or df.empty:
            return -1
        return int(df.shape[0]) * max(1, int(df.shape[1]))

    return max(dfs, key=score)


def _dataframe_from_register_bytes(data: bytes, filename: str) -> tuple[pd.DataFrame, str]:
    """Load first sheet / first HTML table. Returns (df, source_format)."""
    data = _strip_utf8_bom(data)
    head = (data[:1024] if data else b'').lstrip().lower()
    if head.startswith(b'<!doctype') or head.startswith(b'<html'):
        dfs = _read_html_dataframes(data)
        df = _pick_largest_table(dfs)
        if df is None:
            fn = (filename or '').lower()
            try:
                bio = BytesIO(data)
                if fn.endswith(('.xlsx', '.xlsm')):
                    bio.seek(0)
                    df = pd.read_excel(bio, engine='openpyxl', sheet_name=0)
                elif fn.endswith('.xls'):
                    try:
                        bio.seek(0)
                        df = pd.read_excel(bio, engine='xlrd', sheet_name=0)
                    except Exception:
                        bio.seek(0)
                        df = pd.read_excel(bio, engine='openpyxl', sheet_name=0)
                if df is not None and not df.empty:
                    df = df.dropna(axis=1, how='all')
                    return df, 'excel'
            except Exception:
                pass
            raise ValueError(
                'This file could not be read as a location register table. '
                'Export from CAFM as a real .xlsx file, or use a spreadsheet/HTML export that '
                'includes visible grid rows (table or row/cell layout).'
            )
        df = df.dropna(axis=1, how='all')
        return df, 'html_table'

    buf = BytesIO(data)
    fn = (filename or '').lower()
    try:
        buf.seek(0)
        if fn.endswith('.xls') and not head.startswith(b'<'):
            df = pd.read_excel(buf, engine='xlrd', sheet_name=0)
        elif fn.endswith('.csv'):
            last_err = None
            for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    buf.seek(0)
                    df = pd.read_csv(buf, encoding=enc)
                    break
                except Exception as e:
                    last_err = e
            else:
                raise ValueError(f'Could not read CSV: {last_err}') from last_err
        else:
            df = pd.read_excel(buf, engine='openpyxl', sheet_name=0)
    except ValueError:
        raise
    except Exception as e:
        raise ValueError(
            f'Could not read Excel workbook: {e}. '
            'If you saved the page from the browser, export a real .xlsx from CAFM instead.'
        ) from e

    df = df.dropna(axis=1, how='all')
    return df, 'excel'


def _register_row_sort_key(u: dict) -> tuple:
    """Sort by Key ID when present (natural order), else Base Unit."""
    kid = (str(u.get('key_id') or '')).strip()
    bu = (u.get('base_unit') or '').lower()
    if kid:
        return (0, kid.lower(), bu)
    return (1, bu, bu)


def _zone_category_sort_key(z: str) -> tuple:
    """
    Sort zone headings for the Location Register: Common Area, Residential, etc. first,
    then other zones alphabetically, then (No zone) last.
    """
    s = (z or '').strip().lower()
    if not s or s == '(no zone)':
        return (2, 0, '')
    # Longer phrases first so e.g. "common area" wins over bare "common"
    rules: list[tuple[str, int]] = [
        ('common area', 0),
        ('common parts', 1),
        ('common', 2),
        ('residential', 3),
        ('commercial', 4),
        ('retail', 5),
        ('office', 6),
        ('parking', 7),
        ('amenity', 8),
        ('amenities', 8),
        ('service', 9),
        ('utility', 10),
        ('utilities', 10),
        ('mechanical', 11),
        ('electrical', 12),
        ('landscape', 13),
        ('external', 14),
    ]
    for phrase, rank in rules:
        if phrase in s:
            return (0, rank, s)
    return (1, 0, s)


def _loc_register_property_group_key(raw: str) -> str:
    """
    Roll up per-tower / per-block CAFM property names into one heading for the Location Register UI.

    - Any property text containing both "orient" and "tower" → single group **Orient Tower**
    - Any property text containing both "garden" and "city", or a known Garden City CAFM code → **Garden City**
    Headlines prefer **property_display** (Property Name / estate name) when present so codes (e.g. PRC-0036) are not shown alone.
    Row payloads still keep the original `property` (code) and optional `property_display` from CAFM.
    """
    s = (raw or '').strip()
    if not s:
        return '(No property)'
    low = s.lower()
    if 'orient' in low and 'tower' in low:
        return 'Orient Tower'
    if ('garden' in low and 'city' in low) or _blob_contains_garden_city_property_code(s):
        return 'Garden City'
    return s


def _group_register_by_property_zone(rows: list[dict]) -> list[dict]:
    """Group rows by property → zone → sub_zone (when present) for the admin UI."""
    from collections import defaultdict

    has_sub = any((r.get('sub_zone') or '').strip() for r in rows)

    tree: dict[str, dict[str, dict[str, list[dict]]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(list))
    )
    for r in rows:
        p = _loc_register_property_group_key(_property_headline_source_for_row(r))
        z = (r.get('zone') or '').strip() or '(No zone)'
        if has_sub:
            sz = (r.get('sub_zone') or '').strip() or '(No sub-zone)'
        else:
            sz = ''
        tree[p][z][sz].append(r)

    out: list[dict] = []
    for p in sorted(tree.keys(), key=lambda x: x.lower()):
        zd = tree[p]
        zones_out: list[dict] = []
        for z in sorted(zd.keys(), key=_zone_category_sort_key):
            szd = zd[z]
            if has_sub:
                sub_zones: list[dict] = []
                for sz in sorted(szd.keys(), key=lambda x: x.lower()):
                    units = sorted(szd[sz], key=_register_row_sort_key)
                    sub_zones.append({'sub_zone': sz, 'units': units})
                zones_out.append({'zone': z, 'sub_zones': sub_zones})
            else:
                units = sorted(szd.get('', []), key=_register_row_sort_key)
                zones_out.append({'zone': z, 'units': units})
        out.append({'property': p, 'zones': zones_out})
    return out


def parse_location_register_bytes(data: bytes, filename: str) -> dict:
    """
    Parse Location Register: Base Unit name, BU Funct Type, Property, Zone.

    When a **Key ID** column exists (e.g. CAFM “Key ID” / highlighted identifier), it is used as
    the stable row key for deduplication and saved toggles; otherwise rows are deduped by
    (Base Unit, Property, Zone, Sub Zone).

    Supports .xlsx, binary .xls, and CAFM HTML spreadsheet exports (single-file HTML must contain
    a `<table>`; “Save as Web Page” .xls that only references external sheet files may be empty —
    re-save as .xlsx from Excel).

    Returns grouped structure for the Report settings UI.
    """
    df, source_format = _dataframe_from_register_bytes(data, filename)
    if df is None or df.empty:
        return {
            'source_format': source_format,
            'columns_detected': {},
            'apt_like_count': 0,
            'total_rows': 0,
            'rows': [],
            'by_property': [],
            'stats': {},
        }

    df = df.dropna(how='all')
    if df.empty:
        return {
            'source_format': source_format,
            'columns_detected': {},
            'apt_like_count': 0,
            'total_rows': 0,
            'rows': [],
            'by_property': [],
            'stats': {},
        }

    cols = _map_register_columns(df)
    bu_col = cols.get('base_unit')
    if not bu_col or bu_col not in df.columns:
        raise ValueError(
            'Could not find a Base Unit name column. '
            'Expected a column titled like “Base Unit”, “Base Unit Name”, or “BU Name”.'
        )

    ft_col = cols.get('funct_type')
    pr_col = cols.get('property')
    zn_col = cols.get('zone')

    rows: list[dict] = []
    apt_like = 0
    seen_key: set[tuple] = set()

    cols_list = list(df.columns)
    bi = cols_list.index(bu_col)

    def _idx(cname: str | None) -> int | None:
        if not cname or cname not in cols_list:
            return None
        return cols_list.index(cname)

    fti = _idx(ft_col)
    pri = _idx(pr_col)
    prop_display_col = cols.get('property_display')
    propdi = _idx(prop_display_col)
    zni = _idx(zn_col)
    subzi = _idx(cols.get('sub_zone'))
    ftsubi = _idx(cols.get('funct_sub_type'))
    keyidi = _idx(cols.get('key_id'))
    bucodei = _idx(cols.get('base_unit_code'))
    cityi = _idx(cols.get('city'))
    areai = _idx(cols.get('area'))
    agri = _idx(cols.get('area_group'))
    szcodei = _idx(cols.get('sub_zone_code'))
    permi = _idx(cols.get('permit_required'))
    teni = _idx(cols.get('tenantable'))

    def _tup_cell(tup: tuple, ix: int | None) -> str:
        if ix is None:
            return ''
        v = tup[ix]
        if pd.isna(v):
            return ''
        s = str(v).strip()
        return '' if s.lower() == 'nan' else s

    # itertuples() is much faster than iterrows() for large exports
    for tup in df.itertuples(index=False, name=None):
        bu_raw = tup[bi]
        if pd.isna(bu_raw):
            continue
        bu = str(bu_raw).strip()
        if not bu or bu.lower() == 'nan':
            continue
        prop_v = _tup_cell(tup, pri)
        zone_v = _tup_cell(tup, zni)
        sz_v = _tup_cell(tup, subzi)
        kid_raw = _tup_cell(tup, keyidi) if keyidi is not None else ''
        kid = kid_raw.strip()
        if kid:
            dedupe = ('id', kid.lower())
        else:
            dedupe = ('bu', bu.lower(), prop_v.lower(), zone_v.lower(), sz_v.lower())
        if dedupe in seen_key:
            continue
        seen_key.add(dedupe)

        ft_v = _tup_cell(tup, fti)
        is_apt = bool(_APT_NO_WITH_NUMBER_RE.search(bu))
        if is_apt:
            apt_like += 1

        row_obj: dict = {
            'base_unit': bu,
            'funct_type': ft_v,
            'property': prop_v,
            'zone': zone_v,
            'is_apt': is_apt,
        }
        if propdi is not None:
            pd_v = _tup_cell(tup, propdi)
            if pd_v:
                row_obj['property_display'] = pd_v
        if kid:
            row_obj['key_id'] = kid
        if sz_v:
            row_obj['sub_zone'] = sz_v
        if ftsubi is not None:
            row_obj['funct_sub_type'] = _tup_cell(tup, ftsubi)
        if bucodei is not None:
            row_obj['base_unit_code'] = _tup_cell(tup, bucodei)
        if cityi is not None:
            row_obj['city'] = _tup_cell(tup, cityi)
        if areai is not None:
            row_obj['area'] = _tup_cell(tup, areai)
        if agri is not None:
            row_obj['area_group'] = _tup_cell(tup, agri)
        if szcodei is not None:
            row_obj['sub_zone_code'] = _tup_cell(tup, szcodei)
        if permi is not None:
            v = tup[permi] if permi < len(tup) else None
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                row_obj['permit_required'] = v
        if teni is not None:
            v = tup[teni] if teni < len(tup) else None
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                row_obj['tenantable'] = v

        rows.append(row_obj)

    by_property = _group_register_by_property_zone(rows)

    non_apt_n = sum(1 for r in rows if not r.get('is_apt'))
    props = {_loc_register_property_group_key(_property_headline_source_for_row(r)) for r in rows}
    zones = {r.get('zone', '') for r in rows}
    subz = {r.get('sub_zone', '') for r in rows if (r.get('sub_zone') or '').strip()}
    stats = {
        'row_count': len(rows),
        'property_count': len([p for p in props if p]),
        'zone_count': len([z for z in zones if z]),
        'sub_zone_count': len(subz),
        'non_apartment_rows': non_apt_n,
        'apartment_rows': apt_like,
    }

    return {
        'source_format': source_format,
        'columns_detected': {k: v for k, v in cols.items() if v},
        'apt_like_count': apt_like,
        'total_rows': len(rows),
        'rows': rows,
        'by_property': by_property,
        'stats': stats,
    }


def _load_mmr_chargeable_config_from_db() -> dict:
    try:
        from flask import has_app_context
        if not has_app_context():
            return dict(DEFAULT_MMR_CHARGEABLE_CONFIG)
        from app.models import MmrChargeableConfig
        row = MmrChargeableConfig.query.first()
        if row and row.config_json:
            return _merge_mmr_chargeable_config(row.config_json)
    except Exception as e:
        logger.debug('MMR chargeable config DB read skipped: %s', e)
    return dict(DEFAULT_MMR_CHARGEABLE_CONFIG)


def get_mmr_chargeable_config() -> dict:
    """Return merged MMR chargeable settings (cached until invalidate_mmr_chargeable_config_cache)."""
    global _mmr_chargeable_config_cache
    if _mmr_chargeable_config_cache is None:
        _mmr_chargeable_config_cache = _load_mmr_chargeable_config_from_db()
    return dict(_mmr_chargeable_config_cache)


def invalidate_mmr_chargeable_config_cache() -> None:
    global _mmr_chargeable_config_cache
    _mmr_chargeable_config_cache = None


def _apply_mmr_chargeable_overrides(
    base_unit_val: str,
    service_group_val: str,
    client_val: str,
    contract_val: str,
    work_description_val: str,
    specific_area_val: str,
    resolved: str,
    config: dict,
    property_val: str = '',
    site_val: str = '',
    complaint_val: str = '',
) -> str:
    """Admin-defined BaseUnit substring overrides. Longest pattern wins."""
    overrides = config.get('baseunit_overrides') or []
    if not overrides or not (base_unit_val or '').strip():
        return resolved
    sg = (service_group_val or '').strip().lower()
    if _br(config, 'facade_cleaning_non_chargeable') and 'facade cleaning' in sg:
        return resolved
    if _br(config, 'elevator_non_chargeable') and _ELEVATOR_SERVICE_GROUP_RE.search(
        service_group_val or ''
    ):
        return resolved
    if _br(config, 'roof_top_non_chargeable') and _text_indicates_roof_top(
        base_unit_val or '', work_description_val or '', specific_area_val or '',
    ):
        return resolved
    if _br(config, 'garden_city_ac_non_chargeable') and _text_indicates_garden_city_project(
        client_val or '', contract_val or '', property_val or '', site_val or ''
    ) and _text_indicates_ac_hvac_complaint(
        service_group_val or '',
        work_description_val or '',
        specific_area_val or '',
        complaint_val or '',
    ):
        return resolved
    if (
        _br(config, 'residential_hvac_non_chargeable')
        and _text_indicates_garden_city_project(
            client_val or '', contract_val or '', property_val or '', site_val or ''
        )
        and _APT_NO_WITH_NUMBER_RE.search(base_unit_val or '')
        and _text_indicates_ac_hvac_complaint(
            service_group_val or '',
            work_description_val or '',
            specific_area_val or '',
            complaint_val or '',
        )
    ):
        return resolved
    if _br(config, 'apt_number_chargeable') and _APT_NO_WITH_NUMBER_RE.search(
        base_unit_val or ''
    ):
        return resolved
    bu = (base_unit_val or '').strip().lower()
    best_chargeable = None
    best_len = -1
    for o in overrides:
        if not isinstance(o, dict):
            continue
        pat = (o.get('pattern') or '').strip().lower()
        if not pat or pat not in bu:
            continue
        if len(pat) > best_len:
            best_len = len(pat)
            best_chargeable = bool(o.get('chargeable'))
    if best_chargeable is None:
        return resolved
    return 'Chargeable' if best_chargeable else 'Non-Chargeable'


def _resolve_chargeable(space_val: str, base_unit_val: str, client_val: str,
                       service_group_val: str = '', contract_val: str = '',
                       work_description_val: str = '', specific_area_val: str = '',
                       config: dict | None = None,
                       property_val: str = '', site_val: str = '',
                       complaint_val: str = '') -> str:
    """
    Resolve Chargeable vs Non-Chargeable.
    - Facade Cleaning service group: always Non-Chargeable.
    - Elevator system (service group: elevator / common CAFM typo elevater): always Non-Chargeable.
    - Garden City only: AC/HVAC complaints = Non-Chargeable (estate from Client/Contract/Property/Site;
      AC/HVAC from Service Group and/or Work Description / Specific Area / Complaint).
    - BaseUnit / Work Description / Specific Area mentioning roof top (or typo e.g. roof toop) = Non-Chargeable.
    - If BaseUnit is non-empty: **Garden City only** — apartment pattern + HVAC/AC-type Service Group or
      complaint → Non-Chargeable when that rule is on; else "Apt No" + number → Chargeable (before reception/floor rules).
    - Else Non-Chargeable for specific CAFM labels (reception, outside / out side, exit+entry or exit/),
      or if BaseUnit contains the word floor; other non-apartment BaseUnit text follows admin config
      (default: Non-Chargeable to match CAFM exports where only apartments bill).
    - If BaseUnit is empty: Askaan, Ajman Holding, Injaaz default Chargeable; else use Excel Space.
    - Optional admin substring overrides (longest match) via get_mmr_chargeable_config().
    """
    if config is None:
        config = get_mmr_chargeable_config()
    client = (client_val or '').strip().lower()
    contract = (contract_val or '').strip().lower()
    combined = f'{client} {contract}'
    sg = (service_group_val or '').strip().lower()
    base_unit = (base_unit_val or '').strip().lower()

    if _br(config, 'facade_cleaning_non_chargeable') and 'facade cleaning' in sg:
        out = 'Non-Chargeable'
    elif _br(config, 'elevator_non_chargeable') and _ELEVATOR_SERVICE_GROUP_RE.search(sg):
        out = 'Non-Chargeable'
    elif _br(config, 'roof_top_non_chargeable') and _text_indicates_roof_top(
        base_unit_val or '', work_description_val or '', specific_area_val or '',
    ):
        out = 'Non-Chargeable'
    elif _br(config, 'garden_city_ac_non_chargeable') and _text_indicates_garden_city_project(
        client_val or '', contract_val or '', property_val or '', site_val or ''
    ) and _text_indicates_ac_hvac_complaint(
        service_group_val or '',
        work_description_val or '',
        specific_area_val or '',
        complaint_val or '',
    ):
        out = 'Non-Chargeable'
    elif base_unit:
        if (
            _br(config, 'residential_hvac_non_chargeable')
            and _text_indicates_garden_city_project(
                client_val or '', contract_val or '', property_val or '', site_val or ''
            )
            and _APT_NO_WITH_NUMBER_RE.search(base_unit_val or '')
            and _text_indicates_ac_hvac_complaint(
                service_group_val or '',
                work_description_val or '',
                specific_area_val or '',
                complaint_val or '',
            )
        ):
            out = 'Non-Chargeable'
        elif _br(config, 'apt_number_chargeable') and _APT_NO_WITH_NUMBER_RE.search(
            base_unit_val or ''
        ):
            out = 'Chargeable'
        elif _br(config, 'cafm_reception_outside_exit_non_chargeable') and _baseunit_is_non_chargeable_cafm_labels(
            base_unit
        ):
            out = 'Non-Chargeable'
        elif _br(config, 'floor_word_non_chargeable') and 'floor' in base_unit:
            out = 'Non-Chargeable'
        elif config.get('non_apartment_baseunit_non_chargeable', True):
            out = 'Non-Chargeable'
        else:
            out = 'Chargeable'
    elif any(p in combined for p in _ALL_BASEUNITS_CHARGEABLE_CLIENTS):
        out = 'Chargeable'
    else:
        space = (space_val or '').strip().lower()
        if space and space not in ('unknown',):
            n = _normalise_space(space_val)
            if n in ('Chargeable', 'Non-Chargeable'):
                out = n
            else:
                out = 'Non-Chargeable'
        else:
            out = 'Non-Chargeable'

    return _apply_mmr_chargeable_overrides(
        base_unit_val,
        service_group_val,
        client_val,
        contract_val,
        work_description_val,
        specific_area_val,
        out,
        config,
        property_val,
        site_val,
        complaint_val,
    )


def preview_chargeable_for_base_units(
    base_units: list[str],
    merged_config: dict,
) -> list[dict[str, str]]:
    """
    Resolve Chargeable/Non-Chargeable for BaseUnit-only strings (location register preview).
    Uses the same pipeline as MMR reports: built-in rules, non-apartment default, then overrides.
    Space / Service Group / Client are empty (register export has BaseUnit only).
    """
    out: list[dict[str, str]] = []
    for bu in base_units:
        s = bu if isinstance(bu, str) else str(bu or '')
        s = s.strip()
        r = _resolve_chargeable('', s, '', '', '', '', '', merged_config)
        out.append({'base_unit': s, 'resolved': r})
    return out


def _get_resolved_chargeable_series(df: pd.DataFrame) -> pd.Series:
    """Return a Series of Chargeable/Non-Chargeable per row for chargeable analysis."""
    cfg = get_mmr_chargeable_config()
    space_col = df['Space'] if 'Space' in df.columns else pd.Series([''] * len(df))
    base_col = df['BaseUnit'] if 'BaseUnit' in df.columns else pd.Series([''] * len(df))
    client_col = df['Client'] if 'Client' in df.columns else pd.Series([''] * len(df))
    contract_col = df['Contract'] if 'Contract' in df.columns else pd.Series([''] * len(df))
    sg_col = df['Service Group'] if 'Service Group' in df.columns else pd.Series([''] * len(df))
    wd_col = df['Work Description'] if 'Work Description' in df.columns else pd.Series([''] * len(df))
    sa_col = df['Specific Area'] if 'Specific Area' in df.columns else pd.Series([''] * len(df))
    n = len(df)
    return pd.Series([
        _resolve_chargeable(
            space_col.iat[i] if i < len(space_col) else '',
            base_col.iat[i] if i < len(base_col) else '',
            client_col.iat[i] if i < len(client_col) else '',
            sg_col.iat[i] if i < len(sg_col) else '',
            contract_col.iat[i] if i < len(contract_col) else '',
            wd_col.iat[i] if i < len(wd_col) else '',
            sa_col.iat[i] if i < len(sa_col) else '',
            cfg,
            _extra_project_blob_for_chargeable(df, i) if i < n else '',
            '',
            _complaint_text_for_chargeable(df, i) if i < n else '',
        )
        for i in range(n)
    ], index=df.index)


def format_chargeable_summary_for_email(df: pd.DataFrame) -> str:
    """Build a plain-text chargeable table for email body.
    Rows: Service Groups. Cols: Chargeable (Resolved+Pending), Non-Chargeable (Resolved+Pending).
    """
    required = {'Space', 'Status', 'Service Group'}
    if not required.issubset(set(df.columns)):
        return ''
    work = df.copy()
    work['_space'] = _get_resolved_chargeable_series(df)
    work['_status'] = work['Status'].apply(_normalise_status_bucket)

    agg = (
        work.groupby('Service Group', dropna=False)
        .apply(lambda g: pd.Series({
            'chg_res': len(g[(g['_space'] == 'Chargeable') & (g['_status'] == 'Resolved')]),
            'chg_pen': len(g[(g['_space'] == 'Chargeable') & (g['_status'] == 'Pending')]),
            'nchg_res': len(g[(g['_space'] == 'Non-Chargeable') & (g['_status'] == 'Resolved')]),
            'nchg_pen': len(g[(g['_space'] == 'Non-Chargeable') & (g['_status'] == 'Pending')]),
        }), include_groups=False)
        .reset_index()
    )
    agg['Service Group'] = agg['Service Group'].fillna('').astype(str).str.strip()
    agg = agg.sort_values('Service Group')
    if len(agg) == 0:
        return ''

    sg_w = 28
    num_w = 8
    sep = ' | '
    h1 = 'Service Group'.ljust(sg_w) + sep
    h1 += 'Chg Res'.rjust(num_w) + sep + 'Chg Pend'.rjust(num_w) + sep
    h1 += 'NChg Res'.rjust(num_w) + sep + 'NChg Pend'.rjust(num_w)
    rule = '-' * (sg_w + 4 * (num_w + len(sep)))

    lines = ['Chargeable Summary by Service Group:', rule, h1, rule]
    for _, row in agg.iterrows():
        sg = (str(row['Service Group'] or '')[:sg_w]).ljust(sg_w)
        chg_res = int(row['chg_res'])
        chg_pen = int(row['chg_pen'])
        nchg_res = int(row['nchg_res'])
        nchg_pen = int(row['nchg_pen'])
        lines.append(sg + sep + str(chg_res).rjust(num_w) + sep + str(chg_pen).rjust(num_w) +
                     sep + str(nchg_res).rjust(num_w) + sep + str(nchg_pen).rjust(num_w))

    tot_chg_res = int(agg['chg_res'].sum())
    tot_chg_pen = int(agg['chg_pen'].sum())
    tot_nchg_res = int(agg['nchg_res'].sum())
    tot_nchg_pen = int(agg['nchg_pen'].sum())
    lines.append(rule)
    lines.append('Total'.ljust(sg_w) + sep +
                 str(tot_chg_res).rjust(num_w) + sep + str(tot_chg_pen).rjust(num_w) +
                 sep + str(tot_nchg_res).rjust(num_w) + sep + str(tot_nchg_pen).rjust(num_w))
    lines.append(rule)
    return '\n'.join(lines)


def _normalise_status_bucket(val: str) -> str:
    """Bucket raw Status into Resolved / Pending."""
    v = val.strip().lower()
    if v in ('resolved', 'closed', 'completed', 'done'):
        return 'Resolved'
    return 'Pending'


# Tower display names and matching patterns (Client or Contract)
# Askaan + Saqr treated as one project; Orient/Garden get - TFM suffix
_TOWER_CONFIG = [
    ('Askaan + Saqr Projects', ['askaan', 'saqr']),
    ('Orient Tower - TFM', ['orient']),
    ('Garden City - TFM', ['garden']),
    ('C1 Tower', ['c1 tower', 'c1']),
]
def _tower_for_row(client: str, contract: str) -> str | None:
    """Return tower display name if row belongs to a known tower, else None."""
    c = (client or '').strip().lower()
    t = (contract or '').strip().lower()
    combined = f'{c} {t}'
    for display_name, patterns in _TOWER_CONFIG:
        if any(p in combined for p in patterns):
            return display_name
    return None


def format_per_tower_chargeable_html_for_email(df: pd.DataFrame) -> str:
    """Build HTML tables for email body: one table per tower (Askaan+Saqr, Orient, Garden, C1).
    Each table: Service Name | Chargeable Resolved | Chargeable Pending | Total row | Grand total.
    Uses resolved chargeable logic (Garden City AC = Non-Chargeable). Inline CSS for email compatibility.
    """
    required = {'Space', 'Status', 'Service Group', 'Client', 'Contract'}
    if not required.issubset(set(df.columns)):
        return ''
    work = df.copy()
    work['_space'] = _get_resolved_chargeable_series(df)
    work['_status'] = work['Status'].apply(_normalise_status_bucket)
    work['_tower'] = work.apply(lambda r: _tower_for_row(r.get('Client'), r.get('Contract')), axis=1)
    work = work[work['_space'] == 'Chargeable']
    work = work[work['_tower'].notna()]

    if len(work) == 0:
        return ''

    tables_html = []
    for display_name, _ in _TOWER_CONFIG:
        tower_df = work[work['_tower'] == display_name]
        if len(tower_df) == 0:
            continue
        agg = (
            tower_df.groupby('Service Group', dropna=False)
            .apply(lambda g: pd.Series({
                'resolved': len(g[g['_status'] == 'Resolved']),
                'pending': len(g[g['_status'] == 'Pending']),
            }), include_groups=False)
            .reset_index()
        )
        agg['Service Group'] = agg['Service Group'].fillna('').astype(str).str.strip()
        agg = agg.sort_values('Service Group')
        resolved_total = int(agg['resolved'].sum())
        pending_total = int(agg['pending'].sum())
        grand_total = resolved_total + pending_total

        rows_html = []
        _pad = '2px 4px'
        _cell = f'padding:{_pad};border:1px solid #0d3d24;font-size:10px;text-align:center'
        _hdr = f'background:#{ACCENT};font-weight:bold'
        for _, r in agg.iterrows():
            sg = (r['Service Group'] or '').strip()
            if not sg:
                continue
            res = int(r['resolved'])
            pen = int(r['pending'])
            bg = '#ffffff' if len(rows_html) % 2 == 0 else '#f8faf8'
            rows_html.append(
                f'<tr style="background:{bg}">'
                f'<td style="{_cell}">{sg}</td>'
                f'<td style="{_cell}">{res or ""}</td>'
                f'<td style="{_cell}">{pen or ""}</td>'
                f'</tr>'
            )

        header_bg = f'#{PRIMARY}'
        _sub = f'{_cell};border:1px solid #0d3d24;{_hdr}'
        table = (
            f'<table cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse;font-family:Arial,sans-serif;font-size:10px;width:100%;max-width:420px;margin:0 0 8px 0">'
            f'<tr><td colspan="3" style="padding:4px 6px;background:{header_bg};color:#fff;font-weight:bold;text-align:center;border:1px solid #0d3d24;font-size:11px">{display_name}</td></tr>'
            f'<tr>'
            f'<td style="{_sub}">Service Name</td>'
            f'<td colspan="2" style="{_sub}">Chargeable</td>'
            f'</tr>'
            f'<tr>'
            f'<td style="{_sub}"></td>'
            f'<td style="{_sub}">Resolved</td>'
            f'<td style="{_sub}">Pending</td>'
            f'</tr>'
            + ''.join(rows_html)
            + f'<tr style="background:{header_bg};color:#fff;font-weight:bold">'
            f'<td style="{_cell}">Total</td>'
            f'<td style="{_cell}">{resolved_total}</td>'
            f'<td style="{_cell}">{pending_total}</td>'
            f'</tr>'
            f'<tr><td style="{_cell};border:1px solid #0d3d24"></td><td colspan="2" style="{_cell};{_hdr};border:1px solid #0d3d24">{grand_total}</td></tr>'
            f'</table>'
        )
        tables_html.append(table)

    if not tables_html:
        return ''
    # Stack tables vertically with line break between each
    return '<br><br>'.join(tables_html)


# Clients whose contracts must each get their own sheet instead of one
# combined client sheet.  Matching is case-insensitive / fuzzy.
_SPLIT_CLIENTS = {'aqaar community mangement', 'aqar community management',
                  'aqaar community management'}


def _is_split_client(client: str) -> bool:
    return client.strip().lower() in _SPLIT_CLIENTS


def generate_report_excel(df: pd.DataFrame) -> bytes:
    """Build a professional multi-sheet Excel report and return raw bytes.

    Sheet layout:
      1. All Work Orders
      2–N. Client sheets (one per client, all their contracts combined)
           Exception: "Aqar Community Management" → one sheet per *contract*
      Last. Chargeable / Non-Chargeable Analysis (filterable summary table)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Populate Space with resolved Chargeable/Non-Chargeable (from BaseUnit when empty)
    work = df.copy()
    if 'Space' in work.columns and 'BaseUnit' in work.columns:
        work['Space'] = _get_resolved_chargeable_series(work)

    # Sheet 1 – All Work Orders
    _write_data_sheet(wb, work, 'All Work Orders')

    # Sheet 2 – Dashboard (KPI + charts)
    _write_dashboard_sheet(wb, work)

    # Client-wise sheets
    clients = sorted(work['Client'].replace('', None).dropna().unique().tolist())
    for client in clients:
        client_df = work[work['Client'] == client]
        if _is_split_client(client):
            # Split into per-contract sheets
            contracts = sorted(
                client_df['Contract'].replace('', None).dropna().unique().tolist()
            )
            for contract in contracts:
                name = str(contract)[:31]
                _write_data_sheet(
                    wb,
                    client_df[client_df['Contract'] == contract].copy(),
                    name,
                )
        else:
            name = str(client)[:31]
            _write_data_sheet(wb, client_df.copy(), name)

    # Chargeable / Non-Chargeable Analysis (filterable summary)
    _write_chargeable_analysis(wb, work)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Styling helpers
# ──────────────────────────────────────────────────────────────────────────────

def _thin(color=BORDER):
    return Side(style='thin', color=color)

def _border(color=BORDER):
    s = _thin(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _header_fill():
    return PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type='solid')

def _section_fill():
    return PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type='solid')

def _alt_fill():
    return PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type='solid')

def _write_logo_header(ws, title: str, span_cols: int) -> int:
    """Writes logo + title block; returns next free row number.
    Logo is centered in column A with a little padding."""
    _LOGO_PADDING_PX = 6
    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.merge_cells('A1:A3')
    col_a = ws.column_dimensions.get('A')
    col_a_width = col_a.width if col_a and col_a.width is not None else None
    if col_a_width is None:
        ws.column_dimensions['A'].width = 14
        col_a_width = 14

    if os.path.exists(LOGO_PATH):
        try:
            img = XLImage(LOGO_PATH)
            img.width = 72
            img.height = 72
            p2e = pixels_to_EMU
            size = XDRPositiveSize2D(p2e(72), p2e(72))
            cell_w_px = col_a_width * 7.5
            cell_h_px = (48 + 18 + 18) * (96 / 72)
            col_off_px = _LOGO_PADDING_PX + max(0, (cell_w_px - 2 * _LOGO_PADDING_PX - 72) / 2)
            row_off_px = _LOGO_PADDING_PX + max(0, (cell_h_px - 2 * _LOGO_PADDING_PX - 72) / 2)
            marker = AnchorMarker(col=0, colOff=p2e(col_off_px), row=0, rowOff=p2e(row_off_px))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception:
            pass

    end = get_column_letter(span_cols)
    title_cell = ws.cell(row=1, column=2, value=title)
    title_cell.font = Font(bold=True, size=16, color=PRIMARY, name='Calibri')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    if span_cols > 2:
        ws.merge_cells(f'B1:{end}1')

    sub_cell = ws.cell(row=2, column=2, value='INJAAZ PLATFORM – Report Generation')
    sub_cell.font = Font(bold=True, size=9, color=SECONDARY, name='Calibri')
    sub_cell.alignment = Alignment(horizontal='left', vertical='center')
    if span_cols > 2:
        ws.merge_cells(f'B2:{end}2')

    date_cell = ws.cell(row=3, column=2, value=f'Generated: {datetime.now().strftime("%d %b %Y  %H:%M")}')
    date_cell.font = Font(size=8, color='888888', italic=True, name='Calibri')
    date_cell.alignment = Alignment(horizontal='left', vertical='center')
    if span_cols > 2:
        ws.merge_cells(f'B3:{end}3')

    return 5  # first data row (row 4 is a spacer)


def _style_header_row(ws, row_num: int, num_cols: int):
    ws.row_dimensions[row_num].height = 32
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(bold=True, color=WHITE, size=10, name='Calibri')
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border(PRIMARY)


def _style_data_row(ws, row_num: int, num_cols: int, alt: bool = False):
    ws.row_dimensions[row_num].height = 22
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.font = Font(size=9, name='Calibri')
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        cell.fill = _alt_fill() if alt else PatternFill(fill_type=None)
        cell.border = _border()


# ──────────────────────────────────────────────────────────────────────────────
# Dashboard sheet – cell-based visuals (renders in Protected View / email)
# ──────────────────────────────────────────────────────────────────────────────

_BAR_CHAR = '\u2588'  # █  full-block used for in-cell bar charts
_BAR_MAX_LEN = 30     # max blocks in the longest bar


def _write_bar_section(ws, current_row, title, counts: dict, bar_color=PRIMARY):
    """Write a titled horizontal-bar section using coloured █ characters.

    Layout (columns A–D):
        A: label   B: bar (█ chars)   C: count   D: % share

    Returns the next free row.
    """
    if not counts:
        return current_row

    # Section title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    t = ws.cell(row=current_row, column=1, value=title)
    t.font = Font(bold=True, size=11, color=PRIMARY, name='Calibri')
    t.fill = _section_fill()
    t.alignment = Alignment(horizontal='left', vertical='center')
    t.border = _border()
    for c in range(2, 5):
        ws.cell(row=current_row, column=c).fill = _section_fill()
        ws.cell(row=current_row, column=c).border = _border()
    ws.row_dimensions[current_row].height = 26
    current_row += 1

    max_val = max(counts.values()) if counts else 1
    total_val = sum(counts.values())

    for ri, (name, cnt) in enumerate(counts.items()):
        bar_len = round(cnt / max_val * _BAR_MAX_LEN) if max_val else 0
        pct = round(cnt / total_val * 100, 1) if total_val else 0

        # Label
        lbl = ws.cell(row=current_row, column=1, value=str(name))
        lbl.font = Font(size=9, name='Calibri', color='333333')
        lbl.alignment = Alignment(horizontal='right', vertical='center')
        lbl.border = _border()
        lbl.fill = _alt_fill() if ri % 2 else PatternFill(fill_type=None)

        # Bar
        bar = ws.cell(row=current_row, column=2, value=_BAR_CHAR * bar_len)
        bar.font = Font(size=9, name='Calibri', color=bar_color)
        bar.alignment = Alignment(horizontal='left', vertical='center')
        bar.border = _border()
        bar.fill = _alt_fill() if ri % 2 else PatternFill(fill_type=None)

        # Count
        cv = ws.cell(row=current_row, column=3, value=cnt)
        cv.font = Font(bold=True, size=9, name='Calibri', color=PRIMARY)
        cv.alignment = Alignment(horizontal='center', vertical='center')
        cv.border = _border()
        cv.fill = _alt_fill() if ri % 2 else PatternFill(fill_type=None)

        # Percentage
        pv = ws.cell(row=current_row, column=4, value=f'{pct}%')
        pv.font = Font(size=9, name='Calibri', color='888888')
        pv.alignment = Alignment(horizontal='center', vertical='center')
        pv.border = _border()
        pv.fill = _alt_fill() if ri % 2 else PatternFill(fill_type=None)

        ws.row_dimensions[current_row].height = 20
        current_row += 1

    return current_row + 1  # blank spacer row


def _write_summary_table(ws, current_row, title, data: list, headers: list,
                         *, col_start=1):
    """Write a small summary table.  `data` is a list of row-tuples.

    Returns the next free row.
    """
    num_cols = len(headers)
    end_col = col_start + num_cols - 1

    # Title
    ws.merge_cells(
        start_row=current_row, start_column=col_start,
        end_row=current_row, end_column=end_col,
    )
    t = ws.cell(row=current_row, column=col_start, value=title)
    t.font = Font(bold=True, size=11, color=PRIMARY, name='Calibri')
    t.fill = _section_fill()
    t.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(col_start, end_col + 1):
        ws.cell(row=current_row, column=c).fill = _section_fill()
        ws.cell(row=current_row, column=c).border = _border()
    ws.row_dimensions[current_row].height = 26
    current_row += 1

    # Headers
    for ci, h in enumerate(headers, col_start):
        cell = ws.cell(row=current_row, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border(PRIMARY)
    ws.row_dimensions[current_row].height = 24
    current_row += 1

    # Data rows
    for ri, row_data in enumerate(data):
        for ci, val in enumerate(row_data, col_start):
            cell = ws.cell(row=current_row, column=ci, value=val)
            cell.font = Font(size=9, name='Calibri')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = _border()
            cell.fill = _alt_fill() if ri % 2 else PatternFill(fill_type=None)
        # First column left-aligned (label)
        ws.cell(row=current_row, column=col_start).alignment = Alignment(
            horizontal='left', vertical='center')
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    return current_row + 1


def _write_dashboard_sheet(wb: openpyxl.Workbook, df: pd.DataFrame):
    """Cell-based visual dashboard — two-panel layout, always renders.

    LEFT  (A–D): Bar sections (Client, Contract, Service Group)
    RIGHT (F–H): Chargeable + Priority summary tables
    """
    ws = wb.create_sheet('Dashboard')

    # ── Column widths ────────────────────────────────────────────────
    # Left panel: A=label, B=bar, C=count, D=%
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    # Gutter
    ws.column_dimensions['E'].width = 3
    # Right panel: F=label, G=count, H=share
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # ── Header ───────────────────────────────────────────────────────
    TOTAL_COLS = 8  # A–H
    current_row = _write_logo_header(ws, 'Dashboard Overview', TOTAL_COLS)

    # ── KPI CARDS (full width, 4 per row across A–H) ────────────────
    total   = len(df)
    pending = len(df[df['Status'].str.lower() == 'pending'])    if 'Status'   in df.columns else 0
    p1      = len(df[df['Priority'].str.upper() == 'P1'])       if 'Priority' in df.columns else 0
    p3      = len(df[df['Priority'].str.upper() == 'P3'])       if 'Priority' in df.columns else 0
    n_cli   = df['Client'].replace('', None).dropna().nunique() if 'Client'   in df.columns else 0
    n_ctr   = df['Contract'].replace('', None).dropna().nunique() if 'Contract' in df.columns else 0

    kpis = [
        (total,   'WORK ORDERS',  PRIMARY),
        (pending, 'PENDING',      SECONDARY),
        (p1,      'PRIORITY P1',  'E65100'),
        (p3,      'PRIORITY P3',  '43A047'),
        (n_cli,   'CLIENTS',      PRIMARY),
        (n_ctr,   'CONTRACTS',    PRIMARY),
    ]

    kpi_bg = PatternFill(start_color='F0FAF4', end_color='F0FAF4', fill_type='solid')
    kpi_bdr = Border(
        left=Side(style='thin', color='C8E6C9'),
        right=Side(style='thin', color='C8E6C9'),
        top=Side(style='thin', color='C8E6C9'),
        bottom=Side(style='thin', color='C8E6C9'),
    )

    # 2 rows × 3 KPIs each — columns A, C, F (with merge for wider cards)
    kpi_cols = [
        (1, 2),   # merge A:B
        (3, 4),   # merge C:D
        (6, 8),   # merge F:H  (skip E gutter)
    ]
    for batch in range(2):
        vr = current_row
        lr = current_row + 1
        for slot, (sc, ec) in enumerate(kpi_cols):
            idx = batch * 3 + slot
            if idx >= len(kpis):
                break
            val, label, clr = kpis[idx]
            ws.merge_cells(start_row=vr, start_column=sc, end_row=vr, end_column=ec)
            ws.merge_cells(start_row=lr, start_column=sc, end_row=lr, end_column=ec)

            v = ws.cell(row=vr, column=sc, value=val)
            v.font = Font(bold=True, size=26, color=clr, name='Calibri')
            v.alignment = Alignment(horizontal='center', vertical='bottom')

            lb = ws.cell(row=lr, column=sc, value=label)
            lb.font = Font(bold=True, size=8, color='777777', name='Calibri')
            lb.alignment = Alignment(horizontal='center', vertical='top')

            for r in (vr, lr):
                for c in range(sc, ec + 1):
                    ws.cell(row=r, column=c).fill = kpi_bg
                    ws.cell(row=r, column=c).border = kpi_bdr

        ws.row_dimensions[vr].height = 48
        ws.row_dimensions[lr].height = 18
        current_row += 2

    current_row += 1  # spacer

    # ── Remember where the two-panel area starts ─────────────────────
    panel_start = current_row

    # ── LEFT PANEL: Bar sections (columns A–D) ──────────────────────
    def _counts(col, max_items=20):
        if col not in df.columns:
            return {}
        s = df[col].replace('', None).dropna().value_counts().head(max_items)
        return {str(k): int(v) for k, v in s.items()}

    current_row = _write_bar_section(
        ws, current_row, 'WORK ORDERS BY CLIENT',
        _counts('Client'), bar_color=PRIMARY)

    current_row = _write_bar_section(
        ws, current_row, 'WORK ORDERS BY CONTRACT',
        _counts('Contract'), bar_color=SECONDARY)

    current_row = _write_bar_section(
        ws, current_row, 'BY SERVICE GROUP',
        _counts('Service Group'), bar_color='43A047')

    # ── RIGHT PANEL: Summary tables (columns F–H) ───────────────────
    right_row = panel_start
    RC = 6  # column F

    # Chargeable vs Non-Chargeable (resolved from Space + BaseUnit when Space empty)
    if 'Space' in df.columns:
        if 'BaseUnit' in df.columns:
            space_counts = _get_resolved_chargeable_series(df).value_counts()
        else:
            space_counts = df['Space'].apply(_normalise_space).value_counts()
        total_space = space_counts.sum()
        rows = []
        for name, cnt in space_counts.items():
            pct = round(cnt / total_space * 100, 1) if total_space else 0
            rows.append((str(name), cnt, f'{pct}%'))
        rows.append(('Total', total_space, '100%'))
        right_row = _write_summary_table(
            ws, right_row, 'CHARGEABLE vs NON-CHARGEABLE',
            rows, ['Type', 'Count', 'Share'], col_start=RC)

    # Work Orders by Priority
    if 'Priority' in df.columns:
        prio_counts = df['Priority'].replace('', None).dropna().value_counts()
        total_prio = prio_counts.sum()
        rows = []
        for name, cnt in prio_counts.items():
            pct = round(cnt / total_prio * 100, 1) if total_prio else 0
            rows.append((str(name), int(cnt), f'{pct}%'))
        rows.append(('Total', int(total_prio), '100%'))
        right_row = _write_summary_table(
            ws, right_row, 'WORK ORDERS BY PRIORITY',
            rows, ['Priority', 'Count', 'Share'], col_start=RC)

    # Status breakdown
    if 'Status' in df.columns:
        status_counts = df['Status'].replace('', None).dropna().value_counts()
        total_st = status_counts.sum()
        rows = []
        for name, cnt in status_counts.items():
            pct = round(cnt / total_st * 100, 1) if total_st else 0
            rows.append((str(name), int(cnt), f'{pct}%'))
        rows.append(('Total', int(total_st), '100%'))
        right_row = _write_summary_table(
            ws, right_row, 'WORK ORDERS BY STATUS',
            rows, ['Status', 'Count', 'Share'], col_start=RC)

    # ── Print settings ───────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1


# ──────────────────────────────────────────────────────────────────────────────
# Sheet writers
# ──────────────────────────────────────────────────────────────────────────────

def _write_data_sheet(wb: openpyxl.Workbook, df: pd.DataFrame, title: str):
    ws = wb.create_sheet(title)

    cols = [c for c in EXPECTED_COLS if c in df.columns]
    n = len(cols)
    if n == 0:
        ws.cell(row=1, column=1, value='No data')
        return

    current_row = _write_logo_header(ws, title, n)

    # Header
    for ci, col in enumerate(cols, 1):
        ws.cell(row=current_row, column=ci, value=col)
    _style_header_row(ws, current_row, n)
    ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
    current_row += 1

    # Data
    for ri, (_, row) in enumerate(df.iterrows()):
        for ci, col in enumerate(cols, 1):
            val = row.get(col, '')
            if pd.isna(val):
                val = ''
            elif isinstance(val, pd.Timestamp):
                val = val.strftime('%Y-%m-%d')
            ws.cell(row=current_row, column=ci, value=str(val).strip())
        _style_data_row(ws, current_row, n, alt=(ri % 2 == 1))
        current_row += 1

    # Auto-width (keep col A at least 14 for centered logo with padding)
    for ci, col in enumerate(cols, 1):
        max_len = max(len(col), 8)
        for row in ws.iter_rows(min_row=6, max_row=min(ws.max_row, 200),
                                 min_col=ci, max_col=ci):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 60))
        w = min(max_len + 4, 60)
        if ci == 1:
            w = max(w, 14)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.right.text = "Page &P of &N"


def _write_chargeable_analysis(wb: openpyxl.Workbook, df: pd.DataFrame):
    """Chargeable / Non-Chargeable summary with Client dropdown (AutoFilter).

    One flat table with columns:
        Client | Contract | Service Group |
        Chargeable Resolved | Chargeable Pending |
        Non-Chargeable Resolved | Non-Chargeable Pending | Total

    Excel AutoFilter on every column gives the user a click-to-filter
    dropdown on the Client (and every other) column.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    ws = wb.create_sheet('Chargeable Analysis')

    required = {'Space', 'Status', 'Contract', 'Service Group', 'Client'}
    if not required.issubset(set(df.columns)):
        ws.cell(row=1, column=1, value='Insufficient columns for chargeable analysis')
        return

    work = df.copy()
    work['_space']  = _get_resolved_chargeable_series(df)
    work['_status'] = work['Status'].apply(_normalise_status_bucket)

    DATA_COLS = [
        ('Chargeable',     'Resolved'),
        ('Chargeable',     'Pending'),
        ('Non-Chargeable', 'Resolved'),
        ('Non-Chargeable', 'Pending'),
    ]

    HEADERS = [
        'Client', 'Contract', 'Service Group',
        'Chargeable\nResolved', 'Chargeable\nPending',
        'Non-Chargeable\nResolved', 'Non-Chargeable\nPending',
        'Total',
    ]
    NUM_COLS = len(HEADERS)
    col_widths = [28, 28, 26, 16, 16, 18, 18, 10]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Accent fills
    chg_fill  = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    nchg_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    total_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    current_row = _write_logo_header(ws, 'Chargeable / Non-Chargeable Analysis', NUM_COLS)

    # ── Category sub-header row ──────────────────────────────────────
    for c in range(1, NUM_COLS + 1):
        ws.cell(row=current_row, column=c).border = _border()

    # Merge cells A-C (blank area above Client/Contract/ServiceGroup)
    ws.merge_cells(f'A{current_row}:C{current_row}')

    # Chargeable group header (D-E)
    ws.merge_cells(f'D{current_row}:E{current_row}')
    ch = ws.cell(row=current_row, column=4, value='Chargeable')
    ch.font = Font(bold=True, size=10, color=SECONDARY, name='Calibri')
    ch.fill = chg_fill
    ch.alignment = Alignment(horizontal='center', vertical='center')
    ch.border = _border()
    ws.cell(row=current_row, column=5).fill = chg_fill
    ws.cell(row=current_row, column=5).border = _border()

    # Non-Chargeable group header (F-G)
    ws.merge_cells(f'F{current_row}:G{current_row}')
    nc = ws.cell(row=current_row, column=6, value='Non-Chargeable')
    nc.font = Font(bold=True, size=10, color='E65100', name='Calibri')
    nc.fill = nchg_fill
    nc.alignment = Alignment(horizontal='center', vertical='center')
    nc.border = _border()
    ws.cell(row=current_row, column=7).fill = nchg_fill
    ws.cell(row=current_row, column=7).border = _border()

    ws.cell(row=current_row, column=8).fill = total_fill
    ws.cell(row=current_row, column=8).border = _border()
    ws.row_dimensions[current_row].height = 22
    current_row += 1

    # ── Column header row ────────────────────────────────────────────
    short_headers = [
        'Client', 'Contract', 'Service Group',
        'Resolved', 'Pending', 'Resolved', 'Pending', 'Total',
    ]
    header_row = current_row
    for ci, h in enumerate(short_headers, 1):
        cell = ws.cell(row=current_row, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, size=9, name='Calibri')
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border(PRIMARY)
    ws.row_dimensions[current_row].height = 28
    current_row += 1

    # ── Build aggregation: Client × Contract × Service Group ─────────
    group_keys = ['Client', 'Contract', 'Service Group']
    combos = (
        work.groupby(group_keys)
            .apply(lambda g: pd.Series({
                'chg_res':  len(g[(g['_space'] == 'Chargeable')     & (g['_status'] == 'Resolved')]),
                'chg_pen':  len(g[(g['_space'] == 'Chargeable')     & (g['_status'] == 'Pending')]),
                'nchg_res': len(g[(g['_space'] == 'Non-Chargeable') & (g['_status'] == 'Resolved')]),
                'nchg_pen': len(g[(g['_space'] == 'Non-Chargeable') & (g['_status'] == 'Pending')]),
            }), include_groups=False)
            .reset_index()
            .sort_values(group_keys)
    )

    # Write data rows
    prev_client = None
    for ri, (_, row) in enumerate(combos.iterrows()):
        client  = str(row['Client']).strip()
        contract = str(row['Contract']).strip()
        sg      = str(row['Service Group']).strip()
        counts  = [int(row['chg_res']), int(row['chg_pen']),
                   int(row['nchg_res']), int(row['nchg_pen'])]
        total   = sum(counts)

        ws.cell(row=current_row, column=1, value=client)
        ws.cell(row=current_row, column=2, value=contract)
        ws.cell(row=current_row, column=3, value=sg)
        for ci, val in enumerate(counts, 4):
            ws.cell(row=current_row, column=ci, value=val)
        ws.cell(row=current_row, column=NUM_COLS, value=total)

        _style_data_row(ws, current_row, NUM_COLS, alt=(ri % 2 == 1))

        # Visual separator: bold + light fill when client changes
        if client != prev_client:
            c1 = ws.cell(row=current_row, column=1)
            c1.font = Font(bold=True, size=9, color=PRIMARY, name='Calibri')
            c1.fill = _section_fill()
            prev_client = client

        current_row += 1

    last_data_row = current_row - 1
    first_data_row = header_row + 1

    # Blank row between data and totals so Excel AutoFilter does NOT include totals
    # (Excel extends filter to contiguous rows; blank row breaks contiguity)
    blank_row = current_row
    current_row += 1
    ws.row_dimensions[blank_row].height = 6

    # ── Grand Total row (SUBTOTAL so it updates when user filters) ────
    grand_total_row = current_row
    ws.cell(row=grand_total_row, column=1, value='GRAND TOTAL')
    ws.merge_cells(f'A{grand_total_row}:C{grand_total_row}')
    # Use SUBTOTAL(109, ...) so filtering shows sum of visible rows only
    for ci in range(4, NUM_COLS + 1):
        col_letter = get_column_letter(ci)
        ws.cell(row=grand_total_row, column=ci, value=f'=SUBTOTAL(109,{col_letter}{first_data_row}:{col_letter}{last_data_row})')

    for c in range(1, NUM_COLS + 1):
        cell = ws.cell(row=grand_total_row, column=c)
        cell.font = Font(bold=True, size=10, color=WHITE, name='Calibri')
        cell.fill = PatternFill(start_color=SECONDARY, end_color=SECONDARY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = _border(SECONDARY)
    ws.cell(row=grand_total_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[grand_total_row].height = 28

    # ── Chargeable Total & Non-Chargeable Total rows (below grand total) ───
    gt = grand_total_row
    d_l, e_l, f_l, g_l = get_column_letter(4), get_column_letter(5), get_column_letter(6), get_column_letter(7)
    chg_total_row = gt + 1
    ws.cell(row=chg_total_row, column=1, value='Chargeable Total')
    ws.merge_cells(f'A{chg_total_row}:C{chg_total_row}')
    ws.merge_cells(f'D{chg_total_row}:E{chg_total_row}')
    ws.cell(row=chg_total_row, column=4, value=f'={d_l}{gt}+{e_l}{gt}')
    ws.merge_cells(f'F{chg_total_row}:G{chg_total_row}')
    ws.cell(row=chg_total_row, column=6, value='')
    ws.cell(row=chg_total_row, column=8, value=f'={d_l}{gt}+{e_l}{gt}')
    for c in range(1, NUM_COLS + 1):
        cell = ws.cell(row=chg_total_row, column=c)
        cell.font = Font(bold=True, size=9, name='Calibri')
        cell.fill = chg_fill
        cell.alignment = Alignment(horizontal='center' if c > 3 else 'left', vertical='center')
        cell.border = _border()
    ws.cell(row=chg_total_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[chg_total_row].height = 22

    nchg_total_row = gt + 2
    ws.cell(row=nchg_total_row, column=1, value='Non-Chargeable Total')
    ws.merge_cells(f'A{nchg_total_row}:C{nchg_total_row}')
    ws.merge_cells(f'D{nchg_total_row}:E{nchg_total_row}')
    ws.cell(row=nchg_total_row, column=4, value='')
    ws.merge_cells(f'F{nchg_total_row}:G{nchg_total_row}')
    ws.cell(row=nchg_total_row, column=6, value=f'={f_l}{gt}+{g_l}{gt}')
    ws.cell(row=nchg_total_row, column=8, value=f'={f_l}{gt}+{g_l}{gt}')
    for c in range(1, NUM_COLS + 1):
        cell = ws.cell(row=nchg_total_row, column=c)
        cell.font = Font(bold=True, size=9, name='Calibri')
        cell.fill = nchg_fill
        cell.alignment = Alignment(horizontal='center' if c > 3 else 'left', vertical='center')
        cell.border = _border()
    ws.cell(row=nchg_total_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[nchg_total_row].height = 22

    # ── AutoFilter (dropdown arrows on every column) ─────────────────
    end_col = get_column_letter(NUM_COLS)
    ws.auto_filter.ref = f'A{header_row}:{end_col}{last_data_row}'

    # Freeze panes just below the header row so it stays visible
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Column widths ────────────────────────────────────────────────
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Print settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.oddFooter.right.text = "Page &P of &N"
